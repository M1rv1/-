from tkinter import ttk
from tkinter import *
from sqlite3 import *
from os import environ, makedirs, path
import os
from ortools.sat.python import cp_model
from sqlite3 import *
import random
from openpyxl import Workbook
from itertools import product
from openpyxl.styles import Alignment, PatternFill, Font, Border, Side
from openpyxl.utils import get_column_letter
import threading

#<Глав меню> - 1
class main_menu:
    def __init__(self, root):
    #Данные корня
        self.root_menu_1 = root
        self.root_menu_1.style = ttk.Style()
        root.style.theme_use('clam')
        self.root_menu_1.title("Cоздание расписания Аничкого лицея")
        self.root_menu_1.geometry(f"{int(self.root_menu_1.winfo_screenwidth() * 0.64)}x{int(self.root_menu_1.winfo_screenheight()*0.7)}")
        self.root_menu_1.resizable(width=False, height=False) 
    ###

    #БД
        #Подключение баз данных
        database_folder = "DataBase"
        makedirs(database_folder, exist_ok=True)
        db_path = path.join(database_folder, "InputData.db")
        self.all_data_base = connect(f"{db_path}", check_same_thread=False)
        self.con_all_data_base = self.all_data_base.cursor()
        ###

    #Создание таблицы параллелей 

        #Таблица для параллели
        self.con_all_data_base.execute(""" CREATE TABLE IF NOT EXISTS parallels (id INTEGER PRIMARY KEY, 
                                        Letter TEXT, Number TEXT);""")
        self.all_data_base.commit()

        self.con_all_data_base.execute("""CREATE TABLE IF NOT EXISTS lessons (id INTEGER PRIMARY KEY, 
                                        Id_parallel TEXT, Subject TEXT, Hours TEXT, Id_teacher TEXT);""")
        self.all_data_base.commit()

        self.con_all_data_base.execute("""CREATE TABLE IF NOT EXISTS pe (id INTEGER PRIMARY KEY, 
                                        Id_parallel TEXT, Teacher TEXT, Day TEXT, Lesson TEXT);""")
        self.all_data_base.commit()
        # дополнительная таблица для внеурочек (только преподаватель)
        self.con_all_data_base.execute("""CREATE TABLE IF NOT EXISTS extra (id INTEGER PRIMARY KEY,
                                        Id_parallel TEXT, Teacher TEXT);""")
        self.all_data_base.commit()
        ###


        #Таблица для кабинета
        self.con_all_data_base.execute(""" CREATE TABLE IF NOT EXISTS rooms (id INTEGER PRIMARY KEY,
                                        Number TEXT, Graph_or_Lyceum TEXT, Big_or_Small TEXT, Subject TEXT);""")
        self.all_data_base.commit()
        self.con_all_data_base.execute(
            "SELECT * FROM rooms WHERE Number = ? AND Graph_or_Lyceum = ?",
            ("ОТ_1", "Лицей")
        )
        if not self.con_all_data_base.fetchone():
            self.con_all_data_base.execute(
                "INSERT INTO rooms (Number, Graph_or_Lyceum, Big_or_Small, Subject) VALUES (?, ?, ?, ?)",
                ("ОТ_1", "Лицей", "Маленький", "Инфа")
            )
            self.all_data_base.commit()

        #ОТ_2
        self.con_all_data_base.execute(
            "SELECT * FROM rooms WHERE Number = ? AND Graph_or_Lyceum = ?",
            ("ОТ_2", "Лицей")
        )
        if not self.con_all_data_base.fetchone():
            self.con_all_data_base.execute(
                "INSERT INTO rooms (Number, Graph_or_Lyceum, Big_or_Small, Subject) VALUES (?, ?, ?, ?)",
                ("ОТ_2", "Лицей", "Маленький", "Инфа")
            )
            self.all_data_base.commit()

        #УОО
        self.con_all_data_base.execute(
            "SELECT * FROM rooms WHERE Number = ? AND Graph_or_Lyceum = ?",
            ("УОО", "Лицей")
        )
        if not self.con_all_data_base.fetchone():
            self.con_all_data_base.execute(
                "INSERT INTO rooms (Number, Graph_or_Lyceum, Big_or_Small, Subject) VALUES (?, ?, ?, ?)",
                ("УОО", "Лицей", "Большой", "Физра")
            )
            self.all_data_base.commit()



        #СЗ
        self.con_all_data_base.execute(
            "SELECT * FROM rooms WHERE Number = ? AND Graph_or_Lyceum = ?",
            ("СЗ", "Графский")
        )
        if not self.con_all_data_base.fetchone():
            self.con_all_data_base.execute(
                "INSERT INTO rooms (Number, Graph_or_Lyceum, Big_or_Small, Subject) VALUES (?, ?, ?, ?)",
                ("СЗ", "Графский", "Большой", "Физра")
            )
            self.all_data_base.commit()
        
        #401
        self.con_all_data_base.execute(
            "SELECT * FROM rooms WHERE Number = ? AND Graph_or_Lyceum = ?",
            ("401", "Графский")
        )
        if not self.con_all_data_base.fetchone():
            self.con_all_data_base.execute(
                "INSERT INTO rooms (Number, Graph_or_Lyceum, Big_or_Small, Subject) VALUES (?, ?, ?, ?)",
                ("401", "Графский", "Большой", "Инфа")
            )
            self.all_data_base.commit()
        
        #407
        self.con_all_data_base.execute(
            "SELECT * FROM rooms WHERE Number = ? AND Graph_or_Lyceum = ?",
            ("407", "Графский")
        )
        if not self.con_all_data_base.fetchone():
            self.con_all_data_base.execute(
                "INSERT INTO rooms (Number, Graph_or_Lyceum, Big_or_Small, Subject) VALUES (?, ?, ?, ?)",
                ("407", "Графский", "Большой", "Инфа")
            )
            self.all_data_base.commit()


        #Таблицы для учителя
        self.con_all_data_base.execute(""" CREATE TABLE IF NOT EXISTS teachers (id INTEGER PRIMARY KEY,
                                        Surname TEXT, Name TEXT, Patrony TEXT, Trans TEXT);""")
        self.all_data_base.commit()

        self.con_all_data_base.execute(""" CREATE TABLE IF NOT EXISTS exceptions (id INTEGER PRIMARY KEY,
                                        Rel TEXT, Day TEXT, Lessons TEXT);""")
        self.all_data_base.commit()

        self.con_all_data_base.execute(""" CREATE TABLE IF NOT EXISTS prio (id INTEGER PRIMARY KEY,
                                        Rel TEXT, Rooms Text, Building Text);""")
        self.all_data_base.commit()
        ###

    ###

    #Проверка открытия окна
        self.menu_1 = None
        self.menu_2 = None
        self.menu_3 = None
        self.menu_4 = None
        self.menu_5 = None
        self.menu_6 = None
        self.menu_7 = None
    ###

        style_0 = ttk.Style()

        style_0.theme_use('clam')

        style_0.configure("Make_T.TButton", font=("Helvetica", 30, "bold"), background="#DCDCDC", foreground="orange",)

        style_0.configure("Make_Par.TButton", font=("Helvetica", 28, ), background="#DCDCDC", foreground="blue",)

        style_0.configure("Check_Par.TButton", font=("Helvetica", 28, ), background="#DCDCDC", foreground="blue",)

        style_0.configure("Make_Room.TButton", font=("Helvetica", 28, ), background="#DCDCDC", foreground="purple",)

        style_0.configure("Check_Room.TButton", font=("Helvetica", 28, ), background="#DCDCDC", foreground="purple",)

        style_0.configure("Make_Teacher.TButton", font=("Helvetica", 28, ), background="#DCDCDC", foreground="green",)

        style_0.configure("Check_Teacher.TButton", font=("Helvetica", 28, ), background="#DCDCDC", foreground="green",)

        self.root_menu_1.configure(bg="#e0dcd4")

#Виджеты - глав меню ++
    #Блок создания расписания
        self.creating_timetabe =  ttk.Button(
            self.root_menu_1,
            text = "Составление расписания",
            style="Make_T.TButton",
            command = self.open_making_settings
            )
        self.creating_timetabe.place(relx=0, rely=0, relwidth = 1, relheight=0.3)
    ###

    #Блок создания и проверки параллели
        creating_parallel = ttk.Button(
            self.root_menu_1,
            text = "Добавление класса",
            style="Make_Par.TButton",
            command = self.open_screen_add_parallel
            )
        creating_parallel.place(relx=0, rely = 0.3, relwidth = 1, relheight = 0.3)

        check_parallel = ttk.Button(
            self.root_menu_1,
            text = str(chr(int("2968", 16))),
            style="Check_Par.TButton",
            command = self.open_screen_check_parallel
            )
        check_parallel.place(relx = 0, rely = 0.6, relwidth = 1, relheight= 0.1)
    ###

    #Блок создания и проверки кабинетов
        creating_room = ttk.Button(
            self.root_menu_1,
            text = "Добавление кабинета",
            style="Make_Room.TButton",
            command = self.open_screen_add_room
            )
        creating_room.place(relx = 0, rely = 0.7, relwidth = 0.5, relheight = 0.2)

        check_room = ttk.Button(
            self.root_menu_1,
            text = str(chr(int("2968", 16))),
            style="Check_Room.TButton",
            command = self.open_screen_check_room
            )
        check_room.place(relx = 0, rely = 0.9, relwidth = 0.5, relheight = 0.1)
    ###

    #Блок создания и проверки учителей
        creating_teacher = ttk.Button(
            self.root_menu_1,
            text = "Добавление учителя",
            style="Make_Teacher.TButton",
            command = self.open_screen_add_teacher
            )
        creating_teacher.place(relx = 0.5, rely = 0.7, relwidth = 0.5, relheight=0.2)

        check_teacher = ttk.Button(
            self.root_menu_1,
            text = str(chr(int("2968", 16))),
            style = "Check_Teacher.TButton",
            command = self.open_screen_check_teacher
            )
        check_teacher.place(relx = 0.5, rely = 0.9, relwidth = 0.5, relheight = 0.1)
    ###

#Открытие окон
    #Открытие составления параллелей - 2
    def open_making_settings(self):
        if self.menu_1 is None or not self.menu_1.root_menu_1.winfo_exists(): #Проверка на открытость окна
            self.menu_1 = making_settings(self.root_menu_1, self.all_data_base, self.con_all_data_base, self)
            self.menu_1.root_menu_1.mainloop()
        else:
            self.menu_1.root_menu_1.focus()
    ###

    #Открытие составления параллелей - 2
    def open_screen_add_parallel(self):
        if self.menu_2 is None or not self.menu_2.root_menu_2.winfo_exists(): #Проверка на открытость окна
            self.menu_2 = menu_add_parallel(self.root_menu_1, self.all_data_base, self.con_all_data_base)
            self.menu_2.root_menu_2.mainloop()
        else:
            self.menu_2.root_menu_2.focus()
    ###

    #Открытие проверки параллелей - 3
    def open_screen_check_parallel(self):
        if self.menu_3 is None or not self.menu_3.root_menu_3.winfo_exists(): #Проверка на открытость окна
            self.menu_3 = menu_check_parallel(self.root_menu_1, self.all_data_base,  self.con_all_data_base)
            self.menu_3.root_menu_3.mainloop()
        else:
            self.menu_3.root_menu_3.focus()
    ###

    #Открытие создания кабинета - 4
    def open_screen_add_room(self):
        if self.menu_4 is None or not self.menu_4.root_menu_4.winfo_exists(): #Проверка на открытость окна
            self.menu_4 = menu_add_room(self.root_menu_1, self.all_data_base,  self.con_all_data_base)
            self.menu_4.root_menu_4.mainloop()
        else:
            self.menu_4.root_menu_4.focus()
    ###

    #Открытие проверки кабинета - 5
    def open_screen_check_room(self):
        if self.menu_5 is None or not self.menu_5.root_menu_5.winfo_exists(): #Проверка на открытость окна
            self.menu_5 = menu_check_room(self.root_menu_1, self.all_data_base, self.con_all_data_base)
            self.menu_5.root_menu_5.mainloop()
        else:
            self.menu_5.root_menu_5.focus()
    ###

    #Открытие создания учителя - 6
    def open_screen_add_teacher(self):  
        if self.menu_6 is None or not self.menu_6.root_menu_6.winfo_exists(): #Проверка на открытость окна
            self.menu_6 = menu_add_teacher(self.root_menu_1, self.all_data_base, self.con_all_data_base)
            self.menu_6.root_menu_6.mainloop()
        else:
            self.menu_6.root_menu_6.focus()
    ###

    #Открытие проверки учителей - 7
    def open_screen_check_teacher(self):    
        if self.menu_7 is None or not self.menu_7.root_menu_7.winfo_exists(): #Проверка на открытость окна
            self.menu_7 = menu_check_teacher(self.root_menu_1, self.all_data_base, self.con_all_data_base)
            self.menu_7.root_menu_7.mainloop()
        else:
            self.menu_7.root_menu_7.focus()
    ###

###






#<Настр перед сос> - 1 
class making_settings:
    def __init__(self, root_parent, base, con, parent_menu):
    #Данные корня
        self.base_menu_1 = base
        self.con_menu_1 = con
        self.root_parent_1 = root_parent
        self.root_menu_1 = Toplevel(self.root_parent_1)
        self.root_menu_1.title("Настройки составления")
        self.root_menu_1.geometry(f"{int(self.root_menu_1.winfo_screenwidth() * 0.4)}x{int(self.root_menu_1.winfo_screenheight()*0.44)}")
        self.root_menu_1.resizable(width=False, height=False) 
        self.parent_menu = parent_menu
    ###
        style_1 = ttk.Style()

        style_1.theme_use('clam')

        style_1.configure("TLabel", font=("Helvetica", 30, "italic"))

        style_1.configure("TEntry", fieldbackground="#DCDCDC")

        style_1.configure("TCombobox", fieldbackground="#DCDCDC", arrowsize = 0)
        style_1.map("TCombobox", fieldbackground=[("readonly", "#DCDCDC")])


        style_1.configure("Main_1.TButton", font=("Helvetica", 30, "bold"), background="#696969", foreground="white",)
        style_1.map("Main_1.TButton", background=[("active", "#D2691E")])

        self.root_menu_1.configure(bg="#e0dcd4")
    #Виджеты насторек перед сос ++ 
    #Перегородочки
        separator_1 = Frame(
            self.root_menu_1, 
            height=3,         
            bg='grey',         
            relief=FLAT,
            bd=1              
        )
        separator_1.place(relx=0, rely=0.245, relwidth=1, relheight=0.007) #Горизонтальная

    ###

    #Блок выбора пар классов
        self.options_1 = ["8", "9", "10", "11"]

        self.con_between_1 =ttk.Label(
            self.root_menu_1,
            text = "Корпуса:",
            style="TLabel",
            anchor="center"
            )
        self.con_between_1.place(relx=0, rely=0, relwidth=0.3, relheight=0.24)

        self.number_1 = ttk.Combobox(
            self.root_menu_1,
            values=self.options_1,
            style="TCombobox",
            font=("Helvetica", 27),
            state="readonly", 
            foreground="#4D4D4D"
        )
        self.number_1.place(relx=0.3, rely=0.04, relwidth=0.11, relheight=0.16)

        self.con_between_1 =ttk.Label(
            self.root_menu_1,
            text = "+",
            style="TLabel",
            anchor="center"
            )
        self.con_between_1.place(relx=0.41, rely=0, relwidth=0.08, relheight=0.24)

        self.number_2 = ttk.Combobox(
            self.root_menu_1,
            values=self.options_1,
            style="TCombobox",
            font=("Helvetica", 27),
            state="readonly", 
            foreground="#4D4D4D"

        )
        self.number_2.place(relx=0.49, rely=0.04, relwidth=0.11, relheight=0.16)


        self.con_between_2 =ttk.Label(
            self.root_menu_1,
            text = "И",
            style="TLabel",
            anchor="center"
            )
        self.con_between_2.place(relx=0.6, rely=0, relwidth=0.1, relheight=0.24)


        self.number_3 = ttk.Combobox(
            self.root_menu_1,
            values=self.options_1,
            style="TCombobox",
            font=("Helvetica", 27),
            state="readonly", 
            foreground="#4D4D4D"
        )
        self.number_3.place(relx=0.7, rely=0.04, relwidth=0.11, relheight=0.16)

        self.con_between_3 = ttk.Label(
            self.root_menu_1,
            text = "+",
            style="TLabel",
            anchor="center"
            )
        self.con_between_3.place(relx=0.81, rely=0, relwidth=0.08, relheight=0.24)

        self.number_4 = ttk.Combobox(
            self.root_menu_1,
            values=self.options_1,
            style="TCombobox",
            font=("Helvetica", 27),
            state="readonly", 
            foreground="#4D4D4D"
        )
        self.number_4.place(relx=0.89, rely=0.04, relwidth=0.11, relheight=0.16)
    ###

    #Блок выбора количесва перебежек
        self.num_dash = ttk.Label(
            self.root_menu_1,
            text="Время работы программы:",
            style="TLabel",
            anchor="center"
        )
        self.num_dash.place(relx=0, rely=0.25, relwidth=1, relheight=0.25)

        self.multi = ttk.Entry(
            self.root_menu_1,
            style="TEntry",
            font=("Helvetica", 27),
            foreground="#4D4D4D"
        )
        self.multi.place(relx=0.2, rely=0.44, relwidth= 0.6, relheight=0.2)
    ###

    ###

    #Кнопка начала составления расписаня
        self.begin_makin = ttk.Button(
            self.root_menu_1,
            text="Начать создание",
            style="Main_1.TButton",
            command = self.start_making
        )
        self.begin_makin.place(relx=0, rely=0.75, relwidth=1, relheight=0.25)
    ###

    #Функционал
    def start_making(self):
    #Входные данные
        input_class_1 = self.number_1.get()
        input_class_2 = self.number_2.get()
        input_class_3 = self.number_3.get()
        input_class_4 = self.number_4.get()

        time = self.multi.get()
    ###

    #Проверки перед дальнейшей программой
        #Проверки корпусов
        if input_class_1 not in self.options_1 or input_class_2 not in self.options_1 or input_class_3 not in self.options_1 or input_class_4 not in self.options_1:
            warning = eror_popup(self.root_menu_1, "Ошибка в корпусах")
            warning.root.mainloop()
            return
        if input_class_1 in (input_class_2, input_class_3, input_class_4) or input_class_2 in (input_class_3, input_class_4) or input_class_4 == input_class_3:
            warning = eror_popup(self.root_menu_1, "Ошибка в корпусах")
            warning.root.mainloop()
            return
        ###

        all_numbers = "0987654321"
        #Проверка множителя
        for i in time:
            if i not in all_numbers:
                warning = eror_popup(self.root_menu_1, "Ошибка в времени работы программы")
                warning.root.mainloop()
                return  
        ###
        if int(time) < 50:
            warning = eror_popup(self.root_menu_1, "Слишком мало времени")
            warning.root.mainloop()
            return  
    

        
            
        data_settings = ((input_class_1, input_class_2), (input_class_3, input_class_4), time)

        creating_timetable(self.con_menu_1, self.base_menu_1, data_settings, self.root_parent_1, self.parent_menu, self.root_menu_1, self.parent_menu)
        ###
    ###






#СОЗДАНИЕ РАСПИСАНИЯ
class creating_timetable:
    def __init__(self, con, base, settings, main_menu_root, main_menu, parent_root, main_menu_self):
        self.root = Toplevel(main_menu_root)
        self.root.title("Создание расписания")
        self.root.geometry(f"{int(self.root.winfo_screenwidth() * 0.32)}x{int(self.root.winfo_screenheight()*0.4)}")
        self.root.resizable(width=False, height=False)
        self.root.grab_set()
        self.root.focus_set()
        self.root.transient(main_menu_root)
        self.con = con
        self.base = base
        self.settingss = settings

        self.main_menu_root = main_menu_root
        self.main_menu_itself = main_menu
        parent_root.destroy()
        self.number_of_timetables_created = 0
        self.self_parent_menu = main_menu_self

        style_1_0 = ttk.Style()

        style_1_0.theme_use('clam')

        style_1_0.configure("1.TLabel", font=("Helvetica", 30, "italic"))

        style_1_0.configure("2.TLabel", font=("Helvetica", 60))

        style_1_0.configure("3.TLabel", font=("Helvetica", 30, "italic"))

        style_1_0.configure("TEntry", fieldbackground="#DCDCDC")

        style_1_0.configure("TCombobox", fieldbackground="#DCDCDC", arrowsize = 0)
        style_1_0.map("TCombobox", fieldbackground=[("readonly", "#DCDCDC")])


        style_1_0.configure("Main_1_0.TButton", font=("Helvetica", 30, "bold"), background="#696969", foreground="white",)
        style_1_0.map("Main_1_0.TButton", background=[("active", "#D2691E")])

        self.root.configure(bg="#e0dcd4")
    #Виджеты создания
        
        # Получаем множитель времени в начале
        self.time = int(self.settingss[2])
        
        self.in_process = ttk.Label(
            self.root,
            text = "Расписание создается!",
            style="1.TLabel",
            anchor="center"
            )
        self.in_process.place(relx=0, rely=0.1, relwidth=1, relheight=0.15)

    #Прогресс-бар
        style_1_0.configure("Progress.Horizontal.TProgressbar", 
                           length=300, 
                           mode='determinate',
                           background="#D2691E")
        
        self.progress_var = DoubleVar()
        self.progress_bar = ttk.Progressbar(
            self.root,
            variable=self.progress_var,
            maximum=100,
            length=300,
            mode='determinate',
            style="Progress.Horizontal.TProgressbar"
        )
        self.progress_bar.place(relx=0.1, rely=0.28, relwidth=0.8, relheight=0.1)
        
        # Текст процента
        self.progress_text = ttk.Label(
            self.root,
            text="0%",
            style="3.TLabel",
            anchor="center"
        )
        self.progress_text.place(relx=0.1, rely=0.38, relwidth=0.8, relheight=0.1)
        
        # Инициализация прогресса
        self.progress_step = 0
        self.max_progress_steps = max(self.time * 10, 100)  # количество шагов для прогресса
        
    ###

    
    #Кнопка отмены 
        self.add_class = ttk.Button(
            self.root,
            text="Отмена",
            style="Main_1_0.TButton",
            command = self.root.destroy
        )
        self.add_class.place(relx=0, rely=0.75, relwidth=1, relheight=0.25)
    ###
 
    #Данные из окна
        settings = (self.settingss[0], self.settingss[1]) # кнопочка в меню

        self.group_1 = (int(settings[0][0]), int(settings[0][1]))
        self.group_2 = (int(settings[1][0]), int(settings[1][1]))
    #
        
        # Запуск прогресс-бара с учётом выбранного времени
        self.start_progress_bar()
        # Запуск создания расписания в отдельном потоке
        timetable_thread = threading.Thread(target=self.make_timetable, daemon=True)
        timetable_thread.start()
    
    def start_progress_bar(self):
        """Запускает автоматическое заполнение прогресс-бара в зависимости от времени (time в секундах)"""
        self.progress_step = 0
        self.max_progress_steps = self.time * 10  # 10 шагов в секунду (интервал 100 мс)
        self.auto_update_enabled = True  # Флаг для автоматического обновления
        self.root.after(100, self.update_progress_bar)
    
    def update_progress_bar(self):
        """Обновляет прогресс-бар и планирует следующее обновление"""
        # Если автоматическое обновление отключено, не обновляем
        if not self.auto_update_enabled:
            return
            
        if self.progress_step < self.max_progress_steps:
            progress_percent = (self.progress_step / self.max_progress_steps) * 100
            self.progress_var.set(progress_percent)
            self.progress_text.config(text=f"{int(progress_percent)}%")
            
            self.progress_step += 1
            self.root.after(100, self.update_progress_bar)
        else:
            self.progress_var.set(100)
            self.progress_text.config(text="100%")
    
    def update_progress_manual(self, percent):
        """Ручное обновление прогресс-бара из алгоритма создания расписания (0-100)"""
        self.auto_update_enabled = False  # Отключаем автоматическое обновление
        self.progress_var.set(percent)
        self.progress_text.config(text=f"{int(percent)}%")
        self.root.update()  # Обновляем UI немедленно
    
    def _show_success_popup(self):
        """Безопасно показать popup успеха в главном потоке"""
        warning = popup(self.root, "Откройте входные данные", "Успех")
        warning.root.mainloop()
    
    def _show_error_popup(self):
        """Безопасно показать popup ошибки в главном потоке"""
        warning = eror_popup(self.root, "Конфликтующие вход. данные")
        warning.root.mainloop()
    ###
    def make_timetable(self):
    #Входные данные
        #Кабинеты
        rooms = [] #####
   
        # id, Number TEXT, Graph_or_Lyceum TEXT, Big_or_Small TEXT, Subject TEXT
        self.con.execute("""SELECT * FROM rooms""")
        values = self.con.fetchall()

        for r in values:
            if r[2] == "Графский":
                b = 2
            else:
                b = 1
            if r[4] == "Нет приоритетных кабинетов":
                r_ = "Нет прио"
            else:
                r_ = r[4]
            rooms.append(
                {
                    "name": r[1],
                    "building": b,
                    "prio": r_,
                    "size": r[3]
                }
            )
        ##

        #Физра в слотах
        ful_day_to_short = {
            "Понедельник": "Пн",
            "Вторник": "Вт",
            "Среда": "Ср",
            "Четверг":"Чт",
            "Пятница":"Пт",
            "Суббота":"Сб",
        }


        busy_phy = {} #######
        self.con.execute("""SELECT * FROM parallels""")
        values_p = self.con.fetchall()
        
        id_to_clas = {}
        for c in values_p:
            id_to_clas[str(c[0])] = f"{c[2]}{c[1]}"

        for c in values_p:
            busy_phy[f"{c[2]}{c[1]}"] = {
                "Пн": [],
                "Вт": [],
                "Ср": [], 
                "Чт": [],
                "Пт": [],
                "Сб": []
                }
        self.con.execute("""SELECT * FROM pe""")
        values = self.con.fetchall()    
        for slot_pe in values:
            busy_phy[id_to_clas[slot_pe[1]]][ful_day_to_short[slot_pe[3]]].append(int(slot_pe[4]))
        ###
        # print(busy_phy)
        #Доп уроки
        extra_les = {}
        for c in values_p:
            extra_les[f"{c[2]}{c[1]}"] = []

        
        self.con.execute("""SELECT * FROM extra""")
        values_e = self.con.fetchall()    
        for ext in values_e:
            extra_les[id_to_clas[str(ext[1])]].append(ext[2])
        # print(extra_les)   
        ###

        #Учителя
        teachers = export_teachers_to_schedule_format(self.base)
        
        # print(teachers)
        #Классы
        classes = export_classes_to_schedule_format(self.base, (self.group_1, self.group_2))
        # print(classes)
        print(classes)

        buildings = ["1", "2"]
#Лицей, Графский

        days = ["Пн", "Вт", "Ср", "Чт", "Пт", "Сб"]

        periods_per_day = 8

        # --------------------------------------------
        # 2. Создаем модель
        # --------------------------------------------
        model = cp_model.CpModel()

        # x[учитель, класс, корпус, день, урок]


        x = {}
        for c in classes:
            for s in c['subjects']:
                for d in days:
                    for p in range(1, periods_per_day + 1):
                        x[(s["teacher"], c["name"], c["buildings"][d], d, p)] = model.NewBoolVar(f"{s["teacher"]}_{c['name']}_{c["buildings"][d]}_{d}_{p}")

        # --------------------------------------------
        # 3. Ограничения
        # --------------------------------------------

        # 3.1 Учебный план (все часы должны быть выполнены)
        for c in classes:
            for s in c["subjects"]:
                model.Add(
                    sum(
                        x[(s["teacher"], c["name"], c['buildings'][d], d, p)]
                        for d in days
                        for p in range(1, periods_per_day + 1)
                    ) == s['hours']
                )

        for c in classes:
            for d in days:
                for p in range(1, periods_per_day + 1):
                    model.Add(
                        sum(
                            x[(s['teacher'], c["name"], c['buildings'][d], d, p)]
                            for s in c['subjects']
                        ) <= 1
                    )
        ##


        # 3.3 Учитель не может вести два урока одновременно
        # Сначала соберём, кто в каких классах преподаёт
        teacher_to_classes = {}
        for c in classes:
            for s in c["subjects"]:
                if s['teacher'][0] != "-":
                    teacher_to_classes.setdefault(s["teacher"], []).append(c)

        # Теперь для каждого учителя и каждого слота (день, урок) не более одного урока
        for tname, class_list in teacher_to_classes.items():
            for d in days:
                for p in range(1, periods_per_day + 1):
                
                    vars_ = [x[(tname, c["name"], c["buildings"][d], d, p)] for c in class_list]

                    if (tname == ('Пятибратова К.В.',) and d == "Пт" and p == 1) or (tname == ('Пятибратова К.В.',) and d == "Сб" and p == 8):
                        model.Add(sum(vars_) <= 2)   # разрешили 2
                    else:
                        model.Add(sum(vars_) <= 1)   # всем остальным 1


        ###


        # 3.4 Учитель не может вести при его исключении
        for c in classes:
            for s in c['subjects']:
                for t1 in s['teacher']:
                    for t2 in teachers:
                        if t2['name'] == t1:
                            for d, periods in t2['exce'].items():
                                for p in periods:
                                    model.Add(x[(s['teacher'], c['name'], c["buildings"][d], d, p)] == 0)
        #

        # # # 3.5 Не может быть окон у классов
        lesson = {}
        for c in classes:
            cname = c["name"]
            for d in days:
                # 1) Индикация «есть ли урок в слоте p» для класса cname в день d
                for p in range(1, periods_per_day + 1):
                    lesson[(cname, d, p)] = model.NewBoolVar(f"lesson_{cname}_{d}_{p}")
                    model.Add(
                        sum(
                            x[(s["teacher"], cname, c["buildings"][d], d, p)]
                            for s in c["subjects"]
                        )
                        == lesson[(cname, d, p)]
                    )
                # 2) Для любой пары слотов i<j с хотя бы одним промежуточным k
                #    требуем, чтобы если в i и в j уроки, 
                #    то во всех k между ними тоже был урок
                for i in range(1, periods_per_day + 1):
                    for j in range(i + 2, periods_per_day + 1):
                        for k in range(i + 1, j):
                            # если lesson[i]=lesson[j]=1, то LHS=1 => lesson[k] ≥1
                            model.Add(
                                lesson[(cname, d, i)]
                              + lesson[(cname, d, j)]
                              - 1
                              <= lesson[(cname, d, k)]
                            )

        # 3.6 Не может быть больше 2 предметов одной группы
        for c in classes:
            dic_lesson = {}
            for s in c['subjects']:
                # s["teacher"] всегда кортеж (см. структуру данных)
                if s['name'].endswith("В"):
                    name = s['name'][:-1]
                else:
                    name = s['name']
                dic_lesson.setdefault(name, []).append(s["teacher"])  # просто s["teacher"], без ()

            for d in days:
                for sub, list_t in dic_lesson.items():
                    model.Add(
                        sum(
                            x[(t, c['name'], c["buildings"][d], d, p)]
                            for p in range(1, periods_per_day + 1)
                            for t in list_t
                        ) <= 2
                    )

        #3.7 Как миниум 6 уроков в день
        # for c in classes:
        #     for d in days:
        #         model.Add(
        #             sum(
        #                 x[(s['teacher'], c["name"], c["buildings"][d], d, p)]
        #                 for p in range(1, periods_per_day + 1)
        #                 for s in c['subjects']
        #             ) >= 5
        #         )


        # 3.8 Нет первых двух уроков в пон
        for c in classes:
            for s in c["subjects"]:
                model.Add(
                    sum(
                        x[(s["teacher"], c["name"], c['buildings']["Пн"], "Пн", p)]
                        for p in range(1, 3)
                    ) == 0
                )

        # 3.10 Ограничение на жесткие пары у англ

        pair_engl = {}
        for c in classes:
            cname = c["name"]
            
            # Находим учителя англ для этого класса
            teacher = None
            for s in c["subjects"]:
                if s["name"] == "Англ":
                    teacher = s['teacher']
                    break
            
            # Если англ не найден, пропускаем класс
            if teacher is None:
                continue
            
            for d in days:
                for p in range(2, periods_per_day + 1):
                    pair_engl[(cname, d, p, p - 1)] = model.NewBoolVar(f"clas_{cname}_day_{d}_p2_{p}_p1_{p - 1}")
                    model.Add(
                        sum(
                            x[(teacher, cname, c["buildings"][d], d, k)]
                            for k in range(p - 1, p + 1)
                        ) == 2
                    ).OnlyEnforceIf(pair_engl[(cname, d, p, p - 1)])
                    model.Add(
                        sum(
                            x[(teacher, cname, c["buildings"][d], d, r)]
                            for r in range(p - 1, p + 1)
                        ) < 2
                    ).OnlyEnforceIf(pair_engl[(cname, d, p, p - 1)].Not())
            model.Add(
                sum(
                    pair_engl[(cname, d, p, p - 1)]
                    for d in days
                    for p in range(2, periods_per_day + 1)
                ) == 1
            )


        # 3.11 Физра в конкретных слотах

        for c in classes:
            # Находим предмет "Физра" для этого класса
            phy_subject = None
            for s in c['subjects']:
                if s['name'] == "Физра":
                    phy_subject = s
                    break
                
            if phy_subject is None:
                continue  # Пропускаем классы без физры
            
            teacher_phy = phy_subject['teacher']
            num_phy = phy_subject['hours']

            # Проверяем, есть ли класс в busy_phy
            if c['name'] not in busy_phy:
                continue
            
            # Создаем список всех слотов для физры
            all_phy_slots = []
            for d in busy_phy[c['name']]:
                for p in busy_phy[c['name']][d]:
                    all_phy_slots.append((d, p))

            # Добавляем ограничение: физра должна быть только в указанные слоты
            # и количество уроков должно соответствовать num_phy
            model.Add(
                sum(
                    x[(teacher_phy, c["name"], c['buildings'][d], d, p)]
                    for d, p in all_phy_slots
                ) == num_phy
            )

            # Запрещаем физру в другие слоты
            for d in days:
                for p in range(1, periods_per_day + 1):
                    if (d, p) not in all_phy_slots:
                        model.Add(
                            x[(teacher_phy, c["name"], c['buildings'][d], d, p)] == 0
                        )
        #

        # 3.12 в пон уроки с 3
        for c in classes:
            model.Add(
                sum(
                    x[(s["teacher"], c["name"], c['buildings']["Пн"], "Пн", 3)]
                    for s in c["subjects"]
                ) == 1
            )

        # 3.13 макс 1 урок инфы в корпусе 
        for b in buildings:  
            for d in days:  
                for p in range(1, periods_per_day + 1):  
                    infa_vars = []

                    for c in classes:
                        # берём только те классы, которые в этот день реально в этом корпусе
                        if c["buildings"][d] != b:
                            continue
                        
                        for s in c["subjects"]:
                            if "Инфа" in s["name"]:
                                infa_vars.append(x[(s["teacher"], c["name"], b, d, p)])

                    if infa_vars:
                        model.Add(sum(infa_vars) <= 1)

        # 3.14 Не должно быть разрывов в парах

        #Инициализация хэштаблицы самой важной здесь
        pair_or_not_pair = {}

        for c in classes:
            cname = c["name"]
            for s in c["subjects"]:
                name_t = s['teacher']
                for d in days:
                    for p in range(1, periods_per_day - 1):
                        pair_or_not_pair[(cname, d, name_t, p, p + 1)] = model.NewBoolVar(f"clas_{cname}_day_{d}_name_l{name_t}_p1_{p}_p2_{p + 1}")

        for c in classes:
            cname = c["name"]
            for s in c["subjects"]:
                t = s["teacher"]
                for d in days:
                    b = c["buildings"][d]
                    for p in range(1, periods_per_day - 1):
                        model.Add(
                            sum(
                                 x[(t, cname, b, d, z)]
                                 for z in range(p, p + 2)
                            ) == 2
                        ).OnlyEnforceIf(pair_or_not_pair[(cname, d, t, p, p + 1)])
                        model.Add(
                            sum(
                                 x[(t, cname, b, d, z)]
                                 for z in range(p, p + 2)
                            ) != 2
                        ).OnlyEnforceIf(pair_or_not_pair[(cname, d, t, p, p + 1)].Not())
        ###

        #Хештаблица ровно 1 урок в день у учителя
        have_a_les = {}
        for c in classes:
            cname = c["name"]
            for s in c["subjects"]:
                name_t = s['teacher']
                for d in days:
                        have_a_les[(cname, d, name_t)] = model.NewBoolVar(f"clas_{cname}_day_{d}_name_l{name_t}")

        for c in classes:
            cname = c["name"]
            for s in c["subjects"]:
                t = s["teacher"]
                for d in days:
                    b = c["buildings"][d]
                    model.Add(
                        sum(
                             x[(t, cname, b, d, p)]
                             for p in range(1, periods_per_day + 1)
                        ) == 1
                    ).OnlyEnforceIf(have_a_les[(cname, d, t)])
                    model.Add(
                        sum(
                             x[(t, cname, b, d, p)]
                             for p in range(1, periods_per_day + 1)
                        ) != 1
                    ).OnlyEnforceIf(have_a_les[(cname, d, t)].Not())
        ###

        #Хештаблица 0 уроков в день
        not_have_a_les = {}
        for c in classes:
            cname = c["name"]
            for s in c["subjects"]:
                name_t = s['teacher']
                for d in days:
                        not_have_a_les[(cname, d, name_t)] = model.NewBoolVar(f"clas_{cname}_day_{d}_name_l{name_t}")

        for c in classes:
            cname = c["name"]
            for s in c["subjects"]:
                t = s["teacher"]
                for d in days:
                    b = c["buildings"][d]
                    model.Add(
                        sum(
                             x[(t, cname, b, d, p)]
                             for p in range(1, periods_per_day + 1)
                        ) == 0
                    ).OnlyEnforceIf(not_have_a_les[(cname, d, t)])
                    model.Add(
                        sum(
                             x[(t, cname, b, d, p)]
                             for p in range(1, periods_per_day + 1)
                        ) != 0
                    ).OnlyEnforceIf(not_have_a_les[(cname, d, t)].Not())
        ###


        for c in classes:
            cname = c["name"]
            for s in c["subjects"]:
                for d in days:
                    model.Add(sum
                            (
                    pair_or_not_pair[(cname, d, s["teacher"], p, p + 1)]
                    for p in range(1, periods_per_day - 1)
                        ) == 1
                    ).OnlyEnforceIf([have_a_les[(cname, d, s["teacher"])].Not(), not_have_a_les[(cname, d, s["teacher"])].Not()])

        # 3.15 Если это доп урок, то можно иметь 8 уроков

        num_of_lessons = {}
        for c in classes:
            cname = c["name"]
            for d in days:
                num_of_lessons[(cname, d)] = model.NewIntVar(0, periods_per_day, f"lesson_{cname}_{d}")
                model.Add(
                    sum(
                        x[(s["teacher"], cname, c["buildings"][d], d, p)]
                        for s in c["subjects"]
                        for p in range(1, periods_per_day + 1)
                    ) == num_of_lessons[(cname, d)])

        num_of_teacher = {}
        for c in classes:
            cname = c["name"]
            for d in days:
                for s in c["subjects"]:
                    num_of_teacher[(cname, d, s["teacher"][0])] = model.NewBoolVar(f"lesson_{cname}_{d}_{s["teacher"][0]}")
                    model.Add(
                        sum(
                            x[(s["teacher"], cname, c["buildings"][d], d, p)]
                            for p in range(1, periods_per_day + 1)
                        ) >= 1).OnlyEnforceIf(num_of_teacher[(cname, d, s["teacher"][0])])
                    model.Add(
                        sum(
                            x[(s["teacher"], cname, c["buildings"][d], d, p)]
                            for p in range(1, periods_per_day + 1)
                        ) == 0).OnlyEnforceIf(num_of_teacher[(cname, d, s["teacher"][0])].Not())

        eight_or_not = {}
        state = {}#!!!!!!!!!!!!
        add = {}#!!!!!!!!!!!!
        for c in classes:
            cname = c["name"]
            if cname not in extra_les:
                continue
            
            
            ###Инициализация
            stage = len(extra_les[cname])
            for i_t in range(len(extra_les[cname])):
                teacher = extra_les[cname][i_t]

                add[(i_t, cname)] = model.NewBoolVar(f"add_{i_t}")
                state[(i_t, cname)] = {
                d: model.NewBoolVar(f"state_{cname}_{teacher}_{d}")
                for d in days
                }
                for d in days:
                    model.Add(state[(i_t, cname)][d] <= num_of_teacher[(cname, d, teacher)])

            #Монотонность
            for i_t in range(1, len(extra_les[cname])):
                for d in days:
                    model.Add(state[(i_t, cname)][d] >= state[(i_t - 1, cname)][d])

            ###Преобразования
            for i_t in range(1, len(extra_les[cname])):
                model.Add(
                    sum(state[(i_t, cname)][d] for d in days)
                    - sum(state[(i_t - 1, cname)][d] for d in days)
                    == add[(i_t, cname)]
                )


            for d in days:
                eight_or_not[(d, cname)] = model.NewBoolVar(f"class_{cname}_day_{d}")
                model.Add(sum(
                            state[(i_t, cname)][d]
                            for i_t in range(len(extra_les[cname]))
                        )>= 1).OnlyEnforceIf(eight_or_not[(d, cname)])
                model.Add(sum(
                            state[(i_t, cname)][d]
                            for i_t in range(len(extra_les[cname]))
                        )==0).OnlyEnforceIf(eight_or_not[(d, cname)].Not())

            for i_t in range(1, len(extra_les[cname])):
                for d in days:
                    model.Add(num_of_lessons[(cname, d)] <= 7).OnlyEnforceIf(eight_or_not[(d, cname)].Not())
                    model.Add(num_of_lessons[(cname, d)] <= 8).OnlyEnforceIf(eight_or_not[(d, cname)])

        #3.16 Икспб в конкретных слотах
        # model.Add(x[(('Пятибратова К.В.',), "8А", "1", "Чт", 8)] == 1)
        # model.Add(x[(('Пятибратова К.В.',), "8Б", "1", "Сб", 8)] == 1)
        # model.Add(x[(('Пятибратова К.В.',), "8В", "1", "Сб", 8)] == 1)

        # model.Add(x[(('Пятибратова К.В.',), "9А", "1", "Пт", 1)] == 1)
        # model.Add(x[(('Пятибратова К.В.',), "9Б", "1", "Пт", 1)] == 1)
        # model.Add(x[(('Пятибратова К.В.',), "9В", "2", "Чт", 1)] == 1)

        # --------------------------------------------
        # 4.0 Вспомогалки для оптимизации
        # --------------------------------------------

        # 1) Разворачиваем назначение: 
        #    teacher_to_assignments[teacher_name] = list of (teacher_key, class_obj)
        teacher_to_assignments = {}
        for c in classes:
            for s in c["subjects"]: # может быть строка или кортеж
                for t in s["teacher"]:
                    teacher_to_assignments.setdefault(t, []).append((s["teacher"], c))

        # 2) Индикатор “есть ли урок у teacher_name в день d, период p”
        lesson_t = {}
        for teacher_name, assigns in teacher_to_assignments.items():
            for d in days:
                for p in range(1, periods_per_day + 1):
                    lt = model.NewBoolVar(f"lesson_{teacher_name}_{d}_{p}")
                    lesson_t[(teacher_name, d, p)] = lt

                    s = sum(
                        x[(tkey, cl["name"], cl["buildings"][d], d, p)]
                        for (tkey, cl) in assigns
                    )
                    model.Add(s >= 1).OnlyEnforceIf(lt)
                    model.Add(s == 0).OnlyEnforceIf(lt.Not())


        # 3.17 Жёсткое ограничение: для каждого преподавателя с trans_1 == 1
        # запрещается менять корпус сразу подряд (p → p+1) — т.е. переход без окна
        for t in teachers:
            if t.get("trans_1", 0) != 1:
                continue
            teacher_name = t["name"]
            assigns = teacher_to_assignments.get(teacher_name, [])
            if not assigns:
                continue
            for d in days:
                for p in range(1, periods_per_day):
                    for (tkey1, c1), (tkey2, c2) in product(assigns, assigns):
                        b1 = c1["buildings"][d]
                        b2 = c2["buildings"][d]
                        if b1 == b2:
                            continue  # не смена корпуса
                        x1 = x[(tkey1, c1["name"], b1, d, p)]
                        x2 = x[(tkey2, c2["name"], b2, d, p + 1)]
                        # запретить одновременно урок в корпусе b1 на p и в другом корпусе b2 на p+1
                        model.AddBoolOr([x1.Not(), x2.Not()])




        # --------------------------------------------
        # 4.1 Оптимизация переходов
        # --------------------------------------------

        # Штрафные переменные transitions[(teacher_name, d, p, q)]
        transitions = {}
        for teacher_name, assigns in teacher_to_assignments.items():
            plus = 1
            for t in teachers:
                if t['name'] == teacher_name:
                    if t.get('trans_1', 0) == 1:
                        plus = 2  # требует хотя бы одного пустого периода между p и q
                    break
                
            for d in days:
                # для каждой пары уроков p<q, где между ними нет других уроков
                for p in range(1, periods_per_day + 1):
                    for q in range(p + plus, periods_per_day + 1):
                        gap_clear = model.NewBoolVar(f"gap_clear_{teacher_name}_{d}_{p}_{q}")
                        model.Add(
                            sum(lesson_t[(teacher_name, d, k)] for k in range(p+1, q)) == 0
                        ).OnlyEnforceIf(gap_clear)
                        model.Add(
                            sum(lesson_t[(teacher_name, d, k)] for k in range(p+1, q)) > 0
                        ).OnlyEnforceIf(gap_clear.Not())

                        # b) is_next = 1 ⇔ урок в p, урок в q и gap_clear
                        is_next = model.NewBoolVar(f"is_next_{teacher_name}_{d}_{p}_{q}")
                        model.AddBoolAnd([
                            lesson_t[(teacher_name, d, p)],
                            lesson_t[(teacher_name, d, q)],
                            gap_clear
                        ]).OnlyEnforceIf(is_next)
                        model.AddBoolOr([
                            lesson_t[(teacher_name, d, p)].Not(),
                            lesson_t[(teacher_name, d, q)].Not(),
                            gap_clear.Not()
                        ]).OnlyEnforceIf(is_next.Not())

                        # c) переход между корпусами
                        tr = model.NewBoolVar(f"trans_{teacher_name}_{d}_{p}_{q}")
                        transitions[(teacher_name, d, p, q)] = tr

                        # d) проверяем смену корпуса через все сочетания назначений
                        ors = []
                        for (tkey1, c1), (tkey2, c2) in product(assigns, assigns):
                            b1 = c1["buildings"][d]
                            b2 = c2["buildings"][d]
                            if b1 == b2:
                                continue
                            x1 = x[(tkey1, c1["name"], b1, d, p)]
                            x2 = x[(tkey2, c2["name"], b2, d, q)]
                            tmp = model.NewBoolVar(f"tmp_tr_{teacher_name}_{d}_{p}_{q}_{b1}_{b2}")
                            model.AddBoolAnd([is_next, x1, x2]).OnlyEnforceIf(tmp)
                            model.AddBoolOr([is_next.Not(), x1.Not(), x2.Not()]).OnlyEnforceIf(tmp.Not())
                            ors.append(tmp)

                        if ors:
                            model.AddBoolOr(ors).OnlyEnforceIf(tr)
                            model.AddBoolAnd([lit.Not() for lit in ors]).OnlyEnforceIf(tr.Not())
        # --------------------------------------------
        # 4.2 Оптимизация «окон»
        # --------------------------------------------

        windows = {}

        for teacher_name, assigns in teacher_to_assignments.items():
            for d in days:
                # для каждой пары уроков p<q, где между ними нет других уроков
                for p in range(1, periods_per_day + 1):
                    for q in range(p + 1, periods_per_day + 1):
                        if q != p + 1:
                            # a) gap_clear = 1 ⇔ нет уроков между p и q
                            gap_clear = model.NewBoolVar(f"gap_clear_{teacher_name}_{d}_{p}_{q}")
                            model.Add(
                                sum(lesson_t[(teacher_name, d, k)] for k in range(p+1, q)) == 0
                            ).OnlyEnforceIf(gap_clear)
                            model.Add(
                                sum(lesson_t[(teacher_name, d, k)] for k in range(p+1, q)) > 0
                            ).OnlyEnforceIf(gap_clear.Not())

                            # b) is_next = 1 ⇔ урок в p, урок в q, и gap_clear
                            is_next = model.NewBoolVar(f"is_next_{teacher_name}_{d}_{p}_{q}")
                            model.AddBoolAnd([
                                lesson_t[(teacher_name, d, p)],
                                lesson_t[(teacher_name, d, q)],
                                gap_clear
                            ]).OnlyEnforceIf(is_next)
                            model.AddBoolOr([
                                lesson_t[(teacher_name, d, p)].Not(),
                                lesson_t[(teacher_name, d, q)].Not(),
                                gap_clear.Not()
                            ]).OnlyEnforceIf(is_next.Not())

                            # c) окнонная переменная
                            wd = model.NewBoolVar(f"trans_{teacher_name}_{d}_{p}_{q}")
                            windows[(teacher_name, d, p, q)] = wd

                            # d) для каждого сочетания назначений проверяем смену корпуса
                            ors = []
                            for (tkey1, c1), (tkey2, c2) in product(assigns, assigns):
                                b1 = c1["buildings"][d]
                                b2 = c2["buildings"][d]
                                x1 = x[(tkey1, c1["name"], b1, d, p)]
                                x2 = x[(tkey2, c2["name"], b2, d, q)]
                                if b1 != b2 and q == p + 2: #Не считать оконо если 1 урок под переход
                                    continue
                                tmp = model.NewBoolVar(f"tmp_tr_{teacher_name}_{d}_{p}_{q}_{b1}_{b2}")
                                model.AddBoolAnd([is_next, x1, x2]).OnlyEnforceIf(tmp)
                                model.AddBoolOr([is_next.Not(), x1.Not(), x2.Not()]).OnlyEnforceIf(tmp.Not())
                                ors.append(tmp)

                            # e) устанавливаем tr ↔ OR(ors)
                            if ors:
                                model.AddBoolOr(ors).OnlyEnforceIf(wd)
                                model.AddBoolAnd([lit.Not() for lit in ors]).OnlyEnforceIf(wd.Not())

        # --------------------------------------------
        # 4.3 Оптимизация пар уроков
        # --------------------------------------------
        pairs = {}

        teachers_to_assignments = {}
        for c in classes:
            for s in c["subjects"]: 
                if s['hours'] != 1:
                    teachers_to_assignments.setdefault(s['teacher'], []).append(c)

        # 2.2) Индикатор “есть ли урок у teacher_name в день d, период p”
        for teacher_name, class_list in teachers_to_assignments.items():
            for d in days:
                for p in range(1, periods_per_day):
                    pair_var = model.NewBoolVar(f"pair_{teacher_name}_{d}_{p}_{p+1}")
                    pairs[(teacher_name, d, p)] = pair_var

                    # Пусть l_p = lesson_t[(teacher_name,d,p)], l_q = lesson_t[(teacher_name,d,p+1)]
                    l_p = sum(x[(teacher_name, cl["name"], cl["buildings"][d], d, p)]
                              for cl in class_list)
                    l_q = sum(x[(teacher_name, cl["name"], cl["buildings"][d], d, p+1)]
                              for cl in class_list)

                    # Линейная связь pair_var ⇔ (l_p ≥1 ∧ l_q ≥1):
                    # 1) pair_var ≤ (l_p ≥1)  и  pair_var ≤ (l_q ≥1)
                    model.Add(l_p >= 1).OnlyEnforceIf(pair_var)
                    model.Add(l_q >= 1).OnlyEnforceIf(pair_var)
                    model.Add(pair_var <= l_p)    # если pair_var=1, l_p must be ≥1
                    model.Add(pair_var <= l_q)    # if pair_var=1, l_q must be ≥1

                    # 2) Если оба есть, то pair_var может быть 1, иначе он принудительно 0:
                    #    l_p + l_q - 1 ≤ pair_var * 2
                    #    но проще: l_p + l_q - 1 ≤ pair_var*2  ⟹ если l_p=l_q=1, RHS=2 ⇒ pair_var may be 1
                    model.Add(l_p + l_q - 1 <= pair_var * 2)


        # Целевая функция: минимизируем переходы, окна у учителей + макимизируем пары 
        model.Minimize(
            sum(transitions.values()) * 10 +
            sum(windows.values()) * 4 - 
            sum(pairs.values()) * 1
        )


        # --------------------------------------------
        # 5. Решение 
        # --------------------------------------------
        solver = cp_model.CpSolver()
        solver.parameters.max_time_in_seconds = self.time - 10
        solver.parameters.log_search_progress = True
        solver.parameters.cp_model_presolve = True
        solver.parameters.num_search_workers = 7
        solver.parameters.linearization_level = 2
        solver.parameters.cp_model_probing_level = 0
        status = solver.Solve(model)

        if status == cp_model.OPTIMAL or status == cp_model.FEASIBLE:
            self.root.after(0, lambda: self.update_progress_manual(100))
            self.auto_update_enabled = False
            self.root.after(0, lambda: self.in_process.config(text="Расписание создано"))
            
            print("Расписание составлено успешно!")

            txt_output_windows = []
            txt_output_trans = []

            for (teacher, d, p, q), wind in windows.items():
                if solver.Value(wind):
                    txt_output_windows.append((teacher, d, p, q))
            print(f"Окна у учителей: {len(txt_output_windows)}")
            print(txt_output_windows)

            for (teacher, d, p, q), trans in transitions.items():
                if solver.Value(trans):
                    txt_output_trans.append((teacher, d, p, q))
            print(f"Переходы между корпусами: {len(txt_output_trans)}")
            print(txt_output_trans)
            print("-"*70)

            # Собираем расписание в таблицу
            schedule = []
            for d in days:
                for p in range(1, periods_per_day + 1):
                    for c in classes:
                        for s in c['subjects']:
                            if solver.Value(x[(s["teacher"], c["name"], c["buildings"][d], d, p)]) == 1:
                                            schedule.append({
                                                "День": d,
                                                "Урок": p,
                                                "Класс": c["name"],
                                                "Учитель": s["teacher"],
                                                "Корпус": c["buildings"][d],
                                                "Предмет": s['name']
                                            })
        else:
            # Останавливаем прогресс-бар при ошибке
            self.auto_update_enabled = False
            print("Ошибка: Конфликтующие входные данные")
            self.root.after(0, lambda: self._show_error_popup())
            return

        count = 0
        for a in schedule:
            # print(a)
            count += 1
        print(f"Кол-во уркоов до добавления кабинетов: {count}")

        print("-"*70)
        # --------------------------------------------
        # 6. Выбор кабинета
        # --------------------------------------------



        # ==================== НАСТРОЙКА: фикс-кабинеты информатики ====================
        # Корпус "1" -> ОТ_1 и ОТ_2 (для двух групп), корпус "2" -> 401 и 407
        # Если у тебя реально другие названия — поменяй здесь.
        INF_FIXED_BY_BUILDING = {
            "1": ["ОТ_1", "ОТ_2"],
            "2": ["401", "407"],
        }

        # ==================== НОРМАЛИЗАЦИЯ ДАННЫХ (обязательно) ====================
        def norm_str(x):
            return str(x).strip()

        def norm_teachers(x):
            # Учитель ВСЕГДА list[str]
            if x is None:
                return []
            if isinstance(x, str):
                s = x.strip()
                return [s] if s else []
            return [str(t).strip() for t in x if str(t).strip()]

        # schedule
        for l in schedule:
            l["Класс"] = norm_str(l["Класс"])
            l["День"] = norm_str(l["День"])
            l["Урок"] = int(l["Урок"])                 # int
            l["Корпус"] = norm_str(l["Корпус"])        # str
            l["Предмет"] = norm_str(l["Предмет"])
            l["Учитель"] = norm_teachers(l["Учитель"]) # list[str]

        # classes
        for c in classes:
            c["name"] = norm_str(c["name"])

        # rooms
        for r in rooms:
            r["name"] = norm_str(r["name"])
            r["building"] = norm_str(r["building"])
            r["size"] = norm_str(r.get("size", ""))
            r["prio"] = norm_str(r.get("prio", "Нет прио"))

        # ==================== ordered_schedule[class][day][period] ====================
        ordered_schedule = {}
        for c in classes:
            cname = c["name"]
            ordered_schedule[cname] = {}
            for d in days:
                ordered_schedule[cname][d] = {}
                for p in range(1, periods_per_day + 1):
                    ordered_schedule[cname][d][p] = 0

        for l in schedule:
            cname = l["Класс"]
            day = l["День"]
            p = int(l["Урок"])
            ordered_schedule[cname][day][p] = [l["Учитель"], str(l["Корпус"]), l["Предмет"]]

        # ==================== busy_classes[day][period_str] = [(room_name, building), ...] ====================
        busy_classes = {d: {str(p): [] for p in range(1, periods_per_day + 1)} for d in days}

        # ==================== ВСПОМОГАТЕЛЬНОЕ: получить dict кабинета (или создать заглушку) ====================
        def get_room_dict(room_name: str, building: str):
            room_name = norm_str(room_name)
            building = norm_str(building)
            rr = next((r for r in rooms if r["name"] == room_name and r["building"] == building), None)
            if rr is not None:
                return rr
            # если кабинет не найден в rooms — создаём заглушку, чтобы код не падал
            rr = {"name": room_name, "building": building, "size": "Маленький", "prio": "Инфа"}
            rooms.append(rr)
            return rr

        def reserve_rooms(day: str, period: int, chosen_rooms: list):
            """chosen_rooms = list[dict], бронируем в busy_classes"""
            for r in chosen_rooms:
                busy_classes[day][str(period)].append((r["name"], r["building"]))

        # ==================== ПРЕДПРОГОНКА: СНАЧАЛА СТАВИМ ИНФОРМАТИКЕ НУЖНЫЕ КАБИНЕТЫ ====================
        # Это делается ДО общего подбора, и сразу бронирует кабинеты.
        for c in classes:
            cname = c["name"]
            for d in days:
                p = 1
                while p <= periods_per_day:
                    entry = ordered_schedule[cname][d][p]
                    if entry == 0:
                        p += 1
                        continue
                    
                    subject = entry[2]
                    building = entry[1]
                    teachers_list = entry[0]

                    # ТОЛЬКО информатика
                    if "Инфа" not in subject:
                        p += 1
                        continue
                    
                    need = 1 if len(teachers_list) <= 1 else 2
                    fixed_names = INF_FIXED_BY_BUILDING.get(building, [])
                    fixed_names = fixed_names[:need]

                    # если в маппинге нет нужного количества кабинетов — просто пропускаем (пусть общий подбор решит)
                    if len(fixed_names) != need:
                        p += 1
                        continue
                    
                    chosen = [get_room_dict(rn, building) for rn in fixed_names]

                    # если это пара (на следующий урок тоже Инфа в том же корпусе) — бронируем сразу на два урока
                    is_pair = False
                    if p < periods_per_day:
                        nxt = ordered_schedule[cname][d][p + 1]
                        if nxt != 0 and ("Инфа" in nxt[2]) and (nxt[1] == building):
                            need2 = 1 if len(nxt[0]) <= 1 else 2
                            # важно: чтобы на обоих уроках одинаково 1/2 кабинета
                            if need2 == need:
                                is_pair = True

                    # проверим занятость (если уже занято — не ставим жёстко, пусть дальше best-effort попробует)
                    def free_for(period_check: int):
                        return all((r["name"], r["building"]) not in busy_classes[d][str(period_check)] for r in chosen)

                    if is_pair:
                        if free_for(p) and free_for(p + 1):
                            # записываем выбранные кабинеты в ОБА урока и бронируем
                            if len(ordered_schedule[cname][d][p]) < 4:
                                ordered_schedule[cname][d][p].append(chosen)
                            else:
                                ordered_schedule[cname][d][p][3] = chosen

                            if len(ordered_schedule[cname][d][p + 1]) < 4:
                                ordered_schedule[cname][d][p + 1].append(chosen)
                            else:
                                ordered_schedule[cname][d][p + 1][3] = chosen

                            reserve_rooms(d, p, chosen)
                            reserve_rooms(d, p + 1, chosen)
                            p += 2
                            continue
                        # если на пару не свободно — попробуем хотя бы на текущий
                        if free_for(p):
                            if len(ordered_schedule[cname][d][p]) < 4:
                                ordered_schedule[cname][d][p].append(chosen)
                            else:
                                ordered_schedule[cname][d][p][3] = chosen
                            reserve_rooms(d, p, chosen)
                        p += 1
                        continue
                    else:
                        if free_for(p):
                            if len(ordered_schedule[cname][d][p]) < 4:
                                ordered_schedule[cname][d][p].append(chosen)
                            else:
                                ordered_schedule[cname][d][p][3] = chosen
                            reserve_rooms(d, p, chosen)
                        p += 1
                        continue
                    
                    
        # ==================== ВЫБОР КАБИНЕТОВ: строгий -> слабее -> ещё слабее ====================
        def pick_rooms_core(ordered_schedule, rooms, busy_classes, k, keys, check_periods,
                            allow_other_buildings=False, ignore_subject_prio=False, allow_ot=False):
            """
            Возвращает список из k кабинетов (list[dict]) или [] если не найдено.
            """
            cname, day, p = keys
            entry = ordered_schedule[cname][day][p]
            teach = entry[0]          # list[str]
            building = entry[1]       # str
            subject = entry[2]        # str

            def is_free(room_name, bld):
                for pp in check_periods:
                    if (room_name, bld) in busy_classes[day][str(pp)]:
                        return False
                return True

            # --- ПРАВИЛО РАЗМЕРА ---
            # урок для всего класса -> только "Большой"
            # урок для групп -> можно "Маленький"
            is_group_lesson = (len(teach) > 1)
            sizes = ("Маленький", "Большой") if is_group_lesson else ("Большой",)

            # предметный приоритет (англ/инфа)
            if "Инфа" in subject:
                subj_prio = None if ignore_subject_prio else "Инфа"
            elif "Англ" in subject:
                subj_prio = None if ignore_subject_prio else "Англ"
            else:
                subj_prio = None

            def ok_room(r):
                if not allow_other_buildings and r["building"] != building:
                    return False
                if not allow_ot and r["name"] in ("ОТ_1", "ОТ_2"):
                    return False
                if r.get("prio") == "Физра":
                    return False
                if not is_free(r["name"], r["building"]):
                    return False
                return True

            def subject_ok(r):
                if subj_prio is None:
                    return True
                return subj_prio in r["prio"].split("/")

            # ---------- 1) teacher prio ----------
            prio_lists = []
            for tname in teach:
                pref = []
                for t_g in teachers:
                    if norm_str(t_g.get("name")) == tname:
                        pmap = t_g.get("prio", {})
                        pval = pmap.get(building)
                        if isinstance(pval, str) and pval.strip():
                            pref = [pval.strip()]
                        elif isinstance(pval, list):
                            pref = [norm_str(x) for x in pval if norm_str(x)]
                        break
                prio_lists.append(pref)

            chosen_names = []
            for pref in prio_lists:
                found = None
                for rname in pref:
                    if rname not in chosen_names and is_free(rname, building):
                        found = rname
                        break
                if found:
                    chosen_names.append(found)

            if len(chosen_names) >= k:
                out = []
                for rname in chosen_names[:k]:
                    rr = next((r for r in rooms if r["name"] == rname and r["building"] == building), None)
                    if rr is None:
                        out = []
                        break
                    out.append(rr)
                if len(out) == k:
                    return out

            # ---------- 2) подбор по size + прио предмета ----------
            cands = []
            for sz in sizes:
                cands += [r for r in rooms if ok_room(r) and r["size"] == sz and subject_ok(r)]
            if len(cands) >= k:
                return random.sample(cands, k)

            # ---------- 3) подбор по size без прио предмета ----------
            cands = []
            for sz in sizes:
                cands += [r for r in rooms if ok_room(r) and r["size"] == sz]
            if len(cands) >= k:
                return random.sample(cands, k)

            # ---------- 4) финальный fallback: любые свободные, НО size НЕ отпускаем ----------
            cands = [r for r in rooms if ok_room(r) and r["size"] in sizes]
            if len(cands) >= k:
                return random.sample(cands, k)

            return []


        def pick_rooms_best_effort(ordered_schedule, rooms, busy_classes, k, keys, check_periods):
            chosen = pick_rooms_core(ordered_schedule, rooms, busy_classes, k, keys, check_periods,
                                     allow_other_buildings=False, ignore_subject_prio=False, allow_ot=False)
            if chosen:
                return chosen

            chosen = pick_rooms_core(ordered_schedule, rooms, busy_classes, k, keys, check_periods,
                                     allow_other_buildings=False, ignore_subject_prio=True, allow_ot=False)
            if chosen:
                return chosen

            chosen = pick_rooms_core(ordered_schedule, rooms, busy_classes, k, keys, check_periods,
                                     allow_other_buildings=True, ignore_subject_prio=True, allow_ot=False)
            if chosen:
                return chosen

            return pick_rooms_core(ordered_schedule, rooms, busy_classes, k, keys, check_periods,
                                   allow_other_buildings=True, ignore_subject_prio=True, allow_ot=True)


        # ==================== СБОР final_timetable (пары бронируем на 2 урока сразу) ====================
        final_timetable = []

        for c in classes:
            cname = c["name"]
            for d in days:
                for p in range(1, periods_per_day + 1):
                    entry = ordered_schedule[cname][d][p]
                    if entry == 0:
                        continue
                    
                    teachers_list = entry[0]
                    building = entry[1]
                    subject = entry[2]

                    # физра
                    if subject == "Физра":
                        room_field = "УОО" if building == "1" else "СЗ"
                        final_timetable.append({
                            "День": d, "Урок": p, "Класс": cname,
                            "Учитель": ", ".join(teachers_list),
                            "Корпус": building, "Предмет": subject, "Кабинет": room_field
                        })
                        continue
                    
                    need = 1 if len(teachers_list) <= 1 else 2

                    # уже назначено ранее (например второй урок пары или пред-прогонка информатики)
                    if len(entry) >= 4 and entry[3]:
                        chosen = entry[3]
                        room_field = chosen[0]["name"] if need == 1 else f"{chosen[0]['name']}, {chosen[1]['name']}"
                        final_timetable.append({
                            "День": d, "Урок": p, "Класс": cname,
                            "Учитель": ", ".join(teachers_list),
                            "Корпус": building, "Предмет": subject, "Кабинет": room_field
                        })
                        continue
                    
                    # пара: p и p+1 одинаковый предмет и корпус
                    is_pair = False
                    if p < periods_per_day:
                        nxt = ordered_schedule[cname][d][p + 1]
                        if nxt != 0 and nxt[2] == subject and nxt[1] == building:
                            need2 = 1 if len(nxt[0]) <= 1 else 2
                            if need2 == need and (len(nxt) < 4 or not nxt[3]):
                                is_pair = True

                    if is_pair:
                        chosen = pick_rooms_best_effort(ordered_schedule, rooms, busy_classes, need, (cname, d, p), [p, p + 1])

                        if chosen:
                            for pp in (p, p + 1):
                                for r in chosen:
                                    busy_classes[d][str(pp)].append((r["name"], r["building"]))
                                if len(ordered_schedule[cname][d][pp]) < 4:
                                    ordered_schedule[cname][d][pp].append(chosen)
                                else:
                                    ordered_schedule[cname][d][pp][3] = chosen

                            room_field = chosen[0]["name"] if need == 1 else f"{chosen[0]['name']}, {chosen[1]['name']}"
                        else:
                            # если на пару не получилось — ставим хотя бы на текущий
                            chosen1 = pick_rooms_best_effort(ordered_schedule, rooms, busy_classes, need, (cname, d, p), [p])
                            if chosen1:
                                for r in chosen1:
                                    busy_classes[d][str(p)].append((r["name"], r["building"]))
                                if len(ordered_schedule[cname][d][p]) < 4:
                                    ordered_schedule[cname][d][p].append(chosen1)
                                else:
                                    ordered_schedule[cname][d][p][3] = chosen1
                                room_field = chosen1[0]["name"] if need == 1 else f"{chosen1[0]['name']}, {chosen1[1]['name']}"
                            else:
                                room_field = "НЕТ КАБ"
                    else:
                        chosen = pick_rooms_best_effort(ordered_schedule, rooms, busy_classes, need, (cname, d, p), [p])
                        if chosen:
                            for r in chosen:
                                busy_classes[d][str(p)].append((r["name"], r["building"]))
                            if len(ordered_schedule[cname][d][p]) < 4:
                                ordered_schedule[cname][d][p].append(chosen)
                            else:
                                ordered_schedule[cname][d][p][3] = chosen
                            room_field = chosen[0]["name"] if need == 1 else f"{chosen[0]['name']}, {chosen[1]['name']}"
                        else:
                            room_field = "НЕТ КАБ"

                    final_timetable.append({
                        "День": d, "Урок": p, "Класс": cname,
                        "Учитель": ", ".join(teachers_list),
                        "Корпус": building, "Предмет": subject, "Кабинет": room_field
                    })



        count = 0
        for a in final_timetable:
            # print(a)
            count += 1

        print(f"Кол-во уроков после добваления кабинетов: {count}")

        print("-"*70)


        count = 0
        for c in classes:
            count_c = 0
            for s in c["subjects"]:
                count_c += s['hours']
            count += count_c
            print(f"{c['name']}: {count_c} часов")

        print("-"*70)

        print(f"Нормативно должно быть: {count}")


        # # --------------------------------------------
        # # 7. Вывод расписания в excel 
        # # --------------------------------------------

        wb = Workbook()
        list_of_classes = []

        for items in final_timetable:
            if items["Класс"] in list_of_classes:
                continue
            else:
                list_of_classes.append(items["Класс"])


        t_to_color = {}

        for t in teachers:
            t_to_color[t['name']] = t['color']



        # s_list_of_classes = sorted(list_of_classes)
        weight_letters_ = {
            "8А": 1,
            "8Б": 2,
            "8В": 3,
            "8Г": 4,
            "8Д": 5,

            "9А": 6,
            "9Б": 7,
            "9В": 8,
            "9Г": 9,
            "9Д": 10,

            "10А": 11,
            "10Б": 12,
            "10В": 13,
            "10Г": 14,
            "10Д": 15,

            "11А": 16,
            "11Б": 17,
            "11В": 18,
            "11Г": 19,
            "11Д": 20,

        }

        n = 1
        while n < len(list_of_classes):
            for i in range(len(list_of_classes)-n):
                if weight_letters_[list_of_classes[i]] > weight_letters_[list_of_classes[i + 1]]:
                    list_of_classes[i],  list_of_classes[i + 1] = list_of_classes[i + 1], list_of_classes[i]
            n += 1

        s_list_of_classes = list_of_classes

        # print(s_list_of_classes)

        ws = wb.active


        help_class = {} 
        for i in range(0, len(list_of_classes)): 
            help_class[list_of_classes[i]] = i

        #Дни недели 
        vertical_text = Alignment(text_rotation=90, horizontal='center',  
                                  vertical='center') 

        help_days_fully = { 
            "Пн": ("Понедельник",  
                   PatternFill(start_color='F0E68C', fill_type='solid') ,  
                    Font(color='000000', size=13, bold = True), 1), 
            "Вт": ("Вторник",  
                   PatternFill(start_color='6A5ACD', fill_type='solid') ,  
                    Font(color='000000', size=13, bold = True), 2), 
            "Ср": ("Среда",  
                   PatternFill(start_color='1E90FF', fill_type='solid') ,  
                    Font(color='000000', size=13, bold = True), 3), 
            "Чт": ("Четверг",  
                   PatternFill(start_color='FF69B4', fill_type='solid') ,  
                    Font(color='000000', size=13, bold = True), 4), 
            "Пт": ("Пятница",  
                   PatternFill(start_color='EE82EE', fill_type='solid') ,  
                    Font(color='000000', size=13, bold = True), 5), 
            "Сб": ("Суббота",  
                   PatternFill(start_color='2E8B57', fill_type='solid') ,  
                    Font(color='000000', size=13, bold = True), 6), 
        } 

        for idx, (day, fill, text, n) in enumerate(help_days_fully.values(), start=1): 
            # Получаем букву колонки 
            column_1 = get_column_letter(1)  
            row = (idx - 1) * 9 + (idx - 1) * 2 + 3 

            # Записываем день недели 
            cell = ws[f"{column_1}{row}"] 
            cell.value = day 

            cell.fill = fill  # Фон 
            cell.font = text 
            # Применяем вертикальную ориентацию 
            cell.alignment = vertical_text 

            # Увеличиваем ширину колонки 
            ws.column_dimensions[column_1].width = 5 

            # Объединяем ячейки по вертикали если нужно 
            ws.merge_cells(f"{column_1}{row}:{column_1}{row + 8}") 
        ### 

        #Вспомогалочка 
        period_to_time = { 
            "1": "8:35-  9:20", 
            "2": "9:30-10:15", 
            "3": "10:25-11:10", 
            "4": "11:30-12:15", 
            "5": "12:35-13:20", 
            "6": "13:40-14:25", 
            "7": "14:35-15:20", 
            "8": "15:30-16:15", 
            "9": "16:25-17:10" 
        } 
        # 

        #Cкелетик 

        thin_border = Border( 
            left=Side(style='thin', color='808080'), 
            right=Side(style='thin', color='808080'), 
            top=Side(style='thin', color='808080'), 
            bottom=Side(style='thin', color='808080') 
        ) 
        ### 

        separator_fill = PatternFill(start_color='D3D3D3',  # Светло-серый цвет 
                                       end_color='D3D3D3', 
                                       fill_type='solid') 
        ### 

        time_alignment = Alignment( 
            horizontal='center',     # Выравнивание по горизонтали 
            vertical='center',       # Выравнивание по вертикали 
            wrap_text=True,         # Перенос текста 
            shrink_to_fit=True,     # Уменьшение текста для вмещения 
        ) 
        time_text = Font(color='000000',  
                         size=11, bold = False) 
        ### 

        for c in help_class: 
            num = help_class[c] 
            #Название класса 
            column_3 = get_column_letter(5 + 8 * num)  
            row = 1  # Явно указываем строку 
            cell = ws[f"{column_3}{row}"] 
            cell.value = c 
            cell.font = Font(color='000000', size=15, bold=True) 
            cell.alignment = Alignment(horizontal='center', vertical='center') 
            # Объединение ячеек после установки значения 
            ws.merge_cells(f"{column_3}{row}:{get_column_letter(9 + 8 * num)}{row}") 
            # 

            #Перегородочки 

            ws.column_dimensions[f'{get_column_letter(2 + num * 8)}'].width = 1.7 
            ws.column_dimensions[f'{get_column_letter(4 + num * 8)}'].width = 1.7 
            column_2 = get_column_letter(3 + num * 8)

            for row in range(1, 67): 
                cell = ws[f"{column_2}{row}"] 
                cell.fill = separator_fill 
            # Устанавливаем высоту строки перегородки 
            ws.column_dimensions[column_2].width = 0.4 
            ### 

            #Время 
            for idx, (day, fill, text, n) in enumerate(help_days_fully.values(), start=1): 
                for i in range(0, 9): 
                    column = get_column_letter(5 + num * 8)

                    cell = ws[f'{column}{2 * (idx - 1) + (idx - 1) * 9 + i + 3}'] 
                    cell.value = period_to_time[str(i + 1)] 
                    cell.alignment = time_alignment 
                    cell.border = thin_border 
                    cell.font = time_text 
            ### 

            #Уроки
            for idx, (day, fill, text, n) in enumerate(help_days_fully.values(), start=1): 
                for i in range(0, 9): 
                    column = get_column_letter(6 + num * 8)
                    row = 2 * (idx - 1) + (idx - 1) * 9 + i + 3 

                    # Сначала записываем значение 
                    cell = ws[f'{column}{row}'] 
                    cell.value = i+1 
                    cell.alignment = time_alignment 
                    cell.border = thin_border 
                    cell.font = time_text 
                    ws.column_dimensions[column].width = 1.8 
            ### 

            #Корпуса 
            n_to_b = { 
                "1": "Лицей",  
                "2": "Графский" 
            } 

            for _ in classes: 
                if _['name'] == c: 
                    for d in _['buildings']: 
                        column_4 = get_column_letter(4 + 8 * num) 
                        row_4 = 2 + 11 * (help_days_fully[d][3] - 1) 
                        cell = ws[f'{column_4}{row_4}'] 
                        cell.value = n_to_b[_['buildings'][d]] 

                        cell.alignment = Alignment( 
                            horizontal='center',     # Выравнивание по горизонтали 
                            vertical='center',       # Выравнивание по вертикали 
                        ) 

                        cell.font = Font(color='000000',  
                            size=13, bold = False, italic=True 
                        ) 

                        ws.merge_cells(f"{column_4}{row_4}:{get_column_letter(9 + 8 * num)}{row_4}")    
            ### 

            #Обязательный РОВ

            #Предмет
            row = 4 
            column = get_column_letter(7 + help_class[c] * 8)
            ws.column_dimensions[column].width = 15 
            cell = ws[f'{column}{row}'] 
            cell.value = "РОВ"
            cell.alignment = Alignment( 
                            horizontal='center',     # Выравнивание по горизонтали 
                            vertical='center', 
                            wrap_text=False,        # Отключаем перенос текста 
                            shrink_to_fit=False       # Выравнивание по вертикали 
                        ) 
            cell.border = thin_border 
            cell.font = Font(color='FFFFFF', size=11, bold = False)  
            cell.fill = PatternFill(start_color='b00149', fill_type='solid') 
            ### 

            #Учитель 
            row = 4
            column = get_column_letter(9 + help_class[c] * 8) 
            ws.column_dimensions[column].width = 15 
            cell = ws[f'{column}{row}'] 
            cell.value = "--"
            cell.alignment = Alignment( 
                            horizontal='center',     # Выравнивание по горизонтали 
                            vertical='center', 
                            wrap_text=True,        # Отключаем перенос текста 
                            shrink_to_fit=False       # Выравнивание по вертикали 
                        ) 
            cell.border = thin_border 
            cell.font = Font(color='FFFFFF', size=11, bold = False)  
            cell.fill = PatternFill(start_color='b00149', fill_type='solid') 
            # 

            #Кабинеты 
            row = 4
            column = get_column_letter(8 + help_class[c] * 8)
            ws.column_dimensions[column].width = 15 
            cell = ws[f'{column}{row}'] 
            cell.value = "ККЗ"
            cell.alignment = Alignment( 
                            horizontal='center',     # Выравнивание по горизонтали 
                            vertical='center', 
                            wrap_text=False,        # Отключаем перенос текста 
                            shrink_to_fit=False        # Выравнивание по вертикали 
                        ) 
            cell.border = thin_border 
            cell.font = Font(color='FFFFFF', size=11, bold = False)  
            cell.fill = PatternFill(start_color='b00149', fill_type='solid') 
            #


        for prec_p in final_timetable: 
        
            d = prec_p['День'] 
            p = prec_p['Урок'] 
            c = prec_p['Класс'] 
            t = prec_p['Учитель'] 
            s = prec_p['Предмет'] 
            r = prec_p['Кабинет'] 

            fill = PatternFill(start_color= t_to_color[t.split(",")[0]], fill_type='solid') 
            #Предмет 
            row = 3 + (help_days_fully[d][3] - 1) * 11 + (p - 1) 
            column = get_column_letter(7 + help_class[c] * 8)
            ws.column_dimensions[column].width = 15 
            cell = ws[f'{column}{row}'] 
            cell.value = s 
            cell.alignment = Alignment( 
                            horizontal='center',     # Выравнивание по горизонтали 
                            vertical='center', 
                            wrap_text=False,        # Отключаем перенос текста 
                            shrink_to_fit=False       # Выравнивание по вертикали 
                        ) 
            cell.border = thin_border 
            cell.font = time_text 
            cell.fill = fill
            ### 

            #Учитель 
            row = 3 + (help_days_fully[d][3] - 1) * 11 + (p - 1) 
            column = get_column_letter(9 + help_class[c] * 8) 
            ws.column_dimensions[column].width = 15 
            cell = ws[f'{column}{row}'] 
            cell.value = t.replace(",", ",\n") 
            cell.alignment = Alignment( 
                            horizontal='center',     # Выравнивание по горизонтали 
                            vertical='center', 
                            wrap_text=True,        # Отключаем перенос текста 
                            shrink_to_fit=False       # Выравнивание по вертикали 
                        ) 
            cell.border = thin_border 
            cell.font = time_text 
            cell.fill = fill
            # 

            #Кабинеты 
            row = 3 + (help_days_fully[d][3] - 1) * 11 + (p - 1) 
            column = get_column_letter(8 + help_class[c] * 8)
            ws.column_dimensions[column].width = 15 
            cell = ws[f'{column}{row}'] 
            cell.value = r 
            cell.alignment = Alignment( 
                            horizontal='center',     # Выравнивание по горизонтали 
                    vertical='center', 
                    wrap_text=False,        # Отключаем перенос текста 
                    shrink_to_fit=False        # Выравнивание по вертикали 
                ) 
            cell.border = thin_border 
            cell.font = time_text 

        # Создаём папку output если её нет
        output_path = os.path.join(os.getcwd(), "Выходные данные")
        os.makedirs(output_path, exist_ok=True)
        
        wb.save(os.path.join(output_path, "Итоговое_расписание.xlsx"))
        # --------------------------------------------
        # 8. Вывод окон и переходов в txt
        # --------------------------------------------
        file_name = os.path.join(output_path, "Переходы_и_окна_учителей.txt")
        with open(file_name, "w", encoding="UTF-8") as file:
            file.write("**Переходы и окна учитеей по физре см в таблице!! \n")
            file.write("\n")
            file.write("Переходы между корпусами у учителей:\n")
            if txt_output_trans:
                for item in txt_output_trans:
                
                    file.write(f"{item[0]}: {item[1]} - {item[2]} по {item[3]}\n")
            else:
                file.write("Отсутств...\n")

            file.write("\n")

            file.write("Окна у учителей:\n")
            if txt_output_windows:
            
                for item in txt_output_windows:       
                    file.write(f"{item[0]}: {item[1]} - {item[2]} по {item[3]}\n")
            else:
                file.write("Отсутств...\n")

        # Показываем popup успеха в главном потоке
        self.root.after(0, lambda: self._show_success_popup())

        ###

        #Алгоритм
        #

#<Созд пара> - 2 
class menu_add_parallel:
    def __init__(self, parent, base, connect):
        
    #Данные корня
        self.base_menu_2 = base
        self.con_menu_2 = connect
        self.root_menu_2 = Toplevel(parent)
        self.root_menu_2.title("Добавление класса")
        self.root_menu_2.geometry(f"{int(self.root_menu_2.winfo_screenwidth() * 0.56)}x{int(self.root_menu_2.winfo_screenheight()*0.6)}")
        self.root_menu_2.resizable(width=False, height=False)
        self.full_data_class = []
        self.clas_buttons = []

        self.data_PE= []
        self.extra_les = []                     # хранит список внеурочек (только ФИО учителя)
        ###
    ###

        style_2 = ttk.Style()

        style_2.theme_use('clam')

        style_2.configure("TLabel", font=("Helvetica", 30, "italic"))

        style_2.configure("TEntry", fieldbackground="#DCDCDC")

        style_2.configure("TCombobox", fieldbackground="#DCDCDC", arrowsize = 0)
        style_2.map("TCombobox", fieldbackground=[("readonly", "#DCDCDC")])


        style_2.configure("Main_2.TButton", font=("Helvetica", 30, "bold"), background="#696969", foreground="white",)
        style_2.map("Main_2.TButton", background=[("active", "#0000CD")])

        style_2.configure("Sec_2.TButton", font=("Helvetica", 13, "bold"), background="#696969", foreground="white",)
        style_2.map("Sec_2.TButton", background=[("active", "#4169E1")])

        self.root_menu_2.configure(bg="#e0dcd4")

        self.PE_les = []
        self.extra_les = []


    #Виджеты - создания параллели +++
    #Границы 
        separator_1 = Frame(
            self.root_menu_2, 
            height=3,         
            bg='grey',         
            relief=FLAT,
            bd=1              
        )
        separator_1.place(relx=0, rely=0.14, relwidth=1, relheight=0.007) #Горизонтальная

        separator_1 = Frame(
            self.root_menu_2, 
            height=3,          
            bg='grey',        
            relief=FLAT,
            bd=1               
        )
        separator_1.place(relx=0.29, rely=0.14, relwidth= 0.004, relheight= 1) #Вертикальная

        separator_1 = Frame(
            self.root_menu_2, 
            height=3,          
            bg='grey',        
            relief=FLAT,
            bd=1               
        )
        separator_1.place(relx=0.73, rely=0, relwidth= 0.004, relheight= 0.14) #Вертикальная


    ###

    #Блок выбора номера
        self.creating_number = ttk.Label(
            self.root_menu_2,
            text = "Номер:",
            style="TLabel",
            )
        self.creating_number.place(relx=0.1, rely=0, relwidth=0.20, relheight=0.14)

        
        self.options_1 = ["8", "9", "10", "11"]
        self.number = ttk.Combobox(
            self.root_menu_2,
            values=self.options_1,
            style="TCombobox",
            font=("Helvetica", 29),
            state="readonly", 
            foreground="#4D4D4D"
        )
        self.number.place(relx=0.3, rely=0, relwidth=0.08, relheight=0.14)
    ###

    #Блок выбора буквы
        self.creating_letter = ttk.Label(
            self.root_menu_2,
            text = "Буква:",
            style = "TLabel"
            )
        self.creating_letter.place(relx=0.45, rely=0, relwidth=0.20, relheight=0.14)

        options_2 = ["А", "Б", "В", "Г", "Д"]
        self.letter = ttk.Combobox(
            self.root_menu_2,
            values=options_2,
            style="TCombobox",
            font=("Helvetica", 29),
            state="readonly", 
            foreground="#4D4D4D"
        )
        self.letter.place(relx=0.65, rely=0, relwidth=0.08, relheight=0.14)
    ###

    #Блок выбора доп инфы 
        self.add_infa = ttk.Button(
            self.root_menu_2,
            text="Доп инфа",
            style = "Main_2.TButton",
            command= self.open_dop_infa
        )
        self.add_infa.place(relx=0.734, rely=0, relwidth=0.266, relheight=0.14)

    ###

    #Блок выбора урока
        self.creating_subject = ttk.Label(
            self.root_menu_2,
            text = "Урок:",
            style ="TLabel",
            anchor="center"
            )
        self.creating_subject.place(relx=0, rely=0.15, relwidth=0.29, relheight=0.1)

        self.options_3 = ["Алгебра",
            "Геометрия",
            "ТеорВер",
            "Физика",
            "Русский",
            "Литература",
            "История",
            "Информатика_1",
            "Информатика_2",
            "География",
            "Английский язык_1",
            "Английский язык_2",
            "Физра",
            "Биология",
            "Химия",
            "Обществознание",
            "ОБЖ",
            "Технология"]
        
        self.subject = ttk.Combobox(
            self.root_menu_2,
            values=self.options_3,
            style="TCombobox",
            font=("Helvetica", 27),
            state="readonly", 
            foreground="#4D4D4D"
        )
        self.subject.place(relx=0, rely=0.25, relwidth=0.29, relheight=0.1)
    ###

    #Блок выбора учителя
        self.options_teacher = []
        self.con_menu_2.execute("""SELECT * FROM teachers""")
        values = self.con_menu_2.fetchall() # Проверка на предварительное наличие кабинета 

        for i in values:
            self.options_teacher.append(f"{i[1]} {i[2][0]}. {i[3][0]}.")

        self.creating_teacher = ttk.Label(
            self.root_menu_2,
            text = "Учитель:",
            style="TLabel",
            anchor="center"
            )
        self.creating_teacher.place(relx=0, rely=0.35, relwidth=0.29, relheight=0.1)
        
        self.teacher = ttk.Combobox(
            self.root_menu_2,
            values=self.options_teacher,
            style="TCombobox",
            font=("Helvetica", 27),
            state="readonly", 
            foreground="#4D4D4D"
        )
        self.teacher.place(relx=0, rely=0.45, relwidth=0.29, relheight=0.1)
    ###

    #Блок выбора количества часов
        self.creating_hours = ttk.Label(
            self.root_menu_2,
            text = "Кол-во часов:",
            style="TLabel",
            anchor="center"
            )
        self.creating_hours.place(relx=0, rely=0.55, relwidth=0.29, relheight=0.1)
        
        self.hours = ttk.Entry(
            self.root_menu_2,
            style="TEntry",
            font=("Helvetica", 27),
            foreground="#4D4D4D"
        )
        self.hours.place(relx=0, rely=0.65, relwidth=0.29, relheight=0.1)
    ###

    #Футер
    #Кнопка добавления урока 
        self.add_clas = ttk.Button(
            self.root_menu_2,
            text="Добавить",
            style = "Main_2.TButton",
            command = self.add_class
        )
        self.add_clas.place(relx=0, rely=0.75, relwidth=0.29, relheight=0.1)
    ###

    #Кнопка добавления параллели
        self.add_parallel = ttk.Button(
            self.root_menu_2,
            text="Добавить параллель",
            style = "Main_2.TButton",
            command= self.addd_parallel
        )
        self.add_parallel.place(relx=0, rely=0.85, relwidth=1, relheight=0.15)
    ###

    #Функционал
    #Добавение параллели
    def addd_parallel(self):
        letter = self.letter.get()
        num = self.number.get()

        # Проверка существования параллели в БД
        self.con_menu_2.execute(
            "SELECT * FROM parallels WHERE Letter = ? AND Number = ?",
            (letter, num)
        )
        if self.con_menu_2.fetchall():
            warning = eror_popup(self.root_menu_2, "Такая параллель уже есть")
            warning.root.mainloop()
            return
        ###

        #Проверка на 2 групппы
        list_sub = []
        for _ in self.full_data_class:
            if _:
                list_sub.append(_[0])

        if ("Информатика_1" in list_sub and "Информатика_2" not in list_sub) or ("Информатика_1" not in list_sub and "Информатика_2" in list_sub):
            warning = eror_popup(self.root_menu_2, "Созданы не 2 группы")
            warning.root.mainloop()
            return
        
        if ("Английский язык_1" in list_sub and "Английский язык_2" not in list_sub) or ("Английский язык_1" not in list_sub and "Английский язык_2" in list_sub):
            warning = eror_popup(self.root_menu_2, "Созданы не 2 группы")
            warning.root.mainloop()
            return



        # Добавляем параллель в таблицу parallels
        
        self.con_menu_2.execute(
            "INSERT INTO parallels (Letter, Number) VALUES (?, ?)",
            (letter, num)
        )
        self.base_menu_2.commit()
        ###

        # Получаем id добавленной параллели
        self.con_menu_2.execute(
            "SELECT id FROM parallels WHERE Letter = ? AND Number = ?",
            (letter, num)
        )
        parallel_id = self.con_menu_2.fetchone()[0]
        ###

        # Добавляем уроки в таблицу lessons с id учителя
        for subject_data in self.full_data_class:
            if subject_data:  # Проверяем, что данные не пустые

                # Получаем id учителя по ФИО
                teacher_fio = subject_data[1].split()  # Разбиваем "Фамилия И. О." на части
                surname = teacher_fio[0]
                name = teacher_fio[1][0]  # Первая буква имени
                patrony = teacher_fio[2][0]  # Первая буква отчества

                # Получаем id учителя из таблицы teachers
                self.con_menu_2.execute(
                    "SELECT id FROM teachers WHERE Surname = ? AND Name LIKE ? AND Patrony LIKE ?",
                    (surname, f"{name}%", f"{patrony}%")
                )
                teacher_id = self.con_menu_2.fetchone()[0]

                # Вставляем данные с Id_parallel и Id_teacher
                self.con_menu_2.execute(
                    "INSERT INTO lessons (Id_parallel, Subject, Hours, Id_teacher) VALUES (?, ?, ?, ?)",
                    (parallel_id, subject_data[0], subject_data[2], teacher_id)
                )
        self.base_menu_2.commit() 

        for data in self.data_PE:
            if data:
                self.con_menu_2.execute(
                        "INSERT INTO pe (Id_parallel, Teacher, Day, Lesson) VALUES (?, ?, ?, ?)",
                        (parallel_id, data[0], data[1], data[2])
                    )
                self.base_menu_2.commit()
        # и внеурочки (только преподаватели)
        for teacher in self.extra_les:
            if teacher:
                self.con_menu_2.execute(
                        "INSERT INTO extra (Id_parallel, Teacher) VALUES (?, ?)",
                        (parallel_id, teacher)
                    )
        self.base_menu_2.commit()


        # Очищаем форму
        self.letter.set('')
        self.number.set('')
        self.hours.delete(0, END)
        self.hours.insert(0, '')
        self.teacher.set('')
        self.subject.set('')
        self.full_data_class.clear()
        self.display_class()
        self.full_data_class = []
        self.data_PE = []
        self.extra_les = []
        popup(self.root_menu_2, f"Параллель {num}{letter} создана!", "Успех").root.mainloop()
    ###
    def open_dop_infa(self):
        extra_info_class(self, self.root_menu_2, self.con_menu_2, self.base_menu_2)




#См инфы об уроке
    def display_class(self):

    # Сначала удалим старые кнопки
        for btn in self.clas_buttons:
            btn.destroy()
        self.clas_buttons = []

        # Создаем список только непустых элементов
        active_classes = [clas for clas in self.full_data_class if clas]

        # Теперь используем индексы только активных классов
        for i, clas in enumerate(active_classes):
            pos_x = (i % 4) * 0.175 + 0.3
            pos_y = (i // 4) * 0.14 + 0.15

            btn = ttk.Button(
                self.root_menu_2,
                text=f"{clas[0]}",  
                command=lambda cl=(i, clas): self.open_popup_class(cl),
                style = "Sec_2.TButton"
            )
            btn.place(relx=pos_x, rely=pos_y, relwidth=0.175, relheight=0.14)
            self.clas_buttons.append(btn)

    #Открыть изменение предмета
    def open_popup_class(self, clas):
        popup_subject_1(clas[0], clas[1], self.root_menu_2, self.base_menu_2, self.con_menu_2, self)
    ###

    #Проверка на добавление урока
    def add_class(self):
        mass = []
        sub = self.subject.get()
        teacher = self.teacher.get()
        hours = self.hours.get()

        #Проверка урока
        if sub not in self.options_3:
            warning = eror_popup(self.root_menu_2, "Ошибка предмете")
            warning.root.mainloop()
            return
        mass.append(sub)

        
        for i in self.full_data_class:
            if i and i[0] == sub:  # Сначала проверяем, что i не пустой, затем сравниваем
                warning = eror_popup(self.root_menu_2, "Уже есть предмет")
                warning.root.mainloop()
                return
        ###

        #Проверка учителя
        if teacher not in self.options_teacher:
            warning = eror_popup(self.root_menu_2, "Нет такого учителя")
            warning.root.mainloop()
            return
        mass.append(teacher)
        ###

        #Проверка количества часов
        char = "0123456789"
        if hours == "":
            warning = eror_popup(self.root_menu_2, "Нет кол-ва часов")
            warning.root.mainloop()
            return

        for i in hours:
            if i not in char:
                warning = eror_popup(self.root_menu_2, "Ошибка кол-ве часов")
                warning.root.mainloop()
                return
        
        if int(hours) > 30:
            warning = eror_popup(self.root_menu_2, "Слишком много часов")
            warning.root.mainloop()
            return
        ###
        self.full_data_class.append([sub, teacher, hours])
        self.display_class()
        warning = popup(self.root_menu_2, "Предмет добавлен", f"{sub}")
###
#<Окно доп инфы> - 
class extra_info_class:
    def __init__(self, parent, paren_root, con, base):
        self.parent = parent
        self.root_extra_info = Toplevel(paren_root)
        self.con = con
        self.base = base
        self.root_extra_info.title("Доп инфа")
        self.root_extra_info.geometry(f"{int(self.root_extra_info.winfo_screenwidth() * 0.54)}x{int(self.root_extra_info.winfo_screenheight()*0.58)}")
        self.root_extra_info.resizable(width=False, height=False)

        self.data_PE_les = list(self.parent.data_PE)
        self.PE_buttons = []
        # дополнительный список для внеурочек
        self.data_extra = list(getattr(self.parent, 'extra_les', []))
        self.extra_buttons = []
        
        # Очищаем невалидные данные (удаленные учителя)
        self._clean_invalid_teachers()

        style_extra_info_class = ttk.Style()

        style_extra_info_class.theme_use('clam')

        style_extra_info_class.configure("TLabel", font=("Helvetica", 30, "italic"))

        style_extra_info_class.configure("TEntry", fieldbackground="#DCDCDC")

        style_extra_info_class.configure("TCombobox", fieldbackground="#DCDCDC", arrowsize = 0)
        style_extra_info_class.map("TCombobox", fieldbackground=[("readonly", "#DCDCDC")])


        style_extra_info_class.configure("Main_2.TButton", font=("Helvetica", 30, "bold"), background="#696969", foreground="white",)
        style_extra_info_class.map("Main_2.TButton", background=[("active", "#0000CD")])

        style_extra_info_class.configure("Sec_2.TButton", font=("Helvetica", 13, "bold"), background="#696969", foreground="white",)
        style_extra_info_class.map("Sec_2.TButton", background=[("active", "#4169E1")])

        self.root_extra_info.configure(bg="#e0dcd4")

    #Граница
        separator_1 = Frame(
            self.root_extra_info, 
            height=3,          
            bg='grey',        
            relief=FLAT,
            bd=1               
        )
        separator_1.place(relx=0.498, rely=0, relwidth= 0.004, relheight= 0.86) #Вертикальная
    ###

    #Блок физкультур
        self.PE_les = ttk.Label(
            self.root_extra_info,
            text = "Физра:",
            style="TLabel",
            anchor="center"
            )
        self.PE_les.place(relx=0, rely=0, relwidth=0.498, relheight=0.14)

        self.teacher = ttk.Label(
            self.root_extra_info,
            text = "Учитель:",
            style="TLabel",
            anchor="center",
            font=("Helvetica", 26),
            )
        
        self.teacher.place(relx=0, rely=0.14, relwidth=0.2, relheight=0.14)

        self.options_teacher = []
        self.con.execute("""SELECT * FROM teachers""")
        values = self.con.fetchall() # Проверка на предварительное наличие кабинета 

        for i in values:
            self.options_teacher.append(f"{i[1]} {i[2][0]}. {i[3][0]}.")
        
        self.teacher_in = ttk.Combobox(
            self.root_extra_info,
            values=self.options_teacher,
            style="TCombobox",
            font=("Helvetica", 25),
            state="readonly", 
            foreground="#4D4D4D"
        )
        self.teacher_in.place(relx=0.2, rely=0.14, relwidth=0.298, relheight=0.14)


        self.day = ttk.Label(
            self.root_extra_info,
            text = "День:",
            style="TLabel",
            anchor="center",
            font=("Helvetica", 26),
            )
        self.day.place(relx=0, rely=0.28, relwidth=0.1, relheight=0.14)

        self.day_option = ["Понедельник", "Вторник", "Среда", "Четверг", "Пятница", "Суббота"]
        self.days = ttk.Combobox(
            self.root_extra_info,
            values=self.day_option,
            style="TCombobox",
            font=("Helvetica", 24),
            state="readonly", 
            foreground="#4D4D4D"
        )
        self.days.place(relx=0.1, rely=0.28, relwidth=0.2, relheight=0.14)


        self.lesson = ttk.Label(
            self.root_extra_info,
            text = "Урок:",
            style="TLabel",
            anchor="center",
            font=("Helvetica", 26),
            )
        self.lesson.place(relx=0.3, rely=0.28, relwidth=0.1, relheight=0.14)

        self.les = ["1", "2", "3", "4", "5", "6", "7", "8"]
        self.les_option = ttk.Combobox(
            self.root_extra_info,
            values=self.les,
            style="TCombobox",
            font=("Helvetica", 24),
            state="readonly", 
            foreground="#4D4D4D"
        )
        self.les_option.place(relx=0.4, rely=0.28, relwidth=0.098, relheight=0.14)

        #Кнопка добавления урока 
        self.add_clas = ttk.Button(
            self.root_extra_info,
            text="Добавить",
            style = "Main_2.TButton",
            command = self.add_PE
        )
        self.add_clas.place(relx=0, rely=0.42, relwidth=0.498, relheight=0.1)
    ###

    ###

    #Блок доп уроков
        self.extra_les = ttk.Label(
            self.root_extra_info,
            text = "Внеурочка:",
            style="TLabel",
            anchor="center",
            )
        self.extra_les.place(relx=0.52, rely=0, relwidth=0.498, relheight=0.14)

        self.teacher_e = ttk.Label(
            self.root_extra_info,
            text = "Учитель:",
            style="TLabel",
            anchor="center",
            font=("Helvetica", 26),
            )
        
        self.teacher_e.place(relx=0.52, rely=0.2, relwidth=0.2, relheight=0.14)

        self.options_teacher = []

        for data in self.parent.full_data_class:
            if data:  # Проверяем, что элемент не пустой
                self.options_teacher.append(data[1])
        
        self.teacher_in_1 = ttk.Combobox(
            self.root_extra_info,
            values=self.options_teacher,
            style="TCombobox",
            font=("Helvetica", 25),
            state="readonly", 
            foreground="#4D4D4D"
        )
        self.teacher_in_1.place(relx=0.72, rely=0.2, relwidth=0.28, relheight=0.14)

        # кнопка добавления внеурочки
        self.add_extra_btn = ttk.Button(
            self.root_extra_info,
            text="Добавить",
            style = "Main_2.TButton",
            command = self.add_extra
        )
        self.add_extra_btn.place(relx=0.502, rely=0.42, relwidth=0.498, relheight=0.1)
    ###
    
    #Футер
        self.cancel_1 = ttk.Button(
            self.root_extra_info,
            text="Отмена",
            style = "Main_2.TButton",
            command = self.root_extra_info.destroy
        )
        self.cancel_1.place(relx=0, rely=0.86, relwidth=0.5, relheight=0.14)

        self.save = ttk.Button(
            self.root_extra_info,
            text="Сохранить",
            style = "Main_2.TButton",
            command = self.save_extra_info
        )
        self.save.place(relx=0.5, rely=0.86, relwidth=0.5, relheight=0.14)

        # Очищаем невалидные данные перед отображением
        self._clean_invalid_teachers()
        
        # сразу показываем уже имеющиеся записи, если они есть
        self.display_PE()
        self.display_extra()

    def save_extra_info(self):
        # сохранить изменения в родителе (копируем, чтобы не было алиасов)
        self.parent.data_PE = list(self.data_PE_les)
        self.parent.extra_les = list(self.data_extra)
        # закроем окно до показа сообщения, чтобы при повторном открытии
        # здесь не осталось старого виджета
        self.root_extra_info.destroy()
        warning = popup(self.parent.root_menu_2, "Доп инфа сохранена!", "Успех")

    #Добавление физры на экран
    def add_PE(self):
        self.t = self.teacher_in.get()
        self.d = self.days.get()
        self.l = self.les_option.get()

        #Проверка на наличие хоть чего-то 
        if self.t == "":
            warning = eror_popup(self.root_extra_info, "Учитель физры не выбран")
            warning.root.mainloop()
            return
        if self.d == "":
            warning = eror_popup(self.root_extra_info, "День физры не выбран")
            warning.root.mainloop()
            return
        if self.l == "":
            warning = eror_popup(self.root_extra_info, "Урок физры не выбран")
            warning.root.mainloop()
            return
        #
        #(учитель, день, урок)
        #Проверка на уже существование урока
        if len(self.data_PE_les) != 0:
            for les in self.data_PE_les:
                if les[1] == self.d and les[2] == self.l:
                    warning = eror_popup(self.root_extra_info, "Есть физра в этом месте")
                    warning.root.mainloop()
                    return
        ###
        if len(self.data_PE_les) == 6:
            warning = eror_popup(self.root_extra_info, "Достигнут лимит кол-ва")
            warning.root.mainloop()
            return
        
        self.data_PE_les.append([self.t, self.d, self.l])
        self.display_PE()
        warning = popup(self.root_extra_info, "Физра добвлена, еще?", "Успех")

    def display_PE(self):
        for btn in self.PE_buttons:
            btn.destroy()
        self.PE_buttons = []
        for i, lesson in enumerate(self.data_PE_les):
            # Пропускаем пустые записи (например, удалённые)
            if not lesson: 
                continue
            pos_x = (i % 2) * 0.249  # левые колонки
            pos_y = (i // 2) * 0.11 + 0.52
            btn = ttk.Button(
                self.root_extra_info,
                text=f"{lesson[1]}", 
                command=lambda idx=i, l=self.data_PE_les, : self.open_popup_PE(idx, l),
                style = "Sec_2.TButton"
            )
            btn.place(relx=pos_x, rely=pos_y, relwidth=0.249, relheight=0.11)
            self.PE_buttons.append(btn)
    
    def add_extra(self):
        teacher = self.teacher_in_1.get()
        if teacher == "":
            warning = eror_popup(self.root_extra_info, "Учитель не выбран")
            warning.root.mainloop()
            return
        if teacher in self.data_extra:
            warning = eror_popup(self.root_extra_info, "Учитель уже добавлен")
            warning.root.mainloop()
            return
        self.data_extra.append(teacher)
        self.display_extra()
        self.teacher_in_1.set("")
        warning = popup(self.root_extra_info, "Учитель добавлен, еще?", "Успех")
    
    def display_extra(self):
        for btn in self.extra_buttons:
            btn.destroy()
        self.extra_buttons = []
        for i, teacher in enumerate(self.data_extra):
            if not teacher:
                continue
            pos_x = 0.502 + (i % 2) * 0.249
            pos_y = (i // 2) * 0.11 + 0.52
            btn = ttk.Button(
                self.root_extra_info,
                text=f"{teacher}",
                command=lambda idx=i: self.open_popup_extra(idx),
                style="Sec_2.TButton"
            )
            btn.place(relx=pos_x, rely=pos_y, relwidth=0.249, relheight=0.11)
            self.extra_buttons.append(btn)
    
    def open_popup_extra(self, idx):
        popup_extra_1((idx, self.data_extra), self)

#     # При нажатии на кнопку исключения открывается окно редактирования/удаления
    def open_popup_PE(self, idx, l):
        popup_PE_1((idx, l), self)

    def _clean_invalid_teachers(self):
        """Удаляет из списков данные с удаленными учителями"""
        # Получаем актуальный список учителей из БД
        self.con.execute("SELECT Surname, Name, Patrony FROM teachers")
        valid_teachers = []
        for row in self.con.fetchall():
            teacher_fio = f"{row[0]} {row[1][0]}. {row[2][0]}."
            valid_teachers.append(teacher_fio)
        
        # Очищаем данные физры от невалидных учителей
        self.data_PE_les[:] = [
            lesson for lesson in self.data_PE_les 
            if not lesson or lesson[0] in valid_teachers
        ]
        
        # Очищаем данные внеурочки от невалидных учителей
        self.data_extra[:] = [
            teacher for teacher in self.data_extra 
            if not teacher or teacher in valid_teachers
        ]

#Окно изменения физры
class popup_PE_1:
    def __init__(self, data, parent_menu):
        self.con = parent_menu.con
        self.idx, self.list_PE = data
        self.parent_menu = parent_menu
        self.root_PE = Toplevel(parent_menu.root_extra_info)
        self.root_PE.title("Изменить исключение")
        self.root_PE.geometry(f"{int(self.root_PE.winfo_screenwidth() * 0.27)}x{int(self.root_PE.winfo_screenheight()*0.3)}")
        self.root_PE.resizable(width=False, height=False)

        style_extra_info_class = ttk.Style()

        style_extra_info_class.theme_use('clam')

        style_extra_info_class.configure("TLabel", font=("Helvetica", 30, "italic"))

        style_extra_info_class.configure("TEntry", fieldbackground="#DCDCDC")

        style_extra_info_class.configure("TCombobox", fieldbackground="#DCDCDC", arrowsize = 0)
        style_extra_info_class.map("TCombobox", fieldbackground=[("readonly", "#DCDCDC")])


        style_extra_info_class.configure("Main_67.TButton", font=("Helvetica", 23, "bold"), background="#696969", foreground="white",)
        style_extra_info_class.map("Main_67.TButton", background=[("active", "#0000CD")])

        style_extra_info_class.configure("Sec_67.TButton", font=("Helvetica", 13, "bold"), background="#696969", foreground="white",)
        style_extra_info_class.map("Sec_67.TButton", background=[("active", "#4169E1")])

        self.root_PE.configure(bg="#e0dcd4")


        self.teacher = ttk.Label(
            self.root_PE,
            text = "Учитель:",
            style="TLabel",
            anchor="center",
            font=("Helvetica", 26),
            )
        
        self.teacher.place(relx=0, rely=0, relwidth=0.4, relheight=0.33)

        self.options_teacher = []
        self.con.execute("""SELECT * FROM teachers""")
        values = self.con.fetchall() # Проверка на предварительное наличие кабинета 

        for i in values:
            self.options_teacher.append(f"{i[1]} {i[2][0]}. {i[3][0]}.")
        
        self.teacher_in = ttk.Combobox(
            self.root_PE,
            values=self.options_teacher,
            style="TCombobox",
            font=("Helvetica", 25),
            state="readonly", 
            foreground="#4D4D4D"
        )
        self.teacher_in.place(relx=0.4, rely=0, relwidth=0.6, relheight=0.33)
        self.teacher_in.set(self.list_PE[self.idx][0])


        self.day = ttk.Label(
            self.root_PE,
            text = "День:",
            style="TLabel",
            anchor="center",
            font=("Helvetica", 26),
            )
        self.day.place(relx=0, rely=0.33, relwidth=0.2, relheight=0.33)

        self.day_option = ["Понедельник", "Вторник", "Среда", "Четверг", "Пятница", "Суббота"]
        self.days = ttk.Combobox(
            self.root_PE,
            values=self.day_option,
            style="TCombobox",
            font=("Helvetica", 25),
            state="readonly", 
            foreground="#4D4D4D"
        )
        self.days.place(relx=0.2, rely=0.33, relwidth=0.45, relheight=0.33)
        self.days.set(self.list_PE[self.idx][1])


        self.lesson = ttk.Label(
            self.root_PE,
            text = "Урок:",
            style="TLabel",
            anchor="center",
            font=("Helvetica", 26),
            )
        self.lesson.place(relx=0.65, rely=0.33, relwidth=0.2, relheight=0.33)

        self.les = ["1", "2", "3", "4", "5", "6", "7", "8"]
        self.les_option = ttk.Combobox(
            self.root_PE,
            values=self.les,
            style="TCombobox",
            font=("Helvetica", 25),
            state="readonly", 
            foreground="#4D4D4D"
        )
        self.les_option.place(relx=0.85, rely=0.33, relwidth=0.15, relheight=0.33)
        self.les_option.set(self.list_PE[self.idx][2])



        # Футер (кнопки)
        self.save_button = ttk.Button(
            self.root_PE,
            text="Сохранить",
            style="Main_67.TButton",
            command=self.save_data  
        )
        self.save_button.place(relx=0, rely=0.66, relwidth=0.35, relheight=0.34)

        self.cancel_button = ttk.Button(
            self.root_PE,
            text="Отмена",
            style="Main_67.TButton",
            command=self.root_PE.destroy
        )
        self.cancel_button.place(relx=0.35, rely=0.66, relwidth=0.3, relheight=0.34)

        self.delete_button = ttk.Button(
            self.root_PE,
            text="Удалить",
            style="Main_67.TButton",
            command=self.delete_data
        )
        self.delete_button.place(relx=0.65, rely=0.66, relwidth=0.35, relheight=0.34)


    def delete_data(self):
        if self.idx < len(self.parent_menu.data_PE_les):
            self.parent_menu.data_PE_les.pop(self.idx)
        self.parent_menu.display_PE()
        self.root_PE.destroy()
        popup(self.parent_menu.root_extra_info, "Физра удалена!", "Успех").root.mainloop()

    #Сохранение инфы
    def save_data(self):
        new_t = self.teacher_in.get()        # ← teacher combobox, not les_option
        new_d = self.days.get()
        new_l = self.les_option.get()
    
        if self.parent_menu.data_PE_les[self.idx] == [new_t, new_d, new_l]:
            warning = eror_popup(self.root_PE, "Вы ничего не поменяли")
            warning.root.mainloop()
            return

        if new_t == "":
            warning = eror_popup(self.root_PE, "Не выбран учитель")
            warning.root.mainloop()
            return
        
        if new_d == "":
            warning = eror_popup(self.root_PE, "Не выбран день")
            warning.root.mainloop()
            return
    
        if new_l == "":
            warning = eror_popup(self.root_PE, "Не выбран урок")
            warning.root.mainloop()
            return

# ------------------------------
# popup для редактирования/удаления внеурочки
class popup_extra_1:
    def __init__(self, data, parent_menu):
        self.idx, self.list_extra = data
        self.parent_menu = parent_menu
        self.root_ex = Toplevel(parent_menu.root_extra_info)
        self.root_ex.title("Изменить внеурочку")
        self.root_ex.geometry(f"{int(self.root_ex.winfo_screenwidth() * 0.27)}x{int(self.root_ex.winfo_screenheight()*0.25)}")
        self.root_ex.resizable(width=False, height=False)

        style = ttk.Style()
        style.theme_use('clam')
        style.configure("TLabel", font=("Helvetica", 30, "italic"))
        style.configure("TCombobox", fieldbackground="#DCDCDC", arrowsize=0)
        style.map("TCombobox", fieldbackground=[("readonly", "#DCDCDC")])
        style.configure("Main_67.TButton", font=("Helvetica", 23, "bold"), background="#696969", foreground="white")
        style.map("Main_67.TButton", background=[("active", "#0000CD")])

        self.teacher_lbl = ttk.Label(
            self.root_ex,
            text="Учитель:",
            style="TLabel",
            anchor="center",
            font=("Helvetica", 26)
        )
        self.teacher_lbl.place(relx=0, rely=0, relwidth=0.4, relheight=0.5)

        self.teacher_box = ttk.Combobox(
            self.root_ex,
            values=parent_menu.options_teacher,
            style="TCombobox",
            font=("Helvetica", 25),
            state="readonly",
            foreground="#4D4D4D"
        )
        self.teacher_box.place(relx=0.4, rely=0, relwidth=0.6, relheight=0.5)
        self.teacher_box.set(self.list_extra[self.idx])

        self.save_btn = ttk.Button(
            self.root_ex,
            text="Сохранить",
            style="Main_67.TButton",
            command=self.save_data
        )
        self.save_btn.place(relx=0, rely=0.5, relwidth=0.35, relheight=0.5)

        self.cancel_btn = ttk.Button(
            self.root_ex,
            text="Отмена",
            style="Main_67.TButton",
            command=self.root_ex.destroy
        )
        self.cancel_btn.place(relx=0.35, rely=0.5, relwidth=0.3, relheight=0.5)

        self.delete_btn = ttk.Button(
            self.root_ex,
            text="Удалить",
            style="Main_67.TButton",
            command=self.delete_data
        )
        self.delete_btn.place(relx=0.65, rely=0.5, relwidth=0.35, relheight=0.5)

    def save_data(self):
        new_t = self.teacher_box.get()
        if new_t == "":
            warning = eror_popup(self.root_ex, "Учитель не выбран")
            warning.root.mainloop()
            return
        if new_t == self.list_extra[self.idx]:
            warning = eror_popup(self.root_ex, "Ничего не поменяли")
            warning.root.mainloop()
            return
        if new_t in self.list_extra:
            warning = eror_popup(self.root_ex, "Такой учитель уже добавлен")
            warning.root.mainloop()
            return
        self.list_extra[self.idx] = new_t
        self.parent_menu.display_extra()
        popup(self.parent_menu.root_extra_info, "Учитель изменён", "Успех").root.mainloop()
        self.root_ex.destroy()

    def delete_data(self):
        self.list_extra.pop(self.idx)
        self.parent_menu.display_extra()
        self.root_ex.destroy()
        popup(self.parent_menu.root_extra_info, "Учитель удалён", "Успех").root.mainloop()


        

# les_option
# day_option
# teacher_in


#<Изменение урока> - # 
class popup_subject_1:
    def __init__(self, id, clas, parent, base, connect, parent_con):
    #Данные корня
        self.root_subject = Toplevel(parent)
        self.root_subject.geometry(f"{int(self.root_subject.winfo_screenwidth() * 0.22)}x{int(self.root_subject.winfo_screenheight()*0.41)}")
        self.root_subject.resizable(width=False, height=False)
        self.root_subject.title(f"Предмет {clas[0]}")
        self.base_subject = base
        self.con_subject = connect
        self.id = id
        self.clas = clas
        self.parent_con = parent_con
    ###
        style_2_0 = ttk.Style()

        style_2_0.theme_use('clam')

        style_2_0.configure("TLabel", font=("Helvetica", 30, "italic"))

        style_2_0.configure("TEntry", fieldbackground="#DCDCDC")

        style_2_0.configure("TCombobox", fieldbackground="#DCDCDC", arrowsize = 0)
        style_2_0.map("TCombobox", fieldbackground=[("readonly", "#DCDCDC")])


        style_2_0.configure("Main_2_0.TButton", font=("Helvetica", 16, "bold"), background="#696969", foreground="white",)
        style_2_0.map("Main_2_0.TButton", background=[("active", "#0000CD")])

        self.root_subject.configure(bg="#e0dcd4")

    #Виджеты изменения урока +++
    #Блок выбора урока
        self.creating_subject = ttk.Label(
            self.root_subject,
            text = "Урок:",
            style = "TLabel",
            anchor="center"
            )
        self.creating_subject.place(relx=0, rely=0, relwidth=1, relheight=0.14)

        self.options_3 = ["Алгебра",
            "Геометрия",
            "ТеорВер",
            "Физика",
            "Русский",
            "Литература",
            "История",
            "Информатика_1",
            "Информатика_2",
            "География",
            "Английский язык_1",
            "Английский язык_2",
            "Физра",
            "Биология",
            "Химия",
            "Обществознание",
            "ОБЖ",
            "Технология"]
        
        self.subject = ttk.Combobox(
            self.root_subject,
            values=self.options_3,
            style="TCombobox",
            font=("Helvetica", 27),
            state="readonly", 
            foreground="#4D4D4D"
        )
        self.subject.place(relx=0, rely=0.14, relwidth=1, relheight=0.14)
        self.subject.set(self.clas[0])
    ###

    #Блок выбора учителя
        self.options_teacher = []
        self.con_subject.execute("""SELECT * FROM teachers""")
        values = self.con_subject.fetchall() # Проверка на предварительное наличие кабинета 

        for i in values:
            self.options_teacher.append(f"{i[1]} {i[2][0]}. {i[3][0]}.")

        self.creating_teacher =ttk.Label(
            self.root_subject,
            text = "Учитель:",
            style = "TLabel",
            anchor="center"
            )
        self.creating_teacher.place(relx=0, rely=0.28, relwidth=1, relheight=0.14)
        
        self.teacher = ttk.Combobox(
            self.root_subject,
            values=self.options_teacher,
            style="TCombobox",
            font=("Helvetica", 27),
            state="readonly", 
            foreground="#4D4D4D"
        )
        self.teacher.place(relx=0, rely=0.42, relwidth=1, relheight=0.14)
        self.teacher.set(self.clas[1])
    ###

    #Блок выбора количества часов
        self.creating_hours = ttk.Label(
            self.root_subject,
            text = "Кол-во часов:",
            style="TLabel",
            anchor="center"
            )
        self.creating_hours.place(relx=0, rely=0.56, relwidth=1, relheight=0.14)
        
        self.hours = ttk.Entry(
            self.root_subject,
            style="TEntry",
            font=("Helvetica", 27),
            foreground="#4D4D4D"
        )
        self.hours.insert(0, self.clas[2])
        self.hours.place(relx=0, rely=0.7, relwidth=1, relheight=0.14)
    ###

    #Футер
        self.save_button = ttk.Button(
            self.root_subject,
            text="Сохранить",
            style = "Main_2_0.TButton",
            command=self.save_data  
        )
        self.save_button.place(relx=0, rely=0.84, relwidth=0.3, relheight=0.16)

        self.cancel_button = ttk.Button(
            self.root_subject,
            style = "Main_2_0.TButton",
            text="Отмена",
            command=self.root_subject.destroy
        )
        self.cancel_button.place(relx=0.3, rely=0.84, relwidth=0.4, relheight=0.16)

        self.delete_button = ttk.Button(
            self.root_subject,
            text="Удалить",
            style = "Main_2_0.TButton",
            command=self.delete_data
        )
        self.delete_button.place(relx=0.7, rely=0.84, relwidth=0.3, relheight=0.16)
    ###

    #Функуионал
    #Удалие данных
    def delete_data(self):
        lisstt = self.parent_con.full_data_class
        for i in range(0, len(lisstt)):
            if lisstt[i][0] == self.clas[0]: 
                self.parent_con.full_data_class[i] = []
                self.parent_con.display_class()
                self.root_subject.destroy()
                popup(self.parent_con.root_menu_2, "Предмет удален!", "Успех").root.mainloop()
                break
    ###

    #Изменение данных
    def save_data(self):
        # Получаем значения из полей ввода
        new_subject = self.subject.get()
        new_teacher = self.teacher.get()
        new_hours = self.hours.get()

        # Проверка на хоть какое-то изменение
        if ([new_subject, new_teacher, new_hours] == [self.clas[0], self.clas[1], self.clas[2]]):
            warning = eror_popup(self.root_subject, "Ничего не поменяли")
            warning.root.mainloop()
            return

        # Проверка урока
        if new_subject not in self.options_3:
            warning = eror_popup(self.root_subject, "Ошибка в предмете")
            warning.root.mainloop()
            return

        # Проверка на существование такого предмета (кроме текущего)
        for i, subject in enumerate(self.parent_con.full_data_class):
            if i != self.id and subject and subject[0] == new_subject:
                warning = eror_popup(self.root_subject, "Уже есть такой предмет")
                warning.root.mainloop()
                return

        # Проверка учителя
        if new_teacher not in self.options_teacher:
            warning = eror_popup(self.root_subject, "Нет такого учителя")
            warning.root.mainloop()
            return

        # Проверка количества часов
        char = "0123456789"
        if new_hours == "":
            warning = eror_popup(self.root_subject, "Нет кол-ва часов")
            warning.root.mainloop()
            return

        if not all(c in char for c in new_hours):
            warning = eror_popup(self.root_subject, "Ошибка в кол-ве часов")
            warning.root.mainloop()
            return

        if int(new_hours) > 30:
            warning = eror_popup(self.root_subject, "Слишком много часов")
            warning.root.mainloop()
            return

        # Обновляем значение в списке
        self.parent_con.full_data_class[self.id] = [new_subject, new_teacher, new_hours]

    # Обновляем отображение
        self.parent_con.display_class()
        self.root_subject.destroy()
        popup(self.parent_con.root_menu_2, "Предмет обновлен!", "Успех").root.mainloop()
        

    ###
        
    ###






# <СМ пара> - 3 
class menu_check_parallel:
    def __init__(self, parent, base, con):
    #Данные корня
        self.root_menu_3 = Toplevel(parent)
        self.root_menu_3.title("Просмотр параллелей")
        self.root_menu_3.geometry(f"{int(self.root_menu_3.winfo_screenwidth() * 0.56)}x{int(self.root_menu_3.winfo_screenheight()*0.6)}")
        self.root_menu_3.resizable(width=False, height=False) 
        self.con_root_menu_3 = con
        self.base_root_menu_3 = base
    ###
        style_3 = ttk.Style()

        style_3.theme_use('clam')

        style_3.configure("TLabel", font=("Helvetica", 40, "bold", "italic"))

        style_3.configure("Sec_3.TButton", font=("Helvetica", 30, "bold"), background="#696969", foreground="white",)
        style_3.map("Sec_3.TButton", background=[("active", "#4169E1")])

        self.root_menu_3.configure(bg="#e0dcd4")

    #Виджеты - просмотра параллелей +++ 
    #Границы 
        separator_1 = Frame(
            self.root_menu_3, 
            height=3,         
            bg='grey',         
            relief=FLAT,
            bd=1              
        )
        separator_1.place(relx=0, rely=0.5, relwidth=1, relheight=0.007) #Горизонтальная

        separator_1 = Frame(
            self.root_menu_3, 
            height=3,          
            bg='grey',        
            relief=FLAT,
            bd=1               
        )
        separator_1.place(relx=0.5, rely=0, relwidth= 0.004, relheight= 1) #Вертикальная
    ###

    #Лейблы
        self.creating_number = ttk.Label(
                self.root_menu_3,
                text = "8",
                style="TLabel",
                anchor="center"
                )
        self.creating_number.place(relx=0, rely=0, relwidth=0.49, relheight=0.1)

        self.creating_number = ttk.Label(
                self.root_menu_3,
                text = "9",
                style = "TLabel",
                anchor="center"
                )
        self.creating_number.place(relx=0.51, rely=0, relwidth=0.49, relheight=0.1)

        self.creating_number = ttk.Label(
                self.root_menu_3,
                text = "10",
                style = "TLabel",
                anchor="center"
                )
        self.creating_number.place(relx=0, rely=0.51, relwidth=0.49, relheight=0.1)

        self.creating_number = ttk.Label(
                self.root_menu_3,
                text = "11",
                style = "TLabel",
                anchor="center"
                )
        self.creating_number.place(relx=0.51, rely=0.51, relwidth=0.49, relheight=0.1)
    ###

    #Отображение параллелей
        self.display_parallel("8")
        self.display_parallel("9")
        self.display_parallel("10")
        self.display_parallel("11")
    ###

    #Функционал
    #Отображение параллелей
    def display_parallel(self, number):
        # Получаем параллели с указанным номером из БД
        self.con_root_menu_3.execute(
            "SELECT DISTINCT Letter FROM parallels WHERE Number = ? ORDER BY Letter",
            (number,)
        )
        parallels = self.con_root_menu_3.fetchall()
        # Определяем начальные координаты в зависимости от номера параллели
        if number == "8":
            start_y = 0.1
            start_x = 0
        elif number == "9":
            start_y = 0.1
            start_x = 0.51
        elif number == "10":
            start_y = 0.61
            start_x = 0
        else:  # для 11 класса
            start_y = 0.61
            start_x = 0.51

        # Создаем кнопки для каждой параллели
        for i, parallel in enumerate(parallels):
            letter = parallel[0]
            pos_x = start_x + (i % 4) * 0.12  # 4 кнопки в ряд
            pos_y = start_y + (i // 4) * 0.1   # новый ряд каждые 4 кнопки

            data = (number, letter, self.base_root_menu_3, self.con_root_menu_3, self, self.root_menu_3)
            
            btn = ttk.Button(
                self.root_menu_3,
                text=f"{number}{letter}",
                command=lambda data = data: self.open_parallel_info(data),
                style="Sec_3.TButton"
            )
            btn.place(relx=pos_x, rely=pos_y, relwidth=0.11, relheight=0.09)
    ###

    def update_buttons(self):
        for widget in self.root_menu_3.winfo_children():
            if widget.winfo_class() == "TButton":
                widget.destroy()
        self.display_parallel("8")
        self.display_parallel("9")
        self.display_parallel("10")
        self.display_parallel("11")

    #Открытие данных о параллели
    def open_parallel_info(self, data):
        change_parallel(data)
    ###
        
###



#<Изменение параллели> - # +++
class change_parallel:
    def __init__(self, data):
    #Данные корня
        self.root_popup_2 = Toplevel(data[5])
        self.num = data[0]
        self.letter = data[1]
        self.base = data[2]
        self.con = data[3]
        self.parent_con = data[4]
        self.root_popup_2.title(f"Параллель {self.num}{self.letter}")
        self.root_popup_2.geometry(f"{int(self.root_popup_2.winfo_screenwidth() * 0.56)}x{int(self.root_popup_2.winfo_screenheight()*0.6)}")
        self.root_popup_2.resizable(width=False, height=False)
        

        style_3_1 = ttk.Style()

        style_3_1.theme_use('clam')

        style_3_1.configure("TLabel", font=("Helvetica", 30, "italic"))

        style_3_1.configure("TEntry", fieldbackground="#DCDCDC")

        style_3_1.configure("TCombobox", fieldbackground="#DCDCDC", arrowsize = 0)
        style_3_1.map("TCombobox", fieldbackground=[("readonly", "#DCDCDC")])


        style_3_1.configure("Main_2_1.TButton", font=("Helvetica", 30, "bold"), background="#696969", foreground="white",)
        style_3_1.map("Main_2_1.TButton", background=[("active", "#0000CD")])

        style_3_1.configure("Sec_2_1.TButton", font=("Helvetica", 13, "bold"), background="#696969", foreground="white",)
        style_3_1.map("Sec_2_1.TButton", background=[("active", "#4169E1")])

        self.root_popup_2.configure(bg="#e0dcd4")
        # Данные для параллели
        self.full_data_class = []  # для хранения предметов
        self.clas_buttons = []     # для хранения кнопок
        # собственные списки доп. информации
        self.data_PE = []          # список физры
        self.extra_les = []        # список внеурочек (только учителя)
        # варианты учителей для комбобоксов
        self.options_teacher = []
        self.con.execute("""SELECT * FROM teachers""")
        for teacher in self.con.fetchall():
            self.options_teacher.append(f"{teacher[1]} {teacher[2][0]}. {teacher[3][0]}.")
        
        # Получаем данные из БД
        # сначала определим id параллели, чтобы потом подгрузить pe/extra
        self.con.execute(
            "SELECT id FROM parallels WHERE Number = ? AND Letter = ?",
            (self.num, self.letter)
        )
        row = self.con.fetchone()
        parallel_id = row[0] if row else None

        self.con.execute(
           """
            SELECT 
                l.Subject,
                t.Surname || ' ' || substr(t.Name, 1, 1) || '. ' || substr(t.Patrony, 1, 1) || '.' AS Teacher,
                l.Hours
            FROM lessons l
            JOIN parallels p ON l.Id_parallel = p.id
            JOIN teachers t ON l.Id_teacher = t.id
            WHERE p.Number = ? 
              AND p.Letter = ?
            """,
            (self.num, self.letter)
        )
        lessons_data = self.con.fetchall()
        for lesson in lessons_data:
            self.full_data_class.append(list(lesson))
        # загрузка pe и extra в собственные списки
        if parallel_id is not None:
            self.con.execute("SELECT Teacher, Day, Lesson FROM pe WHERE Id_parallel = ?", (parallel_id,))
            for row in self.con.fetchall():
                self.data_PE.append(list(row))
            self.con.execute("SELECT Teacher FROM extra WHERE Id_parallel = ?", (parallel_id,))
            for row in self.con.fetchall():
                self.extra_les.append(row[0])
    ###

    #Виджеты изменения параллели
    #Границы
        separator_1 = Frame(
            self.root_popup_2, 
            height=3,
            bg='grey',
            relief=FLAT,
            bd=1
        )
        separator_1.place(relx=0, rely=0.14, relwidth=1, relheight=0.007)

        separator_2 = Frame(
            self.root_popup_2, 
            height=3,
            bg='grey',
            relief=FLAT,
            bd=1
        )
        separator_2.place(relx=0.29, rely=0.14, relwidth=0.004, relheight=1)

        separator_3 = Frame(
            self.root_popup_2, 
            height=3,          
            bg='grey',        
            relief=FLAT,
            bd=1               
        )
        separator_3.place(relx=0.73, rely=0, relwidth= 0.004, relheight= 0.14) #Вертикальная

    ###

    # Блок номера
        self.creating_number = ttk.Label(
            self.root_popup_2,
            text="Номер:",
            style="TLabel"
        )
        self.creating_number.place(relx=0.1, rely=0, relwidth=0.20, relheight=0.14)

        self.number = ttk.Combobox(
            self.root_popup_2,
            values=["8", "9", "10", "11"],
            style="TCombobox",
            font=("Helvetica", 27),
            state="readonly", 
            foreground="#4D4D4D"
        )
        self.number.set(self.num)
        self.number.place(relx=0.3, rely=0, relwidth=0.08, relheight=0.14)
    ###

    #Блок буквы
        self.creating_letter = ttk.Label(
            self.root_popup_2,
            text="Буква:",
            style="TLabel"
        )
        self.creating_letter.place(relx=0.45, rely=0, relwidth=0.20, relheight=0.14)

        self.letter_box = ttk.Combobox(
            self.root_popup_2,
            values=["А", "Б", "В", "Г", "Д"],
            style="TCombobox",
            font=("Helvetica", 27),
            state="readonly", 
            foreground="#4D4D4D"
        )
        self.letter_box.set(self.letter)
        self.letter_box.place(relx=0.65, rely=0, relwidth=0.08, relheight=0.14)
    ###

    # Блок доп инфы (PE / extra)
        self.add_infa = ttk.Button(
            self.root_popup_2,
            text="Доп инфа",
            style = "Main_2_1.TButton",
            command = self.open_dop_infa
        )
        self.add_infa.place(relx=0.734, rely=0, relwidth=0.266, relheight=0.14)
    ###

    #Блок выбора урока
        self.creating_subject =ttk.Label(
            self.root_popup_2,
            text = "Урок:",
            style="TLabel",
            anchor="center"
            )
        self.creating_subject.place(relx=0, rely=0.15, relwidth=0.29, relheight=0.1)

        self.options_3 = ["Алгебра",
            "Геометрия",
            "ТеорВер",
            "Физика",
            "Русский",
            "Литература",
            "История",
            "Информатика_1",
            "Информатика_2",
            "География",
            "Английский язык_1",
            "Английский язык_2",
            "Физра",
            "Биология",
            "Химия",
            "Обществознание",
            "ОБЖ",
            "Технология"]
        self.subject = ttk.Combobox(
            self.root_popup_2,
            values=self.options_3,
            style="TCombobox",
            font=("Helvetica", 27),
            state="readonly", 
            foreground="#4D4D4D"
        )
        self.subject.place(relx=0, rely=0.25, relwidth=0.29, relheight=0.1)
    ###

    #Блок выбора учителя
        self.options_teacher = []
        self.con.execute("""SELECT * FROM teachers""")
        values = self.con.fetchall() # Проверка на предварительное наличие кабинета 

        for i in values:
            self.options_teacher.append(f"{i[1]} {i[2][0]}. {i[3][0]}.")

        self.creating_teacher = ttk.Label(
            self.root_popup_2,
            text = "Учитель:",
            font=("Helvetica", 30),
            
            anchor="center"
            )
        self.creating_teacher.place(relx=0, rely=0.35, relwidth=0.29, relheight=0.1)
        
        self.teacher = ttk.Combobox(
            self.root_popup_2,
            values=self.options_teacher,
            state="readonly", 
            font=("Helvetica", 30),
        )
        self.teacher.place(relx=0, rely=0.45, relwidth=0.29, relheight=0.1)
    ###

    #Блок выбора количества часов
        self.creating_hours =ttk.Label(
            self.root_popup_2,
            text = "Кол-во часов:",
            style = "TLabel",
            anchor="center"
            )
        self.creating_hours.place(relx=0, rely=0.55, relwidth=0.29, relheight=0.1)
        
        self.hours = ttk.Entry(
            self.root_popup_2,
            style="TEntry",
            font=("Helvetica", 27),
            foreground="#4D4D4D"
        )
        self.hours.place(relx=0, rely=0.65, relwidth=0.29, relheight=0.1)
    ###

    #Футер
    #Кнопка добавления урока 
        self.add_class = ttk.Button(
            self.root_popup_2,
            text="Добавить",
            style = "Main_2_1.TButton",
            command = self.add_class
        )
        self.add_class.place(relx=0, rely=0.75, relwidth=0.29, relheight=0.1)
    ###

    #Кнопка сохранения изменений
        self.save_button = ttk.Button(
            self.root_popup_2,
            text="Сохранить",
            style = "Main_2_1.TButton",
            command=self.save_data
        )
        self.save_button.place(relx=0, rely=0.84, relwidth=0.292, relheight=0.16)
    ###

    #Кнопка отмены
        self.cancel_button = ttk.Button(
            self.root_popup_2,
            text = "Отмена",
            style = "Main_2_1.TButton",
            command= self.root_popup_2.destroy
        )
        self.cancel_button.place(relx=0.292, rely=0.84, relwidth=0.416, relheight=0.16)
    ###

    #Кнопка удаления параллели
        self.delete_parallel = ttk.Button(
            self.root_popup_2,
            text="Удалить",
            style = "Main_2_1.TButton",
            command=self.delete_data
        )
        self.delete_parallel.place(relx=0.708, rely=0.84, relwidth=0.292, relheight=0.16)
    ###

        self.display_class() #Отображаем предметы
    
    # метод открытия окна доп. инфы
    def open_dop_infa(self):
        extra_info_class_2(self, self.root_popup_2, self.con, self.base)
    
    #Функционал
    #Добавление предмета
    def add_class(self):
        mass = []
        sub = self.subject.get()
        teacher = self.teacher.get()
        hours = self.hours.get()

        #
        self.con.execute("""SELECT * FROM parallels WHERE Letter = ? and Number = ?""", (self.letter, self.num, ))
        Id_parallel = self.con.fetchone()[0]

        self.con.execute("""SELECT * FROM lessons WHERE Subject = ? and Id_parallel = ?""", (sub, Id_parallel,))
        if self.con.fetchall():
            warning = eror_popup(self.root_popup_2, "Уже есть такой предмет")
            warning.root.mainloop()
            return

        # Проверка урока
        if sub not in self.options_3:
            warning = eror_popup(self.root_popup_2, "Ошибка в предмете")
            warning.root.mainloop()
            return
        mass.append(sub)

        # Проверка на существование такого предмета
        for i in self.full_data_class:
            if i and i[0] == sub:
                warning = eror_popup(self.root_popup_2, "Уже есть такой предмет")
                warning.root.mainloop()
                return

        # Проверка учителя
        if teacher not in self.options_teacher:
            warning = eror_popup(self.root_popup_2, "Нет такого учителя")
            warning.root.mainloop()
            return
        mass.append(teacher)

        # Проверка количества часов
        char = "0123456789"
        if hours == "":
            warning = eror_popup(self.root_popup_2, "Нет кол-ва часов")
            warning.root.mainloop()
            return

        for i in hours:
            if i not in char:
                warning = eror_popup(self.root_popup_2, "Ошибка в кол-ве часов")
                warning.root.mainloop()
                return

        if int(hours) > 30:
            warning = eror_popup(self.root_popup_2, "Слишком много часов")
            warning.root.mainloop()
            return

        mass.append(hours)
        self.full_data_class.append(mass)
        self.display_class()
        popup(self.root_popup_2, "Предмет добавлен, еще?", "Успех").root.mainloop()

    def save_data(self):
        # Получаем новые значения
        old_num = self.num
        new_num = self.number.get()
        new_letter = self.letter_box.get()
    
        # Проверка номера класса
        if new_num not in ["8", "9", "10", "11"]:
            warning = eror_popup(self.root_popup_2, "Нет такого класса")
            warning.root.mainloop()
            return
    
        # Проверка буквы класса
        if new_letter not in ["А", "Б", "В", "Г", "Д"]:
            warning = eror_popup(self.root_popup_2, "Нет такой буквы")
            warning.root.mainloop()
            return
    
        # Проверка на изменения (номер, буква или предметы)
        has_changes = False
        if new_num != self.num or new_letter != self.letter:
            has_changes = True
        
        #Проверка на 2 групппы
        list_sub = []
        for _ in self.full_data_class:
            if _:
                list_sub.append(_[0])

        if ("Информатика_1" in list_sub and "Информатика_2" not in list_sub) or ("Информатика_1" not in list_sub and "Информатика_2" in list_sub):
            warning = eror_popup(self.root_menu_2, "Созданы не 2 группы")
            warning.root.mainloop()
            return
        
        if ("Английский язык_1" in list_sub and "Английский язык_2" not in list_sub) or ("Английский язык_1" not in list_sub and "Английский язык_2" in list_sub):
            warning = eror_popup(self.root_menu_2, "Созданы не 2 группы")
            warning.root.mainloop()
            return


        # Проверяем изменения в предметах
        self.con.execute(
            """SELECT l.Subject, 
                      (SELECT t.Surname || ' ' || substr(t.Name,1,1) || '. ' || substr(t.Patrony,1,1) || '.'
                       FROM teachers t WHERE t.id = l.Id_teacher) as Teacher,
                      l.Hours
               FROM lessons l
               JOIN parallels p ON l.Id_parallel = p.id 
               WHERE p.Number = ? AND p.Letter = ?""",
            (self.num, self.letter)
        )
        original_lessons = self.con.fetchall()
        original_lessons = [list(lesson) for lesson in original_lessons]

        if sorted(original_lessons) != sorted(self.full_data_class):
            has_changes = True

        if not has_changes:
            warning = eror_popup(self.root_popup_2, "Вы ничего не поменяли")
            warning.root.mainloop()
            return
    
        # Получаем id текущей параллели
        self.con.execute(
            "SELECT id FROM parallels WHERE Number = ? AND Letter = ?",
            (self.num, self.letter)
        )
        parallel_id = self.con.fetchone()[0]
    
        # Обновляем параллель в таблице parallels
        self.con.execute(
            "UPDATE parallels SET Number = ?, Letter = ? WHERE id = ?",
            (new_num, new_letter, parallel_id)
        )
    
        # Удаляем старые уроки
        self.con.execute("DELETE FROM lessons WHERE Id_parallel = ?", (parallel_id,))
        # очистка pe и extra
        self.con.execute("DELETE FROM pe WHERE Id_parallel = ?", (parallel_id,))
        self.con.execute("DELETE FROM extra WHERE Id_parallel = ?", (parallel_id,))
        
        # Добавляем новые уроки
        for subject_data in self.full_data_class:
            if subject_data:  # Проверяем, что данные не пустые
                # Получаем id учителя по ФИО
                teacher_fio = subject_data[1].split()
                surname = teacher_fio[0]
                name = teacher_fio[1][0]
                patrony = teacher_fio[2][0]
    
                self.con.execute(
                    "SELECT id FROM teachers WHERE Surname = ? AND Name LIKE ? AND Patrony LIKE ?",
                    (surname, f"{name}%", f"{patrony}%")
                )
                teacher_id = self.con.fetchone()[0]
    
                # Вставляем данные в таблицу lessons
                self.con.execute(
                    "INSERT INTO lessons (Id_parallel, Subject, Hours, Id_teacher) VALUES (?, ?, ?, ?)",
                    (parallel_id, subject_data[0], subject_data[2], teacher_id)
                )
        # добавляем pe из собственного списка
        for data in self.data_PE:
            if data:
                self.con.execute(
                    "INSERT INTO pe (Id_parallel, Teacher, Day, Lesson) VALUES (?, ?, ?, ?)",
                    (parallel_id, data[0], data[1], data[2])
                )
        # добавляем extra из собственного списка
        for teacher in self.extra_les:
            self.con.execute(
                "INSERT INTO extra (Id_parallel, Teacher) VALUES (?, ?)",
                (parallel_id, teacher)
            )
        self.base.commit()

        self.parent_con.update_buttons()  

        popup(self.root_popup_2, f"Параллель {self.num}{self.letter} изменена!", "Успех").root.mainloop()
        self.root_popup_2.destroy()

    def display_class(self):
        # Удаляем старые кнопки
        for btn in self.clas_buttons:
            btn.destroy()
        self.clas_buttons = []

        # Создаем список только непустых элементов
        active_classes = [clas for clas in self.full_data_class if clas]

        # Размещаем кнопки для каждого предмета
        for i, clas in enumerate(active_classes):
            pos_x = (i % 4) * 0.175 + 0.3
            pos_y = (i // 4) * 0.14 + 0.15

            btn = ttk.Button(
                self.root_popup_2,
                text=f"{clas[0]}",
                command=lambda cl=(i, clas): self.open_popup_class(cl),
                style="Sec_2_1.TButton"
            )
            btn.place(relx=pos_x, rely=pos_y, relwidth=0.175, relheight=0.14)
            self.clas_buttons.append(btn)

    def open_popup_class(self, clas):
        popup_subject_2(clas[0], clas[1], self.root_popup_2, self.base, self.con, self)

    def delete_data(self):
        # Получаем id параллели
        self.con.execute(
            "SELECT id FROM parallels WHERE Number = ? AND Letter = ?",
            (self.num, self.letter)
        )
        parallel_id = self.con.fetchone()[0]

        # Удаляем все уроки этой параллели
        self.con.execute("DELETE FROM lessons WHERE Id_parallel = ?", (parallel_id,))
        
        # Удаляем физру этой параллели
        self.con.execute("DELETE FROM pe WHERE Id_parallel = ?", (parallel_id,))
        
        # Удаляем внеурочки этой параллели
        self.con.execute("DELETE FROM extra WHERE Id_parallel = ?", (parallel_id,))
        
        # Удаляем саму параллель
        self.con.execute("DELETE FROM parallels WHERE id = ?", (parallel_id,))
        
        self.base.commit()

        self.parent_con.update_buttons()  # Обновляем отображение в родительском окне
        self.root_popup_2.destroy()
        popup(self.parent_con.root_menu_3, f"Параллель {self.num}{self.letter} удалена!", "Успех").root.mainloop()


#<Окно доп инфы для change_parallel> - 
class extra_info_class_2:
    def __init__(self, parent, paren_root, con, base):
        self.parent = parent
        self.root_extra_info = Toplevel(paren_root)
        self.con = con
        self.base = base
        self.root_extra_info.title("Доп инфа")
        self.root_extra_info.geometry(f"{int(self.root_extra_info.winfo_screenwidth() * 0.54)}x{int(self.root_extra_info.winfo_screenheight()*0.58)}")
        self.root_extra_info.resizable(width=False, height=False)

        self.data_PE_les = list(self.parent.data_PE)
        self.PE_buttons = []
        # дополнительный список для внеурочек
        self.data_extra = list(getattr(self.parent, 'extra_les', []))
        self.extra_buttons = []
        
        # Очищаем невалидные данные (удаленные учителя)
        self._clean_invalid_teachers()

        style_extra_info_class = ttk.Style()

        style_extra_info_class.theme_use('clam')

        style_extra_info_class.configure("TLabel", font=("Helvetica", 30, "italic"))

        style_extra_info_class.configure("TEntry", fieldbackground="#DCDCDC")

        style_extra_info_class.configure("TCombobox", fieldbackground="#DCDCDC", arrowsize = 0)
        style_extra_info_class.map("TCombobox", fieldbackground=[("readonly", "#DCDCDC")])


        style_extra_info_class.configure("Main_2.TButton", font=("Helvetica", 30, "bold"), background="#696969", foreground="white",)
        style_extra_info_class.map("Main_2.TButton", background=[("active", "#0000CD")])

        style_extra_info_class.configure("Sec_2.TButton", font=("Helvetica", 13, "bold"), background="#696969", foreground="white",)
        style_extra_info_class.map("Sec_2.TButton", background=[("active", "#4169E1")])

        self.root_extra_info.configure(bg="#e0dcd4")

    #Граница
        separator_1 = Frame(
            self.root_extra_info, 
            height=3,          
            bg='grey',        
            relief=FLAT,
            bd=1               
        )
        separator_1.place(relx=0.498, rely=0, relwidth= 0.004, relheight= 0.86) #Вертикальная
    ###

    #Блок физкультур
        self.PE_les = ttk.Label(
            self.root_extra_info,
            text = "Физра:",
            style="TLabel",
            anchor="center"
            )
        self.PE_les.place(relx=0, rely=0, relwidth=0.498, relheight=0.14)

        self.teacher = ttk.Label(
            self.root_extra_info,
            text = "Учитель:",
            style="TLabel",
            anchor="center",
            font=("Helvetica", 26),
            )
        
        self.teacher.place(relx=0, rely=0.14, relwidth=0.2, relheight=0.14)

        self.options_teacher = []
        self.con.execute("""SELECT * FROM teachers""")
        values = self.con.fetchall() 

        for i in values:
            self.options_teacher.append(f"{i[1]} {i[2][0]}. {i[3][0]}.")
        
        self.teacher_in = ttk.Combobox(
            self.root_extra_info,
            values=self.options_teacher,
            style="TCombobox",
            font=("Helvetica", 25),
            state="readonly", 
            foreground="#4D4D4D"
        )
        self.teacher_in.place(relx=0.2, rely=0.14, relwidth=0.298, relheight=0.14)


        self.day = ttk.Label(
            self.root_extra_info,
            text = "День:",
            style="TLabel",
            anchor="center",
            font=("Helvetica", 26),
            )
        self.day.place(relx=0, rely=0.28, relwidth=0.1, relheight=0.14)

        self.day_option = ["Понедельник", "Вторник", "Среда", "Четверг", "Пятница", "Суббота"]
        self.days = ttk.Combobox(
            self.root_extra_info,
            values=self.day_option,
            style="TCombobox",
            font=("Helvetica", 24),
            state="readonly", 
            foreground="#4D4D4D"
        )
        self.days.place(relx=0.1, rely=0.28, relwidth=0.2, relheight=0.14)


        self.lesson = ttk.Label(
            self.root_extra_info,
            text = "Урок:",
            style="TLabel",
            anchor="center",
            font=("Helvetica", 26),
            )
        self.lesson.place(relx=0.3, rely=0.28, relwidth=0.1, relheight=0.14)

        self.les = ["1", "2", "3", "4", "5", "6", "7", "8"]
        self.les_option = ttk.Combobox(
            self.root_extra_info,
            values=self.les,
            style="TCombobox",
            font=("Helvetica", 24),
            state="readonly", 
            foreground="#4D4D4D"
        )
        self.les_option.place(relx=0.4, rely=0.28, relwidth=0.098, relheight=0.14)

        #Кнопка добавления урока 
        self.add_clas = ttk.Button(
            self.root_extra_info,
            text="Добавить",
            style = "Main_2.TButton",
            command = self.add_PE
        )
        self.add_clas.place(relx=0, rely=0.42, relwidth=0.498, relheight=0.1)
    ###

    ###

    #Блок доп уроков
        self.extra_les = ttk.Label(
            self.root_extra_info,
            text = "Внеурочка:",
            style="TLabel",
            anchor="center",
            )
        self.extra_les.place(relx=0.52, rely=0, relwidth=0.498, relheight=0.14)

        self.teacher_e = ttk.Label(
            self.root_extra_info,
            text = "Учитель:",
            style="TLabel",
            anchor="center",
            font=("Helvetica", 26),
            )
        
        self.teacher_e.place(relx=0.52, rely=0.2, relwidth=0.2, relheight=0.14)

        self.options_teacher = []

        for data in self.parent.full_data_class:
            if data:  # Проверяем, что элемент не пустой
                self.options_teacher.append(data[1])
        
        self.teacher_in_1 = ttk.Combobox(
            self.root_extra_info,
            values=self.options_teacher,
            style="TCombobox",
            font=("Helvetica", 25),
            state="readonly", 
            foreground="#4D4D4D"
        )
        self.teacher_in_1.place(relx=0.72, rely=0.2, relwidth=0.28, relheight=0.14)

        # кнопка добавления внеурочки
        self.add_extra_btn = ttk.Button(
            self.root_extra_info,
            text="Добавить",
            style = "Main_2.TButton",
            command = self.add_extra
        )
        self.add_extra_btn.place(relx=0.502, rely=0.42, relwidth=0.498, relheight=0.1)
    ###
    
    #Футер
        self.cancel_1 = ttk.Button(
            self.root_extra_info,
            text="Отмена",
            style = "Main_2.TButton",
            command = self.root_extra_info.destroy
        )
        self.cancel_1.place(relx=0, rely=0.86, relwidth=0.5, relheight=0.14)

        self.save = ttk.Button(
            self.root_extra_info,
            text="Сохранить",
            style = "Main_2.TButton",
            command = self.save_extra_info
        )
        self.save.place(relx=0.5, rely=0.86, relwidth=0.5, relheight=0.14)

        # Очищаем невалидные данные перед отображением
        self._clean_invalid_teachers()
        
        # сразу показываем уже имеющиеся записи, если они есть
        self.display_PE()
        self.display_extra()

    def save_extra_info(self):
        # сохранить изменения в родителе (копируем, чтобы не было алиасов)
        self.parent.data_PE = list(self.data_PE_les)
        self.parent.extra_les = list(self.data_extra)
        
        # Сохраняем в базу данных
        # Получаем parallel_id по номеру и букве класса
        self.con.execute(
            "SELECT id FROM parallels WHERE Number = ? AND Letter = ?",
            (self.parent.num, self.parent.letter)
        )
        row = self.con.fetchone()
        parallel_id = row[0] if row else None
        
        if parallel_id is not None:
            # Удаляем старые записи физры и внеурочки
            self.con.execute("DELETE FROM pe WHERE Id_parallel = ?", (parallel_id,))
            self.con.execute("DELETE FROM extra WHERE Id_parallel = ?", (parallel_id,))
            
            # Добавляем новые записи физры
            for data in self.data_PE_les:
                if data:  # Проверяем, что данные не пустые
                    self.con.execute(
                        "INSERT INTO pe (Id_parallel, Teacher, Day, Lesson) VALUES (?, ?, ?, ?)",
                        (parallel_id, data[0], data[1], data[2])
                    )
            
            # Добавляем новые записи внеурочки
            for teacher in self.data_extra:
                if teacher:  # Проверяем, что данные не пустые
                    self.con.execute(
                        "INSERT INTO extra (Id_parallel, Teacher) VALUES (?, ?)",
                        (parallel_id, teacher)
                    )
            
            # Коммитим изменения
            self.base.commit()
        
        # закроем окно до показа сообщения, чтобы при повторном открытии
        # здесь не осталось старого виджета
        self.root_extra_info.destroy()
        warning = popup(self.parent.root_popup_2, "Доп инфа сохранена!", "Успех")

    #Добавление физры на экран
    def add_PE(self):
        self.t = self.teacher_in.get()
        self.d = self.days.get()
        self.l = self.les_option.get()

        #Проверка на наличие хоть чего-то 
        if self.t == "":
            warning = eror_popup(self.root_extra_info, "Учитель физры не выбран")
            warning.root.mainloop()
            return
        if self.d == "":
            warning = eror_popup(self.root_extra_info, "День физры не выбран")
            warning.root.mainloop()
            return
        if self.l == "":
            warning = eror_popup(self.root_extra_info, "Урок физры не выбран")
            warning.root.mainloop()
            return
        #
        #(учитель, день, урок)
        #Проверка на уже существование урока
        if len(self.data_PE_les) != 0:
            for les in self.data_PE_les:
                if les[1] == self.d and les[2] == self.l:
                    warning = eror_popup(self.root_extra_info, "Есть физра в этом месте")
                    warning.root.mainloop()
                    return
        ###
        if len(self.data_PE_les) == 6:
            warning = eror_popup(self.root_extra_info, "Достигнут лимит кол-ва")
            warning.root.mainloop()
            return
        
        self.data_PE_les.append([self.t, self.d, self.l])
        self.display_PE()
        # Очищаем поля ввода
        self.teacher_in.set("")
        self.days.set("")
        self.les_option.set("")
        warning = popup(self.root_extra_info, "Физра добвлена, еще?", "Успех").root.mainloop()

    def display_PE(self):
        for btn in self.PE_buttons:
            btn.destroy()
        self.PE_buttons = []
        for actual_idx, lesson in enumerate(self.data_PE_les):
            # Пропускаем пустые записи (например, удалённые)
            if not lesson: 
                continue
            pos_x = (actual_idx % 2) * 0.249
            pos_y = (actual_idx // 2) * 0.11 + 0.52
            btn = ttk.Button(
                self.root_extra_info,
                text=f"{lesson[1]}", 
                command=lambda idx=actual_idx, l=self.data_PE_les: self.open_popup_PE(idx, l),
                style = "Sec_2.TButton"
            )
            btn.place(relx=pos_x, rely=pos_y, relwidth=0.249, relheight=0.11)
            self.PE_buttons.append(btn)
    
    def add_extra(self):
        teacher = self.teacher_in_1.get()
        if teacher == "":
            warning = eror_popup(self.root_extra_info, "Учитель не выбран")
            warning.root.mainloop()
            return
        if teacher in self.data_extra:
            warning = eror_popup(self.root_extra_info, "Учитель уже добавлен")
            warning.root.mainloop()
            return
        self.data_extra.append(teacher)
        self.display_extra()
        self.teacher_in_1.set("")
        warning = popup(self.root_extra_info, "Учитель добавлен, еще?", "Успех").root.mainloop()
    
    def display_extra(self):
        for btn in self.extra_buttons:
            btn.destroy()
        self.extra_buttons = []
        for actual_idx, teacher in enumerate(self.data_extra):
            if not teacher:
                continue
            pos_x = 0.502 + (actual_idx % 2) * 0.249
            pos_y = (actual_idx // 2) * 0.11 + 0.52
            btn = ttk.Button(
                self.root_extra_info,
                text=f"{teacher}",
                command=lambda idx=actual_idx: self.open_popup_extra(idx),
                style="Sec_2.TButton"
            )
            btn.place(relx=pos_x, rely=pos_y, relwidth=0.249, relheight=0.11)
            self.extra_buttons.append(btn)
    
    def open_popup_extra(self, idx):
        popup_extra_2((idx, self.data_extra), self)

    def open_popup_PE(self, idx, l):
        popup_PE_2((idx, l), self)

    def _clean_invalid_teachers(self):
        """Удаляет из списков данные с удаленными учителями"""
        # Получаем актуальный список учителей из БД
        self.con.execute("SELECT Surname, Name, Patrony FROM teachers")
        valid_teachers = []
        for row in self.con.fetchall():
            teacher_fio = f"{row[0]} {row[1][0]}. {row[2][0]}."
            valid_teachers.append(teacher_fio)
        
        # Очищаем данные физры от невалидных учителей
        self.data_PE_les[:] = [
            lesson for lesson in self.data_PE_les 
            if not lesson or lesson[0] in valid_teachers
        ]
        
        # Очищаем данные внеурочки от невалидных учителей
        self.data_extra[:] = [
            teacher for teacher in self.data_extra 
            if not teacher or teacher in valid_teachers
        ]


#Окно изменения физры для change_parallel
class popup_PE_2:
    def __init__(self, data, parent_menu):
        self.con = parent_menu.con
        self.idx, self.list_PE = data
        self.parent_menu = parent_menu
        self.root_PE = Toplevel(parent_menu.root_extra_info)
        self.root_PE.title("Изменить исключение")
        self.root_PE.geometry(f"{int(self.root_PE.winfo_screenwidth() * 0.27)}x{int(self.root_PE.winfo_screenheight()*0.3)}")
        self.root_PE.resizable(width=False, height=False)

        style_extra_info_class = ttk.Style()

        style_extra_info_class.theme_use('clam')

        style_extra_info_class.configure("TLabel", font=("Helvetica", 30, "italic"))

        style_extra_info_class.configure("TEntry", fieldbackground="#DCDCDC")

        style_extra_info_class.configure("TCombobox", fieldbackground="#DCDCDC", arrowsize = 0)
        style_extra_info_class.map("TCombobox", fieldbackground=[("readonly", "#DCDCDC")])


        style_extra_info_class.configure("Main_67.TButton", font=("Helvetica", 23, "bold"), background="#696969", foreground="white",)
        style_extra_info_class.map("Main_67.TButton", background=[("active", "#0000CD")])

        style_extra_info_class.configure("Sec_67.TButton", font=("Helvetica", 13, "bold"), background="#696969", foreground="white",)
        style_extra_info_class.map("Sec_67.TButton", background=[("active", "#4169E1")])

        self.root_PE.configure(bg="#e0dcd4")


        self.teacher = ttk.Label(
            self.root_PE,
            text = "Учитель:",
            style="TLabel",
            anchor="center",
            font=("Helvetica", 26),
            )
        
        self.teacher.place(relx=0, rely=0, relwidth=0.4, relheight=0.33)

        self.options_teacher = []
        self.con.execute("""SELECT * FROM teachers""")
        values = self.con.fetchall() 

        for i in values:
            self.options_teacher.append(f"{i[1]} {i[2][0]}. {i[3][0]}.")
        
        self.teacher_in = ttk.Combobox(
            self.root_PE,
            values=self.options_teacher,
            style="TCombobox",
            font=("Helvetica", 25),
            state="readonly", 
            foreground="#4D4D4D"
        )
        self.teacher_in.place(relx=0.4, rely=0, relwidth=0.6, relheight=0.33)
        self.teacher_in.set(self.list_PE[self.idx][0])


        self.day = ttk.Label(
            self.root_PE,
            text = "День:",
            style="TLabel",
            anchor="center",
            font=("Helvetica", 26),
            )
        self.day.place(relx=0, rely=0.33, relwidth=0.2, relheight=0.33)

        self.day_option = ["Понедельник", "Вторник", "Среда", "Четверг", "Пятница", "Суббота"]
        self.days = ttk.Combobox(
            self.root_PE,
            values=self.day_option,
            style="TCombobox",
            font=("Helvetica", 25),
            state="readonly", 
            foreground="#4D4D4D"
        )
        self.days.place(relx=0.2, rely=0.33, relwidth=0.45, relheight=0.33)
        self.days.set(self.list_PE[self.idx][1])


        self.lesson = ttk.Label(
            self.root_PE,
            text = "Урок:",
            style="TLabel",
            anchor="center",
            font=("Helvetica", 26),
            )
        self.lesson.place(relx=0.65, rely=0.33, relwidth=0.2, relheight=0.33)

        self.les = ["1", "2", "3", "4", "5", "6", "7", "8"]
        self.les_option = ttk.Combobox(
            self.root_PE,
            values=self.les,
            style="TCombobox",
            font=("Helvetica", 25),
            state="readonly", 
            foreground="#4D4D4D"
        )
        self.les_option.place(relx=0.85, rely=0.33, relwidth=0.15, relheight=0.33)
        self.les_option.set(self.list_PE[self.idx][2])



        # Футер (кнопки)
        self.save_button = ttk.Button(
            self.root_PE,
            text="Сохранить",
            style="Main_67.TButton",
            command=self.save_data  
        )
        self.save_button.place(relx=0, rely=0.66, relwidth=0.35, relheight=0.34)

        self.cancel_button = ttk.Button(
            self.root_PE,
            text="Отмена",
            style="Main_67.TButton",
            command=self.root_PE.destroy
        )
        self.cancel_button.place(relx=0.35, rely=0.66, relwidth=0.3, relheight=0.34)

        self.delete_button = ttk.Button(
            self.root_PE,
            text="Удалить",
            style="Main_67.TButton",
            command=self.delete_data
        )
        self.delete_button.place(relx=0.65, rely=0.66, relwidth=0.35, relheight=0.34)


    def delete_data(self):
        if self.idx < len(self.parent_menu.data_PE_les):
            self.parent_menu.data_PE_les.pop(self.idx)
        self.parent_menu.display_PE()
        self.root_PE.destroy()
        popup(self.parent_menu.root_extra_info, "Физра удалена!", "Успех").root.mainloop()

    #Сохранение инфы
    def save_data(self):
        # Проверка правильности индекса
        if self.idx >= len(self.parent_menu.data_PE_les):
            warning = eror_popup(self.root_PE, "Ошибка индекса")
            warning.root.mainloop()
            self.root_PE.destroy()
            return

        new_t = self.teacher_in.get()        
        new_d = self.days.get()
        new_l = self.les_option.get()
    
        if self.parent_menu.data_PE_les[self.idx] == [new_t, new_d, new_l]:
            warning = eror_popup(self.root_PE, "Вы ничего не поменяли")
            warning.root.mainloop()
            return

        if new_t == "":
            warning = eror_popup(self.root_PE, "Не выбран учитель")
            warning.root.mainloop()
            return
        
        if new_d == "":
            warning = eror_popup(self.root_PE, "Не выбран день")
            warning.root.mainloop()
            return
    
        if new_l == "":
            warning = eror_popup(self.root_PE, "Не выбран урок")
            warning.root.mainloop()
            return

        # Обновляем данные
        self.parent_menu.data_PE_les[self.idx] = [new_t, new_d, new_l]
        self.parent_menu.display_PE()
        popup(self.parent_menu.root_extra_info, "Физра изменена!", "Успех").root.mainloop()
        self.root_PE.destroy()


#Окно изменения внеурочки для change_parallel
class popup_extra_2:
    def __init__(self, data, parent_menu):
        self.idx, self.list_extra = data
        self.parent_menu = parent_menu
        self.root_ex = Toplevel(parent_menu.root_extra_info)
        self.root_ex.title("Изменить внеурочку")
        self.root_ex.geometry(f"{int(self.root_ex.winfo_screenwidth() * 0.27)}x{int(self.root_ex.winfo_screenheight()*0.25)}")
        self.root_ex.resizable(width=False, height=False)

        style = ttk.Style()
        style.theme_use('clam')
        style.configure("TLabel", font=("Helvetica", 30, "italic"))
        style.configure("TCombobox", fieldbackground="#DCDCDC", arrowsize=0)
        style.map("TCombobox", fieldbackground=[("readonly", "#DCDCDC")])
        style.configure("Main_67.TButton", font=("Helvetica", 23, "bold"), background="#696969", foreground="white")
        style.map("Main_67.TButton", background=[("active", "#0000CD")])

        self.teacher_lbl = ttk.Label(
            self.root_ex,
            text="Учитель:",
            style="TLabel",
            anchor="center",
            font=("Helvetica", 26)
        )
        self.teacher_lbl.place(relx=0, rely=0, relwidth=0.4, relheight=0.5)

        self.teacher_box = ttk.Combobox(
            self.root_ex,
            values=parent_menu.options_teacher,
            style="TCombobox",
            font=("Helvetica", 25),
            state="readonly",
            foreground="#4D4D4D"
        )
        self.teacher_box.place(relx=0.4, rely=0, relwidth=0.6, relheight=0.5)
        self.teacher_box.set(self.list_extra[self.idx])

        self.save_btn = ttk.Button(
            self.root_ex,
            text="Сохранить",
            style="Main_67.TButton",
            command=self.save_data
        )
        self.save_btn.place(relx=0, rely=0.5, relwidth=0.35, relheight=0.5)

        self.cancel_btn = ttk.Button(
            self.root_ex,
            text="Отмена",
            style="Main_67.TButton",
            command=self.root_ex.destroy
        )
        self.cancel_btn.place(relx=0.35, rely=0.5, relwidth=0.3, relheight=0.5)

        self.delete_btn = ttk.Button(
            self.root_ex,
            text="Удалить",
            style="Main_67.TButton",
            command=self.delete_data
        )
        self.delete_btn.place(relx=0.65, rely=0.5, relwidth=0.35, relheight=0.5)

    def save_data(self):
        # Проверка правильности индекса
        if self.idx >= len(self.parent_menu.data_extra):
            warning = eror_popup(self.root_ex, "Ошибка индекса")
            warning.root.mainloop()
            self.root_ex.destroy()
            return
        
        new_t = self.teacher_box.get()
        if new_t == "":
            warning = eror_popup(self.root_ex, "Учитель не выбран")
            warning.root.mainloop()
            return
        if new_t == self.parent_menu.data_extra[self.idx]:
            warning = eror_popup(self.root_ex, "Ничего не поменяли")
            warning.root.mainloop()
            return
        if new_t in self.parent_menu.data_extra:
            warning = eror_popup(self.root_ex, "Такой учитель уже добавлен")
            warning.root.mainloop()
            return
        self.parent_menu.data_extra[self.idx] = new_t
        self.parent_menu.display_extra()
        popup(self.parent_menu.root_extra_info, "Учитель изменён", "Успех").root.mainloop()
        self.root_ex.destroy()

    def delete_data(self):
        if self.idx < len(self.parent_menu.data_extra):
            self.parent_menu.data_extra.pop(self.idx)
        self.parent_menu.display_extra()
        self.root_ex.destroy()
        popup(self.parent_menu.root_extra_info, "Учитель удалён", "Успех").root.mainloop()


#<Изменение предмета> - # +++
class popup_subject_2:
    def __init__(self, id, clas, parent, base, connect, parent_con):
        #Данные корня
        self.root_subject = Toplevel(parent)
        self.root_subject.geometry(f"{int(self.root_subject.winfo_screenwidth() * 0.22)}x{int(self.root_subject.winfo_screenheight()*0.41)}")
        self.root_subject.resizable(width=False, height=False)
        self.root_subject.title(f"Предмет {clas[0]}")
        self.base_subject = base
        self.con_subject = connect
        self.id = id
        self.clas = clas
        self.parent_con = parent_con

        style_3_2 = ttk.Style()

        style_3_2.theme_use('clam')

        style_3_2.configure("TLabel", font=("Helvetica", 30, "italic"))

        style_3_2.configure("TEntry", fieldbackground="#DCDCDC")

        style_3_2.configure("TCombobox", fieldbackground="#DCDCDC", arrowsize = 0)
        style_3_2.map("TCombobox", fieldbackground=[("readonly", "#DCDCDC")])


        style_3_2.configure("Main_3_2.TButton", font=("Helvetica", 15, "bold"), background="#696969", foreground="white",)

        style_3_2.map("Main_3_2.TButton", background=[("active", "#0000CD")])

        self.root_subject.configure(bg="#e0dcd4")

        #Виджеты изменения урока
        # Блок выбора урока
        self.creating_subject = ttk.Label(
            self.root_subject,
            text="Урок:",
            style="TLabel",
            anchor="center"
        )
        self.creating_subject.place(relx=0, rely=0, relwidth=1, relheight=0.14)

        self.options_3 = ["Алгебра",
            "Геометрия",
            "ТеорВер",
            "Физика",
            "Русский",
            "Литература",
            "История",
            "Информатика_1",
            "Информатика_2",
            "География",
            "Английский язык_1",
            "Английский язык_2",
            "Физра",
            "Биология",
            "Химия",
            "Обществознание",
            "ОБЖ",
            "Технология"]
        
        self.subject = ttk.Combobox(
            self.root_subject,
            values=self.options_3,
            style="TCombobox",
            font=("Helvetica", 27),
            state="readonly", 
            foreground="#4D4D4D"
        )
        self.subject.place(relx=0, rely=0.14, relwidth=1, relheight=0.14)
        self.subject.set(self.clas[0])

        # Блок выбора учителя
        self.options_teacher = []
        self.con_subject.execute("""SELECT * FROM teachers""")
        values = self.con_subject.fetchall()

        for i in values:
            self.options_teacher.append(f"{i[1]} {i[2][0]}. {i[3][0]}.")

        self.creating_teacher = ttk.Label(
            self.root_subject,
            text="Учитель:",
            style="TLabel",
            anchor="center"
        )
        self.creating_teacher.place(relx=0, rely=0.28, relwidth=1, relheight=0.14)
        
        self.teacher = ttk.Combobox(
            self.root_subject,
            values=self.options_teacher,
            style="TCombobox",
            font=("Helvetica", 27),
            state="readonly", 
            foreground="#4D4D4D"
        )
        self.teacher.place(relx=0, rely=0.42, relwidth=1, relheight=0.14)
        self.teacher.set(self.clas[1])

        # Блок выбора количества часов
        self.creating_hours = ttk.Label(
            self.root_subject,
            text="Кол-во часов:",
            style="TLabel",
            anchor="center"
        )
        self.creating_hours.place(relx=0, rely=0.56, relwidth=1, relheight=0.14)
        
        self.hours = ttk.Entry(
            self.root_subject,
            style="TEntry",
            font=("Helvetica", 27),
            foreground="#4D4D4D"

        )
        self.hours.insert(0, self.clas[2])
        self.hours.place(relx=0, rely=0.7, relwidth=1, relheight=0.14)

        # Футер
        self.save_button = ttk.Button(
            self.root_subject,
            text="Сохранить",
            command=self.save_data,
            style = "Main_3_2.TButton"
        )
        self.save_button.place(relx=0, rely=0.84, relwidth=0.3, relheight=0.16)

        self.cancel_button = ttk.Button(
            self.root_subject,
            text="Отмена",
            style = "Main_3_2.TButton",
            command=self.root_subject.destroy
        )
        self.cancel_button.place(relx=0.3, rely=0.84, relwidth=0.4, relheight=0.16)

        self.delete_button = ttk.Button(
            self.root_subject,
            text="Удалить",
            style = "Main_3_2.TButton",
            command=self.delete_data
        )
        self.delete_button.place(relx=0.7, rely=0.84, relwidth=0.3, relheight=0.16)

    #Функционал
    def delete_data(self):
        if self.id < len(self.parent_con.full_data_class):
            self.parent_con.full_data_class[self.id] = []
            self.parent_con.display_class()
            self.root_subject.destroy()
            popup(self.parent_con.root_popup_2, "Предмет удален!", "Успех").root.mainloop()

    def save_data(self):
        new_subject = self.subject.get()
        new_teacher = self.teacher.get()
        new_hours = self.hours.get()

        # Проверка на изменения
        if [new_subject, new_teacher, new_hours] == [self.clas[0], self.clas[1], self.clas[2]]:
            warning = eror_popup(self.root_subject, "Ничего не поменяли")
            warning.root.mainloop()
            return

        # Проверка предмета
        if new_subject not in self.options_3:
            warning = eror_popup(self.root_subject, "Ошибка в предмете")
            warning.root.mainloop()
            return

        # Проверка на существование такого предмета
        for i, subject in enumerate(self.parent_con.full_data_class):
            if i != self.id and subject and subject[0] == new_subject:
                warning = eror_popup(self.root_subject, "Уже есть такой предмет")
                warning.root.mainloop()
                return

        # Проверка учителя
        if new_teacher not in self.options_teacher:
            warning = eror_popup(self.root_subject, "Нет такого учителя")
            warning.root.mainloop()
            return

        # Проверка часов
        char = "0123456789"
        if new_hours == "":
            warning = eror_popup(self.root_subject, "Нет кол-ва часов")
            warning.root.mainloop()
            return

        if not all(c in char for c in new_hours):
            warning = eror_popup(self.root_subject, "Ошибка в кол-ве часов")
            warning.root.mainloop()
            return

        if int(new_hours) > 30:
            warning = eror_popup(self.root_subject, "Слишком много часов")
            warning.root.mainloop()
            return

        # Обновляем значение в списке
        self.parent_con.full_data_class[self.id] = [new_subject, new_teacher, new_hours]
        
        # Обновляем отображение
        self.parent_con.display_class()
        popup(self.root_subject, "Предмет обновлен!", "Успех").root.mainloop()
        self.root_subject.destroy()






#<Созд каба> - 4 +++
class menu_add_room:
    def __init__(self, parent, base, connect):
    #Данные корня
        self.base_menu_4 = base
        self.con_menu_4 = connect
        self.root_menu_4 = Toplevel(parent)
        self.root_menu_4.title("Добавление кабинета")
        self.root_menu_4.geometry(f"{int(self.root_menu_4.winfo_screenwidth() * 0.56)}x{int(self.root_menu_4.winfo_screenheight()*0.6)}")
        self.root_menu_4.resizable(width=False, height=False)
        self.tup_base_con = (self.base_menu_4, self.con_menu_4)
    ###
    
        style_4 = ttk.Style()

        style_4.theme_use('clam')

        style_4.configure("TLabel", font=("Helvetica", 30, "italic"))

        style_4.configure("TEntry", fieldbackground="#DCDCDC")

        style_4.configure("TCombobox", fieldbackground="#DCDCDC", arrowsize = 0)
        style_4.map("TCombobox", fieldbackground=[("readonly", "#DCDCDC")])


        style_4.configure("Main_4.TButton", font=("Helvetica", 30, "bold"), background="#696969", foreground="white",)
        style_4.map("Main_4.TButton", background=[("active", "#9932CC")])

        self.root_menu_4.configure(bg="#e0dcd4")

    #Виджеты добавления кабинета ++
    # Границы вертикальная
        separator_1 = Frame(
            self.root_menu_4, 
            height=3,           
            bg='grey',         
            relief=FLAT,
            bd=1               
        )
        separator_1.place(relx=0, rely=0.45, relwidth=1, relheight=0.007)

        # Блок номера кабинета
        self.number_room_label = ttk.Label(
            self.root_menu_4,
            text="Номер кабинета:",
            style="TLabel",
            anchor="center"

        )
        
        self.number_room_label.place(relx=0, rely=0, relwidth=0.5, relheight=0.15)

        self.input_number_room = ttk.Entry(
            self.root_menu_4,
            style = "TEntry",
            font=("Helvetica", 27),
            foreground="#4D4D4D"
        )
        self.input_number_room.place(relx=0.5, rely=0, relwidth=0.5, relheight=0.15)

        # Блок графский или нет
        options = ["Графский", "Лицей"]
        self.graphskiy_or_not_label = ttk.Label(
            self.root_menu_4,
            text="Графский или лицей:",
            style="TLabel",
            anchor="center"
        )
        self.graphskiy_or_not_label.place(relx=0, rely=0.15, relwidth=0.5, relheight=0.15)

        self.graphskiy_or_not = ttk.Combobox(
            self.root_menu_4,
            values= options,
            style="TCombobox",
            font=("Helvetica", 27),
            state="readonly", 
            foreground="#4D4D4D"
              
        )
        self.graphskiy_or_not.place(relx=0.5, rely=0.15, relwidth=0.5, relheight=0.15)
        self.graphskiy_or_not.set("Графский или нет")

        
        # Блок большой/маленький
        options = ["Маленький", "Большой"]
        self.small_or_big_label = ttk.Label(
            self.root_menu_4,
            text="Маленький или большой:",
            style="TLabel",
            anchor="center"
        )
        self.small_or_big_label.place(relx=0, rely=0.3, relwidth=0.5, relheight=0.15)

        self.small_or_big = ttk.Combobox(
            self.root_menu_4,
            values=options,
            style="TCombobox",
            font=("Helvetica", 27),
            state="readonly", 
            foreground="#4D4D4D"
        )
        self.small_or_big.place(relx=0.5, rely=0.3, relwidth=0.5, relheight=0.15)
        self.small_or_big.set("Маленький или большой каб.")
        # Блок выбора приоритетного кабинета
        
        self.subjects = [
            "Нет приоритетных кабинетов",
            "Алгебра/Геометрия/ТеорВер",
            "Физика",
            "Русский/Литература",
            "История",
            "Информатика",
            "География",
            "Английский язык",
            "Физра",
            "Биология",
            "Химия",
            "Обществознание"
        ]

        self.prio_subjects_label = ttk.Label(
            self.root_menu_4,
            text="Приоритетные уроки",
            style="TLabel",
            anchor="center"
        )
        self.prio_subjects_label.place(relx=0, rely=0.457, relwidth=1, relheight=0.097)

        self.prio_subjects = Listbox(
            self.root_menu_4,
            selectmode=SINGLE,
            background="#DCDCDC",
            selectbackground="#4b6985", 
            font=("Helvetica", 27), 
            foreground="#4D4D4D", 
            relief="groove",
            bd=2
        )
        self.prio_subjects.place(relx=0.2, rely=0.55, relwidth=0.6, relheight=0.3)

        for item in self.subjects:
            self.prio_subjects.insert(END, item)

        # Кнопка добавления кабинета
        self.add_button = ttk.Button(
            self.root_menu_4,
            text="Добавить кабинет",
            style = "Main_4.TButton",
            command=lambda base_con=self.tup_base_con: self.add_room(base_con)
        )
        self.add_button.place(relx=0, rely=0.85, relwidth=1, relheight=0.15)
    ###

#Функционал
#Вв данных и пров
    def add_room(self, base_plus_con):
    #Вспомогательные штучки
        data_room = []
        self.base_add_room = base_plus_con[0]
        self.con_add_room = base_plus_con[1]
    ###

    #Проверка номера кабинета + ввод в data_room
        all_num = "1234567890"
        room_number = str(self.input_number_room.get())
        count = 0 
        for digit in room_number:
            if digit in all_num:
                count += 1

        if (count == len(room_number)) and (room_number != ""): #Проверка на наличие числа и проверка на ввод данных 
            if int(room_number) // 100 < 6: #Проверка на наличие этажей между 1 - 5
                data_room.append(str(room_number))
            else:
                warning = eror_popup(self.root_menu_4, "Ошибка в номере кабинета")
                warning.root.mainloop()
                return
        else:
            warning = eror_popup(self.root_menu_4, "Ошибка в номере кабинета")
            warning.root.mainloop()
            return
    ###

    #Проверка графского или лицея + ввод
        graphskiy_or_not = self.graphskiy_or_not.get()
        if graphskiy_or_not == "Графский" or graphskiy_or_not == "Лицей": #Проверка на привльности написания ответа
            data_room.append(graphskiy_or_not)
        else:
            warning = eror_popup(self.root_menu_4, "Ошибка в Графский/Лицей")
            warning.root.mainloop()
            return
    ###

    #Проверка большого или маленького кабинета
        small_or_big = self.small_or_big.get()
        if small_or_big == "Маленький" or small_or_big == "Большой":
            data_room.append(small_or_big)
        else:                                       
            warning = eror_popup(self.root_menu_4, "Ошибка в большой/мал.")
            warning.root.mainloop()
            return
    ###

    #Проверка приоклассов + ввод
        selected_indices = self.prio_subjects.curselection()

        if selected_indices == tuple():
            warning = eror_popup(self.root_menu_4, "Ошибка в прио. уроках")
            warning.root.mainloop()
            return
        else:
            data_room.append(self.subjects[selected_indices[0]])
    ###

    #Самая финальная проверка на наличие в таблице + Ввод в таблицу + обновление страницы
        self.con_menu_4.execute("SELECT * FROM rooms WHERE Graph_or_Lyceum = ?", (data_room[1],))
        values = self.con_menu_4.fetchall()  # Проверка на предварительное наличие кабинета 
        all_num_of_rooms = []
        for i in values:
            all_num_of_rooms.append(i[1])

        if data_room[0] not in all_num_of_rooms:
            self.con_menu_4.execute("INSERT INTO rooms (Number, Graph_or_Lyceum, Big_or_Small, Subject) VALUES (?, ?, ?, ?)",
                                    (data_room[0], data_room[1], data_room[2], data_room[3]))
            self.base_menu_4.commit()
        else:
            warning = eror_popup(self.root_menu_4, "Есть кабинет в корпусе")
            warning.root.mainloop()
            return
    ###

    #Проверка на максималное число на экране
        if len(values) + 1 > 36:
            warning = eror_popup(self.root_menu_4, "Лимит классов")
            warning.root.mainloop()
            return
    ###

        self.reset_form()
        warning = popup(self.root_menu_4, f"Кабинет {room_number} создан!", "Успех")
        warning.root.mainloop()

###

#Сброс всех полей ввода
    def reset_form(self):
        
        self.input_number_room.delete(0, END)
        self.graphskiy_or_not.set('Графский или лицей')
        self.small_or_big.set('Маленький или большой каб.')
        self.prio_subjects.selection_clear(0, END)
###
     





#<См каба> - 5 +++
class menu_check_room:
    def __init__(self, parent, base, connect):
    #Данные корня
        self.root_menu_5 = Toplevel(parent)
        self.root_menu_5.title("Просмотр кабинетов")
        self.root_menu_5.geometry(f"{int(self.root_menu_5.winfo_screenwidth() * 0.56)}x{int(self.root_menu_5.winfo_screenheight()*0.6)}")
        self.root_menu_5.resizable(width=False, height=False) 
        self.room_menu_5_base = base
        self.room_menu_5_con = connect
    ###


        style_5 = ttk.Style()

        style_5.theme_use('clam')

        style_5.configure("TLabel", font=("Helvetica", 30, "italic"))

        style_5.configure("TEntry", fieldbackground="#DCDCDC")

        style_5.configure("TCombobox", fieldbackground="#DCDCDC", arrowsize = 0)
        style_5.map("TCombobox", fieldbackground=[("readonly", "#DCDCDC")])

        style_5.configure("Sec_5.TButton", font=("Helvetica", 25, "bold"), background="#696969", foreground="white",)
        style_5.map("Sec_5.TButton", background=[("active", "#BA55D3")])

        self.root_menu_5.configure(bg="#e0dcd4")

    #Виджеты просмотра кабинетов ++
    #Границы
        separator_1 = Frame(
            self.root_menu_5, 
            height=3,           # Увеличиваем высоту
            bg='grey',         # Задаем цвет
            relief=FLAT,
            bd=1               # Увеличиваем толщину бордюра
        )
        separator_1.place(relx=0.5, rely=0, relwidth=0.004, relheight= 1)
    ###

        self.ne_graphskiy = ttk.Label(
            self.root_menu_5,
            text="Графский",
            style="TLabel",
            anchor="center"
        )
        self.ne_graphskiy.place(relx=0.1, rely=0, relwidth=0.3, relheight=0.09)

        self.input_search_g = ttk.Entry(
            self.root_menu_5,
            style="TEntry",
            font=("Helvetica", 25),
            foreground="#4D4D4D"
        )
        self.input_search_g.place(relx=0, rely=0, relwidth=0.1, relheight=0.09)
        self.input_search_g.bind("<KeyRelease>", lambda e, b="Графский": self.on_search_change(e, b))


    
        self.graphskiy = ttk.Label(
            self.root_menu_5,
            text="Лицей",
            style="TLabel",
            anchor="center"
        )
        self.graphskiy.place(relx=0.604, rely=0, relwidth=0.296, relheight=0.09)
        self.input_search_l = ttk.Entry(
            self.root_menu_5,
            style="TEntry",
            font=("Helvetica", 25),
            foreground="#4D4D4D"
        )
        self.input_search_l.place(relx=0.9, rely=0, relwidth=0.1, relheight=0.09)
        self.input_search_l.bind("<KeyRelease>", lambda e, b="Лицей": self.on_search_change(e, b))

        self.search_buttons = {
            "Графский": [],
            "Лицей": []
        }
        self.all_rooms = {
            "Графский": [],
            "Лицей": []
        }

        self.display_buttons("Графский")
        self.display_buttons("Лицей")

    def display_buttons(self, answer):
        self.answ = answer
        self.room_menu_5_con.execute("SELECT DISTINCT Number FROM rooms WHERE Graph_or_Lyceum = ?", (self.answ,))
        self.all_rooms[self.answ] = self.room_menu_5_con.fetchall()

        wid_x = 0.11  # уменьшаем ширину кнопки для 4 кабинетов
        hei_y = 0.1   # высота остается прежней

        if self.answ == "Графский":
            start_x = 0.02  # начальная позиция для Графского
            positions = [0.02, 0.14, 0.26, 0.38]  # 4 позиции для кнопок
        else:
            start_x = 0.52  # начальная позиция для Лицея
            positions = [0.52, 0.64, 0.76, 0.88]  # 4 позиции для кнопок

        for i, room in enumerate(self.all_rooms[self.answ]):
            pos_x = positions[i % 4]  # позиция по x (чередуем 4 позиции)
            pos_y = 0.1 + (i // 4) * 0.1  # позиция по y (новая строка каждые 4 кабинета)

            room_button = ttk.Button(
                self.root_menu_5, 
                text=room[0], 
                command=lambda num=room[0], ans=self.answ: self.open_room_screen(num, ans),
                style = "Sec_5.TButton",
            )
            room_button.place(relx=pos_x, rely=pos_y, relwidth=wid_x, relheight=hei_y)

    def on_search_change(self, event, building):
        if building == "Графский":
            self.input_search_l.delete(0, END)
        else:
            self.input_search_g.delete(0, END)

        search_text = event.widget.get().strip()

        # Очищаем старые кнопки поиска
        for btn in self.search_buttons[building]:
            btn.destroy()
        self.search_buttons[building].clear()

        if building == "Графский":
            for widget in self.root_menu_5.winfo_children():
                if widget.winfo_class() == "TButton":
                    widget.destroy()
            self.display_buttons("Лицей")
        else:
            for widget in self.root_menu_5.winfo_children():
                if widget.winfo_class() == "TButton":
                    widget.destroy()
            self.display_buttons("Графский")
        
        if not search_text:
            self.update_buttons()
            return

        
        # Фильтруем кабинеты по введённому тексту (регистр игнорируется)
        filtered_teachers = [
            r for r in self.all_rooms.get(building, [])
            if search_text.lower() in f"{r[0]}".lower()
        ]
        
        # Если есть результаты, выводим кнопки ниже поля поиска
        if filtered_teachers:
            wid_x = 0.11  # уменьшаем ширину кнопки для 4 кабинетов
            hei_y = 0.1   # высота остается прежней
    
            if building == "Графский":
                positions = [0.02, 0.14, 0.26, 0.38]
            else:
                positions = [0.52, 0.64, 0.76, 0.88]

            for i, room in enumerate(filtered_teachers):
                pos_x = positions[i % 4]
                pos_y = 0.1 + (i // 4) * 0.1

                room_button = ttk.Button(
                    self.root_menu_5,
                    text=room[0],
                    command=lambda num=room[0], ans=building: self.open_room_screen(num, ans),
                    style="Sec_5.TButton",
                )
                room_button.place(relx=pos_x, rely=pos_y, relwidth=wid_x, relheight=hei_y)
                self.search_buttons[building].append(room_button)



    def open_room_screen(self, room_number, answerr):
        room_info(self.root_menu_5, room_number, self.room_menu_5_base, self, answerr, self.root_menu_5)

#Постоянное обновление виджетов страницы       
    def update_buttons(self):
        for widget in self.root_menu_5.winfo_children():
            if widget.winfo_class() == "TButton":
                widget.destroy()
        self.display_buttons("Графский")
        self.display_buttons("Лицей")
###



#<Инфа каба> - # +++
class room_info:
    def __init__(self, root, room_number, dbs, main_screen, asw, root_parent):
    #Данные корня
        self.root_room_info = Toplevel(root)
        self.root_parent = root_parent
        self.main_screen = main_screen
        self.room_number = room_number
        self.root_room_info.title(f"Кабинет {self.room_number}")
        self.root_room_info.geometry(f"{int(self.root_room_info.winfo_screenwidth() * 0.56)}x{int(self.root_room_info.winfo_screenheight()*0.6)}")
        self.root_room_info.resizable(width=False, height=False)
        self.base_room_info = dbs
        self.con_room_info = self.base_room_info.cursor()
        self.answ = asw
        self.con_room_info.execute("SELECT * FROM rooms WHERE Number = ? AND Graph_or_Lyceum = ?", (self.room_number, self.answ))
        full_data = self.con_room_info.fetchall()
        self.def_num = full_data[0][1]
        self.def_graphskiy_not = full_data[0][2]
        self.def_big_not = full_data[0][3]
        self.def_sub = full_data[0][4]
    ###

        style_5_0 = ttk.Style()

        style_5_0.theme_use('clam')

        style_5_0.configure("TLabel", font=("Helvetica", 30, "italic"))

        style_5_0.configure("TEntry", fieldbackground="#DCDCDC")

        style_5_0.configure("TCombobox", fieldbackground="#DCDCDC", arrowsize = 0)
        style_5_0.map("TCombobox", fieldbackground=[("readonly", "#DCDCDC")])


        style_5_0.configure("Main_5_0.TButton", font=("Helvetica", 30, "bold"), background="#696969", foreground="white",)
        style_5_0.map("Main_5_0.TButton", background=[("active", "#9932CC")])

        self.root_room_info.configure(bg="#e0dcd4")


    #Виджеты изменения кабинета
    #Границы
        separator_1 = Frame(
            self.root_room_info, 
            height=3,           # Увеличиваем высоту
            bg='grey',         # Задаем цвет
            relief=FLAT,
            bd=1               # Увеличиваем толщину бордюра
        )
        separator_1.place(relx=0, rely=0.45, relwidth=1, relheight=0.007)
    ###

    #Блок номера кабинета 
        self.number_room_label = ttk.Label(
            self.root_room_info,
            text="Номер кабинета:",
            style="TLabel",
            anchor="center")
        self.number_room_label.place(relx=0, rely=0, relwidth=0.5, relheight=0.15)

        self.input_number_room = ttk.Entry(
            self.root_room_info,
            style="TEntry",
        font=("Helvetica", 27),
        foreground="#4D4D4D"
        )
        self.input_number_room.place(relx=0.5, rely=0, relwidth=0.5, relheight=0.15)
        
        self.input_number_room.insert(0, self.def_num) # Пердварительный ввод номера кабинета
    ###

    #Блок графский или нет
        options = ["Графский", "Лицей"]
        self.graphskiy_or_not_label = ttk.Label(
            self.root_room_info,
            text="Графский или лицей:",
            style="TLabel",
            anchor="center"
        )
        self.graphskiy_or_not_label.place(relx=0, rely=0.15, relwidth=0.5, relheight=0.15)

        self.graphskiy_or_not = ttk.Combobox(
            self.root_room_info,
            values=options,
            style="TCombobox",
            font=("Helvetica", 27),
            state="readonly", 
            foreground="#4D4D4D"
        )
        self.graphskiy_or_not.set(self.def_graphskiy_not) # Предварительный ввод графский/лицей
        self.graphskiy_or_not.place(relx=0.5, rely=0.15, relwidth=0.5, relheight=0.15)
    ###

    #Блок выбора большого/маленького кабинета
        options = ["Маленький", "Большой"]
        self.small_or_big_label = ttk.Label(
            self.root_room_info,
            text="Маленький или большой:",
            style="TLabel",
            anchor="center"
        )
        self.small_or_big_label.place(relx=0, rely=0.3, relwidth=0.5, relheight=0.15)

        self.small_or_big = ttk.Combobox(
            self.root_room_info,
            values=options,
            style="TCombobox",
            font=("Helvetica", 27),
            state="readonly", 
            foreground="#4D4D4D"
        )
        self.small_or_big.set(self.def_big_not)
        self.small_or_big.place(relx=0.5, rely=0.3, relwidth=0.5, relheight=0.15)
    ###

    #Блок выбора приоритетного кабинета
        self.subjects = [
            "Нет приоритетных кабинетов",
            "Алгебра/Геометрия/ТеорВер",
            "Физика",
            "Русский/Литература",
            "История",
            "Информатика",
            "География",
            "Английский язык",
            "Физра",
            "Биология",
            "Химия",
            "Обществознание"
        ]

        self.prio_subjects_label = ttk.Label(
            self.root_room_info,
            text = "Приоритетные уроки",
            style="TLabel",
            anchor="center",
        )
        self.prio_subjects_label.place(relx=0, rely=0.457, relwidth=1, relheight=0.09)

        self.prio_subjects = Listbox(
            self.root_room_info,
            selectmode=SINGLE,
            background="#DCDCDC",
            selectbackground="#4b6985", 
            font=("Helvetica", 27), 
            foreground="#4D4D4D", 
            relief="groove",
            bd=2
        )
        self.prio_subjects.place(relx=0.2, rely=0.55, relwidth=0.6, relheight=0.3)

        for subject in self.subjects:
            self.prio_subjects.insert(END, subject)
        
        for i in range(len(self.subjects)): # Предварительный ввод предмета
            if self.subjects[i] == self.def_sub:
                a = i
                self.prio_subjects.select_set(a)
                break
    ###
        
    #Футер
        self.save_button = ttk.Button(
            self.root_room_info,
            text="Сохранить",
            style="Main_5_0.TButton",
            command=self.save_data # Сохранение данных
        )
        self.save_button.place(relx=0, rely=0.85, relwidth=0.3, relheight=0.15)

        self.cancel_button = ttk.Button(
            self.root_room_info,
            text="Отмена",
            style="Main_5_0.TButton",
            command=self.root_room_info.destroy # Обычный выход из меню
        )
        self.cancel_button.place(relx=0.3, rely=0.85, relwidth=0.4, relheight=0.15)

        self.delete_button = ttk.Button(
            self.root_room_info,
            text="Удалить",
            style="Main_5_0.TButton",
            command=self.delete_data # Удаление значения
        )
        self.delete_button.place(relx=0.7, rely=0.85, relwidth=0.3, relheight=0.15)
    ###
#Функционал
#Cохранение
    def save_data(self):
    # Данные от ввода пользователя
        data_room = []
        input_num = self.input_number_room.get()
        input_graphskiy_not = self.graphskiy_or_not.get()
        input_big_or_small = self.small_or_big.get()
        selected_indices = self.prio_subjects.curselection()
    ###

    #Проверка на ввод чего-либо у прио урока
        if selected_indices:
            input_sub = self.subjects[selected_indices[0]]
        else:                                         
            warning = eror_popup(self.root_room_info, "Ошибка в прио. предметах")
            warning.root.mainloop()
            return
    ###

    # Проверка на какие-либо изменения в вводе   
        if input_num == self.def_num and input_graphskiy_not == self.def_graphskiy_not and input_sub == self.def_sub:
            warning = eror_popup(self.root_room_info, "Вы ничего не поменяли")
            warning.root.mainloop()
            return
    ###

    #Проверка номера кабинета + ввод в data_room
        all_num = "1234567890"
        count = 0 
        for digit in input_num:
            if digit in all_num:
                count += 1

        if (count == len(input_num)) and (input_num != ""): # Проверка на наличие числа и проверка на ввод данных 
            if int(input_num) // 100 < 6: # Проверка на наличие этажей между 1 - 5
                data_room.append(str(input_num))
            else:
                warning = eror_popup(self.root_room_info, "Ошибка в номере кабинета")
                warning.root.mainloop()
                return                                
        else:
            warning = eror_popup(self.root_room_info, "Ошибка в номере кабинета")
            warning.root.mainloop()
            return
    ###

    #Проверка графского или нет + ввод
        if input_graphskiy_not == "Лицей" or input_graphskiy_not == "Графский": # Проверка на правильность написания ответа
            data_room.append(input_graphskiy_not)
        else:
            warning = eror_popup(self.root_room_info, "Ошибка в Графский/лицей")
            warning.root.mainloop()
            return
    ###

    #Проверка большой или маленький
        if input_big_or_small == "Большой" or input_big_or_small == "Маленький":
            data_room.append(input_big_or_small)
        else:
            warning = eror_popup(self.root_room_info, "Ошибка в большой/мал.")
            warning.root.mainloop()
            return
    ###

    #Ввод прио урока в нужном порядке
        data_room.append(input_sub)
    ###

    # Самая финальная проверка на наличие в таблице + Ввод в таблицу + обновление страницы
        self.con_room_info.execute("SELECT * FROM rooms WHERE Number = ? AND Graph_or_Lyceum = ?", (data_room[0], data_room[1]))
        values = self.con_room_info.fetchall() # Проверка на предварительное наличие кабинета 
        if values and (data_room[0] != self.def_num or data_room[1] != self.def_graphskiy_not):
            warning = eror_popup(self.root_room_info, "Уже есть каб. в корпусе")
            warning.root.mainloop()                    
            return

        self.con_room_info.execute(
            "UPDATE rooms SET Number = ?, Graph_or_Lyceum = ?, Big_or_Small = ?, Subject = ? WHERE Number = ? AND Graph_or_Lyceum = ?",
            (input_num, input_graphskiy_not, input_big_or_small, input_sub, self.def_num, self.answ)
        )
        self.base_room_info.commit()
        self.main_screen.update_buttons()
        warning = popup(self.root_room_info, f"Кабинет {self.room_number} изменен!", "Успех")
        warning.root.mainloop()
    ###

###

#Удаление
    def delete_data(self):
        self.con_room_info.execute("DELETE FROM rooms WHERE Number = ? AND Graph_or_Lyceum = ?", (self.room_number, self.answ))
        self.base_room_info.commit()
        self.main_screen.update_buttons()
        self.root_room_info.destroy()
        warning = popup(self.root_parent, f"Кабинет {self.room_number} удален!", "Успех")
        warning.root.mainloop()
###     






#<Созд учи> - 6 +++
class menu_add_teacher:
    def __init__(self, parent, base, con):
    #Даннфые корня
        self.root_menu_6 = Toplevel(parent)
        self.root_menu_6.title("Добавление учителя")
        self.root_menu_6.geometry(f"{int(self.root_menu_6.winfo_screenwidth() * 0.56)}x{int(self.root_menu_6.winfo_screenheight()*0.6)}")
        self.root_menu_6.resizable(width=False, height=False) 
        self.base_menu_6 = base
        self.con_menu_6 = con
        self.data_exce = []  
        self.exce_buttons = []
    ###

        style_6 = ttk.Style()

        style_6.theme_use('clam')

        style_6.configure("TLabel", font=("Helvetica", 30, "italic"))

        style_6.configure("TEntry", fieldbackground="#DCDCDC")

        style_6.configure("TCombobox", fieldbackground="#DCDCDC", arrowsize = 0)
        style_6.map("TCombobox", fieldbackground=[("readonly", "#DCDCDC")])


        style_6.configure("Main_6.TButton", font=("Helvetica", 30, "bold"), background="#696969", foreground="white",)
        style_6.map("Main_6.TButton", background=[("active", "#006400")])

        style_6.configure("Sec_6.TButton", font=("Helvetica", 25, "bold"), background="#696969", foreground="white",)
        style_6.map("Sec_6.TButton", background=[("active", "#228B22")])

        self.root_menu_6.configure(bg="#e0dcd4")


    #Виджеты добавления учителя ++
    #Границы 
        separator_1 = Frame(
            self.root_menu_6, 
            height=3,         
            bg='grey',         
            relief=FLAT,
            bd=1              
        )
        separator_1.place(relx=0, rely=0.45, relwidth=0.5, relheight=0.007) #Горизонтальные1

        separator_1 = Frame(
            self.root_menu_6, 
            height=3,          
            bg='grey',        
            relief=FLAT,
            bd=1               
        )
        separator_1.place(relx=0.5, rely=0, relwidth= 0.004, relheight= 1) #Вертикальные

        separator_3 = Frame(
            self.root_menu_6, 
            height=3,         
            bg='grey',         
            relief=FLAT,
            bd=1              
        )
        separator_3.place(relx=0, rely=0.719, relwidth=0.5, relheight=0.007) #Горизонтальные2
    ###

    #Блок фамилии
        self.surname = ttk.Label(
            self.root_menu_6,
            text="Ф",
            style="TLabel",
            anchor="center"
        )
        self.surname.place(relx=0, rely=0, relwidth=0.1, relheight=0.15)

        self.input_surname = ttk.Entry(
            self.root_menu_6,
            style="TEntry",
            font=("Helvetica", 27),
            foreground="#4D4D4D"       
            )
        self.input_surname.place(relx=0.1, rely=0, relwidth=0.4, relheight=0.15)
    ###

    #Блок имени
        self.name = ttk.Label(
            self.root_menu_6,
            text="И",
            style="TLabel",
            anchor="center"
        )
        self.name.place(relx=0, rely=0.15, relwidth=0.1, relheight=0.15)

        self.input_name = ttk.Entry(
            self.root_menu_6,
            style="TEntry",
            font=("Helvetica", 27),
            foreground="#4D4D4D"
        )
        self.input_name.place(relx=0.1, rely=0.15, relwidth=0.4, relheight=0.15)
    ###

    #Блок отчества
        self.patrony = ttk.Label(
            self.root_menu_6,
            text="О",
            style="TLabel",
            anchor="center"
            )
        self.patrony.place(relx=0, rely=0.3, relwidth=0.1, relheight=0.15)

        self.input_patrony = ttk.Entry(
            self.root_menu_6,
            style="TEntry",
            font=("Helvetica", 27),
            foreground="#4D4D4D"
        )
        self.input_patrony.place(relx=0.1, rely=0.3, relwidth=0.4, relheight=0.15)
    ###

    #Блок исключений
        self.exce = ttk.Label(
            self.root_menu_6,
            text="Исключения:",
            style="TLabel",
            anchor="center"
        )
        self.exce.place(relx=0.504, rely=0, relwidth=0.496, relheight=0.1125)
    ###
        
    #Блок дня
        self.day = ttk.Label(
            self.root_menu_6,
            text="День:",
            style="TLabel",
            anchor="center"
        )
        self.day.place(relx=0.504, rely=0.1125, relwidth=0.186, relheight=0.1125)

        self.day_opt = ["Понедельник", "Вторник", "Среда", "Четверг", "Пятница", "Суббота"]
        self.input_day = ttk.Combobox(
            self.root_menu_6,
            values=self.day_opt,
            style="TCombobox",
            font=("Helvetica", 27),
            state="readonly", 
            foreground="#4D4D4D"
        )
        self.input_day.place(relx=0.69, rely=0.1125, relwidth=0.31, relheight=0.1125)
    ###
        
    #Блок исключений по урокам
        self.lessons = ttk.Label(
            self.root_menu_6,
            text="Не может:",
            style="TLabel",
            anchor="center"
        )
        self.lessons.place(relx=0.504, rely=0.225, relwidth=0.186, relheight=0.1125)

        self.begin = ttk.Entry(
            self.root_menu_6,
            style = "TEntry",
            font=("Helvetica", 25),
            foreground="#4D4D4D",
        )
        self.begin.place(relx=0.7, rely=0.225, relwidth=0.3, relheight=0.1125)
        self.begin.insert(0, "1, 2, ... , n")
    ###

    #Кнопка добавить искючение
        self.add_exce = ttk.Button(
            self.root_menu_6,
            text="Создать",
            style="Main_6.TButton",
            command= self.exce_add
        )
        self.add_exce.place(relx=0.504, rely=0.3375, relwidth=0.496, relheight=0.1125)
    ###

    #Блок выбора жесткого перехода
        options = ["Да", "Нет"]
        self.stricttrans = ttk.Label(
            self.root_menu_6,
            text="1 урок переход:",
            style="TLabel",
            anchor="center"
        )
        self.stricttrans.place(relx=0, rely=0.726, relwidth=0.3, relheight=0.138)

        self.stricttrans_or_not = ttk.Combobox(
            self.root_menu_6,
            values= options,
            style="TCombobox",
            font=("Helvetica", 27),
            state="readonly", 
            foreground="#4D4D4D"
              
        )
        self.stricttrans_or_not.place(relx=0.3, rely=0.726, relwidth=0.2, relheight=0.138)
        self.stricttrans_or_not.set("Да или нет")
    ###

    #Кабинеты прио 

    #Лицей
        self.prio_l = ttk.Label(
            self.root_menu_6,
            text="Каб.лицей:",
            style="TLabel",
            anchor="center",
            font=("Helvetica", 26),

        )
        
        self.prio_l.place(relx=0, rely=0.457, relwidth=0.25, relheight=0.13)

        self.input_prio_l = ttk.Entry(
            self.root_menu_6,
            style = "TEntry",
            font=("Helvetica", 25),
            foreground="#4D4D4D",
            
        )
        self.input_prio_l.place(relx=0.25, rely=0.457, relwidth=0.25, relheight=0.13)
        self.input_prio_l.insert(0, "Каб_1, ... , Каб_n")
    #Графский
        self.prio_g = ttk.Label(
            self.root_menu_6,
            text="Каб.графский:",
            style="TLabel",
            anchor="center",
            font=("Helvetica", 26),

        )
        
        self.prio_g.place(relx=0, rely=0.588, relwidth=0.25, relheight=0.13)

        self.input_prio_g = ttk.Entry(
            self.root_menu_6,
            style = "TEntry",
            font=("Helvetica", 25),
            foreground="#4D4D4D"
        )
        self.input_prio_g.place(relx=0.25, rely=0.588, relwidth=0.25, relheight=0.13)
        self.input_prio_g.insert(0, "Каб_1, ... , Каб_n")


    # Кнопка добавления учителя
        self.add_button = ttk.Button(
            self.root_menu_6,
            text="Добавить учителя",
            style = "Main_6.TButton",
            command=self.add_teacher
        )
        self.add_button.place(relx=0, rely=0.85, relwidth=1, relheight=0.15)
    ###

#Функционал
#Доб искл
    def exce_add(self):
    #Входные данные
        in_day = self.input_day.get()
        exce_less = self.begin.get()
    ###

    #Проверка дня
        if in_day not in self.day_opt:
            warning = eror_popup(self.root_menu_6, "Ошибка в дне искл")
            warning.root.mainloop()
            return
    ###
        char = ("1", "2", "3","4", "5","6","7", "8")

        l_class = []
        if exce_less == "":
            warning = eror_popup(self.root_menu_6, "Не выбраны уроки искл.")
            warning.root.mainloop()
            return
        if exce_less != "1, 2, ... , n" or exce_less != "":
            for l in exce_less.split(", "):
                if l not in char or l in l_class:
                    warning = eror_popup(self.root_menu_6, "Ошибке в искл. уроке")
                    warning.root.mainloop()
                    return
                else:
                    l_class.append(l)

    #

    #Проверка на существование

        for i in self.data_exce:
            if i[0] == in_day:
                warning = eror_popup(self.root_menu_6, "Уже есть такой день")
                warning.root.mainloop()
                return
    ###

    #Проверка на максимальное количество исключений
        if len(self.data_exce) == 6:
            warning = eror_popup(self.root_menu_6, "Макс. количество искл!")
            warning.root.mainloop()                 
            return
    ###

        self.data_exce.append([in_day, exce_less])
        self.display_exce()
        warning = popup(self.root_menu_6, "Искл. добвлено, еще?", "Успех")
                                    
###

#Обнов искл
    def display_exce(self):
        # Сначала удалим старые кнопки
        for btn in self.exce_buttons:
            btn.destroy()
        self.exce_buttons = []
        # Располагаем кнопки для каждого исключения из self.data_exce
        for i, exce in enumerate(self.data_exce):
            # Пропускаем пустые записи (например, удалённые)
            if not exce: 
                continue
            pos_x = (i % 2) * 0.22 + 0.53  # Относительные координаты по X (начало около 0.53)
            pos_y = (i // 2) * 0.12 + 0.47  # Относительные координаты по Y (начало около 0.47)
            btn = ttk.Button(
                self.root_menu_6,
                text=f"{exce[0]}",  # Показываем день исключения
                command=lambda idx=i, ex=exce: self.open_popup_exce(idx, ex),
                style = "Sec_6.TButton"
            )
            btn.place(relx=pos_x, rely=pos_y, relwidth=0.22, relheight=0.12)
            self.exce_buttons.append(btn)
            
    def update_exce(self):
        self.display_exce()

    # При нажатии на кнопку исключения открывается окно редактирования/удаления
    def open_popup_exce(self, idx, ex):
        popup_exce_1((idx, ex[0], ex[1]), self)

###

#Доб учителя
    def add_teacher(self):
    #Данные корня
        name = self.input_name.get()
        surname = self.input_surname.get()
        patrony = self.input_patrony.get()
        stricttr = self.stricttrans_or_not.get()
        data_teacher = []
        char = "абвгдеёжзийклмнопрстуфхцчшщъыьэюя"

        prio_l = self.input_prio_l.get()
        prio_g = self.input_prio_g.get()
    ###

   

    #Проверка имени
        if name == "" or not all(i.lower() in char for i in name):
            warning = eror_popup(self.root_menu_6, "Ошибка в имени")
            warning.root.mainloop()
            return
        else:
            data_teacher.append(name)
    ###

    #Проверка фамилии
        if surname == "" or not all(i.lower() in char for i in surname):
            warning = eror_popup(self.root_menu_6, "Ошибка в фамилии")
            warning.root.mainloop()
            return
        else:
            data_teacher.append(surname)
    ###

    #Проверка отчества
        if patrony == "" or not all(i.lower() in char for i in patrony):
            warning = eror_popup(self.root_menu_6, "Ошибка в отчестве")
            warning.root.mainloop()
            return
        else:
            data_teacher.append(patrony)
    ###

     #Проверка 1 урок переход
        if stricttr == "Да" or stricttr == "Нет":
            data_teacher.append(stricttr)
        else:                                       
            warning = eror_popup(self.root_menu_6, "Не выбран переход")
            warning.root.mainloop()
            return
    ###


    #Проверка на предварительное наличие учителя 
        self.con_menu_6.execute("SELECT * FROM teachers WHERE Surname = ? AND Name = ? AND Patrony = ?",
                            (data_teacher[1], data_teacher[0], data_teacher[2]))
        values = self.con_menu_6.fetchall()  
        if values:
            warning = eror_popup(self.root_menu_6, "Уже есть такой учитель")
            warning.root.mainloop()
            return
    ###

        

    #Проверка прио кабинетов + вставка в таблицу
        self.con_menu_6.execute("""SELECT * FROM rooms""")
        values = self.con_menu_6.fetchall()

        real_rooms= {
            "Лицей":[],
            "Графский":[]
        }

        for v in values:
            real_rooms[v[2]].append(str(v[1]))


        if prio_l != "" and prio_l != "Каб_1, ... , Каб_n":
            list_rooms = prio_l.split(", ")
            container_l_1 = []
            for r in list_rooms:
                if r not in real_rooms["Лицей"]:
                    container_l_1.append(r)

                if len(container_l_1) != 0:
                        warning = eror_popup(self.root_menu_6, f"Не сущ. {"".join(container_l_1)}")
                        warning.root.mainloop()
                        return

        
        if prio_g != "" and prio_g != "Каб_1, ... , Каб_n":
            list_rooms = prio_g.split(", ")
            container_l_2 = []
            for r in list_rooms:
                if r not in real_rooms["Графский"]:
                    container_l_2.append(r)
                if len(container_l_2) != 0:
                        warning = eror_popup(self.root_menu_6, f"Не сущ. {"".join(container_l_2)}")
                        warning.root.mainloop()
                        return
                

    #Вставка данных в таблицу teachers
            
        self.con_menu_6.execute("INSERT INTO teachers (Surname, Name, Patrony, Trans) VALUES (?, ?, ?, ?)",
                                (data_teacher[1], data_teacher[0], data_teacher[2], data_teacher[3]))
        self.base_menu_6.commit()

        self.con_menu_6.execute("SELECT id FROM teachers WHERE Surname = ? AND Name = ? AND Patrony = ? AND Trans = ?",
                            (data_teacher[1], data_teacher[0], data_teacher[2], data_teacher[3]))
        
        teacher_id = self.con_menu_6.fetchone()#Нахождение индекса
    ###


        if prio_l != "" and prio_l != "Каб_1, ... , Каб_n":
            list_rooms = prio_l.split(", ")
            container_l_1 = []
            for r in list_rooms:
                if r not in real_rooms["Лицей"]:
                    container_l_1.append(r)
                self.con_menu_6.execute("INSERT INTO prio (Rel, Rooms, Building) VALUES (?, ?, ?)",
                                (teacher_id[0], prio_l, "Лицей"))
                self.base_menu_6.commit()
        
        if prio_g != "" and prio_g != "Каб_1, ... , Каб_n":
            list_rooms = prio_g.split(", ")
            container_l_2 = []
            for r in list_rooms:
                if r not in real_rooms["Графский"]:
                    container_l_2.append(r)
                self.con_menu_6.execute("INSERT INTO prio (Rel, Rooms, Building) VALUES (?, ?, ?)",
                                (teacher_id[0], prio_g, "Графский"))
                self.base_menu_6.commit()


    #Вставка данных в таблицу исключений
        
        if teacher_id:
            teacher_id = teacher_id[0]
            
            for i in self.data_exce:
                self.con_menu_6.execute("INSERT INTO exceptions (Rel, Day, Lessons) VALUES (?, ?, ?)",
                                        (teacher_id, i[0], i[1]))# Вставка исключений в таблицу exceptions
                self.base_menu_6.commit()  

        else:
            warning = eror_popup(self.root_menu_6, "Ошибка при потере индекса")
            warning.root.mainloop()
            return
    ###

    
        
        warning = popup(self.root_menu_6, "Учитель добавлен!", "Успех")
        self.data_exce = list()
        self.reset_form()
###

#Сброс полей 
    def reset_form(self):
        self.input_surname.delete(0, END)
        self.input_name.delete(0, END)
        self.input_patrony.delete(0, END)
        self.input_day.set('')
        self.begin.delete(0, END)
        self.begin.insert(0, "1, 2, ... , n")
        self.data_exce.clear()
        self.input_prio_g.delete(0, END)
        self.input_prio_g.insert(0, "Каб_1, ... , Каб_n")
        self.input_prio_l.delete(0, END)
        self.input_prio_l.insert(0, "Каб_1, ... , Каб_n")
        self.stricttrans_or_not.set("Да или нет")
        for btn in self.exce_buttons:
            btn.destroy()
        self.exce_buttons.clear()
###



#<Окно изм искл> - 0.1 +++
class popup_exce_1:
    def __init__(self, data, parent_menu):
    #Данные корня
        self.idx, self.day_val, self.val = data
        self.parent_menu = parent_menu
        self.root_exce = Toplevel(parent_menu.root_menu_6)
        self.root_exce.title("Изменить исключение")
        self.root_exce.geometry(f"{int(self.root_exce.winfo_screenwidth() * 0.27)}x{int(self.root_exce.winfo_screenheight()*0.3)}")
        self.root_exce.resizable(width=False, height=False)
    ###

        style_6_0 = ttk.Style()

        style_6_0.theme_use('clam')

        style_6_0.configure("TLabel", font=("Helvetica", 30, "italic"))

        style_6_0.configure("TEntry", fieldbackground="#DCDCDC")

        style_6_0.configure("TCombobox", fieldbackground="#DCDCDC", arrowsize = 0)
        style_6_0.map("TCombobox", fieldbackground=[("readonly", "#DCDCDC")])


        style_6_0.configure("Main.TButton", font=("Helvetica", 19, "bold"), background="#696969", foreground="white",)
        style_6_0.map("Main.TButton", background=[("active", "#006400")])

        self.root_exce.configure(bg="#e0dcd4")
        
    #Виджеты изменения исключения - 1
        label = ttk.Label(
            self.root_exce,
            text=f"Исключение {self.day_val.lower()}:",
            style="TLabel",
            anchor="center"
        )
        label.place(relx=0, rely=0, relwidth=1, relheight=0.25)
        
        # Блок дня
        self.day = ttk.Label(
            self.root_exce,
            text="День:",
            style="TLabel",
        )
        self.day.place(relx=0.05, rely=0.25, relwidth=0.295, relheight=0.25)

        self.day_opt = ["Понедельник", "Вторник", "Среда", "Четверг", "Пятница", "Суббота"]
        self.input_day = ttk.Combobox(
            self.root_exce,
            values=self.day_opt,
            style="TCombobox",
            font=("Helvetica", 27),
            state="readonly", 
            foreground="#4D4D4D"
        )
        self.input_day.place(relx=0.3, rely=0.25, relwidth=0.695, relheight=0.25)
        self.input_day.set(self.day_val)  # Задаем изначальное значение

        # Блок исключений по урокам

        self.lessons = ttk.Label(
            self.root_exce,
            text="Не может:",
            style="TLabel",
        )
        self.lessons.place(relx=0, rely=0.5, relwidth=0.38, relheight=0.25)

        self.begin = ttk.Entry(
            self.root_exce,
            style = "TEntry",
            font=("Helvetica", 25),
            foreground="#4D4D4D",
        )
        self.begin.place(relx=0.38, rely=0.5, relwidth=0.62, relheight=0.25)
        self.begin.insert(0, self.val)

        # Футер (кнопки)
        self.save_button = ttk.Button(
            self.root_exce,
            text="Сохранить",
            style="Main.TButton",
            command=self.save_data  
        )
        self.save_button.place(relx=0, rely=0.75, relwidth=0.3, relheight=0.25)

        self.cancel_button = ttk.Button(
            self.root_exce,
            text="Отмена",
            style="Main.TButton",
            command=self.root_exce.destroy
        )
        self.cancel_button.place(relx=0.3, rely=0.75, relwidth=0.4, relheight=0.25)

        self.delete_button = ttk.Button(
            self.root_exce,
            text="Удалить",
            style="Main.TButton",
            command=self.delete_data
        )
        self.delete_button.place(relx=0.7, rely=0.75, relwidth=0.3, relheight=0.25)

#Функционал
    #Удаление исключения
    def delete_data(self):
        # Удаляем исключение с индексом self.idx
        if self.idx < len(self.parent_menu.data_exce):
            self.parent_menu.data_exce.pop(self.idx)
        self.parent_menu.update_exce()
        self.root_exce.destroy()
        popup(self.parent_menu.root_menu_6, "Исключение удалено!", "Успех").root.mainloop()
    ###

    #Сохранение данных
    #Метод сохранения изменений в исключении
    def save_data(self):
        new_day = self.input_day.get()
        new_val = self.begin.get()
        
        # Проверка дня
        if new_day not in self.day_opt:
            warning = eror_popup(self.root_exce, "Ошибка в дне искл")
            warning.root.mainloop()
            return
        
        # Проверка cлотов
        char = ("1", "2", "3", "4", "5", "6", "7", "8")
        l_class = []
        if new_val == "":
            warning = eror_popup(self.root_menu_6, "Не выбраны уроки искл.")
            warning.root.mainloop()
            return
        if new_val != "1, 2, ... , n" or new_val != "":
            for l in new_val.split(", "):
                if l not in char or l in l_class:
                    warning = eror_popup(self.root_menu_6, "Ошибке в искл. уроке")
                    warning.root.mainloop()
                    return
                else:
                    l_class.append(l)
        #
        
        # Новая проверка: если уже существует другое исключение с таким днем
        for idx, exce in enumerate(self.parent_menu.data_exce):
            if idx != self.idx and exce[0] == new_day:
                warning = eror_popup(self.root_exce, "Уже есть искл этого дня")
                warning.root.mainloop()             
                return

        # Обновляем значение исключения в родительском окне
        self.parent_menu.data_exce[self.idx] = [new_day, new_val]
        self.parent_menu.update_exce()
        popup(self.parent_menu.root_menu_6, "Искл. обновлено!", "Успех").root.mainloop()
        self.root_exce.destroy()

    ###
     
###





        
#<СМ учителя> - 7 +++
class menu_check_teacher:
    def __init__(self, parent, base, connect):
    #Данные корня
        self.root_menu_7 = Toplevel(parent)
        self.root_menu_7.title("Просмотр учителей")
        self.root_menu_7.geometry(f"{int(self.root_menu_7.winfo_screenwidth() * 0.56)}x{int(self.root_menu_7.winfo_screenheight()*0.6)}")
        self.root_menu_7.resizable(width=False, height=False) 
        self.teacher_menu_7_base = base
        self.teacher_menu_7_con = connect

        style_7 = ttk.Style()

        style_7.theme_use('clam')

        style_7.configure("TLabel", font=("Helvetica", 30, "italic"))

        style_7.configure("TEntry", fieldbackground="#DCDCDC")

        style_7.configure("TCombobox", fieldbackground="#DCDCDC", arrowsize = 0)
        style_7.map("TCombobox", fieldbackground=[("readonly", "#DCDCDC")])

        style_7.configure("Sec_7.TButton", font=("Helvetica", 10, "bold"), background="#696969", foreground="white",)
        style_7.map("Sec_7.TButton", background=[("active", "#228B22")])

        self.root_menu_7.configure(bg="#e0dcd4")

    #Виджеты проверки учителей
        self.teacher_label = ttk.Label(
            self.root_menu_7,
            text="Учителя",
            style="TLabel",
            anchor="center"
        )
        self.teacher_label.place(relx=0.2, rely=0, relwidth=0.6, relheight=0.1)

    #Поисковик
        self.input_search = ttk.Entry(
            self.root_menu_7,
            style="TEntry",
            font=("Helvetica", 25),
            foreground="#4D4D4D"
        )
        self.input_search.place(relx=0.7, rely=0, relwidth=0.25, relheight=0.1)
        self.input_search.bind("<KeyRelease>", self.on_search_change)
    #
    
        self.search_buttons = []  # Список для хранения кнопок поиска
        self.all_teachers = []    # Список всех учителей
        self.display_buttons()

    def display_buttons(self):
        self.teacher_menu_7_con.execute("SELECT DISTINCT Surname, Name, Patrony FROM teachers")
        self.all_teachers = self.teacher_menu_7_con.fetchall()

        wid_x = 0.16  # ширина кнопки (4 кнопки + отступы)
        hei_y = 0.08  # высота кнопки

        for i, teacher in enumerate(self.all_teachers):
            pos_x = (i % 6) * 0.16 + 0.01  # позиция по x (6 колонки)
            pos_y = (i // 6) * 0.09 + 0.1   # позиция по y (10 строк)

            teacher_button = ttk.Button(
                self.root_menu_7,
                text=f"{teacher[0]} {teacher[1][0]}. {teacher[2][0]}.",
                command=lambda t=teacher: self.open_teacher_screen(t),
                style = "Sec_7.TButton"
            )
            teacher_button.place(relx=pos_x, rely=pos_y, relwidth=wid_x, relheight=hei_y)
    
    def on_search_change(self, event):
        """Обработчик изменения текста в поле поиска"""
        search_text = self.input_search.get()

        # Очищаем старые кнопки поиска
        for btn in self.search_buttons:
            btn.destroy()
        self.search_buttons.clear()


        for widget in self.root_menu_7.winfo_children():
            if widget.winfo_class() == "TButton":
                widget.destroy()
        
        if not search_text:
            self.update_buttons()
            return

        
        
        # Фильтруем учителей по введённому тексту
        filtered_teachers = [
            teacher for teacher in self.all_teachers
            if search_text in f"{teacher[0]} {teacher[1]} {teacher[2]}"
        ]
        
        # Если есть результаты, выводим кнопки ниже поля поиска
        if filtered_teachers:
            for i, teacher in enumerate(filtered_teachers):
                pos_x = (i % 6) * 0.16 + 0.01  # позиция по x (6 колонки)
                pos_y = (i // 6) * 0.09 + 0.1   # позиция по y (10 строк)
                wid_x = 0.16  # ширина кнопки (4 кнопки + отступы)
                hei_y = 0.08  # высота кнопки


                teacher_button = ttk.Button(
                    self.root_menu_7,
                    text=f"{teacher[0]} {teacher[1][0]}. {teacher[2][0]}.",
                    command=lambda t=teacher: self.open_teacher_screen(t),
                    style = "Sec_7.TButton"
                )
                teacher_button.place(relx=pos_x, rely=pos_y, relwidth=wid_x, relheight=hei_y)
                self.search_buttons.append(teacher_button)

    def open_teacher_screen(self, teacher):
        teacher_info(self.root_menu_7, teacher, self.teacher_menu_7_base, self, self.root_menu_7)

    def update_buttons(self):
        for widget in self.root_menu_7.winfo_children():
            if widget.winfo_class() == "TButton":
                widget.destroy()
        self.display_buttons()



#<Инфа уч> - # +++
class teacher_info:
    def __init__(self, root, teacher, dbs, main_screen, root_parent):
        #Данные корня
        self.root_teacher_info = Toplevel(root)
        self.root_parent = root_parent
        self.main_screen = main_screen
        self.teacher = teacher
        self.root_teacher_info.title(f"Учитель {self.teacher[0]} {self.teacher[1]} {self.teacher[2]}")
        self.root_teacher_info.geometry(f"{int(self.root_teacher_info.winfo_screenwidth() * 0.56)}x{int(self.root_teacher_info.winfo_screenheight()*0.6)}")
        self.root_teacher_info.resizable(width=False, height=False)
        self.base_teacher_info = dbs
        self.con_teacher_info = self.base_teacher_info.cursor()

        style_7_0 = ttk.Style()

        style_7_0.theme_use('clam')

        style_7_0.configure("TLabel", font=("Helvetica", 30, "italic"))

        style_7_0.configure("TEntry", fieldbackground="#DCDCDC")

        style_7_0.configure("TCombobox", fieldbackground="#DCDCDC", arrowsize = 0)
        style_7_0.map("TCombobox", fieldbackground=[("readonly", "#DCDCDC")])

        style_7_0.configure("Main_7_0.TButton", font=("Helvetica", 30, "bold"), background="#696969", foreground="white",)
        style_7_0.map("Main_7_0.TButton", background=[("active", "#006400")])

        style_7_0.configure("Sec_7_0.TButton", font=("Helvetica", 25, "bold"), background="#696969", foreground="white",)
        style_7_0.map("Sec_7_0.TButton", background=[("active", "#228B22")])

        self.root_teacher_info.configure(bg="#e0dcd4")

        self.con_teacher_info.execute(
            "SELECT * FROM teachers WHERE Surname = ? AND Name = ? AND Patrony = ?",
            (self.teacher[0], self.teacher[1], self.teacher[2])
        )
        full_data = self.con_teacher_info.fetchall()
        self.def_surname = full_data[0][1]
        self.def_name = full_data[0][2]
        self.def_patrony = full_data[0][3]
        self.def_trans = full_data[0][4]
        self.teacher_id = full_data[0][0]

        # Загрузка исключений из БД для данного учителя по полю Rel
        self.data_exce = []  # локальный список исключений
        self.exce_buttons = []  # список кнопок для отображения исключений
        self.con_teacher_info.execute(
            "SELECT Day, Lessons FROM exceptions WHERE Rel = ?",
            (self.teacher_id,)
        )
        exceptions = self.con_teacher_info.fetchall()
        for exc in exceptions:
            self.data_exce.append(list(exc))  # приводим к списку для единообразия
        
        self.con_teacher_info.execute(
            "SELECT * FROM prio WHERE Rel = ? AND Building = ?",
            (self.teacher_id, "Лицей")
        )

        self.data_l = self.con_teacher_info.fetchall()

        self.con_teacher_info.execute(
            "SELECT * FROM prio WHERE Rel = ? AND Building = ?" ,
            (self.teacher_id, "Графский")
        )
        self.data_g = self.con_teacher_info.fetchall()

    #Виджеты инфы учителя
        # Границы (аналогично menu_add_teacher)
        separator_1 = Frame(
            self.root_teacher_info, 
            height=3,
            bg='grey',
            relief=FLAT,
            bd=1
        )
        separator_1.place(relx=0, rely=0.45, relwidth=0.5, relheight=0.007)

        separator_2 = Frame(
            self.root_teacher_info, 
            height=3,
            bg='grey',
            relief=FLAT,
            bd=1
        )
        separator_2.place(relx=0.5, rely=0, relwidth=0.004, relheight=1)

        separator_3 = Frame(
            self.root_teacher_info, 
            height=3,         
            bg='grey',         
            relief=FLAT,
            bd=1              
        )
        separator_3.place(relx=0, rely=0.719, relwidth=0.5, relheight=0.007) #Горизонтальные2
        

        # Блок Фамилия
        self.surname = ttk.Label(
            self.root_teacher_info,
            text="Ф",
            style="TLabel",
            anchor="center"
        )
        self.surname.place(relx=0, rely=0, relwidth=0.1, relheight=0.15)

        self.input_surname = ttk.Entry(
            self.root_teacher_info,
            style="TEntry",
            font=("Helvetica", 27),
            foreground="#4D4D4D"
        )
        self.input_surname.place(relx=0.1, rely=0, relwidth=0.4, relheight=0.15)
        self.input_surname.insert(0, self.def_surname)

        # Блок Имя
        self.name = ttk.Label(
            self.root_teacher_info,
            text="И",
            style="TLabel",
            anchor="center"
        )
        self.name.place(relx=0, rely=0.15, relwidth=0.1, relheight=0.15)

        self.input_name = ttk.Entry(
            self.root_teacher_info,
            style="TEntry",
            font=("Helvetica", 27),
            foreground="#4D4D4D"
        )
        self.input_name.place(relx=0.1, rely=0.15, relwidth=0.4, relheight=0.15)
        self.input_name.insert(0, self.def_name)

        # Блок Отчество
        self.patrony = ttk.Label(
            self.root_teacher_info,
            text="О",
            style="TLabel",
            anchor="center"
        )
        self.patrony.place(relx=0, rely=0.3, relwidth=0.1, relheight=0.15)

        self.input_patrony = ttk.Entry(
            self.root_teacher_info,
            style="TEntry",
            font=("Helvetica", 27),
            foreground="#4D4D4D"
        )
        self.input_patrony.place(relx=0.1, rely=0.3, relwidth=0.4, relheight=0.15)
        self.input_patrony.insert(0, self.def_patrony)

    #Блок жесткого перехода
        options = ["Да", "Нет"]
        self.stricttrans = ttk.Label(
            self.root_teacher_info,
            text="1 урок переход:",
            style="TLabel",
            anchor="center"
        )
        self.stricttrans.place(relx=0, rely=0.726, relwidth=0.3, relheight=0.138)

        self.stricttrans_or_not = ttk.Combobox(
            self.root_teacher_info,
            values= options,
            style="TCombobox",
            font=("Helvetica", 27),
            state="readonly", 
            foreground="#4D4D4D"
              
        )
        self.stricttrans_or_not.place(relx=0.3, rely=0.726, relwidth=0.2, relheight=0.138)
        self.stricttrans_or_not.set(self.def_trans)
        ###

        # Блок исключений (заголовок)
        self.exce = ttk.Label(
            self.root_teacher_info,
            text="Исключения:",
            style="TLabel",
            anchor="center"
        )
        self.exce.place(relx=0.504, rely=0, relwidth=0.496, relheight=0.1125)

        # Блок дня для исключений
        self.day = ttk.Label(
            self.root_teacher_info,
            text="День:",
            style="TLabel",
            anchor="center"
        )
        self.day.place(relx=0.504, rely=0.1125, relwidth=0.186, relheight=0.1125)

        self.day_opt = ["Понедельник", "Вторник", "Среда", "Четверг", "Пятница", "Суббота"]
        self.input_day = ttk.Combobox(
            self.root_teacher_info,
            values=self.day_opt,
            style="TCombobox",
            font=("Helvetica", 27),
            state="readonly", 
            foreground="#4D4D4D")
        self.input_day.place(relx=0.69, rely=0.1125, relwidth=0.31, relheight=0.1125)

        # Блок исключений по урокам
        self.num_les = ["0", "1", "2", "3", "4", "5", "6", "7", "8"]

        self.lessons = ttk.Label(
            self.root_teacher_info,
            text="Не может:",
            style="TLabel",
            anchor="center"
        )
        self.lessons.place(relx=0.504, rely=0.225, relwidth=0.186, relheight=0.1125)

        self.lessons = ttk.Entry(
            self.root_teacher_info,
            style = "TEntry",
            font=("Helvetica", 25),
            foreground="#4D4D4D",
        )
        self.lessons.place(relx=0.7, rely=0.225, relwidth=0.3, relheight=0.1125)
        self.lessons.insert(0, "1, 2, ... , n")

    #Кабинеты прио 

    #Лицей
        self.prio_l = ttk.Label(
            self.root_teacher_info,
            text="Каб.лицей:",
            style="TLabel",
            anchor="center",
            font=("Helvetica", 26),

        )
        
        self.prio_l.place(relx=0, rely=0.457, relwidth=0.25, relheight=0.13)

        self.input_prio_l = ttk.Entry(
            self.root_teacher_info,
            style = "TEntry",
            font=("Helvetica", 25),
            foreground="#4D4D4D",
            
        )
        self.input_prio_l.place(relx=0.25, rely=0.457, relwidth=0.25, relheight=0.13)
        if self.data_l:
            self.input_prio_l.insert(0, self.data_l[0][2])
            self.data_l = self.data_l[0][2]
        else:
            self.input_prio_l.insert(0, "Каб_1, ... , Каб_n")
            self.data_l = "Каб_1, ... , Каб_n"
    #Графский
        self.prio_g = ttk.Label(
            self.root_teacher_info,
            text="Каб.графский:",
            style="TLabel",
            anchor="center",
            font=("Helvetica", 26),

        )
        
        self.prio_g.place(relx=0, rely=0.588, relwidth=0.25, relheight=0.13)

        self.input_prio_g = ttk.Entry(
            self.root_teacher_info,
            style = "TEntry",
            font=("Helvetica", 25),
            foreground="#4D4D4D"
        )
        self.input_prio_g.place(relx=0.25, rely=0.588, relwidth=0.25, relheight=0.13)
        if self.data_g: 
            self.input_prio_g.insert(0, self.data_g[0][2])
            self.data_g = self.data_g[0][2]
        else:
            self.input_prio_g.insert(0, "Каб_1, ... , Каб_n")
            self.data_g = "Каб_1, ... , Каб_n"

    ###

        # Кнопка добавления исключения
        self.add_exce = ttk.Button(
            self.root_teacher_info,
            text="Создать",
            style="Main_7_0.TButton",
            command=self.exce_add
        )
        self.add_exce.place(relx=0.504, rely=0.3375, relwidth=0.496, relheight=0.1125)

        # Футер: кнопки "Сохранить", "Отмена", "Удалить", "Доп инфа"
        self.save_button = ttk.Button(
            self.root_teacher_info,
            text="Сохранить",
            style="Main_7_0.TButton",
            command=self.save_data
        )
        self.save_button.place(relx=0, rely=0.85, relwidth=0.25, relheight=0.15)

        self.cancel_button = ttk.Button(
            self.root_teacher_info,
            text="Отмена",
            style="Main_7_0.TButton",
            command=self.root_teacher_info.destroy
        )
        self.cancel_button.place(relx=0.25, rely=0.85, relwidth=0.252, relheight=0.15)

        self.cancel_button = ttk.Button(
            self.root_teacher_info,
            text="Занятость",
            style="Main_7_0.TButton",
            command=self.open_extra_info
        )
        self.cancel_button.place(relx=0.502, rely=0.85, relwidth=0.248, relheight=0.15)

        self.delete_button = ttk.Button(
            self.root_teacher_info,
            text="Удалить",
            style="Main_7_0.TButton",
            command=self.delete_data
        )
        self.delete_button.place(relx=0.75, rely=0.85, relwidth=0.25, relheight=0.15)

        # Отображаем исключения, загруженные из БД
        self.display_exce()
    
    #Функционал
    def open_extra_info(self):
        extra_info(self.teacher_id, self.root_teacher_info, self.con_teacher_info, self.base_teacher_info)
        
    def exce_add(self):
        # Входные данные
        in_day = self.input_day.get()
        in_val = self.lessons.get()

        # Проверка дня
        if in_day not in self.day_opt:
            warning = eror_popup(self.root_teacher_info, "Ошибка в дне искл")
            warning.root.mainloop()
            return
        
        char = ("1", "2", "3","4", "5","6","7", "8")

        l_class = []
        if in_val == "":
            warning = eror_popup(self.root_teacher_info, "Не выбраны уроки искл.")
            warning.root.mainloop()
            return
        if in_val != "1, 2, ... , n" or in_val != "":
            for l in in_val.split(", "):
                if l not in char or l in l_class:
                    warning = eror_popup(self.root_teacher_info, "Ошибке в искл. уроке")
                    warning.root.mainloop()
                    return
                else:
                    l_class.append(l)

        # Проверка наличия исключения для уже заданного дня
        for exc in self.data_exce:
            if exc[0] == in_day:
                warning = eror_popup(self.root_teacher_info, "Уже есть такой день")
                warning.root.mainloop()
                return
        # Ограничение количества исключений
        if len(self.data_exce) >= 6:
            warning = eror_popup(self.root_teacher_info, "Макс. количество искл!")
            warning.root.mainloop()
            return

        

        self.data_exce.append([in_day, in_val])
        self.con_teacher_info.execute(
            """INSERT INTO exceptions (Rel, Day, Lessons) VALUES (?, ?, ?)""",
            (self.teacher_id, in_day, in_val)
        )
        self.base_teacher_info.commit()
        self.display_exce()
        popup(self.root_teacher_info, "Искл. добавлено, можно еще", "Успех").root.mainloop()

    def display_exce(self):
        # Удаляем старые кнопки
        for btn in self.exce_buttons:
            btn.destroy()
        self.exce_buttons = []
        # Размещаем кнопки для каждого исключения из self.data_exce
        for i, exc in enumerate(self.data_exce):
            if not exc:
                continue
            pos_x = (i % 2) * 0.22 + 0.53
            pos_y = (i // 2) * 0.12 + 0.47
            btn = ttk.Button(
                self.root_teacher_info,
                text=f"{exc[0]}",  # отображаем день
                command=lambda idx=i, ex=exc: self.open_popup_exce(idx, ex),
                style = "Sec_7_0.TButton"
            )
            btn.place(relx=pos_x, rely=pos_y, relwidth=0.22, relheight=0.12)
            self.exce_buttons.append(btn)

    def update_exce(self):
        self.display_exce()

    def open_popup_exce(self, idx, exc):
        popup_exce_2((idx, exc[0], exc[1]), self)

    def save_data(self):
        # Пример сохранения данных учителя (реализация уже присутствует)
        data_teacher = []
        input_surname = self.input_surname.get()
        input_name = self.input_name.get()
        input_patrony = self.input_patrony.get()
        input_trans = self.stricttrans_or_not.get()
        input_prio_l = self.input_prio_l.get()
        input_prio_g = self.input_prio_g.get()
        if (input_surname == self.def_surname and input_name == self.def_name and 
            input_patrony == self.def_patrony and input_trans == self.def_trans and 
            (input_prio_g == self.data_g or (input_prio_g == "" and self.data_g == "Каб_1, ... , Каб_n")) and
            (input_prio_l == self.data_l or (input_prio_l == "" and self.data_l == "Каб_1, ... , Каб_n"))):
            warning = eror_popup(self.root_teacher_info, "Вы ничего не поменяли")
            warning.root.mainloop()
            return
        char = "абвгдеёжзийклмнопрстуфхцчшщъыьэюя"
        if input_surname == "" or not all(c.lower() in char for c in input_surname):
            warning = eror_popup(self.root_teacher_info, "Ошибка в фамилии")
            warning.root.mainloop()
            return
        else:
            data_teacher.append(input_surname)
        if input_name == "" or not all(c.lower() in char for c in input_name):
            warning = eror_popup(self.root_teacher_info, "Ошибка в имени")
            warning.root.mainloop()
            return
        else:
            data_teacher.append(input_name)
        if input_patrony == "" or not all(c.lower() in char for c in input_patrony):
            warning = eror_popup(self.root_teacher_info, "Ошибка в отчестве")
            warning.root.mainloop()
            return
        else:
            data_teacher.append(input_patrony)



        real_rooms= {
            "Лицей":[],
            "Графский":[]
        }


        self.con_teacher_info.execute("""SELECT * FROM rooms""")
        values = self.con_teacher_info.fetchall()
    ###
        for v in values:
            real_rooms[v[2]].append(str(v[1]))

        if input_prio_l != "" and input_prio_l != "Каб_1, ... , Каб_n":
            list_rooms = input_prio_l.split(", ")
            container_l_1 = []
            for r in list_rooms:
                if r not in real_rooms["Лицей"]:
                    container_l_1.append(r)

                if len(container_l_1) != 0:
                        warning = eror_popup(self.root_teacher_info, f"Не сущ. {"".join(container_l_1)}")
                        warning.root.mainloop()
                        return

        
        if input_prio_g != "" and input_prio_g != "Каб_1, ... , Каб_n":
            list_rooms = input_prio_g.split(", ")
            container_l_2 = []
            for r in list_rooms:
                if r not in real_rooms["Графский"]:
                    container_l_2.append(r)
                if len(container_l_2) != 0:
                        warning = eror_popup(self.root_teacher_info, f"Не сущ. {"".join(container_l_2)}")
                        warning.root.mainloop()
                        return
        #  input_prio_g == self.data_g and  input_prio_l == self.data_l

        if input_prio_g != self.data_g:
            if input_prio_g == "" or input_prio_g ==  "Каб_1, ... , Каб_n":
                self.con_teacher_info.execute("DELETE FROM prio WHERE Rel = ? and Rooms = ? and Building = ?", (self.teacher_id, self.data_g, "Графский"))
                self.base_teacher_info.commit()
            else:
                self.con_teacher_info.execute(
                "UPDATE prio SET Rooms = ? WHERE Building = ? and Rel = ?", (input_prio_g, "Графский", self.teacher_id))
                self.base_teacher_info.commit()

            if self.data_g == "Каб_1, ... , Каб_n" and (input_prio_g != "" or input_prio_g != "Каб_1, ... , Каб_n") and self.data_g != input_prio_g:
                self.con_teacher_info.execute(
                """INSERT INTO prio (Rel, Rooms, Building) VALUES (?, ?, ?)""", (self.teacher_id, input_prio_g, "Графский"))
                self.base_teacher_info.commit()
        

        if input_prio_l != self.data_l:
            if input_prio_l == "" or input_prio_l == "Каб_1, ... , Каб_n":
                self.con_teacher_info.execute("DELETE FROM prio WHERE Rel = ? and Rooms = ? and Building = ?", (self.teacher_id, self.data_l, "Лицей"))
                self.base_teacher_info.commit()
            else:
                self.con_teacher_info.execute(
                "UPDATE prio SET Rooms = ? WHERE Building = ? and Rel == ?", (input_prio_l,"Лицей", self.teacher_id))
                self.base_teacher_info.commit()

            if self.data_l == "Каб_1, ... , Каб_n" and (input_prio_l != "" and input_prio_l != "Каб_1, ... , Каб_n"):
                self.con_teacher_info.execute(
                """INSERT INTO prio (Rel, Rooms, Building) VALUES (?, ?, ?)""", (self.teacher_id, input_prio_l, "Лицей"))
                self.base_teacher_info.commit()
        
        
        
        
        data_teacher.append(input_trans)
        self.con_teacher_info.execute(
            "UPDATE teachers SET Surname = ?, Name = ?, Patrony = ?, Trans = ? WHERE id = ?",
            (data_teacher[0], data_teacher[1], data_teacher[2], data_teacher[3], self.teacher_id)
        )
        self.base_teacher_info.commit()
        self.data_g = input_prio_g
        self.data_l = input_prio_l
        self.main_screen.update_buttons()
        popup(self.root_teacher_info, "Учитель обновлён!", "Успех").root.mainloop()

    def delete_data(self):

        # Удаляем учителя из таблицы teachers
        self.con_teacher_info.execute("DELETE FROM prio WHERE Rel = ?", (self.teacher_id,))
        self.base_teacher_info.commit()

        # Удаляем учителя из таблицы teachers
        self.con_teacher_info.execute("DELETE FROM teachers WHERE id = ?", (self.teacher_id,))
        self.base_teacher_info.commit()
    
        # Удаляем все исключения, связанные с данным учителем (по полю Rel)
        self.con_teacher_info.execute("DELETE FROM exceptions WHERE Rel = ?", (self.teacher_id,))
        self.base_teacher_info.commit()

        # Удаление всех уроков с таким учителем
        self.con_teacher_info.execute("DELETE FROM lessons WHERE Id_teacher = ?", (self.teacher_id,))
        self.base_teacher_info.commit()
        
        # Удаление физры (pe) с таким учителем
        teacher_fio = f"{self.def_surname} {self.def_name[0]}. {self.def_patrony[0]}."
        self.con_teacher_info.execute("DELETE FROM pe WHERE Teacher = ?", (teacher_fio,))
        self.base_teacher_info.commit()
        
        # Удаление внеурочки (extra) с таким учителем
        self.con_teacher_info.execute("DELETE FROM extra WHERE Teacher = ?", (teacher_fio,))
        self.base_teacher_info.commit()

        self.main_screen.update_buttons()
        self.root_teacher_info.destroy()
        popup(self.root_parent, "Учитель удален!", "Успех").root.mainloop()

#<Окно изм искл> - 0.2 +++
class popup_exce_2:
    def __init__(self, data, parent_menu):
        #Данные корня
        self.idx, self.day_val, self.val = data
        self.parent_menu = parent_menu
        self.root_exce = Toplevel(parent_menu.root_teacher_info)
        self.root_exce.title("Изменить исключение")
        self.root_exce.geometry(f"{int(self.root_exce.winfo_screenwidth() * 0.27)}x{int(self.root_exce.winfo_screenheight()*0.3)}")
        self.root_exce.resizable(width=False, height=False)
        
        style_6_1 = ttk.Style()

        style_6_1.theme_use('clam')

        style_6_1.configure("TLabel", font=("Helvetica", 30, "italic"))

        style_6_1.configure("TEntry", fieldbackground="#DCDCDC")

        style_6_1.configure("TCombobox", fieldbackground="#DCDCDC", arrowsize = 0)
        style_6_1.map("TCombobox", fieldbackground=[("readonly", "#DCDCDC")])


        style_6_1.configure("Main_6_1.TButton", font=("Helvetica", 19, "bold"), background="#696969", foreground="white",)
        style_6_1.map("Main_6_1.TButton", background=[("active", "#006400")])

        self.root_exce.configure(bg="#e0dcd4")

        #Виджеты изменения исключения - 2
        label = ttk.Label(
            self.root_exce,
            text=f"Исключение {self.day_val.lower()}:",
            style="TLabel",
            anchor="center"
        )
        label.place(relx=0, rely=0, relwidth=1, relheight=0.25)
        
        # Блок дня
        self.day = ttk.Label(
            self.root_exce,
            text="День:",
            style="TLabel",
        )
        self.day.place(relx=0.05, rely=0.25, relwidth=0.295, relheight=0.25)

        self.day_opt = ["Понедельник", "Вторник", "Среда", "Четверг", "Пятница", "Суббота"]
        self.input_day = ttk.Combobox(
            self.root_exce,
            values=self.day_opt,
            style="TCombobox",
            font=("Helvetica", 27),
            state="readonly", 
            foreground="#4D4D4D"
        )
        self.input_day.place(relx=0.345, rely=0.25, relwidth=0.655, relheight=0.25)
        self.input_day.set(self.day_val)  # устанавливаем предустановленное значение

        # Блок исключений по урокам
        self.num_les = ["0", "1", "2", "3", "4", "5", "6", "7", "8"]
        self.lessons = ttk.Label(
            self.root_exce,
            text="Не может:",
            style="TLabel",
        )
        self.lessons.place(relx=0, rely=0.5, relwidth=0.38, relheight=0.25)

        self.begin = ttk.Entry(
            self.root_exce,
            style = "TEntry",
            font=("Helvetica", 25),
            foreground="#4D4D4D",
        )
        self.begin.place(relx=0.38, rely=0.5, relwidth=0.62, relheight=0.25)
        self.begin.insert(0, self.val)

        # Футер (кнопки "Сохранить", "Отмена", "Удалить")
        self.save_button = ttk.Button(
            self.root_exce,
            text="Сохранить",
            style = "Main_6_1.TButton",
            command=self.save_data
        )
        self.save_button.place(relx=0, rely=0.75, relwidth=0.3, relheight=0.25)

        self.cancel_button = ttk.Button(
            self.root_exce,
            text="Отмена",
            style = "Main_6_1.TButton",
            command=self.root_exce.destroy
        )
        self.cancel_button.place(relx=0.3, rely=0.75, relwidth=0.4, relheight=0.25)

        self.delete_button = ttk.Button(
            self.root_exce,
            text="Удалить",
            style = "Main_6_1.TButton",
            command=self.delete_data
        )
        self.delete_button.place(relx=0.7, rely=0.75, relwidth=0.3, relheight=0.25)

    #Функционал
    def delete_data(self):
        # Удаляем запись из базы данных (если существует)
        self.parent_menu.con_teacher_info.execute(
            "DELETE FROM exceptions WHERE Rel = ? AND Day = ? AND Lessons = ?",
            (self.parent_menu.teacher_id, self.day_val, self.val)
        )
        self.parent_menu.base_teacher_info.commit()
        
        # Удаляем исключение из локального списка
        if self.idx < len(self.parent_menu.data_exce):
            self.parent_menu.data_exce.pop(self.idx)
        self.parent_menu.update_exce()
        self.root_exce.destroy()
        popup(self.parent_menu.root_teacher_info, "Исключение удалено!", "Успех").root.mainloop()

    def save_data(self):
        new_day = self.input_day.get()
        new_val = self.begin.get()

        if new_day not in self.day_opt:
            warning = eror_popup(self.root_exce, "Ошибка в дне искл")
            warning.root.mainloop()
            return


        char = ("1", "2", "3","4", "5","6","7", "8")

        l_class = []
        if new_val == "":
            warning = eror_popup(self.root_exce, "Не выбраны уроки искл.")
            warning.root.mainloop()
            return
        if new_val != "1, 2, ... , n" or new_val != "":
            for l in new_val.split(", "):
                if l not in char or l in l_class:
                    warning = eror_popup(self.root_exce, "Ошибке в искл. уроке")
                    warning.root.mainloop()
                    return
                else:
                    l_class.append(l)
        #

        for idx, exce in enumerate(self.parent_menu.data_exce):
            if idx != self.idx and exce[0] == new_day:
                warning = eror_popup(self.root_exce, "Уже есть искл. этого дня")
                warning.root.mainloop()               
                return

        self.parent_menu.data_exce[self.idx] = [new_day, new_val]
        self.parent_menu.update_exce()
        popup(self.parent_menu.root_teacher_info, "Искл. обновлено!", "Успех").root.mainloop()
        self.root_exce.destroy()
###

#<Окно доп инфы> - # +++
class extra_info:
    def __init__(self, id, parent_menu, con, base):
        #Данные корня
        self.root_extra_info = Toplevel(parent_menu)
        self.con = con
        self.base = base
        self.parent_id = id
        self.root_extra_info.title("Занятость")
        self.root_extra_info.geometry(f"{int(self.root_extra_info.winfo_screenwidth() * 0.54)}x{int(self.root_extra_info.winfo_screenheight()*0.58)}")
        self.root_extra_info.resizable(width=False, height=False)

        style_7_2 = ttk.Style()

        style_7_2.theme_use('clam')

        style_7_2.configure("Main.TLabel", font=("Helvetica", 30, "italic"))

        style_7_2.configure("Sec.TLabel", font=("Helvetica", 25),
            borderwidth=2,
            relief="solid", 
            bg='white',
            fg='black')

        style_7_2.configure("TEntry", fieldbackground="#DCDCDC")

        style_7_2.configure("TCombobox", fieldbackground="#DCDCDC", arrowsize = 0)
        style_7_2.map("TCombobox", fieldbackground=[("readonly", "#DCDCDC")])


        style_7_2.configure("Main.TButton", font=("Helvetica", 30, "bold"), background="#696969", foreground="white",)
        style_7_2.map("Main.TButton", background=[("active", "#006400")])

        self.root_extra_info.configure(bg="#e0dcd4")
    #Лейблы доп инфы
        #Отбражение классов преподования
        label = ttk.Label(
            self.root_extra_info,
            text="Преподает в:",
            style="Mian.TLabel",
            anchor="center"
        )
        label.place(relx=0, rely=0, relwidth=1, relheight=0.15)

        self.con.execute(
            "SELECT Id_parallel, Subject, Hours FROM lessons WHERE Id_teacher = ?",
            (self.parent_id,)
        )
        info_all_class = self.con.fetchall()

        self.labels = []

        for i, info_cer_class in enumerate(info_all_class):
            pos_x = (i % 4) * 0.25  # 4 кнопки в ряд
            pos_y = 0.15 + (i // 4) * 0.1   # новый ряд каждые 4 кнопки
            
            self.con.execute(
            "SELECT Letter, Number FROM parallels WHERE id = ?",
            (info_cer_class[0],))
            info_cer_parallel = self.con.fetchall()

            lab = ttk.Label(
            self.root_extra_info,
            text=f"{info_cer_parallel[0][1]}{info_cer_parallel[0][0]} - {info_cer_class[1][0:3]}.: {info_cer_class[2]} ч.",
            style = "Sec.TLabel"
            
            )
            lab.place(relx= pos_x, rely=pos_y, relwidth=0.25, relheight=0.1)
            self.labels.append(label)

        #Виджеты
        #Отображение количества часов работы
        label = ttk.Label(
            self.root_extra_info,
            text="Общее количество часов:",
            style = "Main.TLabel",
        )
        label.place(relx=0.2, rely=0.75, relwidth=0.5, relheight=0.1)

        self.con.execute(
            "SELECT Hours FROM lessons WHERE Id_teacher = ?",
            (self.parent_id,)
        )
        num_of_studing = self.con.fetchall()
        count = 0
        for i in num_of_studing:
            count += int(i[0])

        label = ttk.Label(
            self.root_extra_info,
            text=count,
            font=("Helvetica", 25)
        )
        label.place(relx=0.7, rely=0.75, relwidth=0.1, relheight=0.1)

        #Футер
        self.save_button = ttk.Button(
            self.root_extra_info,
            text="Выход",
            command=self.root_extra_info.destroy,
            style = "Main.TButton"
        )
        self.save_button.place(relx=0, rely=0.85, relwidth=1, relheight=0.15)
        





#<Окошко ошибок> +++
class eror_popup:
    def __init__(self, parent, mess):
        self.root = Toplevel(parent)
        self.root.geometry("400x300")
        self.root.geometry(f"{int(self.root.winfo_screenwidth() * 0.2)}x{int(self.root.winfo_screenheight()*0.25)}")
        self.root.resizable(width=False, height=False)

        style_error = ttk.Style()

        style_error.theme_use('clam')

        style_error.configure("Error.TButton", font=("Helvetica", 30, "bold"), background="#696969", foreground="white",)
        style_error.map("Error.TButton", background=[("active", "#E52B50")])


        self.root.title("Ошибка")
        text_eror = Label(
            self.root,
            text = mess,
            font = ("Helvetica", 20,)
        )
        text_eror.place(relx = 0, rely = 0.17, relwidth = 1, relheight = 0.25)

        self.loading_chars = ["xxooo", "oxxoo", "ooxxo", "oooxx", "ooxxo", "oxxoo"]
        self.loading_index = 0
        self.loading_label = Label(
            self.root,
            text=self.loading_chars[0],
            font=("Helvetica", 40),
            fg = "#E52B50"
        )
        self.loading_label.place(relx=0, rely=0.4, relwidth=1, relheight=0.2)
        self.animate_loading()

        button_end = ttk.Button(
            self.root,
            text = "Окей",
            style = "Error.TButton",
            command = self.root.destroy
        )
        button_end.place(relx = 0, rely = 0.72, relwidth = 1, relheight = 0.28)

    def animate_loading(self):
        self.loading_index = (self.loading_index + 1) % len(self.loading_chars)
        self.loading_label.config(text=self.loading_chars[self.loading_index])
        self.root.after(500, self.animate_loading)
###

#<Просто окошко> +++
class popup:
    def __init__(self, parent, mess, title):
        self.root = Toplevel(parent)
        self.root.geometry(f"{int(self.root.winfo_screenwidth() * 0.2)}x{int(self.root.winfo_screenheight()*0.25)}")
        self.root.resizable(width=False, height=False)
        self.root.title(title)

        style_error = ttk.Style()

        style_error.theme_use('clam')

        style_error.configure("Popup.TButton", font=("Helvetica", 30, "bold"), background="#696969", foreground="white",)
        style_error.map("Popup.TButton", background=[("active", "black")])


        text_eror = Label(
            self.root,
            text = mess,
            font = ("Helvetica", 20, )
        )
        text_eror.place(relx = 0, rely = 0.17, relwidth = 1, relheight = 0.25)

        self.loading_chars = ["xxooo", "oxxoo", "ooxxo", "oooxx", "ooxxo", "oxxoo"]
        self.loading_index = 0
        self.loading_label = Label(
            self.root,
            text=self.loading_chars[0],
            font=("Helvetica", 40),
            fg = "Black"
        )
        self.loading_label.place(relx=0, rely=0.4, relwidth=1, relheight=0.2)
        self.animate_loading()

        button_end = ttk.Button(
            self.root,
            text = "Окей",
            style = "Popup.TButton",
            command = self.root.destroy
        )
        button_end.place(relx = 0, rely = 0.72, relwidth = 1, relheight = 0.28)
    def animate_loading(self):
        self.loading_index = (self.loading_index + 1) % len(self.loading_chars)
        self.loading_label.config(text=self.loading_chars[self.loading_index])
        self.root.after(500, self.animate_loading)
###

###
# ФУНКЦИЯ ЭКСПОРТА УЧИТЕЛЕЙ В ФОРМАТ РАСПИСАНИЯ
###

def export_teachers_to_schedule_format(conn):

    cursor = conn.cursor()
    
    teachers_list = []
    
    # Получаем всех учителей из БД
    cursor.execute("SELECT id, Surname, Name, Patrony, Trans FROM teachers ORDER BY Surname, Name")
    teachers_db = cursor.fetchall()
    
    days = ("Пн", "Вт", "Ср", "Чт", "Пт", "Сб")
    ful_day_to_short = {
            "Понедельник": "Пн",
            "Вторник": "Вт",
            "Среда": "Ср",
            "Четверг":"Чт",
            "Пятница":"Пт",
            "Суббота":"Сб",
        }
    for teacher_id, surname, name, patrony, trans in teachers_db:
        # Форматируем ФИО
        full_name = f"{surname} {name[0]}. {patrony[0]}."
        
        # Получаем исключения (дни и уроки, когда учитель недоступен)
        cursor.execute(
            "SELECT Day, Lessons FROM exceptions WHERE Rel = ?",
            (f"{teacher_id}",)
        )
        exceptions_db = cursor.fetchall()
        
        # Преобразуем исключения в формат exce
        exce = {
            "Пн": [], 
            "Вт": [], 
            "Ср": [], 
            "Чт": [],   
            "Пт": [], 
            "Сб": []
        }
        for e in exceptions_db:
            for lesson in e[1].split(", "):
                exce[ful_day_to_short[e[0]]].append(int(lesson))


        # Получаем приоритеты кабинетов
        cursor.execute(
            "SELECT Building, Rooms FROM prio WHERE Rel = ?",
            (f"{teacher_id}",)
        )
        prio_db = cursor.fetchall()
        
        # Преобразуем приоритеты в формат prio
        prio = {"1": [], "2": []}
        for building, rooms_str in prio_db:
            if building == "Лицей":
                b = 1
            else:
                b = 2
            for r in rooms_str.split(", "):
                prio[str(b)].append(int(r))
            
        # Генерируем цвет (случайный красивый цвет)
        colors = [
            "FFD1DC", "EEE6A3", "EFA94A", "EFA94A", "7FB5B5", "5D9B9B",
            "E7ECFF", "77DD77", "FF7514", "FF8C69", "FF9BAA", "FFB28B",
            "FCE883", "BEBD7F", "C6DF90", "99FF99", "AFDAFC", "E6E6FA",
            "B5F2EA", "F5F5DC", "E4717A", "B39F7A", "E6D690", "EAE0C8",
            "F2E8C9", "F2DDC6", "FDF4E3", "C6D8FF", "3EB489", "ACE5EE",
            "A8E4A0", "CCCCFF", "FAE7B5", "FADADD", "AFEEEE", "ACB78E",
            "DAD871", "FFCF48", "A2A2D0", "FFC1CC", "FCD975", "5F9EA0", 
            "FFBD88", "9FE2BF", "71BC78", "E5E4E2", "DCD0FF", "ACE1AF",
            "D8BFD8", "FFBCAD", "FFEFD5"
        ]
        color = random.choice(colors)

        if trans == "Да":
            t = 1
        else:
            t = 0

        # Создаём словарь учителя
        teacher_dict = {
            "name": full_name,
            "exce": exce,
            "prio": prio,
            "trans_1": t,
            "color": color
        }
        
        teachers_list.append(teacher_dict)
    
    return teachers_list


###
# ФУНКЦИЯ ЭКСПОРТА КЛАССОВ В ФОРМАТ РАСПИСАНИЯ
###

def export_classes_to_schedule_format(conn, cher_buildings):

    cher = cher_buildings
    cursor = conn.cursor()

    classes_list = []
    
    # Получаем все классы из БД
    cursor.execute("SELECT id, Number, Letter FROM parallels ORDER BY Number, Letter")
    parallels = cursor.fetchall()
    
    days_ru = ("Пн", "Вт", "Ср", "Чт", "Пт", "Сб")
    
    for parallel_id, number, letter in parallels:
        class_name = f"{number}{letter}"
        
        # Получаем информацию о buildings для этого класса
        if int(number) in cher[0]:
            b = { 
                    "Пн": "2",  
                    "Вт": "1",  
                    "Ср": "2", 
                    "Чт": "1",  
                    "Пт": "2",  
                    "Сб": "1" 
                }      
        else:
            b = { 
                    "Пн": "1",  
                    "Вт": "2",  
                    "Ср": "1", 
                    "Чт": "2",  
                    "Пт": "1",  
                    "Сб": "2" 
                }  
        # Получаем предметы для этого класса
        cursor.execute(
            """SELECT Subject, Hours, Id_teacher FROM lessons WHERE Id_parallel = ? ORDER BY Subject""",
            (str(parallel_id),)
        )
        lessons = cursor.fetchall()
        
        subjects = []

        #Вспомогалка для всего кроме двух групп и физкультур
        teacher_help = {}
        for subject, hours, id_teacher in lessons:
            
            
            # Получаем информацию об учителе
            cursor.execute(
                "SELECT Surname, Name, Patrony FROM teachers WHERE id = ?",
                (id_teacher,)
            )
            teacher_row = cursor.fetchone()
    
            surname, name, patrony = teacher_row
            teacher_name = f"{surname} {name[0]}. {patrony[0]}."    

            teacher_help[teacher_name] = [[], 0]
        
        
        for subject, hours, id_teacher in lessons:
            if subject != "Информатика_1" and subject != "Информатика_2" and subject != "Английский язык_1" and subject != "Английский язык_2":


                # Получаем информацию об учителе
                cursor.execute(
                    "SELECT Surname, Name, Patrony FROM teachers WHERE id = ?",
                    (id_teacher,)
                )
                teacher_row = cursor.fetchone()

                surname, name, patrony = teacher_row
                teacher_name = f"{surname} {name[0]}. {patrony[0]}."    

                teacher_help[teacher_name][0].append(subject)
                teacher_help[teacher_name][1] = teacher_help[teacher_name][1] + int(hours)
        #######

            
            
        #Добавление всего кроме двух групп и физры
        for key, val in teacher_help.items():
            # Пропускаем учителей, которые преподают только инфу и англ
            if not val[0]:
                continue
            
            n = ""
            for i in range(0, len(val[0])):
                if i == len(val[0]) - 1:
                    n += val[0][i]
                else:
                    n += f"{val[0][i]}/"

            subjects.append({
                "name": n,
                "teacher": (key,),
                "hours": val[1]
            })
        

       

        #Добавление физры
        cursor.execute("""SELECT Teacher FROM pe WHERE Id_parallel = ?""",
                       (parallel_id, ))
        values = cursor.fetchall()
        c = 0
        for v in values: 
            c += 1
        if c:
            subjects.append({
                "name": "Физра",
                "teacher": (v[0],),
                "hours": c
            })
        ###

        #Добавление инфы
        l_teachers = []
        h = None
        for subject, hours, id_teacher in lessons:
            if subject == "Информатика_1" or subject == "Информатика_2":


                # Получаем информацию об учителе
                cursor.execute(
                    "SELECT Surname, Name, Patrony FROM teachers WHERE id = ?",
                    (id_teacher,)
                )
                teacher_row = cursor.fetchone()
                surname, name, patrony = teacher_row
                teacher_name = f"{surname} {name[0]}. {patrony[0]}."  
                l_teachers.append(teacher_name)
                h = int(hours)
        if l_teachers:
            subjects.append({
                "name": "Инфа",
                "teacher": (l_teachers[0], l_teachers[1], ),
                "hours": h
            })
        #

        #Добавление англ
        a_teachers = []
        h = None

        for subject, hours, id_teacher in lessons:
            if subject == "Английский язык_1" or subject == "Английский язык_2":


                # Получаем информацию об учителе
                cursor.execute(
                    "SELECT Surname, Name, Patrony FROM teachers WHERE id = ?",
                    (id_teacher,)
                )
                teacher_row = cursor.fetchone()
                surname, name, patrony = teacher_row
                teacher_name = f"{surname} {name[0]}. {patrony[0]}."  
                a_teachers.append(teacher_name)
                h = int(hours)
        if a_teachers:
            subjects.append({
                "name": "Англ",
                "teacher": (a_teachers[0], a_teachers[1], ),
                "hours": h
            })
           
        ###


        # Создаём словарь класса
        class_dict = {
            "name": class_name,
            "buildings": b,
            "subjects": subjects
        }
        
        classes_list.append(class_dict)
        ###
    
    return classes_list

###
#Начало программы
if __name__ == "__main__":

    #Создание первого окна
    root_menu_1 = Tk()
    main_screen_1 = main_menu(root_menu_1)
    ###

    root_menu_1.mainloop()
