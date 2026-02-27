from tkinter import ttk
from tkinter import *
from sqlite3 import *
from os import environ, makedirs, path
from openpyxl import Workbook
import random

# Класс класса
class Class:
    def __init__(self, name, clas, teacher_lessons: dict):
        self.name = name
        self.clas = clas
        self.teacher_lessons = dict(teacher_lessons)
###

# Класс учителя
class Teacher:
    def __init__(self, name, exceptions=None):
        self.name = name
        self.exceptions = exceptions if exceptions else set()
###

#<Глав меню> - 1
class main_menu:
    def __init__(self, root):
    #Данные корня
        self.root_menu_1 = root
        self.root_menu_1.style = ttk.Style()
        root.style.theme_use('clam')
        self.root_menu_1.title("Главное окно")
        self.root_menu_1.geometry(f"{int(self.root_menu_1.winfo_screenwidth() * 0.64)}x{int(self.root_menu_1.winfo_screenheight()*0.7)}")
        self.root_menu_1.resizable(width=False, height=False) 
    ###

    #БД
        #Подключение баз данных
        database_folder = "exe timetable"
        makedirs(database_folder, exist_ok=True)
        db_path = path.join(database_folder, "InputData.db")
        self.all_data_base = connect(f"{db_path}")
        self.con_all_data_base = self.all_data_base.cursor()
        ###

    #Создание таблицы параллелей 

        #Таблица для параллели
        self.con_all_data_base.execute(""" CREATE TABLE IF NOT EXISTS parallels (id INTEGER PRIMARY KEY, 
                                        Letter TEXT, Number TEXT);""")
        self.all_data_base.commit()

        self.con_all_data_base.execute("""CREATE TABLE IF NOT EXISTS lessons (id INTEGER PRIMARY KEY, 
                                        Id_parallel TEXT, Subject TEXT, Hours TEXT, Id_teacher TEXT);""")
        self.all_data_base.commit()              
        ###


        #Таблица для кабинета
        self.con_all_data_base.execute(""" CREATE TABLE IF NOT EXISTS rooms (id INTEGER PRIMARY KEY,
                                        Number TEXT, Graph_or_Lyceum TEXT, Big_or_Small TEXT, Subject TEXT);""")
        self.all_data_base.commit()
        ###

        #Таблицы для учителя
        self.con_all_data_base.execute(""" CREATE TABLE IF NOT EXISTS teachers (id INTEGER PRIMARY KEY,
                                        Surname TEXT, Name TEXT, Patrony TEXT, Lesson TEXT);""")
        self.all_data_base.commit()

        self.con_all_data_base.execute(""" CREATE TABLE IF NOT EXISTS exceptions (id INTEGER PRIMARY KEY,
                                        Rel TEXT, Day TEXT, Begining TEXT, Ending TEXT);""")
        self.all_data_base.commit()
        ###

    ###

    #Проверка открытия окна
        self.menu_1 = None
        self.menu_2 = None
        self.menu_3 = None
        self.menu_4 = None
        self.menu_5 = None
        self.menu_6 = None
        self.menu_7 = None
    ###

        style_0 = ttk.Style()

        style_0.theme_use('clam')

        style_0.configure("Make_T.TButton", font=("Helvetica", 30, "bold"), background="#DCDCDC", foreground="orange",)

        style_0.configure("Make_Par.TButton", font=("Helvetica", 28, ), background="#DCDCDC", foreground="blue",)

        style_0.configure("Check_Par.TButton", font=("Helvetica", 28, ), background="#DCDCDC", foreground="blue",)

        style_0.configure("Make_Room.TButton", font=("Helvetica", 28, ), background="#DCDCDC", foreground="purple",)

        style_0.configure("Check_Room.TButton", font=("Helvetica", 28, ), background="#DCDCDC", foreground="purple",)

        style_0.configure("Make_Teacher.TButton", font=("Helvetica", 28, ), background="#DCDCDC", foreground="green",)

        style_0.configure("Check_Teacher.TButton", font=("Helvetica", 28, ), background="#DCDCDC", foreground="green",)

        self.root_menu_1.configure(bg="#e0dcd4")

#Виджеты - глав меню ++
    #Блок создания расписания
        self.creating_timetabe =  ttk.Button(
            self.root_menu_1,
            text = "Составление расписания",
            style="Make_T.TButton",
            command = self.open_making_settings
            )
        self.creating_timetabe.place(relx=0, rely=0, relwidth = 1, relheight=0.3)
    ###

    #Блок создания и проверки параллели
        creating_parallel = ttk.Button(
            self.root_menu_1,
            text = "Добавление паралелли",
            style="Make_Par.TButton",
            command = self.open_screen_add_parallel
            )
        creating_parallel.place(relx=0, rely = 0.3, relwidth = 1, relheight = 0.3)

        check_parallel = ttk.Button(
            self.root_menu_1,
            text = str(chr(int("2968", 16))),
            style="Check_Par.TButton",
            command = self.open_screen_check_parallel
            )
        check_parallel.place(relx = 0, rely = 0.6, relwidth = 1, relheight= 0.1)
    ###

    #Блок создания и проверки кабинетов
        creating_room = ttk.Button(
            self.root_menu_1,
            text = "Добавление кабинета",
            style="Make_Room.TButton",
            command = self.open_screen_add_room
            )
        creating_room.place(relx = 0, rely = 0.7, relwidth = 0.5, relheight = 0.2)

        check_room = ttk.Button(
            self.root_menu_1,
            text = str(chr(int("2968", 16))),
            style="Check_Room.TButton",
            command = self.open_screen_check_room
            )
        check_room.place(relx = 0, rely = 0.9, relwidth = 0.5, relheight = 0.1)
    ###

    #Блок создания и проверки учителей
        creating_teacher = ttk.Button(
            self.root_menu_1,
            text = "Добавление учителя",
            style="Make_Teacher.TButton",
            command = self.open_screen_add_teacher
            )
        creating_teacher.place(relx = 0.5, rely = 0.7, relwidth = 0.5, relheight=0.2)

        check_teacher = ttk.Button(
            self.root_menu_1,
            text = str(chr(int("2968", 16))),
            style = "Check_Teacher.TButton",
            command = self.open_screen_check_teacher
            )
        check_teacher.place(relx = 0.5, rely = 0.9, relwidth = 0.5, relheight = 0.1)
    ###

#Открытие окон
    #Открытие составления параллелей - 2
    def open_making_settings(self):
        if self.menu_1 is None or not self.menu_1.root_menu_1.winfo_exists(): #Проверка на открытость окна
            self.menu_1 = making_settings(self.root_menu_1, self.all_data_base, self.con_all_data_base, self)
            self.menu_1.root_menu_1.mainloop()
        else:
            self.menu_1.root_menu_1.focus()
    ###

    #Открытие составления параллелей - 2
    def open_screen_add_parallel(self):
        if self.menu_2 is None or not self.menu_2.root_menu_2.winfo_exists(): #Проверка на открытость окна
            self.menu_2 = menu_add_parallel(self.root_menu_1, self.all_data_base, self.con_all_data_base)
            self.menu_2.root_menu_2.mainloop()
        else:
            self.menu_2.root_menu_2.focus()
    ###

    #Открытие проверки параллелей - 3
    def open_screen_check_parallel(self):
        if self.menu_3 is None or not self.menu_3.root_menu_3.winfo_exists(): #Проверка на открытость окна
            self.menu_3 = menu_check_parallel(self.root_menu_1, self.all_data_base,  self.con_all_data_base)
            self.menu_3.root_menu_3.mainloop()
        else:
            self.menu_3.root_menu_3.focus()
    ###

    #Открытие создания кабинета - 4
    def open_screen_add_room(self):
        if self.menu_4 is None or not self.menu_4.root_menu_4.winfo_exists(): #Проверка на открытость окна
            self.menu_4 = menu_add_room(self.root_menu_1, self.all_data_base,  self.con_all_data_base)
            self.menu_4.root_menu_4.mainloop()
        else:
            self.menu_4.root_menu_4.focus()
    ###

    #Открытие проверки кабинета - 5
    def open_screen_check_room(self):
        if self.menu_5 is None or not self.menu_5.root_menu_5.winfo_exists(): #Проверка на открытость окна
            self.menu_5 = menu_check_room(self.root_menu_1, self.all_data_base, self.con_all_data_base)
            self.menu_5.root_menu_5.mainloop()
        else:
            self.menu_5.root_menu_5.focus()
    ###

    #Открытие создания учителя - 6
    def open_screen_add_teacher(self):  
        if self.menu_6 is None or not self.menu_6.root_menu_6.winfo_exists(): #Проверка на открытость окна
            self.menu_6 = menu_add_teacher(self.root_menu_1, self.all_data_base, self.con_all_data_base)
            self.menu_6.root_menu_6.mainloop()
        else:
            self.menu_6.root_menu_6.focus()
    ###

    #Открытие проверки учителей - 7
    def open_screen_check_teacher(self):    
        if self.menu_7 is None or not self.menu_7.root_menu_7.winfo_exists(): #Проверка на открытость окна
            self.menu_7 = menu_check_teacher(self.root_menu_1, self.all_data_base, self.con_all_data_base)
            self.menu_7.root_menu_7.mainloop()
        else:
            self.menu_7.root_menu_7.focus()
    ###

###






#<Настр перед сос> - 1 
class making_settings:
    def __init__(self, root_parent, base, con, parent_menu):
    #Данные корня
        self.base_menu_1 = base
        self.con_menu_1 = con
        self.root_parent_1 = root_parent
        self.root_menu_1 = Toplevel(self.root_parent_1)
        self.root_menu_1.title("Настройки составления")
        self.root_menu_1.geometry(f"{int(self.root_menu_1.winfo_screenwidth() * 0.4)}x{int(self.root_menu_1.winfo_screenheight()*0.44)}")
        self.root_menu_1.resizable(width=False, height=False) 
        self.parent_menu = parent_menu
    ###
        style_1 = ttk.Style()

        style_1.theme_use('clam')

        style_1.configure("TLabel", font=("Helvetica", 30, "italic"))

        style_1.configure("TEntry", fieldbackground="#DCDCDC")

        style_1.configure("TCombobox", fieldbackground="#DCDCDC", arrowsize = 0)
        style_1.map("TCombobox", fieldbackground=[("readonly", "#DCDCDC")])


        style_1.configure("Main_1.TButton", font=("Helvetica", 30, "bold"), background="#696969", foreground="white",)
        style_1.map("Main_1.TButton", background=[("active", "#D2691E")])

        self.root_menu_1.configure(bg="#e0dcd4")
    #Виджеты насторек перед сос ++ 
    #Перегородочки
        separator_1 = Frame(
            self.root_menu_1, 
            height=3,         
            bg='grey',         
            relief=FLAT,
            bd=1              
        )
        separator_1.place(relx=0, rely=0.245, relwidth=1, relheight=0.007) #Горизонтальная

    ###

    #Блок выбора пар классов
        self.options_1 = ["8", "9", "10", "11"]

        self.con_between_1 =ttk.Label(
            self.root_menu_1,
            text = "Корпуса:",
            style="TLabel",
            anchor="center"
            )
        self.con_between_1.place(relx=0, rely=0, relwidth=0.3, relheight=0.24)

        self.number_1 = ttk.Combobox(
            self.root_menu_1,
            values=self.options_1,
            style="TCombobox",
            font=("Helvetica", 27),
            state="readonly", 
            foreground="#4D4D4D"
        )
        self.number_1.place(relx=0.3, rely=0.04, relwidth=0.11, relheight=0.16)

        self.con_between_1 =ttk.Label(
            self.root_menu_1,
            text = "+",
            style="TLabel",
            anchor="center"
            )
        self.con_between_1.place(relx=0.41, rely=0, relwidth=0.08, relheight=0.24)

        self.number_2 = ttk.Combobox(
            self.root_menu_1,
            values=self.options_1,
            style="TCombobox",
            font=("Helvetica", 27),
            state="readonly", 
            foreground="#4D4D4D"

        )
        self.number_2.place(relx=0.49, rely=0.04, relwidth=0.11, relheight=0.16)


        self.con_between_2 =ttk.Label(
            self.root_menu_1,
            text = "И",
            style="TLabel",
            anchor="center"
            )
        self.con_between_2.place(relx=0.6, rely=0, relwidth=0.1, relheight=0.24)


        self.number_3 = ttk.Combobox(
            self.root_menu_1,
            values=self.options_1,
            style="TCombobox",
            font=("Helvetica", 27),
            state="readonly", 
            foreground="#4D4D4D"
        )
        self.number_3.place(relx=0.7, rely=0.04, relwidth=0.11, relheight=0.16)

        self.con_between_3 = ttk.Label(
            self.root_menu_1,
            text = "+",
            style="TLabel",
            anchor="center"
            )
        self.con_between_3.place(relx=0.81, rely=0, relwidth=0.08, relheight=0.24)

        self.number_4 = ttk.Combobox(
            self.root_menu_1,
            values=self.options_1,
            style="TCombobox",
            font=("Helvetica", 27),
            state="readonly", 
            foreground="#4D4D4D"
        )
        self.number_4.place(relx=0.89, rely=0.04, relwidth=0.11, relheight=0.16)
    ###

    #Блок выбора количесва перебежек
        self.num_dash = ttk.Label(
            self.root_menu_1,
            text="Множитель алгортима:",
            style="TLabel",
            anchor="center"
        )
        self.num_dash.place(relx=0, rely=0.25, relwidth=1, relheight=0.25)

        self.multi = ttk.Entry(
            self.root_menu_1,
            style="TEntry",
            font=("Helvetica", 27),
            foreground="#4D4D4D"
        )
        self.multi.place(relx=0.2, rely=0.44, relwidth= 0.6, relheight=0.2)
    ###

    ###

    #Кнопка начала составления расписаня
        self.begin_makin = ttk.Button(
            self.root_menu_1,
            text="Начать создание",
            style="Main_1.TButton",
            command = self.start_making
        )
        self.begin_makin.place(relx=0, rely=0.75, relwidth=1, relheight=0.25)
    ###

    #Функционал
    def start_making(self):
    #Входные данные
        input_class_1 = self.number_1.get()
        input_class_2 = self.number_2.get()
        input_class_3 = self.number_3.get()
        input_class_4 = self.number_4.get()

        multplier = self.multi.get()
    ###

    #Проверки перед дальнейшей программой
        #Проверки корпусов
        if input_class_1 not in self.options_1 or input_class_2 not in self.options_1 or input_class_3 not in self.options_1 or input_class_4 not in self.options_1:
            warning = eror_popup(self.root_menu_1, "Ошибка в корпусах")
            warning.root.mainloop()
            return
        if input_class_1 in (input_class_2, input_class_3, input_class_4) or input_class_2 in (input_class_3, input_class_4) or input_class_4 == input_class_3:
            warning = eror_popup(self.root_menu_1, "Ошибка в корпусах")
            warning.root.mainloop()
            return
        ###

        all_numbers = "0987654321"
        #Проверка множителя
        for i in multplier:
            if i not in all_numbers:
                warning = eror_popup(self.root_menu_1, "Ошибка в кол-ве перебежек")
                warning.root.mainloop()
                return  
        ###

    

        
            
        data_settings = ((input_class_1, input_class_2), (input_class_3, input_class_4), multplier)

        creating_timetable(self.con_menu_1, self.base_menu_1, data_settings, self.root_parent_1, self.parent_menu, self.root_menu_1, self.parent_menu)
        ###
    ###






#СОЗДАНИЕ РАСПИСАНИЯ
class creating_timetable:
    def __init__(self, con, base, settings, main_menu_root, main_menu, parent_root, main_menu_self):
        self.root = Toplevel(main_menu_root)
        self.root.title("Создание расписания")
        self.root.geometry(f"{int(self.root.winfo_screenwidth() * 0.32)}x{int(self.root.winfo_screenheight()*0.4)}")
        self.root.resizable(width=False, height=False)
        self.root.grab_set()
        self.root.focus_set()
        self.root.transient(main_menu_root)
        self.con = con
        self.base = base
        self.settingss = settings
        self.main_menu_root = main_menu_root
        self.main_menu_itself = main_menu
        parent_root.destroy()
        self.number_of_timetables_created = 0
        self.self_parent_menu = main_menu_self

        style_1_0 = ttk.Style()

        style_1_0.theme_use('clam')

        style_1_0.configure("1.TLabel", font=("Helvetica", 30, "italic"))

        style_1_0.configure("2.TLabel", font=("Helvetica", 60))

        style_1_0.configure("3.TLabel", font=("Helvetica", 30, "italic"))

        style_1_0.configure("TEntry", fieldbackground="#DCDCDC")

        style_1_0.configure("TCombobox", fieldbackground="#DCDCDC", arrowsize = 0)
        style_1_0.map("TCombobox", fieldbackground=[("readonly", "#DCDCDC")])


        style_1_0.configure("Main_1_0.TButton", font=("Helvetica", 30, "bold"), background="#696969", foreground="white",)
        style_1_0.map("Main_1_0.TButton", background=[("active", "#D2691E")])

        self.root.configure(bg="#e0dcd4")
    #Виджеты создания
        self.in_process = ttk.Label(
            self.root,
            text = "Расписания создаются!",
            style="1.TLabel",
            anchor="center"
            )
        self.in_process.place(relx=0, rely=0.1, relwidth=1, relheight=0.25)

    #Иконка загрузки
        self.loading_index = 0
        self.loading_label = ttk.Label(
            self.root,
            text= "֎",
            style= "2.TLabel",
            foreground = "#D2691E",
            anchor="center"
        )
        self.loading_label.place(relx=0, rely=0.3, relwidth=1, relheight=0.25)
        
    ###

    
    #Кнопка отмены 
        self.add_class = ttk.Button(
            self.root,
            text="Отмена",
            style="Main_1_0.TButton",
            command = self.root.destroy
        )
        self.add_class.place(relx=0, rely=0.75, relwidth=1, relheight=0.25)
    ###
 
    #Алгоритм
        self.multiplier = int(self.settingss[2])

        settings = (self.settingss[0], self.settingss[1]) # кнопочка в меню

        self.group_1 = (int(settings[0][0]), int(settings[0][1]))
        self.group_2 = (int(settings[1][0]), int(settings[1][1]))

        #Входыне данные из базы
        self.classes = []
        self.con.execute("""SELECT * FROM parallels""")
        list_parallels = self.con.fetchall() #Получение данных о параллелях
        for i in list_parallels:
            dict_lessons = {}
            self.con.execute("SELECT * FROM lessons WHERE Id_parallel = ?", (str(i[0])))
            lessons = self.con.fetchall() #Получение данных об уроках параллели

            list_english = []
            list_it = []
            for j in lessons:
                if j[2][0:len(j[2])-2] == "Информатика":
                    list_it.append((j[4]))
                    list_it.append((j[3]))
                elif j[2][0:len(j[2])-2] == "Английский":
                    list_english.append(j[4])
                    list_english.append(j[3])

            if len(list_english) > 2:
                dict_lessons[(list_english[0], list_english[2])] = list_english[1]
            elif len(list_english) > 0 :
                dict_lessons[(list_english[0], )] = list_english[1]


            if len(list_it) > 2:
                dict_lessons[(list_it[0], list_it[2])] = list_it[1]
            elif len(list_it) > 0:
                dict_lessons[(list_it[0], )] = list_it[1]

            for j in lessons:
                if not(j[2][0:len(j[2])-2] == "Информатика" or j[2][0:len(j[2])-2] == "Английский"):
                    if dict_lessons.get((j[4], ), False) is False:
                        dict_lessons[(j[4], )] = int(j[3])
                    else:
                        dict_lessons[(j[4], )] = dict_lessons[(j[4], )] + int(j[4])
            self.classes.append(Class(str(i[0]), int(i[2]), dict_lessons))
            ###

            #Учителя
        self.teachers = []
        answ = ("Понедельник", "Вторник", "Среда", "Четверг", "Пятница", "Суббота")
        self.con.execute("""SELECT * FROM teachers""")
        list_teachers = self.con.fetchall() #Получение данных об учителях
        for i in list_teachers:
            set_exce = set()
            self.con.execute("SELECT * FROM exceptions WHERE Rel = ?", (str(i[0]), ))
            exceptions = self.con.fetchall() #Получение данных об исключениях учителя
            if exceptions == []:
                self.teachers.append(Teacher(str(i[0]), exceptions= None))
            else:
                for j in exceptions:
                    day = 0
                    for р in range(0, len(answ)):
                        if answ[р] == j[2]:
                            day = р + 1
                            break
                    if j[3] == "0" and j[4] == "0":
                        for g in range(1, 9):
                            set_exce.add((day, g))
                    else:
                        for g in range(int(j[3]), int(j[4]) + 1):
                            set_exce.add((day, g))
                self.teachers.append(Teacher(str(i[0]), exceptions= set_exce))
            ###

        self.mega_end()

    # 1 ЭТАП АЛГОРИТМ_1

    # *) Определим день-/корпус по id
    def class_building_for_day(self, clas, day, group, stage):

        num = int(clas) 

        if stage == 1:
            if int(day) in range(4): #Сначала идет group_1
                if num in group:
                    return 1
                else:
                    return 2
            else:
                if num in group:
                    return 2
                else:
                    return 1

        elif stage == 2:
            if int(day) in range(4): #Сначала идет group_1
                if num not in group:
                    return 1
                else:
                    return 2
            else:
                if num not in group:
                    return 2
                else:
                    return 1

        elif stage == 3:
            if int(day) % 2 == 0:
                if num in group:
                    return 1
                else:
                    return 2
            else:
                if num in group:
                    return 2
                else:
                    return 1

        elif stage == 4:
            if int(day) % 2 != 0:
                if num in group:
                    return 1
                else:
                    return 2
            else:
                if num not in group:
                    return 2
                else:
                    return 1


# *) Определим корпус по слоту учитель
    def get_teacher_building(self, mode_boundary, slot):
        """
        mode=0 -> весь день корпус_1
        mode=1 -> весь день корпус_2
        mode=2 -> слоты <= boundary -> корпус_1, слоты > boundary -> корпус_2
        mode=3 -> слоты <= boundary -> корпус_2, слоты > boundary -> корпус_1
        """
        mode, boundary = mode_boundary
        if mode == 0:
            return 1
        
        elif mode == 1:
            return 2
        
        elif mode == 2:
            if boundary is None:
                boundary = 4
            return 1 if slot <= boundary else 2
        
        elif mode == 3:
            if boundary is None:
                boundary = 4
            return 2 if slot <= boundary else 1

# 1) Генерация пребывания учителей по дням
    def generate_teacher_day_plan(self, teachers, days=6, slots_per_day=8):
        """
        mode=0 -> весь день корпус_1
        mode=1 -> весь день корпус_2
        mode=2 -> слоты <= boundary -> корпус_1, слоты > boundary -> корпус_2
        mode=3 -> слоты <= boundary -> корпус_2, слоты > boundary -> корпус_1
        """
        plan = {}
        for t in teachers: #Выбор учителя
            for d in range(1, days+1):
                mode = random.choice([0, 1, 2, 3, 4, 5]) #Рандомный выбор мода
                if mode == 2 or  mode == 3:
                    boundary = random.randint(1, slots_per_day - 1) #Рандомный выбор слота переходжа 
                elif mode in [0, 1, 4, 5]:
                    boundary = None
                plan[(t.name, d)] = (mode % 2, boundary)
        return plan
        #dict[(t.name, d)] = (mode, boundary)
    
    # 2) Генерация расписания
    def generate_schedule(self, teachers, classes, multiplier, stage,
                                  days=6, slots_per_day=8):

        best_schedule = None
        best_trans = None
        best_teacher_day_plan = None

        tasks = []
        for class_info in classes:
            for t_name, need in class_info.teacher_lessons.items():
                tasks.append((class_info.name, t_name, need))

    
        for _ in range(self.multiplier):

            current_teacher_day_plan = self.generate_teacher_day_plan(teachers, days=6, slots_per_day=8)
            random.shuffle(tasks)

            current_schedule = {}
            current_teacher_day_count = {}
            count_left = 0  # сколько уроков суммарно не удалось поставить
            # Проходим по задачам
            for (cls_id, t_name, needed) in tasks:
                self.con.execute("SELECT * FROM parallels WHERE id = ?", (int(cls_id), ))
                cls_name = self.con.fetchone()[2] #Получение данных об уроках параллели
                left = int(needed)
                # Все day/slot в случайном порядке
                all_day_slots = [(d, s) for d in range(1, days+1) for s in range(1, slots_per_day+1)]
                random.shuffle(all_day_slots)

                for (d, slot) in all_day_slots:
                    if (d, slot) == (0, 0):
                        continue

                    if left <= 0:
                        break
                    # 1) Слот занят классом?
                    if (d, slot, cls_name) in current_schedule:
                        continue


                    Cont = True


                    # 2- 3) Найдем учителя и исключение?


                    for teacher in t_name:
                        tch_obj = None
                        for t in teachers:
                            if t.name == teacher:
                                tch_obj = t
                                break

                        if tch_obj == None:
                            Cont = False

                        if (d, slot) in tch_obj.exceptions:
                            Cont = False

                    # 4) Лимит 8 уроков
                    for teacher in t_name:
                        day_cnt = current_teacher_day_count.get((teacher, d), 0)
                        if day_cnt >= 8:
                            Cont = False


                    # 5) Корпус
                    for teacher in t_name:
                        c_bld = self.class_building_for_day(cls_name, d, self.group_1, stage)
                        mode_boundary = current_teacher_day_plan[(teacher, d)]
                        t_bld = self.get_teacher_building(mode_boundary, slot)
                        if t_bld != c_bld:
                            Cont = False

                    # 6) Учитель не занят тем же слотом
                    for (od, os, ocl), (oteacher, obld) in current_schedule.items():
                        if od == d and os == slot and oteacher in t_name:
                            Cont = False
                            break

                    if not Cont:
                        continue

                    # Ставим урок
                    for i in range(0, len(all_day_slots)):
                        if all_day_slots[i] == (d, slot):
                            all_day_slots[i] = (0, 0)

                    current_schedule[(d, slot, cls_id)] = (t_name, c_bld)
                    current_teacher_day_count[(t_name, d)] = day_cnt + 1
                    left -= 1

                # Если осталось
                count_left += left

            # Теперь смотрим, если count_left==0, то все уроки расставили
            if count_left == 0:
                current_trans = self.schedule_trans(current_schedule)
                if best_schedule is None:
                    best_schedule = current_schedule
                    best_trans = current_trans
                    best_teacher_day_plan = current_teacher_day_plan
                else:
                    if current_trans < best_trans:
                        best_schedule = current_schedule
                        best_trans = current_trans
            # Иначе -- не удалось расставить все, 
            # не учитываем такое решение (или можно запоминать "самое полное"?)

        return best_schedule, best_teacher_day_plan
        #schedule: dict[(day, slot, class_name)] = (teacher_name, building).

# 3) Проверка расписания + подсчёт переходов
    def schedule_trans(self, schedule):
        if schedule == None:
            return None, []

        teacher_trans = []
        trans = 0

        # teacher_day_usage: (teacher, day): [(slot, building)]
        teacher_day_usage = {}
        for (day_n, slot, class_name), (teacher_name, building) in schedule.items():
            # Соберём все уроки этого учителя в этот день
            for teacher in teacher_name:
                if teacher_day_usage.get((teacher, day_n), False) is False:
                    teacher_day_usage[(teacher, day_n)] = [(slot, building)]
                else:
                    teacher_day_usage[(teacher, day_n)].append((slot, building))

        # Теперь для каждого (teacher, day) смотрим, были ли реальные переходы
        for teacher in teacher_day_usage:
            start_point = teacher_day_usage[teacher][0]
            for slot_bulding in teacher_day_usage[teacher]:
                if slot_bulding[1] != start_point[1]:
                    teacher_trans.append((teacher[0], teacher[1]))
                    trans += 1
                    break    
        return trans, teacher_trans



# 2 ЭТАП АЛГОРИТМ_2

# schedule: dict[(day, slot, class_name)] = (teacher_id, building).
# teacher_day_plan: dict[(teacher_id, day)] = (mode, boundary)
# teacher_trans: [(teacher, day)...]
# tasks: [(current_slot, class_id)]

    def local_search(self, schedule, teacher_day_plan, cost, teacher_trans, multiplier, stage, teacherss):
        """
        schedule: dict[(day, slot, class_name)] = (teacher_id, building)
        teacher_day_plan: dict[(teacher_id, day)] = (mode, boundary)
        cost: текущее число переходов
        teacher_trans: список [(teacher_name, day), ...], где teacher имеет переход
        multiplier: число итераций (делим на 3 для каких-то жадных переборов)
        stage: режим, как классы распределяются по корпусам
        teacherss: список объектов Teacher
        """
        # Если уже нет переходов, можем вернуть текущее расписание
        if cost == 0:
            return schedule, cost
        if cost is None:
            return None, None

        # Сохраняем "предыдущее" (исходное) состояние
        prev_schedule = dict(schedule)
        prev_teacher_day_plan = dict(teacher_day_plan)
        prev_cost = cost
        prev_teacher_trans = list(teacher_trans)

        # Локальный проход по всем учителям-дням, у кого есть переход
        # (teacher_tran, day_of_tran)
        local_best_schedule = None
        local_best_cost = None

        for (teacher_tran, day_of_tran) in prev_teacher_trans:


            # Собираем "tasks" = список уроков (slot, class_name) учителя teacher_tran в day_of_tran
            tasks = []
            cont = True
            for (day, slot, class_name), (teachers_id, building) in prev_schedule.items():
                for teacher_id in teachers_id:
                    if teachers_id > 1:
                        cont = False
                    if day == day_of_tran and teachers_id == teacher_tran:
                        tasks.append((slot, class_name))
            if not cont:
                continue

            # Находим объект учителя
            body_teacher = None
            for teacher_obj in teacherss:
                if teacher_obj.name == teacher_tran:
                    body_teacher = teacher_obj
                    break

            # Сортируем по slot
            tasks.sort(key=lambda x: x[0])

            # Старые настройки (mode, boundary) учителя
            settings = prev_teacher_day_plan[(teacher_tran, day_of_tran)]

            # Пробуем new_mode = [0, 1] (весь день в корпус 1 или корпус 2)
            for new_mode in [0, 1]:
                # Делаем копию plan, schedule и счётчиков
                current_teacher_day_plan = dict(prev_teacher_day_plan)
                current_schedule = dict(prev_schedule)
                current_teacher_day_count = {}

                # Ставим учителю новый режим
                current_teacher_day_plan[(teacher_tran, day_of_tran)] = (new_mode, None)

                # В зависимости от старого settings (2,3) и нового (0,1) отбрасываем часть уроков?
                # Ниже - ваша логика, отлаженная чуть лучше
                boundary = settings[1]  # индекс "границы"

                # tasks_... - "часть" уроков. Но будьте осторожны:
                # tasks[:boundary], tasks[boundary:], etc.
                # d - здесь не определено, заменим на day_of_tran при заполнении словаря
                if settings[0] == 2 and new_mode == 0:
                    # Был 2 (1->2), стало 0 (только 1)
                    # Например, "tasks = tasks[boundary:]" ?
                    # Но вы делали tasks[settings[1]::], смотрите внимательно
                    tasks = tasks[ boundary : ]
                elif settings[0] == 2 and new_mode == 1:
                    # Был 2 (1->2), стало 1 (только 2)
                    tasks = tasks[ : boundary ]
                elif settings[0] == 3 and new_mode == 0:
                    # Был 3 (2->1), стало 0 (только 1)
                    tasks = tasks[ : boundary ]
                elif settings[0] == 3 and new_mode == 1:
                    # Был 3 (2->1), стало 1 (только 2)
                    tasks = tasks[ boundary : ]

                # Удаляем из current_schedule уроки, которые были у этого учителя в этот день
                # (раз мы сейчас будем заново их ставить)
                for (slot, cls_name) in tasks:
                    key_ = (day_of_tran, slot, cls_name)
                    if key_ in current_schedule:
                        del current_schedule[key_]

                # Пытаемся заново поставить эти tasks
                # multiplier / 3 -> нужно int(...)


                # Будем считать, сколько из tasks реально поставили
                # (или "count == len(tasks)")

                for _ in range(int(multiplier // 3)):
                    random.shuffle(tasks)
                    placed_count = 0

                    for (slottt, cls_id) in tasks:
                        # Получаем "cls_name" из базы:
                        self.con.execute("SELECT * FROM parallels WHERE id = ?", (int(cls_id[0]), ))
                        row = self.con.fetchone()
                        if not row:
                            # возможно cls_id не найден?
                            continue
                        cls_name = row[2]  # допустим, там в 3-й колонке имя
                        # Перебираем day/slot (все?)
                        all_day_slots = [(ddd, sss) for ddd in range(1, 3) for sss in range(1, 9)]
                        random.shuffle(all_day_slots)

                        for (ddd, sss) in all_day_slots:
                            if (ddd, sss) == (0,0):
                                continue
                            
                            # Если уже все поставили
                            if placed_count == len(tasks):
                                break
                            
                            # 1) Слот занят классом?
                            if (ddd, sss, cls_name) in current_schedule:
                                continue
                            
                            # 2) Исключение?
                            if (ddd, sss) in body_teacher.exceptions:
                                continue
                            
                            # 3) Лимит 8 уроков
                            day_cnt = current_teacher_day_count.get((teacher_tran, ddd), 0)
                            if day_cnt >= 8:
                                continue
                            
                            # 4) Корпус
                            c_bld = self.class_building_for_day(cls_name, ddd, self.group_1, stage)
                            # mode_boundary = (new_mode, None)? 
                            # Но вы используете current_teacher_day_plan[(teacher_tran, ddd)]:
                            mb = current_teacher_day_plan.get((teacher_tran, ddd), (new_mode, None))
                            t_bld = self.get_teacher_building(mb, sss)
                            if t_bld != c_bld:
                                continue
                            
                            # 5) Проверяем, не занят ли учитель (teacher_tran) в (ddd, sss)
                            conflict_found = False
                            for (od, os, ocl), (oteacher, obld) in current_schedule.items():
                                if od == ddd and os == sss and oteacher[0] == teacher_tran :
                                    conflict_found = True
                                    break
                            if conflict_found:
                                continue
                            
                            # Если всё ок:
                            current_schedule[(ddd, sss, cls_name)] = ((teacher_tran), c_bld)
                            current_teacher_day_count[((teacher_tran), ddd)] = day_cnt + 1
                            placed_count += 1
                            break  # выходим из цикла all_day_slots (ставим 1 урок)
                        
                    # После попытки расстановки "tasks":
                    local_cost_now, local_teacher_trans = self.schedule_trans(current_schedule)

                    # Если удалось поставить все:
                    if placed_count == len(tasks):
                        # либо у нас вообще local_best_cost ещё None:
                        if local_best_cost is None or local_cost_now < local_best_cost:
                            local_best_cost = local_cost_now
                            local_best_schedule = dict(current_schedule)
                    else:
                        # Даже если не все, может всё равно cost лучше?
                        if local_best_cost is None or local_cost_now < local_best_cost:
                            local_best_cost = local_cost_now
                            local_best_schedule = dict(current_schedule)

        # После обхода всех учителей/дней/модов – сравним local_best_cost и prev_cost
        if local_best_cost is not None and local_best_cost < prev_cost:
            # Улучшение
            return local_best_schedule, local_best_cost
        else:
            # Нет улучшения – вернём старые
            return prev_schedule, prev_cost

# *) Получение расписания
 
    def getting_the_timetable(self, teachers, classes, multiplier, days = 6, slots_per_day = 8):

        schedule_list = []
        cost_list = []

        # print("1//////////////////")
        best_schedule_1, best_teacher_day_plan_1 = self.generate_schedule(teachers, classes, multiplier, 1, days, slots_per_day)
        cost_1, teacher_trans_1 = self.schedule_trans(best_schedule_1)

        # print(f"Расписание: {best_schedule_1}", end="\n")
        # print(f"Переходов: {cost_1}, Кто именно: {teacher_trans_1}")

        # print("------------------")

        end_schedule_1, end_cost_1  = self.local_search(best_schedule_1, best_teacher_day_plan_1, cost_1, teacher_trans_1, multiplier, 1, teachers)
        # print(f"Расписание: {end_schedule_1}", end="\n")
        # print(f"Переходов: {end_cost_1}")
        if end_schedule_1:
            schedule_list.append(end_schedule_1)
            cost_list.append(end_cost_1)
            if end_cost_1 == 0:
                return end_schedule_1, end_cost_1


        # print("2//////////////////")

        best_schedule_2, best_teacher_day_plan_2 = self.generate_schedule(teachers, classes, multiplier, 2, days=6, slots_per_day=8)
        cost_2, teacher_trans_2 = self.schedule_trans(best_schedule_2)
        # print(f"Расписание: {best_schedule_2}", end="\n")
        # print(f"Переходов: {cost_2}, Кто именно: {teacher_trans_2}")

        # print("------------------")

        end_schedule_2, end_cost_2  = self.local_search(best_schedule_2, best_teacher_day_plan_2, cost_2, teacher_trans_2, multiplier, 1, teachers)
        # print(f"Расписание: {end_schedule_2}", end="\n")
        # print(f"Переходов: {end_cost_2}")
        if end_schedule_2:
            schedule_list.append(end_schedule_2)
            cost_list.append(end_cost_2)
            if end_cost_2 == 0:
                return end_schedule_2, end_cost_2


        # print("3//////////////////")

        best_schedule_3, best_teacher_day_plan_3 = self.generate_schedule(teachers, classes, multiplier, 3, days=6, slots_per_day=8)
        cost_3, teacher_trans_3 = self.schedule_trans(best_schedule_3)
        # print(f"Расписание: {best_schedule_3}", end="\n")
        # print(f"Переходов: {cost_3}, Кто именно: {teacher_trans_3}")

        # print("------------------")

        end_schedule_3, end_cost_3  = self.local_search(best_schedule_3, best_teacher_day_plan_3, cost_3, teacher_trans_3, multiplier, 1, teachers)
        # print(f"Расписание: {end_schedule_3}", end="\n")
        # print(f"Переходов: {end_cost_3}")
        if end_schedule_3:
            schedule_list.append(end_schedule_3)
            cost_list.append(end_cost_3)
            if end_cost_3 == 0:
                return end_schedule_3, end_cost_3


        # print("4//////////////////")

        best_schedule_4, best_teacher_day_plan_4 = self.generate_schedule(teachers, classes, multiplier, 4, days=6, slots_per_day=8)
        cost_4, teacher_trans_4 = self.schedule_trans(best_schedule_4)
        # print(f"Расписание: {best_schedule_4}", end="\n")
        # print(f"Переходов: {cost_4}, Кто именно: {teacher_trans_4}")

        # print("------------------")

        end_schedule_4, end_cost_4  = self.local_search(best_schedule_4, best_teacher_day_plan_4, cost_4, teacher_trans_4, multiplier, 1, teachers)
        # print(f"Расписание: {end_schedule_4}", end="\n")
        # print(f"Переходов: {end_cost_4}")
        if end_schedule_4:
            schedule_list.append(end_schedule_4)
            cost_list.append(end_cost_1)
            if end_cost_4 == 0:
                return end_schedule_4, end_cost_4

        cost_list = [end_cost_1, end_cost_2, end_cost_3, end_cost_4]

        if schedule_list:
            min_cost = min(cost_list)
            for i in range(0, len(cost_list)):
                if cost_list[i] == min_cost:
                    return schedule_list[i], min_cost
        else: 
            return None, None
        


# 3 ЭТАП ВЫБОР УРОКА И КАБА    

    def getting_lessons_and_rooms(self):
        # Шаг 1: получаем schedule
        schedule, cost = self.getting_the_timetable(self.teachers, self.classes, self.multiplier, days= 6, slots_per_day=8)

        ### Все про учителей
        teacher_lessons = {}
        # teacher_lessons: dict[(teacher_id, class_id)] = [(number_hours, lesson)]
        self.con.execute("""SELECT * FROM lessons""")
        list_lessons = self.con.fetchall()
        for (id_lesson, Id_parallel, Subject, Hours, Id_teacher) in list_lessons:
            if teacher_lessons.get((Id_teacher, Id_parallel), False) is False:
                teacher_lessons[(Id_teacher, Id_parallel)] = [(Hours, Subject)]
            else:
                teacher_lessons[(Id_teacher, Id_parallel)].append((Hours, Subject))
        ###

        ### Все про кабинеты
        # room_lesson: dict[((building_name, size),(tuple_of_subjects))] = ["room_1", ... , "room_n"]
        room_lesson = {}

        # list_room_without_subject: dict[((building_name, size),(tuple(...)))] = ["room_1", ... , "room_n"]
        #   здесь "Subject" == "Нет приоритетных кабинетов"
        list_room_without_subject = {}

        self.con.execute("""SELECT * FROM rooms""")
        list_rooms = self.con.fetchall()
        for room in list_rooms:
            room_id       = room[0]
            number        = room[1]  
            building_name = room[2] 
            size          = room[3] 
            subjects_str  = room[4]  

            if subjects_str != "Нет приоритетных кабинетов":
                # Преобразуем список предметов в tuple
                subjects_tuple = tuple(subjects_str.split("/"))
                key = ((building_name, size), subjects_tuple)
                if key not in room_lesson:
                    room_lesson[key] = [number]
                else:
                    room_lesson[key].append(number)
            else:
                # Нет приоритетных кабинетов
                # Чтобы было единообразно, пусть тоже будет tuple
                subjects_tuple = tuple([subjects_str])  # фактически ("Нет приоритетных кабинетов",)
                key = ((building_name, size), subjects_tuple)
                if key not in list_room_without_subject:
                    list_room_without_subject[key] = [number]
                else:
                    list_room_without_subject[key].append(number)

        # Словари для занятых комнат по корпусам
        # вместо occupied_rooms_biulding_1 и occupied_rooms_biulding_2
        occupied_rooms_building = {
            1: {},
            2: {}
        }

        ### Формируем выход
        output_schedule = {} 
        # schedule: dict[(day, slot, class_id)] = (teacher_ids, building)
        # output_schedule: dict[(day, slot, class_id)] = (subject, teacher_ids, rooms, building)

        for (day, slot, class_id), (teacher_ids, building) in schedule.items():
            # 1) Уменьшаем счётчик часов + определяем subject_output
            subject_output = None
            # teacher_ids может быть (teacher_id,) или (teacher1, teacher2)
            # Найдём подходящий предмет
            if len(teacher_ids) > 0:
                t_main = teacher_ids[0]
                # Перебираем в teacher_lessons[(t_main, class_id)] список (Hours, Subject)
                if (t_main, class_id) in teacher_lessons:
                    for i in range(len(teacher_lessons[(t_main, class_id)])):
                        hours_left_str, subj = teacher_lessons[(t_main, class_id)][i]
                        if hours_left_str != "0":
                            subject_output = subj
                            new_hours = str(int(hours_left_str) - 1)
                            teacher_lessons[(t_main, class_id)][i] = (new_hours, subj)
                            break

            # Если не нашли, условно пусть будет "Неизвестный" (или None)
            if not subject_output:
                subject_output = "Неизвестный"

            # 2) Определяем, сколько аудиторий нужно
            needed_rooms = len(teacher_ids)  # 1 или 2
            # 3) Берём словарь занятых аудиторий для нужного корпуса
            current_occupied = occupied_rooms_building[building]
            if (day, slot) not in current_occupied:
                current_occupied[(day, slot)] = []

            # 4) Пытаемся найти подходящие аудитории
            end_1_1 = False
            end_1_2 = False
            end_2_1 = False
            end_2_2 = False

            list_rooms_out_put = []

            # --- ПЕРВЫЙ проход: ищем в приоритетных (room_lesson)
            for data_room, rooms_list in room_lesson.items():
                (bld_name, size) = data_room[0]
                subjects_tuple   = data_room[1]

                # Определим building_room = 1 или 2
                if bld_name == "Графский":
                    building_room = 2
                else:
                    building_room = 1

                # Проверяем тот ли это корпус
                if building_room != building:
                    continue
                # Фильтруем "Маленькие" аудитории (если они не нужны)
                if size == "Маленький":
                    # Если вы не хотите маленькие — пропускаем
                    continue

                # Проверка на предмет (subject_output должен быть в subjects_tuple)
                if subject_output not in subjects_tuple:
                    continue

                # Теперь ищем нужное кол-во свободных комнат
                if needed_rooms == 1:
                    # Ищем первую свободную
                    for r in rooms_list:
                        if r not in current_occupied[(day, slot)]:
                            current_occupied[(day, slot)].append(r)
                            room_output = r
                            end_1_1 = True
                            break
                    if end_1_1:
                        break
                else:
                    # Нужны 2 аудитории
                    for r in rooms_list:
                        if r not in current_occupied[(day, slot)]:
                            list_rooms_out_put.append(r)
                            current_occupied[(day, slot)].append(r)
                            if len(list_rooms_out_put) == 2:
                                end_1_2 = True
                                break
                    if end_1_2:
                        break

            # --- ВТОРОЙ проход, если ещё не набрали (room_lesson не помог)
            if needed_rooms == 1 and not end_1_1:
                # Ищем в list_room_without_subject
                for data_room, rooms_list in list_room_without_subject.items():
                    (bld_name, size) = data_room[0]
                    # здесь data_room[1] обычно = ("Нет приоритетных кабинетов",)

                    if bld_name == "Графский":
                        building_room = 2
                    else:
                        building_room = 1

                    if building_room != building:
                        continue
                    if size == "Маленький":
                        continue

                    for r in rooms_list:
                        if r not in current_occupied[(day, slot)]:
                            current_occupied[(day, slot)].append(r)
                            room_output = r
                            end_2_1 = True
                            break
                    if end_2_1:
                        break

            if needed_rooms == 2 and not end_1_2:
                # Ищем в list_room_without_subject
                for data_room, rooms_list in list_room_without_subject.items():
                    (bld_name, size) = data_room[0]

                    if bld_name == "Графский":
                        building_room = 2
                    else:
                        building_room = 1

                    if building_room != building:
                        continue
                    if size == "Маленький":
                        continue

                    for r in rooms_list:
                        if r not in current_occupied[(day, slot)]:
                            list_rooms_out_put.append(r)
                            current_occupied[(day, slot)].append(r)
                            if len(list_rooms_out_put) == 2:
                                end_2_2 = True
                                break
                    if end_2_2:
                        break

            # --- ТРЕТИЙ проход (fallback): если по-прежнему нет аудитории
            # (или только 1 из 2)
            # Для упрощения – дублируем небольшую логику
            if needed_rooms == 1 and not (end_1_1 or end_2_1):
                # Ищем любую аудиторию в room_lesson (без проверки subject)
                # или повторно перебираем всё что есть
                for data_room, rooms_list in room_lesson.items():
                    (bld_name, size) = data_room[0]
                    if bld_name == "Графский":
                        building_room = 2
                    else:
                        building_room = 1
                    if building_room != building:
                        continue

                    for r in rooms_list:
                        if r not in current_occupied[(day, slot)]:
                            current_occupied[(day, slot)].append(r)
                            room_output = r
                            break
                    else:
                        # если не нашли в этом data_room, идём к следующему
                        continue
                    # если вышли отсюда, значит аудитория найдена
                    break

                output_schedule[(day, slot, class_id)] = (subject_output, teacher_ids, room_output, building)

            elif needed_rooms == 1 and (end_1_1 or end_2_1):
                # Уже нашли одну аудиторию
                output_schedule[(day, slot, class_id)] = (subject_output, teacher_ids, room_output, building)

            elif needed_rooms == 2 and not (end_1_2 or end_2_2):
                # Значит ещё не успели добрать 2 аудитории
                # Пытаемся добрать в room_lesson вообще любые 2
                for data_room, rooms_list in room_lesson.items():
                    (bld_name, size) = data_room[0]
                    if bld_name == "Графский":
                        building_room = 2
                    else:
                        building_room = 1
                    if building_room != building:
                        continue

                    for r in rooms_list:
                        if len(list_rooms_out_put) == 2:
                            break
                        if r not in current_occupied[(day, slot)]:
                            list_rooms_out_put.append(r)
                            current_occupied[(day, slot)].append(r)
                    if len(list_rooms_out_put) == 2:
                        break

                output_schedule[(day, slot, class_id)] = (
                    subject_output, 
                    teacher_ids, 
                    tuple(list_rooms_out_put),  # т.к. две комнаты
                    building
                )
            else:
                # needed_rooms == 2 и (end_1_2 or end_2_2) уже нашли 2 аудитории
                output_schedule[(day, slot, class_id)] = (
                    subject_output, 
                    teacher_ids, 
                    tuple(list_rooms_out_put), 
                    building
                )

        return output_schedule 


# ВЫВОД EXEL
#output_schedule: dict[(day- позиционка, slot- позиционка, class_id- преобраз)] = (subject- +, teacher_id- преобраз, room- +, building- условие)

    def getting_exel(self, schedule):
        wb = Workbook()
        ws = wb.active

        # 1) Вспомогалки
        # Для быстрого нахождения номер + буква
        id_to_LettNum = {}
        list_all_classes = []
        self.con.execute("""SELECT * FROM parallels""")
        list_parallels = self.con.fetchall()
        for parallel in list_parallels:
            id_to_LettNum[str(parallel[0])] = (f"{parallel[2]}{parallel[1]}")
            list_all_classes.append((int(parallel[2]), parallel[1], parallel[0]))
        properly_sorted_all_classes = sorted(list_all_classes, key = lambda x: (x[2], x[1]))

        # Для быстрого нахождения ФИО учителя
        id_to_SNO = {}
        self.con.execute("""SELECT * FROM teachers""")
        list_teachers = self.con.fetchall()
        for teacher in list_teachers:
            id_to_SNO[str(teacher[0])] = f"{teacher[1]} {teacher[2][0]}. {teacher[3][0]}."

        #Каркас для вывода
        day_id_to_word = ("Понедельник", "Вторник", "Среда", "Четверг", "Пятница", "Суббота")

        for start_column in range(0, len(properly_sorted_all_classes)):
            if start_column == 0:
                real_start_column = 1
            else:
                real_start_column = start_column*6

            for day_of_week_id in range(0, 6):
                if day_of_week_id == 0:
                    start_row = 1
                else:
                    start_row = (day_of_week_id * 11) + 1

                for slot_of_day in range(0, 9):
                    if slot_of_day == 0:
                        ws.cell(row= start_row + slot_of_day, column= real_start_column, value= day_id_to_word[day_of_week_id])
                        ws.cell(row= start_row + slot_of_day, column= real_start_column + 2, value= "Кабинет") 
                        ws.cell(row= start_row + slot_of_day, column= real_start_column + 3, value= "Учитель")
                    else:
                        ws.cell(row= start_row + slot_of_day, column= real_start_column, value= slot_of_day)

    #output_schedule: dict[(day- позиционка, slot- позиционка, class_id- преобраз)] = (subject- +, teacher_id- преобраз, room- +, building- условие)

    #ИСПРАВИТТЬ
        for (day, slot, class_id), (subject, teacher_id, room, building) in schedule.items():
            if day == 1:
                start_row = 1 + slot
                head_row = 1
            else:
                start_row = ((day - 1) * 11) + 1 + slot
                head_row = (day - 1) * 11 + 1

            if type(room) == type(tuple()):
                room = f"{room[0]}, {room[1]}"

            if len(teacher_id) > 1:
                teachers_out = f"{id_to_SNO[teacher_id[0]]}, {id_to_SNO[teacher_id[1]]}"
            else:
                teachers_out = f"{id_to_SNO[teacher_id[0]]}"

            count = None
            for i in range(0, len(properly_sorted_all_classes)):
                if str(properly_sorted_all_classes[i][2]) == str(class_id):
                    count = i

            if count == 0:
                start_column = 2
            else:
                start_column = (count * 6) + 1

            if building == 1:
                output_building = "Лицей"
            else:
                output_building = "Графский"

            ws.cell(row= head_row, column= start_column, value= f"{output_building}, {id_to_LettNum[class_id]}") 
            ws.cell(row= start_row, column= start_column, value= subject)
            ws.cell(row= start_row, column= start_column + 1, value= room) 
            ws.cell(row= start_row, column= start_column + 2, value= teachers_out)

        wb.save("Расписание.xlsx")


    def mega_end(self): 
        end_schedule = self.getting_lessons_and_rooms()
        if end_schedule:
            self.getting_exel(end_schedule)
            self.root.destroy()
            popup(self.self_parent_menu.root_menu_1, f"Расписание создано в папке!", "Успех").root.mainloop()
            
            
        else:
            self.root.destroy()
            warning = eror_popup(self.self_parent_menu.root_menu_1, "Расписание не создано")
            warning.root.mainloop()


    ###







#<Созд пара> - 2 
class menu_add_parallel:
    def __init__(self, parent, base, connect):
        
    #Данные корня
        self.base_menu_2 = base
        self.con_menu_2 = connect
        self.root_menu_2 = Toplevel(parent)
        self.root_menu_2.title("Добавление параллели")
        self.root_menu_2.geometry(f"{int(self.root_menu_2.winfo_screenwidth() * 0.56)}x{int(self.root_menu_2.winfo_screenheight()*0.6)}")
        self.root_menu_2.resizable(width=False, height=False)
        self.full_data_class = []
        self.clas_buttons = []
    ###

        style_2 = ttk.Style()

        style_2.theme_use('clam')

        style_2.configure("TLabel", font=("Helvetica", 30, "italic"))

        style_2.configure("TEntry", fieldbackground="#DCDCDC")

        style_2.configure("TCombobox", fieldbackground="#DCDCDC", arrowsize = 0)
        style_2.map("TCombobox", fieldbackground=[("readonly", "#DCDCDC")])


        style_2.configure("Main_2.TButton", font=("Helvetica", 30, "bold"), background="#696969", foreground="white",)
        style_2.map("Main_2.TButton", background=[("active", "#0000CD")])

        style_2.configure("Sec_2.TButton", font=("Helvetica", 13, "bold"), background="#696969", foreground="white",)
        style_2.map("Sec_2.TButton", background=[("active", "#4169E1")])

        self.root_menu_2.configure(bg="#e0dcd4")


    #Виджеты - создания параллели +++
    #Границы 
        separator_1 = Frame(
            self.root_menu_2, 
            height=3,         
            bg='grey',         
            relief=FLAT,
            bd=1              
        )
        separator_1.place(relx=0, rely=0.14, relwidth=1, relheight=0.007) #Горизонтальная

        separator_1 = Frame(
            self.root_menu_2, 
            height=3,          
            bg='grey',        
            relief=FLAT,
            bd=1               
        )
        separator_1.place(relx=0.29, rely=0.14, relwidth= 0.004, relheight= 1) #Вертикальная
    ###

    #Блок выбора номера
        self.creating_number = ttk.Label(
            self.root_menu_2,
            text = "Номер:",
            style="TLabel",
            )
        self.creating_number.place(relx=0.1, rely=0, relwidth=0.20, relheight=0.14)

        
        self.options_1 = ["8", "9", "10", "11"]
        self.number = ttk.Combobox(
            self.root_menu_2,
            values=self.options_1,
            style="TCombobox",
            font=("Helvetica", 27),
            state="readonly", 
            foreground="#4D4D4D"
        )
        self.number.place(relx=0.3, rely=0, relwidth=0.10, relheight=0.14)
    ###

    #Блок выбора буквы
        self.creating_letter = ttk.Label(
            self.root_menu_2,
            text = "Буква:",
            style = "TLabel"
            )
        self.creating_letter.place(relx=0.5, rely=0, relwidth=0.20, relheight=0.14)

        options_2 = ["А", "Б", "В", "Г", "Д"]
        self.letter = ttk.Combobox(
            self.root_menu_2,
            values=options_2,
            style="TCombobox",
            font=("Helvetica", 27),
            state="readonly", 
            foreground="#4D4D4D"
        )
        self.letter.place(relx=0.7, rely=0, relwidth=0.10, relheight=0.14)
    ###

    #Блок выбора урока
        self.creating_subject = ttk.Label(
            self.root_menu_2,
            text = "Урок:",
            style ="TLabel",
            anchor="center"
            )
        self.creating_subject.place(relx=0, rely=0.15, relwidth=0.29, relheight=0.1)

        self.options_3 = ["Алгебра",
            "Геометрия",
            "ТеорВер",
            "Физика",
            "Русский",
            "Литература",
            "История",
            "Информатика_1",
            "Информатика_2",
            "География",
            "Английский язык_1",
            "Английский язык_2",
            "Физра",
            "Биология",
            "Химия",
            "Обществознание",
            "ОБЖ",
            "Технология"]
        
        self.subject = ttk.Combobox(
            self.root_menu_2,
            values=self.options_3,
            style="TCombobox",
            font=("Helvetica", 27),
            state="readonly", 
            foreground="#4D4D4D"
        )
        self.subject.place(relx=0, rely=0.25, relwidth=0.29, relheight=0.1)
    ###

    #Блок выбора учителя
        self.options_teacher = []
        self.con_menu_2.execute("""SELECT * FROM teachers""")
        values = self.con_menu_2.fetchall() # Проверка на предварительное наличие кабинета 

        for i in values:
            self.options_teacher.append(f"{i[1]} {i[2][0]}. {i[3][0]}.")

        self.creating_teacher = ttk.Label(
            self.root_menu_2,
            text = "Учитель:",
            style="TLabel",
            anchor="center"
            )
        self.creating_teacher.place(relx=0, rely=0.35, relwidth=0.29, relheight=0.1)
        
        self.teacher = ttk.Combobox(
            self.root_menu_2,
            values=self.options_teacher,
            style="TCombobox",
            font=("Helvetica", 27),
            state="readonly", 
            foreground="#4D4D4D"
        )
        self.teacher.place(relx=0, rely=0.45, relwidth=0.29, relheight=0.1)
    ###

    #Блок выбора количества часов
        self.creating_hours = ttk.Label(
            self.root_menu_2,
            text = "Кол-во часов:",
            style="TLabel",
            anchor="center"
            )
        self.creating_hours.place(relx=0, rely=0.55, relwidth=0.29, relheight=0.1)
        
        self.hours = ttk.Entry(
            self.root_menu_2,
            style="TEntry",
            font=("Helvetica", 27),
            foreground="#4D4D4D"
        )
        self.hours.place(relx=0, rely=0.65, relwidth=0.29, relheight=0.1)
    ###

    #Футер
    #Кнопка добавления урока 
        self.add_class = ttk.Button(
            self.root_menu_2,
            text="Добавить",
            style = "Main_2.TButton",
            command = self.add_class
        )
        self.add_class.place(relx=0, rely=0.75, relwidth=0.29, relheight=0.1)
    ###

    #Кнопка добавления параллели
        self.add_parallel = ttk.Button(
            self.root_menu_2,
            text="Добавить параллель",
            style = "Main_2.TButton",
            command= self.addd_parallel
        )
        self.add_parallel.place(relx=0, rely=0.85, relwidth=1, relheight=0.15)
    ###

    #Функционал
    #Добавение параллели
    def addd_parallel(self):
        letter = self.letter.get()
        num = self.number.get()

        # Проверка существования параллели в БД
        self.con_menu_2.execute(
            "SELECT * FROM parallels WHERE Letter = ? AND Number = ?",
            (letter, num)
        )
        if self.con_menu_2.fetchall():
            warning = eror_popup(self.root_menu_2, "Такая параллель уже есть")
            warning.root.mainloop()
            return
        ###

        #Проверка на 2 групппы
        list_sub = []
        for _ in self.full_data_class:
            if _:
                list_sub.append(_[0])

        if ("Информатика_1" in list_sub and "Информатика_2" not in list_sub) or ("Информатика_1" not in list_sub and "Информатика_2" in list_sub):
            warning = eror_popup(self.root_menu_2, "Созданы не 2 группы")
            warning.root.mainloop()
            return
        
        if ("Английский язык_1" in list_sub and "Английский язык_2" not in list_sub) or ("Английский язык_1" not in list_sub and "Английский язык_2" in list_sub):
            warning = eror_popup(self.root_menu_2, "Созданы не 2 группы")
            warning.root.mainloop()
            return



        # Добавляем параллель в таблицу parallels
        
        self.con_menu_2.execute(
            "INSERT INTO parallels (Letter, Number) VALUES (?, ?)",
            (letter, num)
        )
        self.base_menu_2.commit()
        ###

        # Получаем id добавленной параллели
        self.con_menu_2.execute(
            "SELECT id FROM parallels WHERE Letter = ? AND Number = ?",
            (letter, num)
        )
        parallel_id = self.con_menu_2.fetchone()[0]
        ###

        

        # Добавляем уроки в таблицу lessons с id учителя
        for subject_data in self.full_data_class:
            if subject_data:  # Проверяем, что данные не пустые

                # Получаем id учителя по ФИО
                teacher_fio = subject_data[1].split()  # Разбиваем "Фамилия И. О." на части
                surname = teacher_fio[0]
                name = teacher_fio[1][0]  # Первая буква имени
                patrony = teacher_fio[2][0]  # Первая буква отчества

                # Получаем id учителя из таблицы teachers
                self.con_menu_2.execute(
                    "SELECT id FROM teachers WHERE Surname = ? AND Name LIKE ? AND Patrony LIKE ?",
                    (surname, f"{name}%", f"{patrony}%")
                )
                teacher_id = self.con_menu_2.fetchone()[0]

                # Вставляем данные с Id_parallel и Id_teacher
                self.con_menu_2.execute(
                    "INSERT INTO lessons (Id_parallel, Subject, Hours, Id_teacher) VALUES (?, ?, ?, ?)",
                    (parallel_id, subject_data[0], subject_data[2], teacher_id)
                )
        self.base_menu_2.commit()

        # Очищаем форму
        self.letter.set('')
        self.number.set('')
        self.hours.delete(0, END)
        self.hours.insert(0, '')
        self.teacher.set('')
        self.subject.set('')
        self.full_data_class.clear()
        self.display_class()

        popup(self.root_menu_2, f"Параллель {num}{letter} создана!", "Успех").root.mainloop()
    ###



#См инфы об уроке
    def display_class(self):

    # Сначала удалим старые кнопки
        for btn in self.clas_buttons:
            btn.destroy()
        self.clas_buttons = []

        # Создаем список только непустых элементов
        active_classes = [clas for clas in self.full_data_class if clas]

        # Теперь используем индексы только активных классов
        for i, clas in enumerate(active_classes):
            pos_x = (i % 4) * 0.175 + 0.3
            pos_y = (i // 4) * 0.14 + 0.15

            btn = ttk.Button(
                self.root_menu_2,
                text=f"{clas[0]}",  
                command=lambda cl=(i, clas): self.open_popup_class(cl),
                style = "Sec_2.TButton"
            )
            btn.place(relx=pos_x, rely=pos_y, relwidth=0.175, relheight=0.14)
            self.clas_buttons.append(btn)

    #Открыть изменение предмета
    def open_popup_class(self, clas):
        popup_subject_1(clas[0], clas[1], self.root_menu_2, self.base_menu_2, self.con_menu_2, self)
    ###

    #Проверка на добавление урока
    def add_class(self):
        mass = []
        sub = self.subject.get()
        teacher = self.teacher.get()
        hours = self.hours.get()

        #Проверка урока
        if sub not in self.options_3:
            warning = eror_popup(self.root_menu_2, "Ошибка предмете")
            warning.root.mainloop()
            return
        mass.append(sub)

        
        for i in self.full_data_class:
            if i and i[0] == sub:  # Сначала проверяем, что i не пустой, затем сравниваем
                warning = eror_popup(self.root_menu_2, "Уже есть предмет")
                warning.root.mainloop()
                return
        ###

        #Проверка учителя
        if teacher not in self.options_teacher:
            warning = eror_popup(self.root_menu_2, "Нет такого учителя")
            warning.root.mainloop()
            return
        mass.append(teacher)
        ###

        #Проверка количества часов
        char = "0123456789"
        if hours == "":
            warning = eror_popup(self.root_menu_2, "Нет кол-ва часов")
            warning.root.mainloop()
            return

        for i in hours:
            if i not in char:
                warning = eror_popup(self.root_menu_2, "Ошибка кол-ве часов")
                warning.root.mainloop()
                return
        
        if int(hours) > 30:
            warning = eror_popup(self.root_menu_2, "Слишком много часов")
            warning.root.mainloop()
            return
        ###
        self.full_data_class.append([sub, teacher, hours])
        self.display_class()
        warning = popup(self.root_menu_2, "Предмет добавлен", f"{sub}")
###



#<Изменение урока> - # 
class popup_subject_1:
    def __init__(self, id, clas, parent, base, connect, parent_con):
    #Данные корня
        self.root_subject = Toplevel(parent)
        self.root_subject.geometry(f"{int(self.root_subject.winfo_screenwidth() * 0.22)}x{int(self.root_subject.winfo_screenheight()*0.41)}")
        self.root_subject.resizable(width=False, height=False)
        self.root_subject.title(f"Предмет {clas[0]}")
        self.base_subject = base
        self.con_subject = connect
        self.id = id
        self.clas = clas
        self.parent_con = parent_con
    ###
        style_2_0 = ttk.Style()

        style_2_0.theme_use('clam')

        style_2_0.configure("TLabel", font=("Helvetica", 30, "italic"))

        style_2_0.configure("TEntry", fieldbackground="#DCDCDC")

        style_2_0.configure("TCombobox", fieldbackground="#DCDCDC", arrowsize = 0)
        style_2_0.map("TCombobox", fieldbackground=[("readonly", "#DCDCDC")])


        style_2_0.configure("Main_2_0.TButton", font=("Helvetica", 16, "bold"), background="#696969", foreground="white",)
        style_2_0.map("Main_2_0.TButton", background=[("active", "#0000CD")])

        self.root_subject.configure(bg="#e0dcd4")

    #Виджеты изменения урока +++
    #Блок выбора урока
        self.creating_subject = ttk.Label(
            self.root_subject,
            text = "Урок:",
            style = "TLabel",
            anchor="center"
            )
        self.creating_subject.place(relx=0, rely=0, relwidth=1, relheight=0.14)

        self.options_3 = ["Алгебра",
            "Геометрия",
            "ТеорВер",
            "Физика",
            "Русский",
            "Литература",
            "История",
            "Информатика_1",
            "Информатика_2",
            "География",
            "Английский язык_1",
            "Английский язык_2",
            "Физра",
            "Биология",
            "Химия",
            "Обществознание",
            "ОБЖ",
            "Технология"]
        
        self.subject = ttk.Combobox(
            self.root_subject,
            values=self.options_3,
            style="TCombobox",
            font=("Helvetica", 27),
            state="readonly", 
            foreground="#4D4D4D"
        )
        self.subject.place(relx=0, rely=0.14, relwidth=1, relheight=0.14)
        self.subject.set(self.clas[0])
    ###

    #Блок выбора учителя
        self.options_teacher = []
        self.con_subject.execute("""SELECT * FROM teachers""")
        values = self.con_subject.fetchall() # Проверка на предварительное наличие кабинета 

        for i in values:
            self.options_teacher.append(f"{i[1]} {i[2][0]}. {i[3][0]}.")

        self.creating_teacher =ttk.Label(
            self.root_subject,
            text = "Учитель:",
            style = "TLabel",
            anchor="center"
            )
        self.creating_teacher.place(relx=0, rely=0.28, relwidth=1, relheight=0.14)
        
        self.teacher = ttk.Combobox(
            self.root_subject,
            values=self.options_teacher,
            style="TCombobox",
            font=("Helvetica", 27),
            state="readonly", 
            foreground="#4D4D4D"
        )
        self.teacher.place(relx=0, rely=0.42, relwidth=1, relheight=0.14)
        self.teacher.set(self.clas[1])
    ###

    #Блок выбора количества часов
        self.creating_hours = ttk.Label(
            self.root_subject,
            text = "Кол-во часов:",
            style="TLabel",
            anchor="center"
            )
        self.creating_hours.place(relx=0, rely=0.56, relwidth=1, relheight=0.14)
        
        self.hours = ttk.Entry(
            self.root_subject,
            style="TEntry",
            font=("Helvetica", 27),
            foreground="#4D4D4D"
        )
        self.hours.insert(0, self.clas[2])
        self.hours.place(relx=0, rely=0.7, relwidth=1, relheight=0.14)
    ###

    #Футер
        self.save_button = ttk.Button(
            self.root_subject,
            text="Сохранить",
            style = "Main_2_0.TButton",
            command=self.save_data  
        )
        self.save_button.place(relx=0, rely=0.84, relwidth=0.3, relheight=0.16)

        self.cancel_button = ttk.Button(
            self.root_subject,
            style = "Main_2_0.TButton",
            text="Отмена",
            command=self.root_subject.destroy
        )
        self.cancel_button.place(relx=0.3, rely=0.84, relwidth=0.4, relheight=0.16)

        self.delete_button = ttk.Button(
            self.root_subject,
            text="Удалить",
            style = "Main_2_0.TButton",
            command=self.delete_data
        )
        self.delete_button.place(relx=0.7, rely=0.84, relwidth=0.3, relheight=0.16)
    ###

    #Функуионал
    #Удалие данных
    def delete_data(self):
        lisstt = self.parent_con.full_data_class
        for i in range(0, len(lisstt)):
            if lisstt[i][0] == self.clas[0]: 
                self.parent_con.full_data_class[i] = []
                self.parent_con.display_class()
                self.root_subject.destroy()
                popup(self.parent_con.root_menu_2, "Предмет удален!", "Успех").root.mainloop()
                break
    ###

    #Изменение данных
    def save_data(self):
        # Получаем значения из полей ввода
        new_subject = self.subject.get()
        new_teacher = self.teacher.get()
        new_hours = self.hours.get()

        # Проверка на хоть какое-то изменение
        if ([new_subject, new_teacher, new_hours] == [self.clas[0], self.clas[1], self.clas[2]]):
            warning = eror_popup(self.root_subject, "Ничего не поменяли")
            warning.root.mainloop()
            return

        # Проверка урока
        if new_subject not in self.options_3:
            warning = eror_popup(self.root_subject, "Ошибка в предмете")
            warning.root.mainloop()
            return

        # Проверка на существование такого предмета (кроме текущего)
        for i, subject in enumerate(self.parent_con.full_data_class):
            if i != self.id and subject and subject[0] == new_subject:
                warning = eror_popup(self.root_subject, "Уже есть такой предмет")
                warning.root.mainloop()
                return

        # Проверка учителя
        if new_teacher not in self.options_teacher:
            warning = eror_popup(self.root_subject, "Нет такого учителя")
            warning.root.mainloop()
            return

        # Проверка количества часов
        char = "0123456789"
        if new_hours == "":
            warning = eror_popup(self.root_subject, "Нет кол-ва часов")
            warning.root.mainloop()
            return

        if not all(c in char for c in new_hours):
            warning = eror_popup(self.root_subject, "Ошибка в кол-ве часов")
            warning.root.mainloop()
            return

        if int(new_hours) > 30:
            warning = eror_popup(self.root_subject, "Слишком много часов")
            warning.root.mainloop()
            return

        # Обновляем значение в списке
        self.parent_con.full_data_class[self.id] = [new_subject, new_teacher, new_hours]

    # Обновляем отображение
        self.parent_con.display_class()
        self.root_subject.destroy()
        popup(self.parent_con.root_menu_2, "Предмет обновлен!", "Успех").root.mainloop()
        

    ###
        
    ###






# <СМ пара> - 3 
class menu_check_parallel:
    def __init__(self, parent, base, con):
    #Данные корня
        self.root_menu_3 = Toplevel(parent)
        self.root_menu_3.title("Просмотр параллелей")
        self.root_menu_3.geometry(f"{int(self.root_menu_3.winfo_screenwidth() * 0.56)}x{int(self.root_menu_3.winfo_screenheight()*0.6)}")
        self.root_menu_3.resizable(width=False, height=False) 
        self.con_root_menu_3 = con
        self.base_root_menu_3 = base
    ###
        style_3 = ttk.Style()

        style_3.theme_use('clam')

        style_3.configure("TLabel", font=("Helvetica", 40, "bold", "italic"))

        style_3.configure("Sec_3.TButton", font=("Helvetica", 30, "bold"), background="#696969", foreground="white",)
        style_3.map("Sec_3.TButton", background=[("active", "#4169E1")])

        self.root_menu_3.configure(bg="#e0dcd4")

    #Виджеты - просмотра параллелей +++ 
    #Границы 
        separator_1 = Frame(
            self.root_menu_3, 
            height=3,         
            bg='grey',         
            relief=FLAT,
            bd=1              
        )
        separator_1.place(relx=0, rely=0.5, relwidth=1, relheight=0.007) #Горизонтальная

        separator_1 = Frame(
            self.root_menu_3, 
            height=3,          
            bg='grey',        
            relief=FLAT,
            bd=1               
        )
        separator_1.place(relx=0.5, rely=0, relwidth= 0.004, relheight= 1) #Вертикальная
    ###

    #Лейблы
        self.creating_number = ttk.Label(
                self.root_menu_3,
                text = "8",
                style="TLabel",
                anchor="center"
                )
        self.creating_number.place(relx=0, rely=0, relwidth=0.49, relheight=0.1)

        self.creating_number = ttk.Label(
                self.root_menu_3,
                text = "9",
                style = "TLabel",
                anchor="center"
                )
        self.creating_number.place(relx=0.51, rely=0, relwidth=0.49, relheight=0.1)

        self.creating_number = ttk.Label(
                self.root_menu_3,
                text = "10",
                style = "TLabel",
                anchor="center"
                )
        self.creating_number.place(relx=0, rely=0.51, relwidth=0.49, relheight=0.1)

        self.creating_number = ttk.Label(
                self.root_menu_3,
                text = "11",
                style = "TLabel",
                anchor="center"
                )
        self.creating_number.place(relx=0.51, rely=0.51, relwidth=0.49, relheight=0.1)
    ###

    #Отображение параллелей
        self.display_parallel("8")
        self.display_parallel("9")
        self.display_parallel("10")
        self.display_parallel("11")
    ###

    #Функционал
    #Отображение параллелей
    def display_parallel(self, number):
        # Получаем параллели с указанным номером из БД
        self.con_root_menu_3.execute(
            "SELECT DISTINCT Letter FROM parallels WHERE Number = ? ORDER BY Letter",
            (number,)
        )
        parallels = self.con_root_menu_3.fetchall()
        # Определяем начальные координаты в зависимости от номера параллели
        if number == "8":
            start_y = 0.1
            start_x = 0
        elif number == "9":
            start_y = 0.1
            start_x = 0.51
        elif number == "10":
            start_y = 0.61
            start_x = 0
        else:  # для 11 класса
            start_y = 0.61
            start_x = 0.51

        # Создаем кнопки для каждой параллели
        for i, parallel in enumerate(parallels):
            letter = parallel[0]
            pos_x = start_x + (i % 4) * 0.12  # 4 кнопки в ряд
            pos_y = start_y + (i // 4) * 0.1   # новый ряд каждые 4 кнопки

            data = (number, letter, self.base_root_menu_3, self.con_root_menu_3, self, self.root_menu_3)
            
            btn = ttk.Button(
                self.root_menu_3,
                text=f"{number}{letter}",
                command=lambda data = data: self.open_parallel_info(data),
                style="Sec_3.TButton"
            )
            btn.place(relx=pos_x, rely=pos_y, relwidth=0.11, relheight=0.09)
    ###

    def update_buttons(self):
        for widget in self.root_menu_3.winfo_children():
            if widget.winfo_class() == "TButton":
                widget.destroy()
        self.display_parallel("8")
        self.display_parallel("9")
        self.display_parallel("10")
        self.display_parallel("11")

    #Открытие данных о параллели
    def open_parallel_info(self, data):
        change_parallel(data)
    ###
        
###



#<Изменение параллели> - # +++
class change_parallel:
    def __init__(self, data):
    #Данные корня
        self.root_popup_2 = Toplevel(data[5])
        self.num = data[0]
        self.letter = data[1]
        self.base = data[2]
        self.con = data[3]
        self.parent_con = data[4]
        self.root_popup_2.title(f"Параллель {self.num}{self.letter}")
        self.root_popup_2.geometry(f"{int(self.root_popup_2.winfo_screenwidth() * 0.56)}x{int(self.root_popup_2.winfo_screenheight()*0.6)}")
        self.root_popup_2.resizable(width=False, height=False)
        

        style_3_1 = ttk.Style()

        style_3_1.theme_use('clam')

        style_3_1.configure("TLabel", font=("Helvetica", 30, "italic"))

        style_3_1.configure("TEntry", fieldbackground="#DCDCDC")

        style_3_1.configure("TCombobox", fieldbackground="#DCDCDC", arrowsize = 0)
        style_3_1.map("TCombobox", fieldbackground=[("readonly", "#DCDCDC")])


        style_3_1.configure("Main_2_1.TButton", font=("Helvetica", 30, "bold"), background="#696969", foreground="white",)
        style_3_1.map("Main_2_1.TButton", background=[("active", "#0000CD")])

        style_3_1.configure("Sec_2_1.TButton", font=("Helvetica", 13, "bold"), background="#696969", foreground="white",)
        style_3_1.map("Sec_2_1.TButton", background=[("active", "#4169E1")])

        self.root_popup_2.configure(bg="#e0dcd4")
        # Данные для параллели
        self.full_data_class = []  # для хранения предметов
        self.clas_buttons = []     # для хранения кнопок
        
        # Получаем данные из БД
        self.con.execute(
           """
            SELECT 
                l.Subject,
                t.Surname || ' ' || substr(t.Name, 1, 1) || '. ' || substr(t.Patrony, 1, 1) || '.' AS Teacher,
                l.Hours
            FROM lessons l
            JOIN parallels p ON l.Id_parallel = p.id
            JOIN teachers t ON l.Id_teacher = t.id
            WHERE p.Number = ? 
              AND p.Letter = ?
            """,
            (self.num, self.letter)
        )
        lessons_data = self.con.fetchall()
        for lesson in lessons_data:
            self.full_data_class.append(list(lesson))
    ###

    #Виджеты изменения параллели
    #Границы
        separator_1 = Frame(
            self.root_popup_2, 
            height=3,
            bg='grey',
            relief=FLAT,
            bd=1
        )
        separator_1.place(relx=0, rely=0.14, relwidth=1, relheight=0.007)

        separator_2 = Frame(
            self.root_popup_2, 
            height=3,
            bg='grey',
            relief=FLAT,
            bd=1
        )
        separator_2.place(relx=0.29, rely=0.14, relwidth=0.004, relheight=1)
    ###

    # Блок номера
        self.creating_number = ttk.Label(
            self.root_popup_2,
            text="Номер:",
            style="TLabel"
        )
        self.creating_number.place(relx=0.1, rely=0, relwidth=0.20, relheight=0.14)

        self.number = ttk.Combobox(
            self.root_popup_2,
            values=["8", "9", "10", "11"],
            style="TCombobox",
            font=("Helvetica", 27),
            state="readonly", 
            foreground="#4D4D4D"
        )
        self.number.set(self.num)
        self.number.place(relx=0.3, rely=0, relwidth=0.10, relheight=0.14)
    ###

    #Блок буквы
        self.creating_letter = ttk.Label(
            self.root_popup_2,
            text="Буква:",
            style="TLabel"
        )
        self.creating_letter.place(relx=0.5, rely=0, relwidth=0.20, relheight=0.14)

        self.letter_box = ttk.Combobox(
            self.root_popup_2,
            values=["А", "Б", "В", "Г", "Д"],
            style="TCombobox",
            font=("Helvetica", 27),
            state="readonly", 
            foreground="#4D4D4D"
        )
        self.letter_box.set(self.letter)
        self.letter_box.place(relx=0.7, rely=0, relwidth=0.10, relheight=0.14)
    ###

    #Блок выбора урока
        self.creating_subject =ttk.Label(
            self.root_popup_2,
            text = "Урок:",
            style="TLabel",
            anchor="center"
            )
        self.creating_subject.place(relx=0, rely=0.15, relwidth=0.28, relheight=0.1)

        self.options_3 = ["Алгебра",
            "Геометрия",
            "ТеорВер",
            "Физика",
            "Русский",
            "Литература",
            "История",
            "Информатика_1",
            "Информатика_2",
            "География",
            "Английский язык_1",
            "Английский язык_2",
            "Физра",
            "Биология",
            "Химия",
            "Обществознание",
            "ОБЖ",
            "Технология"]
        self.subject = ttk.Combobox(
            self.root_popup_2,
            values=self.options_3,
            style="TCombobox",
            font=("Helvetica", 27),
            state="readonly", 
            foreground="#4D4D4D"
        )
        self.subject.place(relx=0, rely=0.25, relwidth=0.29, relheight=0.1)
    ###

    #Блок выбора учителя
        self.options_teacher = []
        self.con.execute("""SELECT * FROM teachers""")
        values = self.con.fetchall() # Проверка на предварительное наличие кабинета 

        for i in values:
            self.options_teacher.append(f"{i[1]} {i[2][0]}. {i[3][0]}.")

        self.creating_teacher = ttk.Label(
            self.root_popup_2,
            text = "Учитель:",
            font=("Helvetica", 30),
            anchor="center"
            )
        self.creating_teacher.place(relx=0, rely=0.35, relwidth=0.29, relheight=0.1)
        
        self.teacher = ttk.Combobox(
            self.root_popup_2,
            values=self.options_teacher,
            font=("Helvetica", 30),
        )
        self.teacher.place(relx=0, rely=0.45, relwidth=0.29, relheight=0.1)
    ###

    #Блок выбора количества часов
        self.creating_hours =ttk.Label(
            self.root_popup_2,
            text = "Кол-во часов:",
            style = "TLabel",
            anchor="center"
            )
        self.creating_hours.place(relx=0, rely=0.55, relwidth=0.29, relheight=0.1)
        
        self.hours = ttk.Entry(
            self.root_popup_2,
            style="TEntry",
            font=("Helvetica", 27),
            foreground="#4D4D4D"
        )
        self.hours.place(relx=0, rely=0.65, relwidth=0.29, relheight=0.1)
    ###

    #Футер
    #Кнопка добавления урока 
        self.add_class = ttk.Button(
            self.root_popup_2,
            text="Добавить",
            style = "Main_2_1.TButton",
            command = self.add_class
        )
        self.add_class.place(relx=0, rely=0.75, relwidth=0.29, relheight=0.1)
    ###

    #Кнопка сохранения изменений
        self.save_button = ttk.Button(
            self.root_popup_2,
            text="Сохранить",
            style = "Main_2_1.TButton",
            command=self.save_data
        )
        self.save_button.place(relx=0, rely=0.84, relwidth=0.292, relheight=0.16)
    ###

    #Кнопка отмены
        self.cancel_button = ttk.Button(
            self.root_popup_2,
            text = "Отмена",
            style = "Main_2_1.TButton",
            command= self.root_popup_2.destroy
        )
        self.cancel_button.place(relx=0.292, rely=0.84, relwidth=0.416, relheight=0.16)
    ###

    #Кнопка удаления параллели
        self.delete_parallel = ttk.Button(
            self.root_popup_2,
            text="Удалить",
            style = "Main_2_1.TButton",
            command=self.delete_data
        )
        self.delete_parallel.place(relx=0.708, rely=0.84, relwidth=0.292, relheight=0.16)
    ###

        self.display_class() #Отображаем предметы
    
    #Функционал
    #Добавление предмета
    def add_class(self):
        mass = []
        sub = self.subject.get()
        teacher = self.teacher.get()
        hours = self.hours.get()

        #
        self.con.execute("""SELECT * FROM parallels WHERE Letter = ? and Number = ?""", (self.letter, self.num, ))
        Id_parallel = self.con.fetchone()[0]

        self.con.execute("""SELECT * FROM lessons WHERE Subject = ? and Id_parallel = ?""", (sub, Id_parallel,))
        if self.con.fetchall():
            warning = eror_popup(self.root_popup_2, "Уже есть такой предмет")
            warning.root.mainloop()
            return

        # Проверка урока
        if sub not in self.options_3:
            warning = eror_popup(self.root_popup_2, "Ошибка в предмете")
            warning.root.mainloop()
            return
        mass.append(sub)

        # Проверка на существование такого предмета
        for i in self.full_data_class:
            if i and i[0] == sub:
                warning = eror_popup(self.root_popup_2, "Уже есть такой предмет")
                warning.root.mainloop()
                return

        # Проверка учителя
        if teacher not in self.options_teacher:
            warning = eror_popup(self.root_popup_2, "Нет такого учителя")
            warning.root.mainloop()
            return
        mass.append(teacher)

        # Проверка количества часов
        char = "0123456789"
        if hours == "":
            warning = eror_popup(self.root_popup_2, "Нет кол-ва часов")
            warning.root.mainloop()
            return

        for i in hours:
            if i not in char:
                warning = eror_popup(self.root_popup_2, "Ошибка в кол-ве часов")
                warning.root.mainloop()
                return

        if int(hours) > 30:
            warning = eror_popup(self.root_popup_2, "Слишком много часов")
            warning.root.mainloop()
            return

        mass.append(hours)
        self.full_data_class.append(mass)
        self.display_class()
        popup(self.root_popup_2, "Предмет добавлен, еще?", "Успех").root.mainloop()

    def save_data(self):
        # Получаем новые значения
        old_num = self.num
        new_num = self.number.get()
        new_letter = self.letter_box.get()
    
        # Проверка номера класса
        if new_num not in ["8", "9", "10", "11"]:
            warning = eror_popup(self.root_popup_2, "Нет такого класса")
            warning.root.mainloop()
            return
    
        # Проверка буквы класса
        if new_letter not in ["А", "Б", "В", "Г", "Д"]:
            warning = eror_popup(self.root_popup_2, "Нет такой буквы")
            warning.root.mainloop()
            return
    
        # Проверка на изменения (номер, буква или предметы)
        has_changes = False
        if new_num != self.num or new_letter != self.letter:
            has_changes = True
        
        #Проверка на 2 групппы
        list_sub = []
        for _ in self.full_data_class:
            if _:
                list_sub.append(_[0])

        if ("Информатика_1" in list_sub and "Информатика_2" not in list_sub) or ("Информатика_1" not in list_sub and "Информатика_2" in list_sub):
            warning = eror_popup(self.root_menu_2, "Созданы не 2 группы")
            warning.root.mainloop()
            return
        
        if ("Английский язык_1" in list_sub and "Английский язык_2" not in list_sub) or ("Английский язык_1" not in list_sub and "Английский язык_2" in list_sub):
            warning = eror_popup(self.root_menu_2, "Созданы не 2 группы")
            warning.root.mainloop()
            return


        # Проверяем изменения в предметах
        self.con.execute(
            """SELECT l.Subject, 
                      (SELECT t.Surname || ' ' || substr(t.Name,1,1) || '. ' || substr(t.Patrony,1,1) || '.'
                       FROM teachers t WHERE t.id = l.Id_teacher) as Teacher,
                      l.Hours
               FROM lessons l
               JOIN parallels p ON l.Id_parallel = p.id 
               WHERE p.Number = ? AND p.Letter = ?""",
            (self.num, self.letter)
        )
        original_lessons = self.con.fetchall()
        original_lessons = [list(lesson) for lesson in original_lessons]

        if sorted(original_lessons) != sorted(self.full_data_class):
            has_changes = True

        if not has_changes:
            warning = eror_popup(self.root_popup_2, "Вы ничего не поменяли")
            warning.root.mainloop()
            return
    
        # Получаем id текущей параллели
        self.con.execute(
            "SELECT id FROM parallels WHERE Number = ? AND Letter = ?",
            (self.num, self.letter)
        )
        parallel_id = self.con.fetchone()[0]
    
        # Обновляем параллель в таблице parallels
        self.con.execute(
            "UPDATE parallels SET Number = ?, Letter = ? WHERE id = ?",
            (new_num, new_letter, parallel_id)
        )
    
        # Удаляем старые уроки
        self.con.execute("DELETE FROM lessons WHERE Id_parallel = ?", (parallel_id,))
    
        # Добавляем новые уроки
        for subject_data in self.full_data_class:
            if subject_data:  # Проверяем, что данные не пустые
                # Получаем id учителя по ФИО
                teacher_fio = subject_data[1].split()
                surname = teacher_fio[0]
                name = teacher_fio[1][0]
                patrony = teacher_fio[2][0]
    
                self.con.execute(
                    "SELECT id FROM teachers WHERE Surname = ? AND Name LIKE ? AND Patrony LIKE ?",
                    (surname, f"{name}%", f"{patrony}%")
                )
                teacher_id = self.con.fetchone()[0]
    
                # Вставляем данные в таблицу lessons
                self.con.execute(
                    "INSERT INTO lessons (Id_parallel, Subject, Hours, Id_teacher) VALUES (?, ?, ?, ?)",
                    (parallel_id, subject_data[0], subject_data[2], teacher_id)
                )


        self.base.commit()

        self.parent_con.update_buttons()  

        popup(self.root_popup_2, f"Параллель {self.num}{self.letter} изменена!", "Успех").root.mainloop()
        self.root_popup_2.destroy()

    def display_class(self):
        # Удаляем старые кнопки
        for btn in self.clas_buttons:
            btn.destroy()
        self.clas_buttons = []

        # Создаем список только непустых элементов
        active_classes = [clas for clas in self.full_data_class if clas]

        # Размещаем кнопки для каждого предмета
        for i, clas in enumerate(active_classes):
            pos_x = (i % 4) * 0.175 + 0.3
            pos_y = (i // 4) * 0.14 + 0.15

            btn = ttk.Button(
                self.root_popup_2,
                text=f"{clas[0]}",
                command=lambda cl=(i, clas): self.open_popup_class(cl),
                style="Sec_2_1.TButton"
            )
            btn.place(relx=pos_x, rely=pos_y, relwidth=0.175, relheight=0.14)
            self.clas_buttons.append(btn)

    def open_popup_class(self, clas):
        popup_subject_2(clas[0], clas[1], self.root_popup_2, self.base, self.con, self)

    def delete_data(self):
        # Получаем id параллели
        self.con.execute(
            "SELECT id FROM parallels WHERE Number = ? AND Letter = ?",
            (self.num, self.letter)
        )
        parallel_id = self.con.fetchone()[0]

        # Удаляем все уроки этой параллели
        self.con.execute("DELETE FROM lessons WHERE Id_parallel = ?", (parallel_id,))
        
        # Удаляем саму параллель
        self.con.execute("DELETE FROM parallels WHERE id = ?", (parallel_id,))
        
        self.base.commit()

        self.parent_con.update_buttons()  # Обновляем отображение в родительском окне
        self.root_popup_2.destroy()
        popup(self.parent_con.root_menu_3, f"Параллель {self.num}{self.letter} удалена!", "Успех").root.mainloop()



#<Изменение предмета> - # +++
class popup_subject_2:
    def __init__(self, id, clas, parent, base, connect, parent_con):
        #Данные корня
        self.root_subject = Toplevel(parent)
        self.root_subject.geometry(f"{int(self.root_subject.winfo_screenwidth() * 0.22)}x{int(self.root_subject.winfo_screenheight()*0.41)}")
        self.root_subject.resizable(width=False, height=False)
        self.root_subject.title(f"Предмет {clas[0]}")
        self.base_subject = base
        self.con_subject = connect
        self.id = id
        self.clas = clas
        self.parent_con = parent_con

        style_3_2 = ttk.Style()

        style_3_2.theme_use('clam')

        style_3_2.configure("TLabel", font=("Helvetica", 30, "italic"))

        style_3_2.configure("TEntry", fieldbackground="#DCDCDC")

        style_3_2.configure("TCombobox", fieldbackground="#DCDCDC", arrowsize = 0)
        style_3_2.map("TCombobox", fieldbackground=[("readonly", "#DCDCDC")])


        style_3_2.configure("Main_3_2.TButton", font=("Helvetica", 15, "bold"), background="#696969", foreground="white",)

        style_3_2.map("Main_3_2.TButton", background=[("active", "#0000CD")])

        self.root_subject.configure(bg="#e0dcd4")

        #Виджеты изменения урока
        # Блок выбора урока
        self.creating_subject = ttk.Label(
            self.root_subject,
            text="Урок:",
            style="TLabel",
            anchor="center"
        )
        self.creating_subject.place(relx=0, rely=0, relwidth=1, relheight=0.14)

        self.options_3 = ["Алгебра",
            "Геометрия",
            "ТеорВер",
            "Физика",
            "Русский",
            "Литература",
            "История",
            "Информатика_1",
            "Информатика_2",
            "География",
            "Английский язык_1",
            "Английский язык_2",
            "Физра",
            "Биология",
            "Химия",
            "Обществознание",
            "ОБЖ",
            "Технология"]
        
        self.subject = ttk.Combobox(
            self.root_subject,
            values=self.options_3,
            style="TCombobox",
            font=("Helvetica", 27),
            state="readonly", 
            foreground="#4D4D4D"
        )
        self.subject.place(relx=0, rely=0.14, relwidth=1, relheight=0.14)
        self.subject.set(self.clas[0])

        # Блок выбора учителя
        self.options_teacher = []
        self.con_subject.execute("""SELECT * FROM teachers""")
        values = self.con_subject.fetchall()

        for i in values:
            self.options_teacher.append(f"{i[1]} {i[2][0]}. {i[3][0]}.")

        self.creating_teacher = ttk.Label(
            self.root_subject,
            text="Учитель:",
            style="TLabel",
            anchor="center"
        )
        self.creating_teacher.place(relx=0, rely=0.28, relwidth=1, relheight=0.14)
        
        self.teacher = ttk.Combobox(
            self.root_subject,
            values=self.options_teacher,
            style="TCombobox",
            font=("Helvetica", 27),
            state="readonly", 
            foreground="#4D4D4D"
        )
        self.teacher.place(relx=0, rely=0.42, relwidth=1, relheight=0.14)
        self.teacher.set(self.clas[1])

        # Блок выбора количества часов
        self.creating_hours = ttk.Label(
            self.root_subject,
            text="Кол-во часов:",
            style="TLabel",
            anchor="center"
        )
        self.creating_hours.place(relx=0, rely=0.56, relwidth=1, relheight=0.14)
        
        self.hours = ttk.Entry(
            self.root_subject,
            style="TEntry",
            font=("Helvetica", 27),
            foreground="#4D4D4D"

        )
        self.hours.insert(0, self.clas[2])
        self.hours.place(relx=0, rely=0.7, relwidth=1, relheight=0.14)

        # Футер
        self.save_button = ttk.Button(
            self.root_subject,
            text="Сохранить",
            command=self.save_data,
            style = "Main_3_2.TButton"
        )
        self.save_button.place(relx=0, rely=0.84, relwidth=0.3, relheight=0.16)

        self.cancel_button = ttk.Button(
            self.root_subject,
            text="Отмена",
            style = "Main_3_2.TButton",
            command=self.root_subject.destroy
        )
        self.cancel_button.place(relx=0.3, rely=0.84, relwidth=0.4, relheight=0.16)

        self.delete_button = ttk.Button(
            self.root_subject,
            text="Удалить",
            style = "Main_3_2.TButton",
            command=self.delete_data
        )
        self.delete_button.place(relx=0.7, rely=0.84, relwidth=0.3, relheight=0.16)

    #Функционал
    def delete_data(self):
        if self.id < len(self.parent_con.full_data_class):
            self.parent_con.full_data_class[self.id] = []
            self.parent_con.display_class()
            self.root_subject.destroy()
            popup(self.parent_con.root_popup_2, "Предмет удален!", "Успех").root.mainloop()

    def save_data(self):
        new_subject = self.subject.get()
        new_teacher = self.teacher.get()
        new_hours = self.hours.get()

        # Проверка на изменения
        if [new_subject, new_teacher, new_hours] == [self.clas[0], self.clas[1], self.clas[2]]:
            warning = eror_popup(self.root_subject, "Ничего не поменяли")
            warning.root.mainloop()
            return

        # Проверка предмета
        if new_subject not in self.options_3:
            warning = eror_popup(self.root_subject, "Ошибка в предмете")
            warning.root.mainloop()
            return

        # Проверка на существование такого предмета
        for i, subject in enumerate(self.parent_con.full_data_class):
            if i != self.id and subject and subject[0] == new_subject:
                warning = eror_popup(self.root_subject, "Уже есть такой предмет")
                warning.root.mainloop()
                return

        # Проверка учителя
        if new_teacher not in self.options_teacher:
            warning = eror_popup(self.root_subject, "Нет такого учителя")
            warning.root.mainloop()
            return

        # Проверка часов
        char = "0123456789"
        if new_hours == "":
            warning = eror_popup(self.root_subject, "Нет кол-ва часов")
            warning.root.mainloop()
            return

        if not all(c in char for c in new_hours):
            warning = eror_popup(self.root_subject, "Ошибка в кол-ве часов")
            warning.root.mainloop()
            return

        if int(new_hours) > 30:
            warning = eror_popup(self.root_subject, "Слишком много часов")
            warning.root.mainloop()
            return

        # Обновляем значение в списке
        self.parent_con.full_data_class[self.id] = [new_subject, new_teacher, new_hours]
        
        # Обновляем отображение
        self.parent_con.display_class()
        popup(self.root_subject, "Предмет обновлен!", "Успех").root.mainloop()
        self.root_subject.destroy()






#<Созд каба> - 4 +++
class menu_add_room:
    def __init__(self, parent, base, connect):
    #Данные корня
        self.base_menu_4 = base
        self.con_menu_4 = connect
        self.root_menu_4 = Toplevel(parent)
        self.root_menu_4.title("Добавление кабинета")
        self.root_menu_4.geometry(f"{int(self.root_menu_4.winfo_screenwidth() * 0.56)}x{int(self.root_menu_4.winfo_screenheight()*0.6)}")
        self.root_menu_4.resizable(width=False, height=False)
        self.tup_base_con = (self.base_menu_4, self.con_menu_4)
    ###
    
        style_4 = ttk.Style()

        style_4.theme_use('clam')

        style_4.configure("TLabel", font=("Helvetica", 30, "italic"))

        style_4.configure("TEntry", fieldbackground="#DCDCDC")

        style_4.configure("TCombobox", fieldbackground="#DCDCDC", arrowsize = 0)
        style_4.map("TCombobox", fieldbackground=[("readonly", "#DCDCDC")])


        style_4.configure("Main_4.TButton", font=("Helvetica", 30, "bold"), background="#696969", foreground="white",)
        style_4.map("Main_4.TButton", background=[("active", "#9932CC")])

        self.root_menu_4.configure(bg="#e0dcd4")

    #Виджеты добавления кабинета ++
    # Границы вертикальная
        separator_1 = Frame(
            self.root_menu_4, 
            height=3,           
            bg='grey',         
            relief=FLAT,
            bd=1               
        )
        separator_1.place(relx=0, rely=0.45, relwidth=1, relheight=0.007)

        # Блок номера кабинета
        self.number_room_label = ttk.Label(
            self.root_menu_4,
            text="Номер кабинета:",
            style="TLabel",
            anchor="center"

        )
        
        self.number_room_label.place(relx=0, rely=0, relwidth=0.5, relheight=0.15)

        self.input_number_room = ttk.Entry(
            self.root_menu_4,
            style = "TEntry",
            font=("Helvetica", 27),
            foreground="#4D4D4D"
        )
        self.input_number_room.place(relx=0.5, rely=0, relwidth=0.5, relheight=0.15)

        # Блок графский или нет
        options = ["Графский", "Лицей"]
        self.graphskiy_or_not_label = ttk.Label(
            self.root_menu_4,
            text="Графский или лицей:",
            style="TLabel",
            anchor="center"
        )
        self.graphskiy_or_not_label.place(relx=0, rely=0.15, relwidth=0.5, relheight=0.15)

        self.graphskiy_or_not = ttk.Combobox(
            self.root_menu_4,
            values= options,
            style="TCombobox",
            font=("Helvetica", 27),
            state="readonly", 
            foreground="#4D4D4D"
              
        )
        self.graphskiy_or_not.place(relx=0.5, rely=0.15, relwidth=0.5, relheight=0.15)

        
        # Блок большой/маленький
        options = ["Маленький", "Большой"]
        self.small_or_big_label = ttk.Label(
            self.root_menu_4,
            text="Маленький или большой:",
            style="TLabel",
            anchor="center"
        )
        self.small_or_big_label.place(relx=0, rely=0.3, relwidth=0.5, relheight=0.15)

        self.small_or_big = ttk.Combobox(
            self.root_menu_4,
            values=options,
            style="TCombobox",
            font=("Helvetica", 27),
            state="readonly", 
            foreground="#4D4D4D"
        )
        self.small_or_big.place(relx=0.5, rely=0.3, relwidth=0.5, relheight=0.15)

        # Блок выбора приоритетного кабинета
        
        self.subjects = [
            "Нет приоритетных кабинетов",
            "Алгебра/Геометрия/ТеорВер",
            "Физика",
            "Русский/Литература",
            "История",
            "Информатика",
            "География",
            "Английский язык",
            "Физра",
            "Биология",
            "Химия",
            "Обществознание"
        ]

        self.prio_subjects_label = ttk.Label(
            self.root_menu_4,
            text="Приоритетные уроки",
            style="TLabel",
            anchor="center"
        )
        self.prio_subjects_label.place(relx=0, rely=0.457, relwidth=1, relheight=0.097)

        self.prio_subjects = Listbox(
            self.root_menu_4,
            selectmode=SINGLE,
            background="#DCDCDC",
            selectbackground="#4b6985", 
            font=("Helvetica", 27), 
            foreground="#4D4D4D", 
            relief="groove",
            bd=2
        )
        self.prio_subjects.place(relx=0.2, rely=0.55, relwidth=0.6, relheight=0.3)

        for item in self.subjects:
            self.prio_subjects.insert(END, item)

        # Кнопка добавления кабинета
        self.add_button = ttk.Button(
            self.root_menu_4,
            text="Добавить кабинет",
            style = "Main_4.TButton",
            command=lambda base_con=self.tup_base_con: self.add_room(base_con)
        )
        self.add_button.place(relx=0, rely=0.85, relwidth=1, relheight=0.15)
    ###

#Функционал
#Вв данных и пров
    def add_room(self, base_plus_con):
    #Вспомогательные штучки
        data_room = []
        self.base_add_room = base_plus_con[0]
        self.con_add_room = base_plus_con[1]
    ###

    #Проверка номера кабинета + ввод в data_room
        all_num = "1234567890"
        room_number = str(self.input_number_room.get())
        count = 0 
        for digit in room_number:
            if digit in all_num:
                count += 1

        if (count == len(room_number)) and (room_number != ""): #Проверка на наличие числа и проверка на ввод данных 
            if int(room_number) // 100 < 6: #Проверка на наличие этажей между 1 - 5
                data_room.append(str(room_number))
            else:
                warning = eror_popup(self.root_menu_4, "Ошибка в номере кабинета")
                warning.root.mainloop()
                return
        else:
            warning = eror_popup(self.root_menu_4, "Ошибка в номере кабинета")
            warning.root.mainloop()
            return
    ###

    #Проверка графского или лицея + ввод
        graphskiy_or_not = self.graphskiy_or_not.get()
        if graphskiy_or_not == "Графский" or graphskiy_or_not == "Лицей": #Проверка на привльности написания ответа
            data_room.append(graphskiy_or_not)
        else:
            warning = eror_popup(self.root_menu_4, "Ошибка в Графский/Лицей")
            warning.root.mainloop()
            return
    ###

    #Проверка большого или маленького кабинета
        small_or_big = self.small_or_big.get()
        if small_or_big == "Маленький" or small_or_big == "Большой":
            data_room.append(small_or_big)
        else:                                       
            warning = eror_popup(self.root_menu_4, "Ошибка в большой/мал.")
            warning.root.mainloop()
            return
    ###

    #Проверка приоклассов + ввод
        selected_indices = self.prio_subjects.curselection()

        if selected_indices == tuple():
            warning = eror_popup(self.root_menu_4, "Ошибка в прио. уроках")
            warning.root.mainloop()
            return
        else:
            data_room.append(self.subjects[selected_indices[0]])
    ###

    #Самая финальная проверка на наличие в таблице + Ввод в таблицу + обновление страницы
        self.con_menu_4.execute("SELECT * FROM rooms WHERE Graph_or_Lyceum = ?", (data_room[1],))
        values = self.con_menu_4.fetchall()  # Проверка на предварительное наличие кабинета 
        all_num_of_rooms = []
        for i in values:
            all_num_of_rooms.append(i[1])

        if data_room[0] not in all_num_of_rooms:
            self.con_menu_4.execute("INSERT INTO rooms (Number, Graph_or_Lyceum, Big_or_Small, Subject) VALUES (?, ?, ?, ?)",
                                    (data_room[0], data_room[1], data_room[2], data_room[3]))
            self.base_menu_4.commit()
        else:
            warning = eror_popup(self.root_menu_4, "Есть кабинет в корпусе")
            warning.root.mainloop()
            return
    ###

    #Проверка на максималное число на экране
        if len(values) + 1 > 36:
            warning = eror_popup(self.root_menu_4, "Лимит классов")
            warning.root.mainloop()
            return
    ###

        self.reset_form()
        warning = popup(self.root_menu_4, f"Кабинет {room_number} создан!", "Успех")
        warning.root.mainloop()

###

#Сброс всех полей ввода
    def reset_form(self):
        
        self.input_number_room.delete(0, END)
        self.graphskiy_or_not.set('Графский или лицей')
        self.small_or_big.set('Маленький или большой кабинет')
        self.prio_subjects.selection_clear(0, END)
###
     





#<См каба> - 5 +++
class menu_check_room:
    def __init__(self, parent, base, connect):
    #Данные корня
        self.root_menu_5 = Toplevel(parent)
        self.root_menu_5.title("Просмотр кабинетов")
        self.root_menu_5.geometry(f"{int(self.root_menu_5.winfo_screenwidth() * 0.56)}x{int(self.root_menu_5.winfo_screenheight()*0.6)}")
        self.root_menu_5.resizable(width=False, height=False) 
        self.room_menu_5_base = base
        self.room_menu_5_con = connect
    ###


        style_5 = ttk.Style()

        style_5.theme_use('clam')

        style_5.configure("TLabel", font=("Helvetica", 30, "italic"))

        style_5.configure("TEntry", fieldbackground="#DCDCDC")

        style_5.configure("TCombobox", fieldbackground="#DCDCDC", arrowsize = 0)
        style_5.map("TCombobox", fieldbackground=[("readonly", "#DCDCDC")])

        style_5.configure("Sec_5.TButton", font=("Helvetica", 25, "bold"), background="#696969", foreground="white",)
        style_5.map("Sec_5.TButton", background=[("active", "#BA55D3")])

        self.root_menu_5.configure(bg="#e0dcd4")

    #Виджеты просмотра кабинетов ++
    #Границы
        separator_1 = Frame(
            self.root_menu_5, 
            height=3,           # Увеличиваем высоту
            bg='grey',         # Задаем цвет
            relief=FLAT,
            bd=1               # Увеличиваем толщину бордюра
        )
        separator_1.place(relx=0.5, rely=0, relwidth=0.004, relheight= 1)
    ###

        self.ne_graphskiy = ttk.Label(
            self.root_menu_5,
            text="Графский",
            style="TLabel",
            anchor="center"
        )
        self.ne_graphskiy.place(relx=0, rely=0, relwidth=0.5, relheight=0.09)
    
        self.graphskiy = ttk.Label(
            self.root_menu_5,
            text="Лицей",
            style="TLabel",
            anchor="center"
        )
        self.graphskiy.place(relx=0.504, rely=0, relwidth=0.496, relheight=0.09)
        
        self.display_buttons("Графский")
        self.display_buttons("Лицей")

    def display_buttons(self, answer):
        self.answ = answer
        self.room_menu_5_con.execute("SELECT DISTINCT Number FROM rooms WHERE Graph_or_Lyceum = ?", (self.answ,))
        rooms = self.room_menu_5_con.fetchall()

        wid_x = 0.11  # уменьшаем ширину кнопки для 4 кабинетов
        hei_y = 0.1   # высота остается прежней

        if self.answ == "Графский":
            start_x = 0.02  # начальная позиция для Графского
            positions = [0.02, 0.14, 0.26, 0.38]  # 4 позиции для кнопок
        else:
            start_x = 0.52  # начальная позиция для Лицея
            positions = [0.52, 0.64, 0.76, 0.88]  # 4 позиции для кнопок

        for i, room in enumerate(rooms):
            pos_x = positions[i % 4]  # позиция по x (чередуем 4 позиции)
            pos_y = 0.1 + (i // 4) * 0.1  # позиция по y (новая строка каждые 4 кабинета)

            room_button = ttk.Button(
                self.root_menu_5, 
                text=room[0], 
                command=lambda num=room[0], ans=self.answ: self.open_room_screen(num, ans),
                style = "Sec_5.TButton",
            )
            room_button.place(relx=pos_x, rely=pos_y, relwidth=wid_x, relheight=hei_y)

    def open_room_screen(self, room_number, answerr):
        room_info(self.root_menu_5, room_number, self.room_menu_5_base, self, answerr, self.root_menu_5)

#Постоянное обновление виджетов страницы       
    def update_buttons(self):
        for widget in self.root_menu_5.winfo_children():
            if widget.winfo_class() == "TButton":
                widget.destroy()
        self.display_buttons("Графский")
        self.display_buttons("Лицей")
###



#<Инфа каба> - # +++
class room_info:
    def __init__(self, root, room_number, dbs, main_screen, asw, root_parent):
    #Данные корня
        self.root_room_info = Toplevel(root)
        self.root_parent = root_parent
        self.main_screen = main_screen
        self.room_number = room_number
        self.root_room_info.title(f"Кабинет {self.room_number}")
        self.root_room_info.geometry(f"{int(self.root_room_info.winfo_screenwidth() * 0.56)}x{int(self.root_room_info.winfo_screenheight()*0.6)}")
        self.root_room_info.resizable(width=False, height=False)
        self.base_room_info = dbs
        self.con_room_info = self.base_room_info.cursor()
        self.answ = asw
        self.con_room_info.execute("SELECT * FROM rooms WHERE Number = ? AND Graph_or_Lyceum = ?", (self.room_number, self.answ))
        full_data = self.con_room_info.fetchall()
        self.def_num = full_data[0][1]
        self.def_graphskiy_not = full_data[0][2]
        self.def_big_not = full_data[0][3]
        self.def_sub = full_data[0][4]
    ###

        style_5_0 = ttk.Style()

        style_5_0.theme_use('clam')

        style_5_0.configure("TLabel", font=("Helvetica", 30, "italic"))

        style_5_0.configure("TEntry", fieldbackground="#DCDCDC")

        style_5_0.configure("TCombobox", fieldbackground="#DCDCDC", arrowsize = 0)
        style_5_0.map("TCombobox", fieldbackground=[("readonly", "#DCDCDC")])


        style_5_0.configure("Main_5_0.TButton", font=("Helvetica", 30, "bold"), background="#696969", foreground="white",)
        style_5_0.map("Main_5_0.TButton", background=[("active", "#9932CC")])

        self.root_room_info.configure(bg="#e0dcd4")


    #Виджеты изменения кабинета
    #Границы
        separator_1 = Frame(
            self.root_room_info, 
            height=3,           # Увеличиваем высоту
            bg='grey',         # Задаем цвет
            relief=FLAT,
            bd=1               # Увеличиваем толщину бордюра
        )
        separator_1.place(relx=0, rely=0.45, relwidth=1, relheight=0.007)
    ###

    #Блок номера кабинета 
        self.number_room_label = ttk.Label(
            self.root_room_info,
            text="Номер кабинета:",
            style="TLabel",
            anchor="center")
        self.number_room_label.place(relx=0, rely=0, relwidth=0.5, relheight=0.15)

        self.input_number_room = ttk.Entry(
            self.root_room_info,
            style="TEntry",
        font=("Helvetica", 27),
        foreground="#4D4D4D"
        )
        self.input_number_room.place(relx=0.5, rely=0, relwidth=0.5, relheight=0.15)
        
        self.input_number_room.insert(0, self.def_num) # Пердварительный ввод номера кабинета
    ###

    #Блок графский или нет
        options = ["Графский", "Лицей"]
        self.graphskiy_or_not_label = ttk.Label(
            self.root_room_info,
            text="Графский или лицей:",
            style="TLabel",
            anchor="center"
        )
        self.graphskiy_or_not_label.place(relx=0, rely=0.15, relwidth=0.5, relheight=0.15)

        self.graphskiy_or_not = ttk.Combobox(
            self.root_room_info,
            values=options,
            style="TCombobox",
            font=("Helvetica", 27),
            state="readonly", 
            foreground="#4D4D4D"
        )
        self.graphskiy_or_not.set(self.def_graphskiy_not) # Предварительный ввод графский/лицей
        self.graphskiy_or_not.place(relx=0.5, rely=0.15, relwidth=0.5, relheight=0.15)
    ###

    #Блок выбора большого/маленького кабинета
        options = ["Маленький", "Большой"]
        self.small_or_big_label = ttk.Label(
            self.root_room_info,
            text="Маленький или большой:",
            style="TLabel",
            anchor="center"
        )
        self.small_or_big_label.place(relx=0, rely=0.3, relwidth=0.5, relheight=0.15)

        self.small_or_big = ttk.Combobox(
            self.root_room_info,
            values=options,
            style="TCombobox",
            font=("Helvetica", 27),
            state="readonly", 
            foreground="#4D4D4D"
        )
        self.small_or_big.set(self.def_big_not)
        self.small_or_big.place(relx=0.5, rely=0.3, relwidth=0.5, relheight=0.15)
    ###

    #Блок выбора приоритетного кабинета
        self.subjects = [
            "Нет приоритетных кабинетов",
            "Алгебра/Геометрия/ТеорВер",
            "Физика",
            "Русский/Литература",
            "История",
            "Информатика",
            "География",
            "Английский язык",
            "Физра",
            "Биология",
            "Химия",
            "Обществознание"
        ]

        self.prio_subjects_label = ttk.Label(
            self.root_room_info,
            text = "Приоритетные уроки",
            style="TLabel",
            anchor="center",
        )
        self.prio_subjects_label.place(relx=0, rely=0.457, relwidth=1, relheight=0.09)

        self.prio_subjects = Listbox(
            self.root_room_info,
            selectmode=SINGLE,
            background="#DCDCDC",
            selectbackground="#4b6985", 
            font=("Helvetica", 27), 
            foreground="#4D4D4D", 
            relief="groove",
            bd=2
        )
        self.prio_subjects.place(relx=0.2, rely=0.55, relwidth=0.6, relheight=0.3)

        for subject in self.subjects:
            self.prio_subjects.insert(END, subject)
        
        for i in range(len(self.subjects)): # Предварительный ввод предмета
            if self.subjects[i] == self.def_sub:
                a = i
                self.prio_subjects.select_set(a)
                break
    ###
        
    #Футер
        self.save_button = ttk.Button(
            self.root_room_info,
            text="Сохранить",
            style="Main_5_0.TButton",
            command=self.save_data # Сохранение данных
        )
        self.save_button.place(relx=0, rely=0.85, relwidth=0.3, relheight=0.15)

        self.cancel_button = ttk.Button(
            self.root_room_info,
            text="Отмена",
            style="Main_5_0.TButton",
            command=self.root_room_info.destroy # Обычный выход из меню
        )
        self.cancel_button.place(relx=0.3, rely=0.85, relwidth=0.4, relheight=0.15)

        self.delete_button = ttk.Button(
            self.root_room_info,
            text="Удалить",
            style="Main_5_0.TButton",
            command=self.delete_data # Удаление значения
        )
        self.delete_button.place(relx=0.7, rely=0.85, relwidth=0.3, relheight=0.15)
    ###
#Функционал
#Cохранение
    def save_data(self):
    # Данные от ввода пользователя
        data_room = []
        input_num = self.input_number_room.get()
        input_graphskiy_not = self.graphskiy_or_not.get()
        input_big_or_small = self.small_or_big.get()
        selected_indices = self.prio_subjects.curselection()
    ###

    #Проверка на ввод чего-либо у прио урока
        if selected_indices:
            input_sub = self.subjects[selected_indices[0]]
        else:                                         
            warning = eror_popup(self.root_room_info, "Ошибка в прио. предметах")
            warning.root.mainloop()
            return
    ###

    # Проверка на какие-либо изменения в вводе   
        if input_num == self.def_num and input_graphskiy_not == self.def_graphskiy_not and input_sub == self.def_sub:
            warning = eror_popup(self.root_room_info, "Вы ничего не поменяли")
            warning.root.mainloop()
            return
    ###

    #Проверка номера кабинета + ввод в data_room
        all_num = "1234567890"
        count = 0 
        for digit in input_num:
            if digit in all_num:
                count += 1

        if (count == len(input_num)) and (input_num != ""): # Проверка на наличие числа и проверка на ввод данных 
            if int(input_num) // 100 < 6: # Проверка на наличие этажей между 1 - 5
                data_room.append(str(input_num))
            else:
                warning = eror_popup(self.root_room_info, "Ошибка в номере кабинета")
                warning.root.mainloop()
                return                                
        else:
            warning = eror_popup(self.root_room_info, "Ошибка в номере кабинета")
            warning.root.mainloop()
            return
    ###

    #Проверка графского или нет + ввод
        if input_graphskiy_not == "Лицей" or input_graphskiy_not == "Графский": # Проверка на правильность написания ответа
            data_room.append(input_graphskiy_not)
        else:
            warning = eror_popup(self.root_room_info, "Ошибка в Графский/лицей")
            warning.root.mainloop()
            return
    ###

    #Проверка большой или маленький
        if input_big_or_small == "Большой" or input_big_or_small == "Маленький":
            data_room.append(input_big_or_small)
        else:
            warning = eror_popup(self.root_room_info, "Ошибка в большой/мал.")
            warning.root.mainloop()
            return
    ###

    #Ввод прио урока в нужном порядке
        data_room.append(input_sub)
    ###

    # Самая финальная проверка на наличие в таблице + Ввод в таблицу + обновление страницы
        self.con_room_info.execute("SELECT * FROM rooms WHERE Number = ? AND Graph_or_Lyceum = ?", (data_room[0], data_room[1]))
        values = self.con_room_info.fetchall() # Проверка на предварительное наличие кабинета 
        if values and (data_room[0] != self.def_num or data_room[1] != self.def_graphskiy_not):
            warning = eror_popup(self.root_room_info, "Уже есть каб. в корпусе")
            warning.root.mainloop()                    
            return

        self.con_room_info.execute(
            "UPDATE rooms SET Number = ?, Graph_or_Lyceum = ?, Big_or_Small = ?, Subject = ? WHERE Number = ? AND Graph_or_Lyceum = ?",
            (input_num, input_graphskiy_not, input_big_or_small, input_sub, self.def_num, self.answ)
        )
        self.base_room_info.commit()
        self.main_screen.update_buttons()
        warning = popup(self.root_room_info, f"Кабинет {self.room_number} изменен!", "Успех")
        warning.root.mainloop()
    ###

###

#Удаление
    def delete_data(self):
        self.con_room_info.execute("DELETE FROM rooms WHERE Number = ? AND Graph_or_Lyceum = ?", (self.room_number, self.answ))
        self.base_room_info.commit()
        self.main_screen.update_buttons()
        self.root_room_info.destroy()
        warning = popup(self.root_parent, f"Кабинет {self.room_number} удален!", "Успех")
        warning.root.mainloop()   
###     






#<Созд учи> - 6 +++
class menu_add_teacher:
    def __init__(self, parent, base, con):
    #Даннфые корня
        self.root_menu_6 = Toplevel(parent)
        self.root_menu_6.title("Добавление учителя")
        self.root_menu_6.geometry(f"{int(self.root_menu_6.winfo_screenwidth() * 0.56)}x{int(self.root_menu_6.winfo_screenheight()*0.6)}")
        self.root_menu_6.resizable(width=False, height=False) 
        self.base_menu_6 = base
        self.con_menu_6 = con
        self.data_exce = []  
        self.exce_buttons = []
    ###

        style_6 = ttk.Style()

        style_6.theme_use('clam')

        style_6.configure("TLabel", font=("Helvetica", 30, "italic"))

        style_6.configure("TEntry", fieldbackground="#DCDCDC")

        style_6.configure("TCombobox", fieldbackground="#DCDCDC", arrowsize = 0)
        style_6.map("TCombobox", fieldbackground=[("readonly", "#DCDCDC")])


        style_6.configure("Main_6.TButton", font=("Helvetica", 30, "bold"), background="#696969", foreground="white",)
        style_6.map("Main_6.TButton", background=[("active", "#006400")])

        style_6.configure("Sec_6.TButton", font=("Helvetica", 25, "bold"), background="#696969", foreground="white",)
        style_6.map("Sec_6.TButton", background=[("active", "#228B22")])

        self.root_menu_6.configure(bg="#e0dcd4")


    #Виджеты добавления учителя ++
    #Границы 
        separator_1 = Frame(
            self.root_menu_6, 
            height=3,         
            bg='grey',         
            relief=FLAT,
            bd=1              
        )
        separator_1.place(relx=0, rely=0.45, relwidth=0.5, relheight=0.007) #Горизонтальные

        separator_1 = Frame(
            self.root_menu_6, 
            height=3,          
            bg='grey',        
            relief=FLAT,
            bd=1               
        )
        separator_1.place(relx=0.5, rely=0, relwidth= 0.004, relheight= 1) #Вертикальные
    ###

    #Блок фамилии
        self.surname = ttk.Label(
            self.root_menu_6,
            text="Ф",
            style="TLabel",
            anchor="center"
        )
        self.surname.place(relx=0, rely=0, relwidth=0.1, relheight=0.15)

        self.input_surname = ttk.Entry(
            self.root_menu_6,
            style="TEntry",
            font=("Helvetica", 27),
            foreground="#4D4D4D"       
            )
        self.input_surname.place(relx=0.1, rely=0, relwidth=0.4, relheight=0.15)
    ###

    #Блок имени
        self.name = ttk.Label(
            self.root_menu_6,
            text="И",
            style="TLabel",
            anchor="center"
        )
        self.name.place(relx=0, rely=0.15, relwidth=0.1, relheight=0.15)

        self.input_name = ttk.Entry(
            self.root_menu_6,
            style="TEntry",
            font=("Helvetica", 27),
            foreground="#4D4D4D"
        )
        self.input_name.place(relx=0.1, rely=0.15, relwidth=0.4, relheight=0.15)
    ###

    #Блок отчества
        self.patrony = ttk.Label(
            self.root_menu_6,
            text="О",
            style="TLabel",
            anchor="center"
            )
        self.patrony.place(relx=0, rely=0.3, relwidth=0.1, relheight=0.15)

        self.input_patrony = ttk.Entry(
            self.root_menu_6,
            style="TEntry",
            font=("Helvetica", 27),
            foreground="#4D4D4D"
        )
        self.input_patrony.place(relx=0.1, rely=0.3, relwidth=0.4, relheight=0.15)
    ###

    #Блок исключений
        self.exce = ttk.Label(
            self.root_menu_6,
            text="Исключения:",
            style="TLabel",
            anchor="center"
        )
        self.exce.place(relx=0.504, rely=0, relwidth=0.496, relheight=0.1125)
    ###
        
    #Блок дня
        self.day = ttk.Label(
            self.root_menu_6,
            text="День:",
            style="TLabel",
            anchor="center"
        )
        self.day.place(relx=0.504, rely=0.1125, relwidth=0.186, relheight=0.1125)

        self.day_opt = ["Понедельник", "Вторник", "Среда", "Четверг", "Пятница", "Суббота"]
        self.input_day = ttk.Combobox(
            self.root_menu_6,
            values=self.day_opt,
            style="TCombobox",
            font=("Helvetica", 27),
            state="readonly", 
            foreground="#4D4D4D"
        )
        self.input_day.place(relx=0.69, rely=0.1125, relwidth=0.31, relheight=0.1125)
    ###
        
    #Блок исключений по урокам
        self.num_les = ["0", "1", "2", "3", "4", "5", "6", "7", "8"]

        self.lessons = ttk.Label(
            self.root_menu_6,
            text="Не может:",
            style="TLabel",
            anchor="center"
        )
        self.lessons.place(relx=0.504, rely=0.225, relwidth=0.186, relheight=0.1125)

        self.begin = ttk.Label(
            self.root_menu_6,
            text="С",
            style="TLabel",
        )
        self.begin.place(relx=0.7, rely=0.225, relwidth=0.04, relheight=0.1125)

        self.input_begin = ttk.Combobox(
            self.root_menu_6,
            values=self.num_les,
            style="TCombobox",
            font=("Helvetica", 27),
            state="readonly", 
            foreground="#4D4D4D"
        )
        self.input_begin.place(relx=0.74, rely=0.225, relwidth=0.09, relheight=0.1125)

        self.end = ttk.Label(
            self.root_menu_6,
            text="До",
            style="TLabel",
            anchor="center"
        )
        self.end.place(relx=0.83, rely=0.225, relwidth=0.08, relheight=0.1125)

        self.input_end = ttk.Combobox(
            self.root_menu_6,
            values=self.num_les,
            style="TCombobox",
            font=("Helvetica", 27),
            state="readonly", 
            foreground="#4D4D4D"
        )
        self.input_end.place(relx=0.91, rely=0.225, relwidth=0.09, relheight=0.1125)
    ###

    #Кнопка добавить искючение
        self.add_exce = ttk.Button(
            self.root_menu_6,
            text="Создать",
            style="Main_6.TButton",
            command= self.exce_add
        )
        self.add_exce.place(relx=0.504, rely=0.3375, relwidth=0.496, relheight=0.1125)
    ###

    #Блок выбора урока преподования
        self.subjects = [
            "Алгебра/Геометрия/ТеорВер",
            "Физика",
            "Русский/Литература",
            "История",
            "Информатика",
            "География",
            "Английский язык",
            "Физра",
            "Биология/Химия",
            "Технология",
            "Обществознание",
            "ОБЖ"
        ] #Существующие уроки

        self.subjects_label = ttk.Label(
            self.root_menu_6,
            text = "Предмет",
            style="TLabel",
            anchor="center"
        )
        self.subjects_label.place(relx=0, rely=0.457, relwidth=0.5, relheight=0.09)

        self.input_subjects = Listbox(
            self.root_menu_6,
            selectmode=SINGLE,
            background="#DCDCDC",
            selectbackground="#4b6985", 
            font=("Helvetica", 27), 
            foreground="#4D4D4D", 
            relief="groove",
            bd=2
        )
        self.input_subjects.place(relx=0, rely=0.547, relwidth=0.5, relheight=0.303)

        for item in self.subjects:
            self.input_subjects.insert(END, item)
    ###

    # Кнопка добавления учителя
        self.add_button = ttk.Button(
            self.root_menu_6,
            text="Добавить учителя",
            style = "Main_6.TButton",
            command=self.add_teacher
        )
        self.add_button.place(relx=0, rely=0.85, relwidth=1, relheight=0.15)
    ###

#Функционал
#Доб искл
    def exce_add(self):
    #Входные данные
        in_day = self.input_day.get()
        in_beg = self.input_begin.get()
        in_end = self.input_end.get()
    ###

    #Проверка дня
        if in_day not in self.day_opt:
            warning = eror_popup(self.root_menu_6, "Ошибка в дне искл")
            warning.root.mainloop()
            return
    ###

    #Проверка начала исключения
        if in_beg not in self.num_les:
            warning = eror_popup(self.root_menu_6, "Ошибка в искл")
            warning.root.mainloop()
            return
    ###

    #Проверка конца исключения
        if in_beg not in self.num_les:
            warning = eror_popup(self.root_menu_6, "Ошибка в искл")
            warning.root.mainloop()
            return
    ###

    #Проверка начал + конца исключения
        if in_beg > in_end:
            warning = eror_popup(self.root_menu_6, "Ошибка в искл")
            warning.root.mainloop()
            return
    ###

    #Проверка навчало не равно конец
        if in_beg == in_end:
            warning = eror_popup(self.root_menu_6, "Ошибка в искл")
            warning.root.mainloop()
            return
    #

    #Проверка на существование

        for i in self.data_exce:
            if i[0] == in_day:
                warning = eror_popup(self.root_menu_6, "Уже есть такой день")
                warning.root.mainloop()
                return
    ###

    #Проверка на максимальное количество исключений
        if len(self.data_exce) == 6:
            warning = eror_popup(self.root_menu_6, "Макс. количество искл!")
            warning.root.mainloop()                 
            return
    ###

        self.data_exce.append([in_day, in_beg, in_end])
        self.display_exce()
        warning = popup(self.root_menu_6, "Искл. добвлено, еще?", "Успех")
                                    
###

#Обнов искл
    def display_exce(self):
        # Сначала удалим старые кнопки
        for btn in self.exce_buttons:
            btn.destroy()
        self.exce_buttons = []
        # Располагаем кнопки для каждого исключения из self.data_exce
        for i, exce in enumerate(self.data_exce):
            # Пропускаем пустые записи (например, удалённые)
            if not exce: 
                continue
            pos_x = (i % 2) * 0.22 + 0.53  # Относительные координаты по X (начало около 0.53)
            pos_y = (i // 2) * 0.12 + 0.47  # Относительные координаты по Y (начало около 0.47)
            btn = ttk.Button(
                self.root_menu_6,
                text=f"{exce[0]}",  # Показываем день исключения
                command=lambda idx=i, ex=exce: self.open_popup_exce(idx, ex),
                style = "Sec_6.TButton"
            )
            btn.place(relx=pos_x, rely=pos_y, relwidth=0.22, relheight=0.12)
            self.exce_buttons.append(btn)
            
    def update_exce(self):
        self.display_exce()

    # При нажатии на кнопку исключения открывается окно редактирования/удаления
    def open_popup_exce(self, idx, ex):
        popup_exce_1((idx, ex[0], ex[1], ex[2]), self)

###

#Доб учителя
    def add_teacher(self):
    #Данные корня
        name = self.input_name.get()
        surname = self.input_surname.get()
        patrony = self.input_patrony.get()
        ind_subject = self.input_subjects.curselection()
        data_teacher = []
        char = "абвгдеёжзийклмнопрстуфхцчшщъыьэюя"
    ###

    #Проверка имени
        if name == "" or not all(i.lower() in char for i in name):
            warning = eror_popup(self.root_menu_6, "Ошибка в имени")
            warning.root.mainloop()
            return
        else:
            data_teacher.append(name)
    ###

    #Проверка фамилии
        if surname == "" or not all(i.lower() in char for i in surname):
            warning = eror_popup(self.root_menu_6, "Ошибка в фамилии")
            warning.root.mainloop()
            return
        else:
            data_teacher.append(surname)
    ###

    #Проверка отчества
        if patrony == "" or not all(i.lower() in char for i in patrony):
            warning = eror_popup(self.root_menu_6, "Ошибка в отчестве")
            warning.root.mainloop()
            return
        else:
            data_teacher.append(patrony)
    ###

    #Проверка предмета 
        if ind_subject == tuple():
            warning = eror_popup(self.root_menu_6, "Ошибка в прио. предметах")
            warning.root.mainloop()     
            return
        else:
            data_teacher.append(self.subjects[ind_subject[0]])
    ###

    #Проверка на предварительное наличие учителя 
        self.con_menu_6.execute("SELECT * FROM teachers WHERE Surname = ? AND Name = ? AND Patrony = ? AND Lesson = ?",
                            (data_teacher[1], data_teacher[0], data_teacher[2], data_teacher[3]))
        values = self.con_menu_6.fetchall()  
        if values:
            warning = eror_popup(self.root_menu_6, "Уже есть такой учитель")
            warning.root.mainloop()
            return
    ###

    #Вставка данных в таблицу teachers
        self.con_menu_6.execute("INSERT INTO teachers (Surname, Name, Patrony, Lesson) VALUES (?, ?, ?, ?)",
                                (data_teacher[1], data_teacher[0], data_teacher[2], data_teacher[3]))
        self.base_menu_6.commit()
    ###

    #Вставка данных в таблицу исключений
        self.con_menu_6.execute("SELECT id FROM teachers WHERE Surname = ? AND Name = ? AND Patrony = ? AND Lesson = ?",
                                (data_teacher[1], data_teacher[0], data_teacher[2], data_teacher[3]))
        teacher_id = self.con_menu_6.fetchone()#Нахождение индекса

        if teacher_id:
            teacher_id = teacher_id[0]
            
            for i in self.data_exce:
                self.con_menu_6.execute("INSERT INTO exceptions (Rel, Day, Begining, Ending) VALUES (?, ?, ?, ?)",
                                        (teacher_id, i[0], i[1], i[2]))# Вставка исключений в таблицу exceptions
                self.base_menu_6.commit()  

        else:
            warning = eror_popup(self.root_menu_6, "Ошибка при потере индекса")
            warning.root.mainloop()
            return
    ###

        warning = popup(self.root_menu_6, "Учитель добавлен!", "Успех")
        self.data_exce = list()
        self.reset_form()

###

#Сброс полей 
    def reset_form(self):
        self.input_surname.delete(0, END)
        self.input_name.delete(0, END)
        self.input_patrony.delete(0, END)
        self.input_day.set('')
        self.input_begin.set('')
        self.input_end.set('')
        self.input_subjects.selection_clear(0, END)
        self.data_exce.clear()
        for btn in self.exce_buttons:
            btn.destroy()
        self.exce_buttons.clear()
###



#<Окно изм искл> - 0.1 +++
class popup_exce_1:
    def __init__(self, data, parent_menu):
    #Данные корня
        self.idx, self.day_val, self.beg_val, self.end_val = data
        self.parent_menu = parent_menu
        self.root_exce = Toplevel(parent_menu.root_menu_6)
        self.root_exce.title("Изменить исключение")
        self.root_exce.geometry(f"{int(self.root_exce.winfo_screenwidth() * 0.27)}x{int(self.root_exce.winfo_screenheight()*0.3)}")
        self.root_exce.resizable(width=False, height=False)
    ###

        style_6_0 = ttk.Style()

        style_6_0.theme_use('clam')

        style_6_0.configure("TLabel", font=("Helvetica", 30, "italic"))

        style_6_0.configure("TEntry", fieldbackground="#DCDCDC")

        style_6_0.configure("TCombobox", fieldbackground="#DCDCDC", arrowsize = 0)
        style_6_0.map("TCombobox", fieldbackground=[("readonly", "#DCDCDC")])


        style_6_0.configure("Main.TButton", font=("Helvetica", 19, "bold"), background="#696969", foreground="white",)
        style_6_0.map("Main.TButton", background=[("active", "#006400")])

        self.root_exce.configure(bg="#e0dcd4")
        
    #Виджеты изменения исключения - 1
        label = ttk.Label(
            self.root_exce,
            text=f"Исключение {self.day_val.lower()}:",
            style="TLabel",
            anchor="center"
        )
        label.place(relx=0, rely=0, relwidth=1, relheight=0.25)
        
        # Блок дня
        self.day = ttk.Label(
            self.root_exce,
            text="День:",
            style="TLabel",
        )
        self.day.place(relx=0.05, rely=0.25, relwidth=0.295, relheight=0.25)

        self.day_opt = ["Понедельник", "Вторник", "Среда", "Четверг", "Пятница", "Суббота"]
        self.input_day = ttk.Combobox(
            self.root_exce,
            values=self.day_opt,
            style="TCombobox",
            font=("Helvetica", 27),
            state="readonly", 
            foreground="#4D4D4D"
        )
        self.input_day.place(relx=0.3, rely=0.25, relwidth=0.695, relheight=0.25)
        self.input_day.set(self.day_val)  # Задаем изначальное значение

        # Блок исключений по урокам
        self.num_les = ["0", "1", "2", "3", "4", "5", "6", "7", "8"]

        self.lessons = ttk.Label(
            self.root_exce,
            text="Не может:",
            style="TLabel",
        )
        self.lessons.place(relx=0, rely=0.5, relwidth=0.38, relheight=0.25)

        self.begin = ttk.Label(
            self.root_exce,
            text="С",
            style="TLabel",
            anchor= "center"
        )
        self.begin.place(relx=0.39, rely=0.5, relwidth=0.09, relheight=0.25)

        self.input_begin = ttk.Combobox(
            self.root_exce,
            values=self.num_les,
            style="TCombobox",
            font=("Helvetica", 27),
            state="readonly", 
            foreground="#4D4D4D"
        )
        self.input_begin.place(relx=0.48, rely=0.5, relwidth=0.18, relheight=0.25)
        self.input_begin.set(self.beg_val)  # Задаем изначальное значение

        self.end = ttk.Label(
            self.root_exce,
            text="До",
            style="TLabel",
            anchor= "center"
        )
        self.end.place(relx=0.66, rely=0.5, relwidth=0.16, relheight=0.25)

        self.input_end = ttk.Combobox(
            self.root_exce,
            values=self.num_les,
            style="TCombobox",
            font=("Helvetica", 27),
            state="readonly", 
            foreground="#4D4D4D"
        )
        self.input_end.place(relx=0.82, rely=0.5, relwidth=0.18, relheight=0.25)
        self.input_end.set(self.end_val)  # Задаем изначальное значение

        # Футер (кнопки)
        self.save_button = ttk.Button(
            self.root_exce,
            text="Сохранить",
            style="Main.TButton",
            command=self.save_data  # Добавьте обработчик, если понадобится
        )
        self.save_button.place(relx=0, rely=0.75, relwidth=0.3, relheight=0.25)

        self.cancel_button = ttk.Button(
            self.root_exce,
            text="Отмена",
            style="Main.TButton",
            command=self.root_exce.destroy
        )
        self.cancel_button.place(relx=0.3, rely=0.75, relwidth=0.4, relheight=0.25)

        self.delete_button = ttk.Button(
            self.root_exce,
            text="Удалить",
            style="Main.TButton",
            command=self.delete_data
        )
        self.delete_button.place(relx=0.7, rely=0.75, relwidth=0.3, relheight=0.25)

#Функционал
    #Удаление исключения
    def delete_data(self):
        # Удаляем исключение с индексом self.idx
        if self.idx < len(self.parent_menu.data_exce):
            self.parent_menu.data_exce.pop(self.idx)
        self.parent_menu.update_exce()
        self.root_exce.destroy()
        popup(self.parent_menu.root_menu_6, "Исключение удалено!", "Успех").root.mainloop()
    ###

    #Сохранение данных
    #Метод сохранения изменений в исключении
    def save_data(self):
        new_day = self.input_day.get()
        new_beg = self.input_begin.get()
        new_end = self.input_end.get()
        
        # Проверка дня
        if new_day not in self.day_opt:
            warning = eror_popup(self.root_exce, "Ошибка в дне искл")
            warning.root.mainloop()
            return
        
        # Проверка уроков
        if new_beg not in self.num_les or new_end not in self.num_les:
            warning = eror_popup(self.root_exce, "Ошибка в искл")
            warning.root.mainloop()
            return
        
        # Проверка последовательности уроков (начало не больше конца)
        if new_beg > new_end:
            warning = eror_popup(self.root_exce, "Ошибка в искл")
            warning.root.mainloop()
            return
        
        #Проверка навчало не равно конец
        if new_beg == new_end:
            warning = eror_popup(self.root_menu_6, "Ошибка в искл")
            warning.root.mainloop()
            return
        #
        
        # Новая проверка: если уже существует другое исключение с таким днем
        for idx, exce in enumerate(self.parent_menu.data_exce):
            if idx != self.idx and exce[0] == new_day:
                warning = eror_popup(self.root_exce, "Уже есть искл этого дня")
                warning.root.mainloop()             
                return

        # Обновляем значение исключения в родительском окне
        self.parent_menu.data_exce[self.idx] = [new_day, new_beg, new_end]
        self.parent_menu.update_exce()
        popup(self.parent_menu.root_menu_6, "Искл. обновлено!", "Успех").root.mainloop()
        self.root_exce.destroy()

    ###
     
###





        
#<СМ учителя> - 7 +++
class menu_check_teacher:
    def __init__(self, parent, base, connect):
    #Данные корня
        self.root_menu_7 = Toplevel(parent)
        self.root_menu_7.title("Просмотр учителей")
        self.root_menu_7.geometry(f"{int(self.root_menu_7.winfo_screenwidth() * 0.56)}x{int(self.root_menu_7.winfo_screenheight()*0.6)}")
        self.root_menu_7.resizable(width=False, height=False) 
        self.teacher_menu_7_base = base
        self.teacher_menu_7_con = connect

        style_7 = ttk.Style()

        style_7.theme_use('clam')

        style_7.configure("TLabel", font=("Helvetica", 30, "italic"))

        style_7.configure("TEntry", fieldbackground="#DCDCDC")

        style_7.configure("TCombobox", fieldbackground="#DCDCDC", arrowsize = 0)
        style_7.map("TCombobox", fieldbackground=[("readonly", "#DCDCDC")])

        style_7.configure("Sec_7.TButton", font=("Helvetica", 10, "bold"), background="#696969", foreground="white",)
        style_7.map("Sec_7.TButton", background=[("active", "#228B22")])

        self.root_menu_7.configure(bg="#e0dcd4")

    #Виджеты проверки учителей
        self.teacher_label = ttk.Label(
            self.root_menu_7,
            text="Учителя",
            style="TLabel",
            anchor="center"
        )
        self.teacher_label.place(relx=0, rely=0, relwidth=1, relheight=0.1)

        self.display_buttons()

    #Функционал
    def display_buttons(self):
        self.teacher_menu_7_con.execute("SELECT DISTINCT Surname, Name, Patrony FROM teachers")
        teachers = self.teacher_menu_7_con.fetchall()

        wid_x = 0.16  # ширина кнопки (4 кнопки + отступы)
        hei_y = 0.08  # высота кнопки

        for i, teacher in enumerate(teachers):
            pos_x = (i % 6) * 0.16 + 0.01  # позиция по x (6 колонки)
            pos_y = (i // 6) * 0.09 + 0.1   # позиция по y (10 строк)

            teacher_button = ttk.Button(
                self.root_menu_7,
                text=f"{teacher[0]} {teacher[1][0]}. {teacher[2][0]}.",
                command=lambda t=teacher: self.open_teacher_screen(t),
                style = "Sec_7.TButton"
            )
            teacher_button.place(relx=pos_x, rely=pos_y, relwidth=wid_x, relheight=hei_y)

    def open_teacher_screen(self, teacher):
        teacher_info(self.root_menu_7, teacher, self.teacher_menu_7_base, self, self.root_menu_7)

    def update_buttons(self):
        for widget in self.root_menu_7.winfo_children():
            if widget.winfo_class() == "TButton":
                widget.destroy()
        self.display_buttons()



#<Инфа уч> - # +++
class teacher_info:
    def __init__(self, root, teacher, dbs, main_screen, root_parent):
        #Данные корня
        self.root_teacher_info = Toplevel(root)
        self.root_parent = root_parent
        self.main_screen = main_screen
        self.teacher = teacher
        self.root_teacher_info.title(f"Учитель {self.teacher[0]} {self.teacher[1]} {self.teacher[2]}")
        self.root_teacher_info.geometry(f"{int(self.root_teacher_info.winfo_screenwidth() * 0.56)}x{int(self.root_teacher_info.winfo_screenheight()*0.6)}")
        self.root_teacher_info.resizable(width=False, height=False)
        self.base_teacher_info = dbs
        self.con_teacher_info = self.base_teacher_info.cursor()

        style_7_0 = ttk.Style()

        style_7_0.theme_use('clam')

        style_7_0.configure("TLabel", font=("Helvetica", 30, "italic"))

        style_7_0.configure("TEntry", fieldbackground="#DCDCDC")

        style_7_0.configure("TCombobox", fieldbackground="#DCDCDC", arrowsize = 0)
        style_7_0.map("TCombobox", fieldbackground=[("readonly", "#DCDCDC")])

        style_7_0.configure("Main_7_0.TButton", font=("Helvetica", 30, "bold"), background="#696969", foreground="white",)
        style_7_0.map("Main_7_0.TButton", background=[("active", "#006400")])

        style_7_0.configure("Sec_7_0.TButton", font=("Helvetica", 25, "bold"), background="#696969", foreground="white",)
        style_7_0.map("Sec_7_0.TButton", background=[("active", "#228B22")])

        self.root_teacher_info.configure(bg="#e0dcd4")

        self.con_teacher_info.execute(
            "SELECT * FROM teachers WHERE Surname = ? AND Name = ? AND Patrony = ?",
            (self.teacher[0], self.teacher[1], self.teacher[2])
        )
        full_data = self.con_teacher_info.fetchall()
        self.def_surname = full_data[0][1]
        self.def_name = full_data[0][2]
        self.def_patrony = full_data[0][3]
        self.def_lesson = full_data[0][4]
        self.teacher_id = full_data[0][0]

        # Загрузка исключений из БД для данного учителя по полю Rel
        self.data_exce = []  # локальный список исключений
        self.exce_buttons = []  # список кнопок для отображения исключений
        self.con_teacher_info.execute(
            "SELECT Day, Begining, Ending FROM exceptions WHERE Rel = ?",
            (self.teacher_id,)
        )
        exceptions = self.con_teacher_info.fetchall()
        for exc in exceptions:
            self.data_exce.append(list(exc))  # приводим к списку для единообразия

    #Виджеты инфы учителя
        # Границы (аналогично menu_add_teacher)
        separator_1 = Frame(
            self.root_teacher_info, 
            height=3,
            bg='grey',
            relief=FLAT,
            bd=1
        )
        separator_1.place(relx=0, rely=0.45, relwidth=0.5, relheight=0.007)

        separator_2 = Frame(
            self.root_teacher_info, 
            height=3,
            bg='grey',
            relief=FLAT,
            bd=1
        )
        separator_2.place(relx=0.5, rely=0, relwidth=0.004, relheight=1)

        # Блок Фамилия
        self.surname = ttk.Label(
            self.root_teacher_info,
            text="Ф",
            style="TLabel",
            anchor="center"
        )
        self.surname.place(relx=0, rely=0, relwidth=0.1, relheight=0.15)

        self.input_surname = ttk.Entry(
            self.root_teacher_info,
            style="TEntry",
            font=("Helvetica", 27),
            foreground="#4D4D4D"
        )
        self.input_surname.place(relx=0.1, rely=0, relwidth=0.4, relheight=0.15)
        self.input_surname.insert(0, self.def_surname)

        # Блок Имя
        self.name = ttk.Label(
            self.root_teacher_info,
            text="И",
            style="TLabel",
            anchor="center"
        )
        self.name.place(relx=0, rely=0.15, relwidth=0.1, relheight=0.15)

        self.input_name = ttk.Entry(
            self.root_teacher_info,
            style="TEntry",
            font=("Helvetica", 27),
            foreground="#4D4D4D"
        )
        self.input_name.place(relx=0.1, rely=0.15, relwidth=0.4, relheight=0.15)
        self.input_name.insert(0, self.def_name)

        # Блок Отчество
        self.patrony = ttk.Label(
            self.root_teacher_info,
            text="О",
            style="TLabel",
            anchor="center"
        )
        self.patrony.place(relx=0, rely=0.3, relwidth=0.1, relheight=0.15)

        self.input_patrony = ttk.Entry(
            self.root_teacher_info,
            style="TEntry",
            font=("Helvetica", 27),
            foreground="#4D4D4D"
        )
        self.input_patrony.place(relx=0.1, rely=0.3, relwidth=0.4, relheight=0.15)
        self.input_patrony.insert(0, self.def_patrony)

        # Блок выбора предмета преподования
        self.subjects = [
            "Алгебра/Геометрия/ТеорВер",
            "Физика",
            "Русский/Литература",
            "История",
            "Информатика",
            "География",
            "Английский язык",
            "Физра",
            "Биология/Химия",
            "Технология",
            "Обществознание",
            "ОБЖ"
        ]
        self.subjects_label = ttk.Label(
            self.root_teacher_info,
            text="Предмет",
            style="TLabel",
            anchor="center"
        )
        self.subjects_label.place(relx=0, rely=0.457, relwidth=0.5, relheight=0.09)

        self.input_subjects = Listbox(
            self.root_teacher_info,
            selectmode=SINGLE,
            background="#DCDCDC",
            selectbackground="#4b6985", 
            font=("Helvetica", 27), 
            foreground="#4D4D4D", 
            relief="groove",
            bd=2)
        
        self.input_subjects.place(relx=0, rely=0.547, relwidth=0.5, relheight=0.303)
        for item in self.subjects:
            self.input_subjects.insert(END, item)
        for i, subj in enumerate(self.subjects):
            if subj == self.def_lesson:
                self.input_subjects.select_set(i)
                break

        # Блок исключений (заголовок)
        self.exce = ttk.Label(
            self.root_teacher_info,
            text="Исключения:",
            style="TLabel",
            anchor="center"
        )
        self.exce.place(relx=0.504, rely=0, relwidth=0.496, relheight=0.1125)

        # Блок дня для исключений
        self.day = ttk.Label(
            self.root_teacher_info,
            text="День:",
            style="TLabel",
            anchor="center"
        )
        self.day.place(relx=0.504, rely=0.1125, relwidth=0.186, relheight=0.1125)

        self.day_opt = ["Понедельник", "Вторник", "Среда", "Четверг", "Пятница", "Суббота"]
        self.input_day = ttk.Combobox(
            self.root_teacher_info,
            values=self.day_opt,
            style="TCombobox",
            font=("Helvetica", 27),
            state="readonly", 
            foreground="#4D4D4D")
        self.input_day.place(relx=0.69, rely=0.1125, relwidth=0.31, relheight=0.1125)

        # Блок исключений по урокам
        self.num_les = ["0", "1", "2", "3", "4", "5", "6", "7", "8"]

        self.lessons = ttk.Label(
            self.root_teacher_info,
            text="Не может:",
            style="TLabel",
            anchor="center"
        )
        self.lessons.place(relx=0.504, rely=0.225, relwidth=0.186, relheight=0.1125)

        self.begin = ttk.Label(
            self.root_teacher_info,
            text="С",
            style="TLabel",
            anchor="center"
        )
        self.begin.place(relx=0.69, rely=0.225, relwidth=0.05, relheight=0.1125)

        self.input_begin = ttk.Combobox(
            self.root_teacher_info,
            values=self.num_les,
            style="TCombobox",
            font=("Helvetica", 27),
            state="readonly", 
            foreground="#4D4D4D"
        )
        self.input_begin.place(relx=0.74, rely=0.225, relwidth=0.09, relheight=0.1125)

        self.end = ttk.Label(
            self.root_teacher_info,
            text="До",
            style="TLabel",
            anchor="center"
        )
        self.end.place(relx=0.83, rely=0.225, relwidth=0.08, relheight=0.1125)

        self.input_end = ttk.Combobox(
            self.root_teacher_info,
            values=self.num_les,
            style="TCombobox",
            font=("Helvetica", 27),
            state="readonly", 
            foreground="#4D4D4D"
        )
        self.input_end.place(relx=0.91, rely=0.225, relwidth=0.09, relheight=0.1125)

        # Кнопка добавления исключения
        self.add_exce = ttk.Button(
            self.root_teacher_info,
            text="Создать",
            style="Main_7_0.TButton",
            command=self.exce_add
        )
        self.add_exce.place(relx=0.504, rely=0.3375, relwidth=0.496, relheight=0.1125)

        # Футер: кнопки "Сохранить", "Отмена", "Удалить", "Доп инфа"
        self.save_button = ttk.Button(
            self.root_teacher_info,
            text="Сохранить",
            style="Main_7_0.TButton",
            command=self.save_data
        )
        self.save_button.place(relx=0, rely=0.85, relwidth=0.25, relheight=0.15)

        self.cancel_button = ttk.Button(
            self.root_teacher_info,
            text="Отмена",
            style="Main_7_0.TButton",
            command=self.root_teacher_info.destroy
        )
        self.cancel_button.place(relx=0.25, rely=0.85, relwidth=0.252, relheight=0.15)

        self.cancel_button = ttk.Button(
            self.root_teacher_info,
            text="Доп инфа",
            style="Main_7_0.TButton",
            command=self.open_extra_info
        )
        self.cancel_button.place(relx=0.502, rely=0.85, relwidth=0.248, relheight=0.15)

        self.delete_button = ttk.Button(
            self.root_teacher_info,
            text="Удалить",
            style="Main_7_0.TButton",
            command=self.delete_data
        )
        self.delete_button.place(relx=0.75, rely=0.85, relwidth=0.25, relheight=0.15)

        # Отображаем исключения, загруженные из БД
        self.display_exce()
    
    #Функционал
    def open_extra_info(self):
        extra_info(self.teacher_id, self.root_teacher_info, self.con_teacher_info, self.base_teacher_info)
        
    def exce_add(self):
        # Входные данные
        in_day = self.input_day.get()
        in_beg = self.input_begin.get()
        in_end = self.input_end.get()
        
        # Проверка дня
        if in_day not in self.day_opt:
            warning = eror_popup(self.root_teacher_info, "Ошибка в дне искл")
            warning.root.mainloop()
            return
        
        # Проверка начальных значений уроков
        if in_beg not in self.num_les:
            warning = eror_popup(self.root_teacher_info, "Ошибка в искл")
            warning.root.mainloop()
            return
        
        if in_end not in self.num_les:
            warning = eror_popup(self.root_teacher_info, "Ошибка в искл")
            warning.root.mainloop()
            return
        
        # Проверка последовательности (начало не больше конца)
        if in_beg > in_end:
            warning = eror_popup(self.root_teacher_info, "Ошибка в искл")
            warning.root.mainloop()
            return
        
        #Проверка навчало не равно конец
        if in_beg == in_end:
            warning = eror_popup(self.root_teacher_info, "Ошибка в искл")
            warning.root.mainloop()
            return
        #

        # Проверка наличия исключения для уже заданного дня
        for exc in self.data_exce:
            if exc[0] == in_day:
                warning = eror_popup(self.root_teacher_info, "Уже есть такой день")
                warning.root.mainloop()
                return
        # Ограничение количества исключений
        if len(self.data_exce) >= 6:
            warning = eror_popup(self.root_teacher_info, "Макс. количество искл!")
            warning.root.mainloop()
            return

        

        self.data_exce.append([in_day, in_beg, in_end])
        self.con_teacher_info.execute(
            """INSERT INTO exceptions (Rel, Day, Begining, Ending) VALUES (?, ?, ?, ?)""",
            (self.teacher_id, in_day, in_beg, in_end, )
        )
        self.base_teacher_info.commit()
        self.display_exce()
        popup(self.root_teacher_info, "Искл. добавлено, можно еще", "Успех").root.mainloop()

    def display_exce(self):
        # Удаляем старые кнопки
        for btn in self.exce_buttons:
            btn.destroy()
        self.exce_buttons = []
        # Размещаем кнопки для каждого исключения из self.data_exce
        for i, exc in enumerate(self.data_exce):
            if not exc:
                continue
            pos_x = (i % 2) * 0.22 + 0.53
            pos_y = (i // 2) * 0.12 + 0.47
            btn = ttk.Button(
                self.root_teacher_info,
                text=f"{exc[0]}",  # отображаем день
                command=lambda idx=i, ex=exc: self.open_popup_exce(idx, ex),
                style = "Sec_7_0.TButton"
            )
            btn.place(relx=pos_x, rely=pos_y, relwidth=0.22, relheight=0.12)
            self.exce_buttons.append(btn)

    def update_exce(self):
        self.display_exce()

    def open_popup_exce(self, idx, exc):
        popup_exce_2((idx, exc[0], exc[1], exc[2]), self)

    def save_data(self):
        # Пример сохранения данных учителя (реализация уже присутствует)
        data_teacher = []
        input_surname = self.input_surname.get()
        input_name = self.input_name.get()
        input_patrony = self.input_patrony.get()
        selected_indices = self.input_subjects.curselection()
        if not selected_indices:
            warning = eror_popup(self.root_teacher_info, "Ошибка в предмете")
            warning.root.mainloop()
            return
        input_subject = self.subjects[selected_indices[0]]
        if (input_surname == self.def_surname and input_name == self.def_name and 
            input_patrony == self.def_patrony and input_subject == self.def_lesson):
            warning = eror_popup(self.root_teacher_info, "Вы ничего не поменяли")
            warning.root.mainloop()
            return
        char = "абвгдеёжзийклмнопрстуфхцчшщъыьэюя"
        if input_surname == "" or not all(c.lower() in char for c in input_surname):
            warning = eror_popup(self.root_teacher_info, "Ошибка в фамилии")
            warning.root.mainloop()
            return
        else:
            data_teacher.append(input_surname)
        if input_name == "" or not all(c.lower() in char for c in input_name):
            warning = eror_popup(self.root_teacher_info, "Ошибка в имени")
            warning.root.mainloop()
            return
        else:
            data_teacher.append(input_name)
        if input_patrony == "" or not all(c.lower() in char for c in input_patrony):
            warning = eror_popup(self.root_teacher_info, "Ошибка в отчестве")
            warning.root.mainloop()
            return
        else:
            data_teacher.append(input_patrony)
        if input_subject not in self.subjects:
            warning = eror_popup(self.root_teacher_info, "Ошибка в предмете")
            warning.root.mainloop()
            return
        else:
            data_teacher.append(input_subject)
        self.con_teacher_info.execute(
            "UPDATE teachers SET Surname = ?, Name = ?, Patrony = ?, Lesson = ? WHERE id = ?",
            (data_teacher[0], data_teacher[1], data_teacher[2], data_teacher[3], self.teacher_id)
        )
        self.base_teacher_info.commit()
        self.main_screen.update_buttons()
        popup(self.root_teacher_info, "Учитель обновлён!", "Успех").root.mainloop()
        
    def delete_data(self):
        # Удаляем учителя из таблицы teachers
        self.con_teacher_info.execute("DELETE FROM teachers WHERE id = ?", (self.teacher_id,))
        self.base_teacher_info.commit()
    
        # Удаляем все исключения, связанные с данным учителем (по полю Rel)
        self.con_teacher_info.execute("DELETE FROM exceptions WHERE Rel = ?", (self.teacher_id,))
        self.base_teacher_info.commit()

        # Удаление всех уроков с таким учителем
        self.con_teacher_info.execute("DELETE FROM lessons WHERE Id_teacher = ?", (self.teacher_id,))
        self.base_teacher_info.commit()

        self.main_screen.update_buttons()
        self.root_teacher_info.destroy()
        popup(self.root_parent, "Учитель c искл и уркоами удалены!", "Успех").root.mainloop()

#<Окно изм искл> - 0.2 +++
class popup_exce_2:
    def __init__(self, data, parent_menu):
        #Данные корня
        self.idx, self.day_val, self.beg_val, self.end_val = data
        self.parent_menu = parent_menu
        self.root_exce = Toplevel(parent_menu.root_teacher_info)
        self.root_exce.title("Изменить исключение")
        self.root_exce.geometry(f"{int(self.root_exce.winfo_screenwidth() * 0.27)}x{int(self.root_exce.winfo_screenheight()*0.3)}")
        self.root_exce.resizable(width=False, height=False)
        
        style_6_1 = ttk.Style()

        style_6_1.theme_use('clam')

        style_6_1.configure("TLabel", font=("Helvetica", 30, "italic"))

        style_6_1.configure("TEntry", fieldbackground="#DCDCDC")

        style_6_1.configure("TCombobox", fieldbackground="#DCDCDC", arrowsize = 0)
        style_6_1.map("TCombobox", fieldbackground=[("readonly", "#DCDCDC")])


        style_6_1.configure("Main_6_1.TButton", font=("Helvetica", 19, "bold"), background="#696969", foreground="white",)
        style_6_1.map("Main_6_1.TButton", background=[("active", "#006400")])

        self.root_exce.configure(bg="#e0dcd4")

        #Виджеты изменения исключения - 2
        label = ttk.Label(
            self.root_exce,
            text=f"Исключение {self.day_val.lower()}:",
            style="TLabel",
            anchor="center"
        )
        label.place(relx=0, rely=0, relwidth=1, relheight=0.25)
        
        # Блок дня
        self.day = ttk.Label(
            self.root_exce,
            text="День:",
            style="TLabel",
        )
        self.day.place(relx=0.05, rely=0.25, relwidth=0.295, relheight=0.25)

        self.day_opt = ["Понедельник", "Вторник", "Среда", "Четверг", "Пятница", "Суббота"]
        self.input_day = ttk.Combobox(
            self.root_exce,
            values=self.day_opt,
            style="TCombobox",
            font=("Helvetica", 27),
            state="readonly", 
            foreground="#4D4D4D"
        )
        self.input_day.place(relx=0.345, rely=0.25, relwidth=0.655, relheight=0.25)
        self.input_day.set(self.day_val)  # устанавливаем предустановленное значение

        # Блок исключений по урокам
        self.num_les = ["0", "1", "2", "3", "4", "5", "6", "7", "8"]
        self.lessons = ttk.Label(
            self.root_exce,
            text="Не может:",
            style="TLabel",
        )
        self.lessons.place(relx=0, rely=0.5, relwidth=0.38, relheight=0.25)

        self.begin = ttk.Label(
            self.root_exce,
            text="С",
            style="TLabel",
            anchor= "center"
        )
        self.begin.place(relx=0.39, rely=0.5, relwidth=0.09, relheight=0.25)

        self.input_begin = ttk.Combobox(
            self.root_exce,
            values=self.num_les,
            style="TCombobox",
            font=("Helvetica", 27),
            state="readonly", 
            foreground="#4D4D4D"
        )
        self.input_begin.place(relx=0.48, rely=0.5, relwidth=0.18, relheight=0.25)
        self.input_begin.set(self.beg_val)  # устанавливаем предустановленное значение

        self.end = ttk.Label(
            self.root_exce,
            text="До",
            style="TLabel",
            anchor= "center"
        )
        self.end.place(relx=0.66, rely=0.5, relwidth=0.16, relheight=0.25)

        self.input_end = ttk.Combobox(
            self.root_exce,
            values=self.num_les,
            style="TCombobox",
            font=("Helvetica", 27),
            state="readonly", 
            foreground="#4D4D4D"
        )
        self.input_end.place(relx=0.82, rely=0.5, relwidth=0.18, relheight=0.25)
        self.input_end.set(self.end_val)  # устанавливаем предустановленное значение

        # Футер (кнопки "Сохранить", "Отмена", "Удалить")
        self.save_button = ttk.Button(
            self.root_exce,
            text="Сохранить",
            style = "Main_6_1.TButton",
            command=self.save_data
        )
        self.save_button.place(relx=0, rely=0.75, relwidth=0.3, relheight=0.25)

        self.cancel_button = ttk.Button(
            self.root_exce,
            text="Отмена",
            style = "Main_6_1.TButton",
            command=self.root_exce.destroy
        )
        self.cancel_button.place(relx=0.3, rely=0.75, relwidth=0.4, relheight=0.25)

        self.delete_button = ttk.Button(
            self.root_exce,
            text="Удалить",
            style = "Main_6_1.TButton",
            command=self.delete_data
        )
        self.delete_button.place(relx=0.7, rely=0.75, relwidth=0.3, relheight=0.25)

    #Функционал
    def delete_data(self):
        # Удаляем запись из базы данных (если существует)
        self.parent_menu.con_teacher_info.execute(
            "DELETE FROM exceptions WHERE Rel = ? AND Day = ? AND Begining = ? AND Ending = ?",
            (self.parent_menu.teacher_id, self.day_val, self.beg_val, self.end_val)
        )
        self.parent_menu.base_teacher_info.commit()
        
        # Удаляем исключение из локального списка
        if self.idx < len(self.parent_menu.data_exce):
            self.parent_menu.data_exce.pop(self.idx)
        self.parent_menu.update_exce()
        self.root_exce.destroy()
        popup(self.parent_menu.root_teacher_info, "Исключение удалено!", "Успех").root.mainloop()

    def save_data(self):
        new_day = self.input_day.get()
        new_beg = self.input_begin.get()
        new_end = self.input_end.get()

        if new_day not in self.day_opt:
            warning = eror_popup(self.root_exce, "Ошибка в дне искл")
            warning.root.mainloop()
            return

        if new_beg not in self.num_les or new_end not in self.num_les:
            warning = eror_popup(self.root_exce, "Ошибка в искл")
            warning.root.mainloop()
            return

        if new_beg > new_end:
            warning = eror_popup(self.root_exce, "Ошибка в искл")
            warning.root.mainloop()
            return

        #Проверка навчало не равно конец
        if new_beg == new_end:
            warning = eror_popup(self.root_menu_6, "Ошибка в искл")
            warning.root.mainloop()
            return
        #

        for idx, exce in enumerate(self.parent_menu.data_exce):
            if idx != self.idx and exce[0] == new_day:
                warning = eror_popup(self.root_exce, "Уже есть искл. этого дня")
                warning.root.mainloop()               
                return

        self.parent_menu.data_exce[self.idx] = [new_day, new_beg, new_end]
        self.parent_menu.update_exce()
        popup(self.parent_menu.root_teacher_info, "Искл. обновлено!", "Успех").root.mainloop()
        self.root_exce.destroy()
###



#<Окно доп инфы> - # +++
class extra_info:
    def __init__(self, id, parent_menu, con, base):
        #Данные корня
        self.root_extra_info = Toplevel(parent_menu)
        self.con = con
        self.base = base
        self.parent_id = id
        self.root_extra_info.title("Доп инфа")
        self.root_extra_info.geometry(f"{int(self.root_extra_info.winfo_screenwidth() * 0.54)}x{int(self.root_extra_info.winfo_screenheight()*0.58)}")
        self.root_extra_info.resizable(width=False, height=False)

        style_7_2 = ttk.Style()

        style_7_2.theme_use('clam')

        style_7_2.configure("Main.TLabel", font=("Helvetica", 30, "italic"))

        style_7_2.configure("Sec.TLabel", font=("Helvetica", 25),
            borderwidth=2,
            relief="solid", 
            bg='white',
            fg='black')

        style_7_2.configure("TEntry", fieldbackground="#DCDCDC")

        style_7_2.configure("TCombobox", fieldbackground="#DCDCDC", arrowsize = 0)
        style_7_2.map("TCombobox", fieldbackground=[("readonly", "#DCDCDC")])


        style_7_2.configure("Main.TButton", font=("Helvetica", 30, "bold"), background="#696969", foreground="white",)
        style_7_2.map("Main.TButton", background=[("active", "#006400")])

        self.root_extra_info.configure(bg="#e0dcd4")
    #Лейблы доп инфы
        #Отбражение классов преподования
        label = ttk.Label(
            self.root_extra_info,
            text="Преподает в:",
            style="Mian.TLabel",
            anchor="center"
        )
        label.place(relx=0, rely=0, relwidth=1, relheight=0.15)

        self.con.execute(
            "SELECT Id_parallel, Subject, Hours FROM lessons WHERE Id_teacher = ?",
            (self.parent_id,)
        )
        info_all_class = self.con.fetchall()

        self.labels = []

        for i, info_cer_class in enumerate(info_all_class):
            pos_x = (i % 4) * 0.25  # 4 кнопки в ряд
            pos_y = 0.15 + (i // 4) * 0.1   # новый ряд каждые 4 кнопки
            
            self.con.execute(
            "SELECT Letter, Number FROM parallels WHERE id = ?",
            (info_cer_class[0],))
            info_cer_parallel = self.con.fetchall()

            lab = ttk.Label(
            self.root_extra_info,
            text=f"{info_cer_parallel[0][1]}{info_cer_parallel[0][0]} - {info_cer_class[1][0:3]}.: {info_cer_class[2]} ч.",
            style = "Sec.TLabel"
            
            )
            lab.place(relx= pos_x, rely=pos_y, relwidth=0.25, relheight=0.1)
            self.labels.append(label)

        #Виджеты
        #Отображение количества часов работы
        label = ttk.Label(
            self.root_extra_info,
            text="Общее количество часов:",
            style = "Main.TLabel",
        )
        label.place(relx=0.2, rely=0.75, relwidth=0.5, relheight=0.1)

        self.con.execute(
            "SELECT Hours FROM lessons WHERE Id_teacher = ?",
            (self.parent_id,)
        )
        num_of_studing = self.con.fetchall()
        count = 0
        for i in num_of_studing:
            count += int(i[0])

        label = ttk.Label(
            self.root_extra_info,
            text=count,
            font=("Helvetica", 25)
        )
        label.place(relx=0.7, rely=0.75, relwidth=0.1, relheight=0.1)

        #Футер
        self.save_button = ttk.Button(
            self.root_extra_info,
            text="Выход",
            command=self.root_extra_info.destroy,
            style = "Main.TButton"
        )
        self.save_button.place(relx=0, rely=0.85, relwidth=1, relheight=0.15)
        





#<Окошко ошибок> +++
class eror_popup:
    def __init__(self, parent, mess):
        self.root = Toplevel(parent)
        self.root.geometry("400x300")
        self.root.geometry(f"{int(self.root.winfo_screenwidth() * 0.2)}x{int(self.root.winfo_screenheight()*0.25)}")
        self.root.resizable(width=False, height=False)

        style_error = ttk.Style()

        style_error.theme_use('clam')

        style_error.configure("Error.TButton", font=("Helvetica", 30, "bold"), background="#696969", foreground="white",)
        style_error.map("Error.TButton", background=[("active", "#E52B50")])


        self.root.title("Ошибка")
        text_eror = Label(
            self.root,
            text = mess,
            font = ("Helvetica", 20,)
        )
        text_eror.place(relx = 0, rely = 0.17, relwidth = 1, relheight = 0.3)

        self.loading_chars = ["xxooo", "oxxoo", "ooxxo", "oooxx", "ooxxo", "oxxoo"]
        self.loading_index = 0
        self.loading_label = Label(
            self.root,
            text=self.loading_chars[0],
            font=("Helvetica", 40),
            fg = "#E52B50"
        )
        self.loading_label.place(relx=0, rely=0.37, relwidth=1, relheight=0.2)
        self.animate_loading()

        button_end = ttk.Button(
            self.root,
            text = "Окей",
            style = "Error.TButton",
            command = self.root.destroy
        )
        button_end.place(relx = 0, rely = 0.72, relwidth = 1, relheight = 0.28)

    def animate_loading(self):
        self.loading_index = (self.loading_index + 1) % len(self.loading_chars)
        self.loading_label.config(text=self.loading_chars[self.loading_index])
        self.root.after(500, self.animate_loading)
###

#<Просто окошко> +++
class popup:
    def __init__(self, parent, mess, title):
        self.root = Toplevel(parent)
        self.root.geometry(f"{int(self.root.winfo_screenwidth() * 0.2)}x{int(self.root.winfo_screenheight()*0.25)}")
        self.root.resizable(width=False, height=False)
        self.root.title(title)

        style_error = ttk.Style()

        style_error.theme_use('clam')

        style_error.configure("Popup.TButton", font=("Helvetica", 30, "bold"), background="#696969", foreground="white",)
        style_error.map("Popup.TButton", background=[("active", "black")])


        text_eror = Label(
            self.root,
            text = mess,
            font = ("Helvetica", 20, )
        )
        text_eror.place(relx = 0, rely = 0.17, relwidth = 1, relheight = 0.3)

        self.loading_chars = ["xxooo", "oxxoo", "ooxxo", "oooxx", "ooxxo", "oxxoo"]
        self.loading_index = 0
        self.loading_label = Label(
            self.root,
            text=self.loading_chars[0],
            font=("Helvetica", 40),
            fg = "Black"
        )
        self.loading_label.place(relx=0, rely=0.37, relwidth=1, relheight=0.2)
        self.animate_loading()

        button_end = ttk.Button(
            self.root,
            text = "Окей",
            style = "Popup.TButton",
            command = self.root.destroy
        )
        button_end.place(relx = 0, rely = 0.72, relwidth = 1, relheight = 0.28)
    def animate_loading(self):
        self.loading_index = (self.loading_index + 1) % len(self.loading_chars)
        self.loading_label.config(text=self.loading_chars[self.loading_index])
        self.root.after(500, self.animate_loading)
###

#Начало программы
if __name__ == "__main__":

    #Создание первого окна
    root_menu_1 = Tk()
    main_screen_1 = main_menu(root_menu_1)
    ###

    root_menu_1.mainloop()
